# -*- coding: utf-8 -*-
"""
Сводная таблица «Статуса комплектации» (живой запрос 2026-08-20, по образцу
заказчика «260819_Статус комплектации + график.xlsx»).

Заказчик выгружает наш плоский перечень комплектации в Excel и строит по
нему сводную: строки — Завод → Договор → Спецификация → Марка, колонки —
плановая дата поставки, свёрнутая по месяцам, в ячейке количество изделий,
срезы по статусу, договору и заводу. Разбор образца показал ровно это
(лист «Комплектация» — наша выгрузка, лист «Планируемые даты» — сводная по
ней). Здесь то же самое считается ВНУТРИ сервиса, чтобы не выгружать файл
ради каждого пересчёта, а сверху добавлены уровни **Тип** и **Подтип**,
внутри которых лежит марка (решение пользователя 2026-08-20).

Это ВТОРОЙ ВИД ТОГО ЖЕ ОТЧЁТА, а не новый отчёт (решение пользователя):
данные, права и галочка «Учитывать текущий фильтр схемы» у перечня и у
сводной общие, переключатель вида стоит в самой форме. Отсюда и общий
эндпоинт `POST /reports/completion` с параметром `view`.

Отличия от «Графика поставки» (app/report_delivery.py), который тоже
раскладывает изделия по календарю, — намеренные, иначе второй отчёт был бы
не нужен:
  • в ячейке ОДНО число (сколько изделий), а не три шкалы: сводная отвечает
    на вопрос «сколько и когда», а не «где план разошёлся с фактом»;
  • шкала ВЫБИРАЕТСЯ одна на всю таблицу (план / факт / требуемая),
    по умолчанию плановая дата — та, по которой сводную строит заказчик;
  • период не задаётся руками, а берётся из самих данных: сводную читают
    «что вообще есть», а не «что в окне».

Что общее — общее по-настоящему: календарная сетка (шаг, границы периода,
подписи колонок) живёт в `app/reports.py` и используется обоими отчётами.
Вторая копия разошлась бы с первой незаметно — числа остались бы верными,
разъехались бы колонки.

Изделия без выбранной даты не отбрасываются, а собираются в колонку «Без
даты» (в образце это «б/д»): иначе итог справа перестал бы сходиться с
числом изделий под фильтром, и переключение шкалы молча меняло бы общий
объём. Стоит она ПОСЛЕДНЕЙ, а не первой, как в образце, — так же, как в
«Графике поставки»: служебные колонки у нас справа, и две привычки на один
сервис хуже одной непривычной.

Данные, как и у остальных отчётов, считает СЕРВЕР — экран, XLSX и PDF
берут результат одной и той же функции и разойтись не могут.
"""

from typing import Optional

from app.contracts import build_document_label
from app.db import visible_elements_clause
from app.models import STATUS_LABELS_RU, STATUS_ORDER, Status
from app.reports import (
    MAX_PERIOD_COLUMNS, STEP_LABELS, STEPS, auto_step, bucket_start,
    build_period_columns, natural_key, parse_iso_date, pdf_text,
)

TITLE = "Статус комплектации"

# Вид отчёта. «Перечень» — плоский список позиций (app/report_completion.py),
# «сводная» — эта таблица. Значение приходит с клиента в `view`.
VIEW_LIST = "list"
VIEW_PIVOT = "pivot"
VIEWS = (VIEW_LIST, VIEW_PIVOT)

ROOT_LABEL = "Изделия"
TOTAL_LABEL = "Итого"

# Колонка «изделия без выбранной даты». Ключ тот же, что у одноимённой
# колонки «Графика поставки», — но модули друг о друге не знают, поэтому
# он объявлен здесь своим.
NONE_KEY = "none"

# Шкала дат: одна на всю таблицу, в отличие от «Графика поставки» с его
# тремя числами в ячейке. Порядок — порядок в переключателе на форме.
SCALES = [
    {"key": "plan", "label": "Плановая дата поставки", "column": "planned_delivery_date"},
    {"key": "fact", "label": "Фактическая дата поставки", "column": "actual_delivery_date"},
    # Требуемая — дата начала СМР: к ней изделие обязано быть на площадке.
    # Не `project_delivery_date`, та означает ЗАВЕРШЕНИЕ работ по фронту
    # (см. app/schedule_import.py) и к поставке отношения не имеет.
    {"key": "need", "label": "Требуемая дата поставки", "column": "project_smr_start_date"},
]
SCALE_KEYS = [s["key"] for s in SCALES]
# По умолчанию — плановая: именно по ней собрана сводная заказчика.
DEFAULT_SCALE = "plan"

# Уровни группировки. Порядок в списке — порядок по умолчанию в форме;
# сам ПОРЯДОК уровней задаёт пользователь и присылает его в group_by (тот же
# приём, что в «Графике поставки»).
GROUPS = [
    {"key": "counterparty", "label": "Завод"},
    {"key": "agreement", "label": "Договор"},
    {"key": "specification", "label": "Спецификация"},
    # Тип и Подтип — ОТДЕЛЬНЫМИ уровнями, а не склеенной подписью «Колонна
    # верхняя», как в «Графике поставки» (решение пользователя 2026-08-20):
    # сводную читают сверху вниз, и подтип обязан сворачиваться в тип.
    {"key": "type", "label": "Тип"},
    {"key": "subtype", "label": "Подтип"},
    # Марка — внутри подтипа: именно по ней изделия взаимозаменяемы, и
    # именно её называют, когда говорят «чего не хватает».
    {"key": "mark", "label": "Марка"},
    {"key": "status", "label": "Статус"},
    {"key": "crane", "label": "Кран"},
    {"key": "stance", "label": "Стоянка"},
]
GROUP_KEYS = [g["key"] for g in GROUPS]
GROUP_LABELS = {g["key"]: g["label"] for g in GROUPS}
DEFAULT_GROUPS = ["counterparty", "agreement", "specification", "type", "subtype", "mark"]

NO_COUNTERPARTY = "Завод не определён"
NO_AGREEMENT = "Договор не назначен"
NO_SPECIFICATION = "Спецификация не назначена"
NO_TYPE = "Тип не определён"
NO_SUBTYPE = "Без подтипа"
NO_MARK = "Без марки"
NO_STATUS = "Статус не задан"
NO_CRANE = "Кран не определён"
NO_STANCE = "Стоянка не определена"

# «Не определено» любого уровня уходит в конец своей группы, а не встаёт по
# алфавиту в середину: это не такое же значение, как остальные, а их
# отсутствие.
_NO_VALUE_LABELS = {NO_COUNTERPARTY, NO_AGREEMENT, NO_SPECIFICATION, NO_TYPE,
                    NO_SUBTYPE, NO_MARK, NO_STATUS, NO_CRANE, NO_STANCE}


def _sort_key(label: str):
    return (1 if label in _NO_VALUE_LABELS else 0, natural_key(label))


def normalize_view(view: Optional[str]) -> str:
    """Вид приходит с клиента; чужое значение — это перечень, а не ошибка:
    отчёт открывается и без параметра вовсе (старая вкладка в браузере)."""
    return VIEW_PIVOT if view == VIEW_PIVOT else VIEW_LIST


def _normalize_groups(group_by: Optional[list]) -> list:
    """Чужие ключи молча отбрасываются, дубли схлопываются, пустой список
    заменяется значением по умолчанию: параметр приходит с клиента, и
    строить дерево по непроверенному списку нельзя."""
    if not group_by:
        return list(DEFAULT_GROUPS)
    seen, out = set(), []
    for key in group_by:
        if key in GROUP_KEYS and key not in seen:
            seen.add(key)
            out.append(key)
    return out or list(DEFAULT_GROUPS)


def _normalize_scale(scale: Optional[str]) -> str:
    return scale if scale in SCALE_KEYS else DEFAULT_SCALE


def _scale(key: str) -> dict:
    return next(s for s in SCALES if s["key"] == key)


def _status_label(code: Optional[str]) -> str:
    """Русская подпись статуса — та же, что на схеме и в остальных отчётах.
    Неизвестный код показывается как есть: молча подменять его пустотой
    значило бы прятать расхождение в данных."""
    if not code:
        return NO_STATUS
    try:
        return STATUS_LABELS_RU[Status(code)]
    except ValueError:
        return code


def _status_order(code: Optional[str]) -> int:
    for i, status in enumerate(STATUS_ORDER):
        if status.value == code:
            return i
    return len(STATUS_ORDER)


def _zone_label(name: Optional[str], number, prefix: str, missing: str) -> str:
    """Зона подписью: имя, если оно есть, иначе «Кран 3» из номера. В
    плоском перечне кран и стоянка выводятся НОМЕРОМ (в образце заказчика
    так), но там это колонка таблицы, а здесь — подпись строки дерева, и
    голая «3» рядом с «Колонна» не читается."""
    if name:
        return name
    if number is not None:
        return f"{prefix} {number}"
    return missing


def _ru(value) -> str:
    d = parse_iso_date(value) if not hasattr(value, "year") else value
    return f"{d.day:02d}.{d.month:02d}.{d.year}" if d else ""


def build_completion_pivot(
    conn,
    source_file: Optional[str] = None,
    element_ids: Optional[list] = None,
    group_by: Optional[list] = None,
    step: Optional[str] = None,
    scale: Optional[str] = None,
) -> dict:
    """Сводная по тем же изделиям, что и плоский перечень.

    element_ids — сужение текущим фильтром схемы (тот же приём, что у
    остальных отчётов: критерии фильтра живут на клиенте, сервер получает
    готовый список id). Галочка у этого отчёта включена по умолчанию —
    и для перечня, и для сводной.
    """
    groups = _normalize_groups(group_by)
    scale_key = _normalize_scale(scale)
    date_column = _scale(scale_key)["column"]

    clauses, params = [visible_elements_clause("e")], []
    if source_file:
        clauses.append("e.source_file = ?")
        params.append(source_file)
    if element_ids is not None:
        if not element_ids:
            clauses.append("1=0")
        else:
            clauses.append(f"e.id IN ({','.join('?' * len(element_ids))})")
            params.extend(element_ids)
    where = f"WHERE {' AND '.join(clauses)}"

    # Один запрос со ВСЕМИ разрезами сразу, а не по запросу на выбранную
    # группировку: набор строк тот же (группировка лишь схлопывает их
    # сильнее или слабее), а так смена порядка уровней не переписывает SQL.
    rows = conn.execute(
        f"""
        SELECT cp.short_name AS cp_name,
               ag.number AS ag_number, ag.agreement_date AS ag_date,
               sp.number AS sp_number, sp.specification_date AS sp_date,
               e.element_type AS element_type, e.subtype AS subtype, e.mark AS mark,
               e.current_status AS status,
               zc.name AS crane_name, zc.number AS crane_number,
               zs.name AS stance_name, zs.number AS stance_number,
               substr(e.{date_column}, 1, 10) AS d,
               COUNT(*) AS n
        FROM elements e
        LEFT JOIN zones zc ON zc.id = e.zone_crane_id
        LEFT JOIN zones zs ON zs.id = e.zone_stance_id
        LEFT JOIN contracts c ON c.id = e.contract_id
        LEFT JOIN specifications sp ON sp.id = c.specification_id
        LEFT JOIN agreements ag ON ag.id = sp.agreement_id
        LEFT JOIN counterparties cp ON cp.id = ag.counterparty_id
        {where}
        GROUP BY cp.short_name, ag.number, ag.agreement_date, sp.number, sp.specification_date,
                 e.element_type, e.subtype, e.mark, e.current_status,
                 zc.name, zc.number, zs.name, zs.number, d
        """,
        params,
    ).fetchall()

    known = [d for d in (parse_iso_date(r["d"]) for r in rows) if d]
    start, end = (min(known), max(known)) if known else (None, None)

    if known:
        span = (end - start).days + 1
        if step not in STEPS:
            # Первое открытие: шаг подбирается по ширине периода — тот же
            # автоподбор, что в «Графике поставки».
            step = auto_step(span)
        period_columns = build_period_columns(start, end, step)
        if len(period_columns) > MAX_PERIOD_COLUMNS:
            raise ValueError(
                f"Данные дают {len(period_columns)} колонок при шаге "
                f"«{STEP_LABELS[step]}» — это больше {MAX_PERIOD_COLUMNS}. "
                f"Укрупните шаг или сузьте отбор фильтром схемы."
            )
    else:
        # Ни у одного изделия выбранной даты нет — календаря не будет, но
        # таблица должна остаться читаемой: остаётся «Без даты» и итог.
        step = step if step in STEPS else "month"
        period_columns = []

    columns = period_columns + [{
        "key": NONE_KEY, "kind": "edge", "label": "Без даты",
        "title": f"{_scale(scale_key)['label']} не заполнена",
    }]
    valid_keys = {c["key"] for c in period_columns}

    def column_of(value: Optional[str]) -> str:
        d = parse_iso_date(value)
        if d is None:
            return NONE_KEY
        key = bucket_start(d, step).isoformat()
        # Страховка: начало периода первой колонки может лежать ЛЕВЕЕ start
        # (неделя или месяц, в который попадает start) — такая дата всё
        # равно относится к первой колонке.
        return key if key in valid_keys else period_columns[0]["key"]

    def new_node(label, level, order=0):
        # values разрежённые (нулевые ячейки не хранятся): при сотне колонок
        # и тысяче строк плотная матрица была бы в основном нулями.
        return {"label": label, "level": level, "order": order,
                "values": {}, "total": 0, "children": {}}

    root = new_node(TOTAL_LABEL, -1)

    def label_and_order(key: str, r):
        """Подпись уровня и ключ его сортировки внутри родителя. Второе
        нужно только статусу: по алфавиту «В производстве» встало бы раньше
        «Запланирован», а порядок у статусов технологический."""
        if key == "counterparty":
            return r["cp_name"] or NO_COUNTERPARTY, 0
        if key == "agreement":
            return (build_document_label(r["ag_number"], r["ag_date"])
                    if r["ag_number"] else NO_AGREEMENT), 0
        if key == "specification":
            return (build_document_label(r["sp_number"], r["sp_date"])
                    if r["sp_number"] else NO_SPECIFICATION), 0
        if key == "type":
            return r["element_type"] or NO_TYPE, 0
        if key == "subtype":
            return r["subtype"] or NO_SUBTYPE, 0
        if key == "mark":
            return r["mark"] or NO_MARK, 0
        if key == "status":
            return _status_label(r["status"]), _status_order(r["status"])
        if key == "crane":
            return _zone_label(r["crane_name"], r["crane_number"], "Кран", NO_CRANE), 0
        return _zone_label(r["stance_name"], r["stance_number"], "Стоянка", NO_STANCE), 0

    def add(node, col, n):
        node["values"][col] = node["values"].get(col, 0) + n
        node["total"] += n

    with_date = 0
    total_elements = 0
    for r in rows:
        n = r["n"]
        total_elements += n
        col = column_of(r["d"])
        if col != NONE_KEY:
            with_date += n
        node = root
        add(node, col, n)
        for level, key in enumerate(groups):
            label, order = label_and_order(key, r)
            node = node["children"].setdefault(label, new_node(label, level, order))
            add(node, col, n)

    def finish(node) -> dict:
        kids = node["children"]
        children = [finish(kids[k]) for k in
                    sorted(kids, key=lambda k: (kids[k]["order"], _sort_key(k)))]
        return {"label": node["label"], "level": node["level"],
                "values": node["values"], "total": node["total"], "children": children}

    tree = finish(root)

    report = {
        "title": TITLE,
        "view": VIEW_PIVOT,
        "root_label": ROOT_LABEL,
        "total_label": TOTAL_LABEL,
        "columns": columns,
        "rows": tree["children"],
        "total": {"label": TOTAL_LABEL, "values": tree["values"], "total": tree["total"]},
        "step": step,
        "step_label": STEP_LABELS[step],
        "scale": scale_key,
        "scale_label": _scale(scale_key)["label"],
        "scales": SCALES,
        "steps": [{"key": k, "label": STEP_LABELS[k]} for k in STEPS],
        "groups": GROUPS,
        "group_by": groups,
        "group_labels": [GROUP_LABELS[k] for k in groups],
        "date_from": start.isoformat() if start else None,
        "date_to": end.isoformat() if end else None,
        # Та же честная пометка, что у «Графика поставки» и «Динамики»: дата
        # заполнена не у всех изделий, и сводная по части объёма внешне
        # неотличима от полной.
        "coverage": {"with_date": with_date, "total": total_elements},
    }
    # Подпись и предупреждение считаются ЗДЕСЬ и уходят готовым текстом: их
    # показывают экран, Excel и PDF, и три копии одной фразы разошлись бы
    # на первой же правке формулировки.
    report["subtitle"] = subtitle(report)
    report["warning"] = coverage_warning(report)
    return report


def subtitle(report: dict) -> str:
    период = (f"период {_ru(report['date_from'])} — {_ru(report['date_to'])}"
              if report["date_from"] else "дат в отобранном нет")
    return (f"Колонки: {report['scale_label'].lower()}, шаг «{report['step_label']}», "
            f"{период}. Группировка: {' → '.join(report['group_labels'])}. "
            f"В ячейке — количество изделий.")


def coverage_warning(report: dict) -> str:
    cov = report["coverage"]
    if cov["with_date"] >= cov["total"]:
        return ""
    return (f"Внимание: {report['scale_label'].lower()} заполнена у {cov['with_date']} "
            f"изделий из {cov['total']} — остальные показаны в колонке «Без даты». "
            f"Сводная по календарю неполная.")


def flatten(report: dict) -> list:
    """Дерево в плоский список строк — для Excel и PDF, где сворачивать
    нечего. Порядок тот же, что на экране у полностью раскрытого дерева."""
    out = []

    def walk(node):
        out.append(node)
        for child in node.get("children", []):
            walk(child)

    for row in report["rows"]:
        walk(row)
    return out


# ---------- выгрузка той же сводной в файлы ----------
#
# Обе функции получают УЖЕ ПОСТРОЕННЫЙ отчёт, а не строят его заново —
# иначе числа на экране, в Excel и в PDF со временем разошлись бы.

def build_completion_pivot_xlsx(report: dict) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    # Имя листа Excel — не длиннее 31 символа, иначе openpyxl обрежет его
    # сам и молча.
    ws.title = "Сводная"

    thin = Side(style="thin", color="D5D8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EEF2F7")
    edge_fill = PatternFill("solid", fgColor="F5F0E6")

    ws.append([report["title"]])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([report["subtitle"]])
    if report["warning"]:
        ws.append([report["warning"]])
        ws.cell(row=ws.max_row, column=1).font = Font(color="C0392B")

    columns = report["columns"]
    # +2, а не append([]) + max_row: пустой append в openpyxl не двигает
    # max_row, и разделительная строка молча пропала бы.
    head_row = ws.max_row + 2
    ws.cell(row=head_row, column=1, value=report["root_label"])
    for i, col in enumerate(columns):
        ws.cell(row=head_row, column=2 + i, value=col["label"])
    total_col = 2 + len(columns)
    ws.cell(row=head_row, column=total_col, value=report["total_label"])
    for i in range(1, total_col + 1):
        cell = ws.cell(row=head_row, column=i)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if i > 1 else "left",
                                   vertical="center", wrap_text=True)

    edge_cols = {2 + i for i, col in enumerate(columns) if col["kind"] == "edge"}

    # Номер строки ведём САМИ. `ws.max_row` в openpyxl — не счётчик, а
    # максимум по всем ячейкам листа, O(n) на каждое обращение: в цикле по
    # строкам это квадратичный рост (см. Docs/backlog.md 2026-08-03), а до
    # уровня марок строк здесь тысячи.
    следующая_строка = head_row + 1

    def write_row(label, values, total, indent):
        nonlocal следующая_строка
        r = следующая_строка
        следующая_строка += 1
        ws.cell(row=r, column=1, value=label)
        # Отступ вложенности — свойством ячейки, а не пробелами в тексте:
        # текст остаётся пригодным для фильтров и формул (как в «Статусах»).
        ws.cell(row=r, column=1).alignment = Alignment(indent=indent)
        for i, col in enumerate(columns):
            # Ноль не пишем вовсе: в сводной на сотню колонок нули — шум,
            # из-за которого не видно самих поставок.
            ws.cell(row=r, column=2 + i, value=values.get(col["key"]) or None)
        ws.cell(row=r, column=total_col, value=total or None)
        for i in range(1, total_col + 1):
            ws.cell(row=r, column=i).border = border
            if i in edge_cols:
                ws.cell(row=r, column=i).fill = edge_fill
        return r

    for node in flatten(report):
        r = write_row(node["label"], node["values"], node["total"], node["level"] * 2)
        # Группировка строк «+/−» — тем же приёмом, что в отчёте «Статусы»:
        # свёрнутость на экране в файл не переносится, зато дерево остаётся
        # деревом.
        if node["level"] > 0:
            ws.row_dimensions[r].outlineLevel = min(node["level"], 7)
        if node["level"] < 2:
            for i in range(1, total_col + 1):
                ws.cell(row=r, column=i).font = Font(bold=True)

    total = report["total"]
    r = write_row(total["label"], total["values"], total["total"], 0)
    for i in range(1, total_col + 1):
        ws.cell(row=r, column=i).font = Font(bold=True)
        ws.cell(row=r, column=i).fill = head_fill

    ws.column_dimensions["A"].width = 46
    for i in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 9
    ws.freeze_panes = ws.cell(row=head_row + 1, column=2)
    ws.sheet_properties.outlinePr.summaryBelow = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Сколько календарных колонок помещается на альбомном A4 читаемым кеглем.
PDF_COLUMNS_PER_BLOCK = 16


def build_completion_pivot_pdf(report: dict, subtitle_extra: str = "") -> bytes:
    """Календарь режется на блоки по PDF_COLUMNS_PER_BLOCK колонок, каждый
    блок — своя таблица с повтором колонки группировки и итога (тот же
    приём, что в «Графике поставки»). Иначе сотню колонок пришлось бы либо
    сжать до нечитаемых, либо обрезать молча — а обрезанный отчёт выглядит
    как полный."""
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.pdf_export import FONT_BOLD, FONT_REGULAR

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
        title=report["title"],
    )
    title_style = ParagraphStyle("t", fontName=FONT_BOLD, fontSize=13, leading=17)
    sub_style = ParagraphStyle("s", fontName=FONT_REGULAR, fontSize=8, leading=11,
                               textColor=colors.HexColor("#666666"))
    warn_style = ParagraphStyle("w", parent=sub_style, textColor=colors.HexColor("#C0392B"))

    story = [Paragraph(pdf_text(report["title"]), title_style)]
    if subtitle_extra:
        story.append(Paragraph(pdf_text(subtitle_extra), sub_style))
    story.append(Paragraph(pdf_text(report["subtitle"]), sub_style))
    if report["warning"]:
        story.append(Paragraph(pdf_text(report["warning"]), warn_style))
    story.append(Spacer(1, 4 * mm))

    nodes = flatten(report) + [{"label": report["total"]["label"], "level": -1,
                                "values": report["total"]["values"],
                                "total": report["total"]["total"]}]
    columns = report["columns"]

    for offset in range(0, len(columns), PDF_COLUMNS_PER_BLOCK):
        block = columns[offset:offset + PDF_COLUMNS_PER_BLOCK]
        data = [[report["root_label"]] + [c["label"] for c in block] + [report["total_label"]]]
        styles = [
            ("FONTNAME", (0, 0), (-1, -1), FONT_REGULAR),
            ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2F7")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5D8DC")),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (1, 0), (-1, -1), 1), ("RIGHTPADDING", (1, 0), (-1, -1), 1),
        ]
        for node in nodes:
            indent = " " * (max(node["level"], 0) * 3)
            data.append([indent + node["label"]]
                        + [node["values"].get(c["key"]) or "" for c in block]
                        + [node["total"] or ""])
            if node["level"] < 2:
                styles.append(("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1), FONT_BOLD))
            if node["level"] < 0:
                styles.append(("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1),
                               colors.HexColor("#EEF2F7")))
        widths = [66 * mm] + [(211 * mm) / (len(block) + 1)] * (len(block) + 1)
        table = Table(data, colWidths=widths, repeatRows=1)
        table.setStyle(TableStyle(styles))
        story.append(table)
        story.append(Spacer(1, 5 * mm))

    doc.build(story)
    return buf.getvalue()
