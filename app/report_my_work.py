"""
Отчёт «Моя работа» и отбор «изменённые за период» (живой запрос 2026-08-03).

Что это. Журнал действий (`activity_log`, app/activity.py) уже писал, кто и
что менял, но добраться до него мог только администратор сервиса и только
через форму поиска по всей системе. Прорабу нужен другой разрез: «что
СДЕЛАЛ Я за сегодня» — и возможность ткнуть в строку и увидеть это изделие
на схеме. Отсюда два потребителя одного и того же отбора:

  * отчёт «Моя работа» — список событий за период (кто, когда, что менял);
  * фильтр рабочей области «Изменения» — какие элементы схемы затронуты.

Оба берут данные ОТСЮДА, одной функцией `_where`: разъехавшись, они
показывали бы разное число изделий на один и тот же вопрос, а именно это
расхождение потом ловится только живым репортом.

Почему источник — журнал, а не `status_history`/`elements.updated_at`.
`status_history` знает автора, но её `changed_at` — это РАБОЧАЯ ДАТА
(backdating, см. CLAUDE.md): смена статуса, сделанная сегодня задним
числом, лежит там прошлой неделей. `elements.updated_at` знает момент, но
не знает автора, а правку реквизитов не отличает от пересчёта привязки к
зонам. Автор + реальный момент есть только в журнале.

Обратная сторона: журнал чистится за период (`POST /activity/cleanup`), и
глубже точки очистки отчёт ничего не покажет. Это осознанно — журнал
наблюдательный, а не бухгалтерский.

Время в журнале — UTC (app/activity._now), а пользователь выбирает период
по своему календарю. Границы приходят от клиента УЖЕ пересчитанными в UTC
(`at_from`/`at_to`, тот же приём, что у `GET /activity`), а `date_from`/
`date_to` остаются местными датами и нужны только для подписи периода в
заголовке — в том числе в Excel и PDF, которые собирает сервер.
"""

from typing import Optional

from app.element_fields import FIELD_LABELS, ru_dates_in_text
from app.models import STATUS_LABELS_RU

TITLE = "Моя работа"

# Сколько строк отдаётся на экран. Массовая смена статуса на реальном файле
# — 9422 события за одну операцию, и показывать их простынёй бессмысленно:
# сводка по действиям сверху отвечает «сколько сделано», список ниже —
# «что именно». Число отброшенных сообщается явно (см. `truncated`), молчаливая
# обрезка читалась бы как «больше ничего и не было».
SCREEN_LIMIT = 1000
FILE_LIMIT = 20000

# Человеческие названия действий. Ключи — те же строки, что кладёт
# app/activity.py; неизвестное действие показывается КАК ЕСТЬ, а не
# прячется: набор действий растёт, и забытая подпись должна выглядеть
# некрасиво, а не превращаться в молча пропавшую работу.
ACTION_TITLES = {
    "status_change": "Смена статуса",
    "status_bulk_edit": "Правка дат статусов (Excel)",
    "history_edit": "Правка записи истории",
    "history_delete": "Удаление записи истории",
    "element_edit": "Правка реквизитов",
    "element_bulk_edit": "Массовая правка реквизитов (Excel)",
    "element_comment": "Комментарий к элементу",
    "planned_date": "Плановая дата поставки",
    # Поэлементные события массовых операций (2026-08-03). Сводка операции
    # пишется отдельной записью с тем же request_id — см. new_request_id.
    "import_dxf_element": "Обновлено чертежом",
    "history_import": "История статусов из файла",
    "schedule_import": "Даты СМР из графика",
    "zone_rebind": "Пересчёт привязки к зонам",
    "import_dxf": "Загрузка чертежа",
    "import_input": "Загрузка из папки Input",
    "import_history": "Импорт истории статусов",
    "import_schedule": "Импорт графика MS Project",
    "import_contracting": "Импорт контрактации",
    "settings_import": "Импорт настроек",
    "status_history_reset": "Сброс всей истории статусов",
    "zone_edit": "Правка зоны",
    "zone_edit_undo": "Откат правки зоны",
    "counterparty_create": "Контрагент создан",
    "counterparty_update": "Контрагент изменён",
    "agreement_create": "Договор создан",
    "agreement_update": "Договор изменён",
    "specification_create": "Спецификация создана",
    "specification_update": "Спецификация изменена",
    "contract_create": "Контракт создан",
    "contract_update": "Контракт изменён",
    "default_contracts": "Контракты по умолчанию",
    "mark_prefix_set": "Префикс марки задан",
    "mark_prefix_delete": "Префикс марки удалён",
    "subtype_add": "Подтип добавлен",
    "subtype_delete": "Подтип удалён",
    "status_colors": "Цвета статусов",
    "zone_colors": "Цвета зон",
    "element_shapes": "Формы маркеров",
    "label_visibility": "Видимость подписей",
    "label_dates_visibility": "Видимость дат в подписях",
    "project_card": "Карточка объекта",
    "late_threshold": "Порог опоздания поставки",
    "report_notes": "События, задачи, вопросы",
    "report_notes_delete": "Удалена редакция событий/задач",
    "user_create": "Пользователь создан",
    "user_update": "Пользователь изменён",
    "user_password": "Пароль изменён",
    "user_label_color": "Цвет подписей марок",
    "user_ui_theme": "Оформление интерфейса",
    "last_object": "Выбран объект",
    "attachment_add": "Вложение добавлено",
    "attachment_delete": "Вложение удалено",
    "project_create": "Проект создан",
    "project_rename": "Проект переименован",
    "project_delete": "Проект удалён",
    "object_create": "Объект создан",
    "object_rename": "Объект переименован",
    "backup_create": "Резервная копия создана",
    "backup_restore": "Восстановление из копии",
    "backup_delete": "Резервная копия удалена",
    "safety_backup": "Страховочная копия",
    "activity_cleanup": "Очистка журнала",
    "access_replace": "Права пользователя изменены",
    "user_auth_method": "Способ входа изменён",
    "ldap_settings": "Настройка доменной авторизации",
}

# Действия, которые меняют РЕКВИЗИТЫ или ИСТОРИЮ СТАТУСА конкретного
# изделия. Это и есть предмет фильтра «Изменения» в рабочей области —
# ровно то, о чём просил пользователь. Список отдельный от подписей выше:
# подпись есть и у «Резервная копия создана», но к изделию она не
# относится.
ELEMENT_CHANGE_ACTIONS = (
    "status_change",
    "status_bulk_edit",
    "history_edit",
    "history_delete",
    "element_edit",
    "element_bulk_edit",
    "element_comment",
    "planned_date",
    # Массовые операции — поэлементными событиями (2026-08-03). Без них
    # переимпорт чертежа, восстановление истории из файла, импорт графика и
    # пересчёт привязки к зонам меняли данные тысяч изделий, а фильтр
    # «Изменения» их не показывал: событие было ОДНО, на всю операцию.
    "import_dxf_element",
    "history_import",
    "schedule_import",
    "zone_rebind",
)

# Что в «Моей работе» НЕ показывается. Правило — денилист, а не вайтлист:
# новое действие должно попадать в отчёт само, иначе очередная функция
# молча не покажет сделанную работу. Здесь только то, что изменением не
# является: вход/выход, выгрузки, служебные события самого журнала.
NON_CHANGE_ACTIONS = {
    "login", "login_failed", "login_blocked", "login_domain_failed", "logout",
    "element_bulk_export",
    "server_start", "schema_migration", "client_batch_truncated", "log_overflow",
}


def action_title(action: str) -> str:
    return ACTION_TITLES.get(action, action)


def value_text(raw: Optional[str]) -> str:
    """Значение «было»/«стало» в человеческом виде.

    Три формата уживаются в одной колонке журнала, и различать их приходится
    по содержимому: код статуса (`delivered`), перечень полей
    (`mark: 1КС1.1; floor: 2` — так пишут element_edit и массовая правка) и
    произвольный текст. Разбор здесь, на сервере, а не на клиенте: те же
    строки уходят в Excel и PDF, а третья копия правил разошлась бы с
    первыми двумя.
    """
    if raw is None or raw == "":
        return ""
    # Даты внутри значения — в русском виде (живой запрос 2026-08-03): в
    # журнале они лежат как 'ГГГГ-ММ-ДД' (так их сравнивают), а читает эту
    # строку человек — и в карточке изделия, и в Excel, и в PDF.
    text = ru_dates_in_text(raw)
    if text in STATUS_LABELS_RU:
        return STATUS_LABELS_RU[text]
    if ": " not in text:
        return text
    части = []
    for кусок in text.split("; "):
        поле, _, значение = кусок.partition(": ")
        подпись = FIELD_LABELS.get(поле.strip())
        # Подписи нет — значит это не «поле: значение», а обычный текст с
        # двоеточием (комментарий, сообщение импорта). Возвращаем как есть,
        # а не разбираем по кускам.
        if подпись is None:
            return text
        части.append(f"{подпись}: {значение if значение != 'None' else '—'}")
    return "; ".join(части)


def _bounds(at_from: Optional[str], at_to: Optional[str],
            date_from: Optional[str], date_to: Optional[str]) -> tuple:
    """Границы отбора в том виде, в каком они сравниваются с `activity_log.at`.

    Обычный путь — готовые UTC-границы от клиента (он один знает часовой
    пояс пользователя). Прямой вызов эндпоинта (curl, внешний скрипт) может
    прислать только даты — тогда трактуем их целыми сутками UTC, как это
    делает `GET /activity`.
    """
    низ = at_from or (f"{date_from} 00:00:00.000" if date_from else None)
    верх = at_to or (f"{date_to} 23:59:59.999" if date_to else None)
    return низ, верх


def _where(at_from: Optional[str], at_to: Optional[str], user_ids: Optional[list],
           object_ids: Optional[set], actions: Optional[tuple]) -> tuple:
    """Условие отбора событий — ОДНО на отчёт и на фильтр рабочей области.

    object_ids=None означает «без ограничения по объектам» (свои события,
    системный администратор), а НЕ «ни одного»: пустое множество здесь
    законно и означает именно «ничего не доступно» — та же развилка, что у
    `accessible_object_ids` (app/access.py).

    user_ids=None — «любой пользователь»; вызывающий обязан сам решить, что
    это законно (см. `_my_work_scope`, app/main.py).
    """
    clauses = ["a.source = 'server'"]
    params: list = []
    if at_from:
        clauses.append("a.at >= ?")
        params.append(at_from)
    if at_to:
        clauses.append("a.at <= ?")
        params.append(at_to)
    if user_ids is not None:
        if not user_ids:
            return "WHERE 0", []
        clauses.append(f"a.user_id IN ({','.join('?' * len(user_ids))})")
        params.extend(user_ids)
    if actions is not None:
        clauses.append(f"a.action IN ({','.join('?' * len(actions))})")
        params.extend(actions)
    else:
        clauses.append(f"a.action NOT IN ({','.join('?' * len(NON_CHANGE_ACTIONS))})")
        params.extend(sorted(NON_CHANGE_ACTIONS))
    if object_ids is not None:
        # Ограничение областью — это ограничение ЭЛЕМЕНТАМИ доступных
        # объектов: событие без элемента (создан контракт, переименован
        # проект) объекта не имеет, и показать его «в рамках объекта» не
        # получится — оно просто не попадает в выборку.
        if not object_ids:
            return "WHERE 0", []
        clauses.append(f"e.object_id IN ({','.join('?' * len(object_ids))})")
        params.extend(sorted(object_ids))
    return "WHERE " + " AND ".join(clauses), params


_FROM = """
    FROM activity_log a
    LEFT JOIN elements e ON a.entity_type = 'element' AND e.id = a.entity_id
"""


def changed_element_ids(conn, *, at_from: Optional[str], at_to: Optional[str],
                        user_ids: Optional[list], object_id: int) -> list:
    """Элементы объекта, чьи реквизиты или история статуса менялись за период.

    Возвращает id ЖИВЫХ элементов (`is_current = 1`): фильтр рабочей области
    отбирает то, что на схеме нарисовано, а элемент, исчезнувший из чертежа
    при переимпорте, там не показывается вовсе.
    """
    where, params = _where(at_from, at_to, user_ids, {object_id}, ELEMENT_CHANGE_ACTIONS)
    rows = conn.execute(
        f"SELECT DISTINCT e.id AS id {_FROM} {where} AND e.is_current = 1", params
    ).fetchall()
    return [r["id"] for r in rows]


def build_my_work_report(conn, *, at_from: Optional[str], at_to: Optional[str],
                         date_from: Optional[str], date_to: Optional[str],
                         user_ids: Optional[list], object_ids: Optional[set],
                         users: Optional[list] = None, limit: int = SCREEN_LIMIT) -> dict:
    """Список изменений за период с разворотом по действиям.

    Строка события несёт СНИМОК изделия на момент действия (тип/подтип/марка
    — так их кладёт журнал: элемент с тех пор мог измениться) и, отдельно,
    ссылку на живой элемент — по ней интерфейс показывает изделие на схеме.
    Одно другое не заменяет: снимок объясняет, ЧТО меняли, ссылка ведёт
    туда, где оно сейчас.
    """
    низ, верх = _bounds(at_from, at_to, date_from, date_to)
    where, params = _where(низ, верх, user_ids, object_ids, None)

    total = conn.execute(f"SELECT COUNT(*) AS n {_FROM} {where}", params).fetchone()["n"]
    by_action = [
        {"action": r["action"], "title": action_title(r["action"]), "count": r["n"]}
        for r in conn.execute(
            f"SELECT a.action AS action, COUNT(*) AS n {_FROM} {where} "
            f"GROUP BY a.action ORDER BY n DESC, a.action", params)
    ]
    rows = conn.execute(
        f"""
        SELECT a.id AS id, a.at AS at, a.user_id AS user_id, a.user_name AS user_name,
               a.action AS action, a.entity_type AS entity_type, a.entity_id AS entity_id,
               a.element_type AS element_type, a.subtype AS subtype, a.mark AS mark,
               a.old_value AS old_value, a.new_value AS new_value,
               e.id AS el_id, e.object_id AS el_object_id, e.is_current AS el_is_current,
               e.address AS el_address, e.floor AS el_floor,
               e.element_type AS el_type, e.subtype AS el_subtype, e.mark AS el_mark
        {_FROM} {where}
        ORDER BY a.at DESC, a.id DESC LIMIT ?
        """,
        params + [limit],
    ).fetchall()

    объекты = {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM objects")}

    out = []
    for r in rows:
        # Снимок из журнала первичен, поля элемента — запасной вариант: у
        # старых событий (и у всего, что логировалось до появления снимка)
        # тип/марка пусты, а показать изделие всё равно надо.
        тип = r["element_type"] or r["el_type"]
        подтип = r["subtype"] or r["el_subtype"]
        марка = r["mark"] or r["el_mark"]
        элемент = None
        if r["el_id"] is not None:
            элемент = {
                "id": r["el_id"],
                "object_id": r["el_object_id"],
                "object_name": объекты.get(r["el_object_id"]),
                "is_current": bool(r["el_is_current"]),
                "address": r["el_address"],
                "floor": r["el_floor"],
            }
        out.append({
            "id": r["id"],
            "at": r["at"],
            "user_id": r["user_id"],
            "user_name": r["user_name"],
            "action": r["action"],
            "action_title": action_title(r["action"]),
            "entity_type": r["entity_type"],
            "entity_id": r["entity_id"],
            "item": " / ".join(p for p in (тип, подтип, марка) if p),
            "old_text": value_text(r["old_value"]),
            "new_text": value_text(r["new_value"]),
            "element": элемент,
        })

    return {
        "title": TITLE,
        "date_from": date_from,
        "date_to": date_to,
        "at_from": низ,
        "at_to": верх,
        "users": users or [],
        "total": total,
        "shown": len(out),
        "truncated": total > len(out),
        "by_action": by_action,
        "rows": out,
    }


def period_subtitle(report: dict) -> str:
    """Подпись периода одной строкой — общая для экрана, Excel и PDF."""
    def ru(iso):
        if not iso:
            return "…"
        г, м, д = iso.split("-")
        return f"{д}.{м}.{г}"

    период = (f"за {ru(report['date_from'])}"
              if report["date_from"] and report["date_from"] == report["date_to"]
              else f"с {ru(report['date_from'])} по {ru(report['date_to'])}")
    кто = ", ".join(u["display_name"] for u in report["users"]) or "все пользователи"
    return f"{период} · {кто}"


def _time_local_text(at: str, tz_offset_minutes: int) -> str:
    """UTC-метка события → местное время строкой.

    Смещение приходит от клиента: сервер стоит в UTC, и «местное» для него
    ничего не значит. Считается арифметикой над строкой (та же, что делает
    браузер в activityTimeLocal), без часовых поясов Python — из журнала
    приходит фиксированный формат.
    """
    from datetime import datetime, timedelta

    try:
        когда = datetime.strptime(at[:19], "%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        return at or ""
    return (когда - timedelta(minutes=tz_offset_minutes)).strftime("%d.%m.%Y %H:%M:%S")


_COLUMNS = ["Время", "Пользователь", "Действие", "Изделие / объект", "Было", "Стало"]


def _file_rows(report: dict, tz_offset_minutes: int) -> list:
    out = []
    for r in report["rows"]:
        предмет = r["item"]
        if r["element"] and r["element"]["object_name"]:
            предмет = f"{предмет} ({r['element']['object_name']})" if предмет else r["element"]["object_name"]
        out.append([
            _time_local_text(r["at"], tz_offset_minutes),
            r["user_name"] or "",
            r["action_title"],
            предмет,
            r["old_text"],
            r["new_text"],
        ])
    return out


def build_my_work_xlsx(report: dict, tz_offset_minutes: int = 0) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Моя работа"

    thin = Side(style="thin", color="D5D8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EEF2F7")

    ws.append([report["title"]])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([period_subtitle(report)])
    ws.append([])

    # Сводка по действиям — тем же листом, над таблицей: в отчёте на тысячи
    # строк она и есть ответ на вопрос «сколько сделано», а отдельный лист
    # ради двух колонок только прятал бы её.
    ws.append(["Действие", "Событий"])
    строка_свода = ws.max_row
    for i in (1, 2):
        c = ws.cell(row=строка_свода, column=i)
        c.font = Font(bold=True)
        c.fill = head_fill
        c.border = border
    # Номер строки ведём САМИ. `ws.max_row` в openpyxl — не счётчик, а
    # максимум по всем ячейкам листа, O(n) на каждое обращение: в цикле по
    # строкам это квадратичный рост. На выгрузке реквизитов такой же цикл
    # стоил 88 секунд вместо 1,7 (см. Docs/backlog.md 2026-08-03), а сюда
    # уходит до 20 000 событий по шесть колонок.
    строка = строка_свода
    for item in report["by_action"]:
        ws.append([item["title"], item["count"]])
        строка += 1
        for i in (1, 2):
            ws.cell(row=строка, column=i).border = border
    ws.append([])

    ws.append(_COLUMNS)
    header_row = ws.max_row
    for i in range(1, len(_COLUMNS) + 1):
        cell = ws.cell(row=header_row, column=i)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", wrap_text=True)

    строка = header_row
    for значения in _file_rows(report, tz_offset_minutes):
        ws.append(значения)
        строка += 1
        for i in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=строка, column=i)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if report["truncated"]:
        ws.append([f"Показаны первые {report['shown']} событий из {report['total']} — "
                   f"сузьте период, чтобы увидеть остальные"])

    for буква, ширина in zip("ABCDEF", (20, 26, 30, 40, 32, 32)):
        ws.column_dimensions[буква].width = ширина
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_my_work_pdf(report: dict, tz_offset_minutes: int = 0) -> bytes:
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.pdf_export import FONT_BOLD, FONT_REGULAR
    from app.reports import pdf_text

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title=report["title"],
    )
    title_style = ParagraphStyle("t", fontName=FONT_BOLD, fontSize=14, leading=18)
    sub_style = ParagraphStyle("s", fontName=FONT_REGULAR, fontSize=9, leading=12,
                               textColor=colors.HexColor("#666666"))
    # Ячейка таблицы — Paragraph, а не строка: значения «было/стало» бывают
    # длиной в несколько полей, и без переноса по словам они уезжают за край
    # страницы (у reportlab строка в ячейке не переносится вовсе).
    cell_style = ParagraphStyle("c", fontName=FONT_REGULAR, fontSize=7, leading=9)

    story = [Paragraph(pdf_text(report["title"]), title_style),
             Paragraph(pdf_text(period_subtitle(report)), sub_style),
             Spacer(1, 5 * mm)]

    сводка = ", ".join(f"{i['title']} — {i['count']}" for i in report["by_action"])
    if сводка:
        story.append(Paragraph(pdf_text(f"Всего событий: {report['total']}. {сводка}"), sub_style))
        story.append(Spacer(1, 4 * mm))

    data = [[Paragraph(f"<b>{pdf_text(c)}</b>", cell_style) for c in _COLUMNS]]
    for значения in _file_rows(report, tz_offset_minutes):
        data.append([Paragraph(pdf_text(v), cell_style) for v in значения])

    table = Table(data, colWidths=[26 * mm, 38 * mm, 40 * mm, 55 * mm, 55 * mm, 55 * mm],
                  repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5D8DC")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2F7")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)

    if report["truncated"]:
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(
            pdf_text(f"Показаны первые {report['shown']} событий из {report['total']} — "
                     f"сузьте период, чтобы увидеть остальные"), sub_style))

    doc.build(story)
    return buf.getvalue()
