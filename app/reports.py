"""
Отчёты (живой запрос 2026-07-29). Пока один — «Статусы»: сводка по
статусам монтажа с иерархией Захватка → Этаж → Тип изделия.

Данные считаются ЗДЕСЬ, на сервере, а не на клиенте, хотя элементы уже
лежат в браузере. Причина: тот же отчёт выгружается в XLSX и PDF, и если
считать его в двух местах, числа в файле и на экране однажды разойдутся.
Экран, Excel и PDF берут результат одной и той же функции.

Структура повторяет сводную таблицу, которую заказчик собирал вручную:
строки — три уровня вложенности, колонки — статусы, ФАКТИЧЕСКИ
встретившиеся в данных (кроме «Запланирован»), затем «Остаток» и «В
проекте». «Остаток» = «В проекте» минус сумма показанных статусов, то есть
сколько изделий ещё не сдвинулось с планового состояния.
"""

from typing import Optional

from app.db import visible_elements_clause
from app.models import STATUS_LABELS_RU, STATUS_ORDER, Status

TITLE = "Статус монтажа"
ROOT_LABEL = "ЖБ изделия"
TOTAL_LABEL = "В проекте"

# ---------- признак «отчёт в разработке» (2026-08-05, живой запрос) ----------
#
# Отчёт бывает выкачен раньше, чем доведён: числа в нём уже показываются, но
# верить им нельзя. Признак живёт ТОЛЬКО В КОДЕ (снимается правкой
# IN_DEVELOPMENT в модуле отчёта), а не настройкой в интерфейсе — это
# состояние разработки, а не решение администратора.
#
# Доступ признак НЕ ограничивает (решение пользователя 2026-08-05): отчёт
# виден всем, кому был виден, — просто с честной пометкой. Ограничение
# доступа и «в разработке» — разные вопросы, и связывать их значило бы
# отбирать у людей отчёт, которым они уже пользуются.
#
# Пометка обязана дойти до ВСЕХ трёх мест сразу: в названии (меню, вкладка),
# на форме отчёта и в выгруженном файле. Файл уходит из системы и живёт
# своей жизнью — без пометки внутри него предупреждение на экране
# бесполезно.
IN_DEVELOPMENT_SUFFIX = "(в разработке)"
IN_DEVELOPMENT_NOTE = (
    "ОТЧЁТ В РАЗРАБОТКЕ: данные могут быть неполными или неверными, "
    "использовать для принятия решений нельзя"
)
# Красный, тот же и в Excel (RGB без решётки), и в PDF, и на экране.
IN_DEVELOPMENT_COLOR = "C00000"


def in_development_title(title: str, in_development: bool) -> str:
    """Название отчёта с пометкой. Одно место на весь проект — иначе пометка
    доехала бы до вкладки и потерялась в имени файла."""
    return f"{title} {IN_DEVELOPMENT_SUFFIX}" if in_development else title

# «Запланирован» намеренно НЕ становится колонкой: он и есть «Остаток».
# Отдельная колонка дублировала бы её, а сумма по строке перестала бы
# сходиться с «В проекте».
BASE_STATUS = Status.PLANNED.value

NO_ZAKHVATKA = "Захватка не определена"
NO_FLOOR = "Этаж не определён"


def pdf_text(value) -> str:
    """Пользовательский текст, уходящий в reportlab.Paragraph.

    `Paragraph` разбирает мини-разметку (`<b>`, `<font>`, `<br/>`), поэтому
    принимает её и из данных: примечания карточки объекта, заголовок,
    открытые вопросы — свободный текст, который правят через интерфейс.
    Одного символа `<` или `&` достаточно, чтобы разбор упал и отчёт
    выдал 500 вместо PDF (аудит безопасности 2026-08-03); намеренная
    разметка искажает вёрстку документа.

    Экранируем на ВХОДЕ в Paragraph, а не при сохранении: в базе должен
    лежать текст, который человек ввёл, — его же показывают экран и Excel,
    где никакой разметки нет. Подписи, рисуемые прямо на холсте
    (`drawCentredString` у вех графика), разметку не разбирают вовсе и
    экранирования не требуют."""
    from xml.sax.saxutils import escape
    return escape("" if value is None else str(value))


def natural_key(text: str):
    """«По-человечески»: Захватка 2 раньше Захватки 10. Общая для всех
    отчётов (используется и «Графиком поставки», app/report_delivery.py) —
    вторая копия однажды разошлась бы с этой."""
    import re
    return [int(p) if p.isdigit() else p.lower() for p in re.split(r"(\d+)", text)]


def _floor_label(floor: Optional[int]) -> str:
    return f"{floor} этаж" if floor is not None else NO_FLOOR


def _item_label(element_type: str, subtype: Optional[str]) -> str:
    """«Колонна нижняя», «Плита перекрытия на отм. +15.000», «Ригель
    периметральный» — тип и подтип через пробел, как в исходной сводной
    таблице заказчика. Без подтипа — только тип."""
    return f"{element_type} {subtype}".strip() if subtype else element_type


def build_status_report(conn, source_file: Optional[str], element_ids: Optional[list] = None) -> dict:
    """element_ids — необязательное сужение до конкретных элементов (тот же
    приём, что у XLS-экспорта: фильтры живут на клиенте, сервер получает
    готовый список id, а не пересчитывает их сам)."""
    clauses, params = [visible_elements_clause("e")], []
    if source_file:
        clauses.append("e.source_file = ?")
        params.append(source_file)
    if element_ids is not None:
        if not element_ids:
            clauses.append("1=0")
        else:
            clauses.append(f"e.id IN ({','.join('?' * len(element_ids))})")
            params.extend(element_ids)
    where = f"WHERE {' AND '.join(clauses)}"

    rows = conn.execute(
        f"""
        SELECT z.name AS zakhvatka, e.floor AS floor, e.element_type AS element_type,
               e.subtype AS subtype, e.current_status AS status, COUNT(*) AS n
        FROM elements e
        LEFT JOIN zones z ON z.id = e.zone_zakhvatka_id
        {where}
        GROUP BY z.name, e.floor, e.element_type, e.subtype, e.current_status
        """,
        params,
    ).fetchall()

    # Какие статусы реально встретились — только они становятся колонками.
    # Показывать все семь смысла нет: пустые колонки занимают ширину и
    # мешают читать (в исходной сводной их тоже не было).
    present = {r["status"] for r in rows if r["status"] != BASE_STATUS}
    statuses = [s.value for s in STATUS_ORDER if s.value in present]
    columns = [{"key": s, "label": STATUS_LABELS_RU[Status(s)]} for s in statuses]

    def empty_values() -> dict:
        return {s: 0 for s in statuses} | {"total": 0}

    tree: dict = {}
    for r in rows:
        zak = r["zakhvatka"] or NO_ZAKHVATKA
        flo = _floor_label(r["floor"])
        item = _item_label(r["element_type"], r["subtype"])
        node = tree.setdefault(zak, {"values": empty_values(), "children": {}})
        floor_node = node["children"].setdefault(flo, {"values": empty_values(), "children": {}})
        item_node = floor_node["children"].setdefault(item, {"values": empty_values()})
        for target in (node, floor_node, item_node):
            target["values"]["total"] += r["n"]
            if r["status"] in target["values"]:
                target["values"][r["status"]] += r["n"]

    def finish(values: dict) -> dict:
        # «Остаток» считается ВЫЧИТАНИЕМ, а не подсчётом «Запланирован»:
        # так сумма по строке всегда сходится с «В проекте», даже если в
        # данных появится статус, которого нет среди колонок.
        used = sum(values[s] for s in statuses)
        return values | {"remainder": values["total"] - used}

    # Сортировка: захватки и этажи — «по-человечески» (Захватка 2 раньше
    # Захватки 10, см. natural_key выше), изделия — по алфавиту, как в
    # исходной сводной.
    out_rows = []
    for zak in sorted(tree, key=natural_key):
        znode = tree[zak]
        zrow = {"label": zak, "level": 0, "values": finish(znode["values"]), "children": []}
        for flo in sorted(znode["children"], key=natural_key):
            fnode = znode["children"][flo]
            frow = {"label": flo, "level": 1, "values": finish(fnode["values"]), "children": []}
            for item in sorted(fnode["children"], key=natural_key):
                frow["children"].append({
                    "label": item, "level": 2,
                    "values": finish(fnode["children"][item]["values"]), "children": [],
                })
            zrow["children"].append(frow)
        out_rows.append(zrow)

    grand = empty_values()
    for r in rows:
        grand["total"] += r["n"]
        if r["status"] in grand:
            grand[r["status"]] += r["n"]

    return {
        "title": TITLE,
        "root_label": ROOT_LABEL,
        "columns": columns + [{"key": "remainder", "label": "Остаток"},
                              {"key": "total", "label": TOTAL_LABEL}],
        "rows": out_rows,
        "total": {"label": TOTAL_LABEL, "values": finish(grand)},
    }


def flatten(report: dict) -> list:
    """Плоский список строк с уровнем вложенности — для XLSX и PDF, где
    дерева нет, а есть отступ и группировка."""
    out = []

    def walk(node):
        out.append(node)
        for child in node.get("children", []):
            walk(child)

    for row in report["rows"]:
        walk(row)
    return out


# ---------- выгрузка того же отчёта в файлы ----------
#
# Обе функции получают УЖЕ ПОСТРОЕННЫЙ отчёт (build_status_report), а не
# строят его заново — иначе числа на экране, в Excel и в PDF со временем
# разошлись бы.

def build_status_report_xlsx(report: dict) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Статус монтажа"

    thin = Side(style="thin", color="D5D8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EEF2F7")

    ws.append([report["title"]])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])

    header = [report["root_label"]] + [c["label"] for c in report["columns"]]
    ws.append(header)
    header_row = ws.max_row
    for i in range(1, len(header) + 1):
        cell = ws.cell(row=header_row, column=i)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if i > 1 else "left", wrap_text=True)

    # Номер строки ведём САМИ. `ws.max_row` в openpyxl — не счётчик, а
    # максимум по всем ячейкам листа, O(n) на каждое обращение: в цикле по
    # строкам это квадратичный рост. На выгрузке реквизитов такой же цикл
    # стоил 88 секунд вместо 1,7 (см. Docs/backlog.md 2026-08-03); здесь
    # строк пока сотни, но иерархия отчёта задаётся данными и растёт.
    row = header_row
    for node in flatten(report):
        # Отступ вложенности — через indent у ячейки, а не пробелами в
        # тексте: Excel умеет отступ сам, и текст остаётся пригодным для
        # фильтров и формул.
        ws.append([node["label"]] + [node["values"].get(c["key"], 0) or None for c in report["columns"]])
        row += 1
        ws.cell(row=row, column=1).alignment = Alignment(indent=node["level"] * 2)
        if node["level"] < 2:
            for i in range(1, len(header) + 1):
                ws.cell(row=row, column=i).font = Font(bold=True)
        # Группировка строк — Excel рисует те же "+/−", что и в исходной
        # сводной таблице заказчика.
        if node["level"] > 0:
            ws.row_dimensions[row].outlineLevel = node["level"]
        for i in range(1, len(header) + 1):
            ws.cell(row=row, column=i).border = border

    total = report["total"]
    ws.append([total["label"]] + [total["values"].get(c["key"], 0) or None for c in report["columns"]])
    row += 1
    for i in range(1, len(header) + 1):
        cell = ws.cell(row=row, column=i)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.border = border

    ws.column_dimensions["A"].width = 44
    for i in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    # Свёрнутый вид по умолчанию — как на исходном скриншоте, где раскрыта
    # только первая захватка. Разворачивает пользователь сам.
    ws.sheet_properties.outlinePr.summaryBelow = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_status_report_pdf(report: dict, subtitle: str = "") -> bytes:
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.pdf_export import FONT_BOLD, FONT_REGULAR

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title=report["title"],
    )
    title_style = ParagraphStyle("t", fontName=FONT_BOLD, fontSize=14, leading=18)
    sub_style = ParagraphStyle("s", fontName=FONT_REGULAR, fontSize=9, leading=12,
                               textColor=colors.HexColor("#666666"))

    story = [Paragraph(pdf_text(report["title"]), title_style)]
    if subtitle:
        story.append(Paragraph(pdf_text(subtitle), sub_style))
    story.append(Spacer(1, 6 * mm))

    data = [[report["root_label"]] + [c["label"] for c in report["columns"]]]
    styles = [
        ("FONTNAME", (0, 0), (-1, -1), FONT_REGULAR),
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2F7")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5D8DC")),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for node in flatten(report):
        # Отступ пробелами — в PDF-таблице нет понятия уровня, а колонка
        # одна; неразрывные пробелы не дают reportlab схлопнуть их.
        data.append([" " * (node["level"] * 4) + node["label"]]
                    + [node["values"].get(c["key"], 0) or "" for c in report["columns"]])
        if node["level"] < 2:
            styles.append(("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1), FONT_BOLD))

    total = report["total"]
    data.append([total["label"]] + [total["values"].get(c["key"], 0) or "" for c in report["columns"]])
    last = len(data) - 1
    styles += [("FONTNAME", (0, last), (-1, last), FONT_BOLD),
               ("BACKGROUND", (0, last), (-1, last), colors.HexColor("#EEF2F7"))]

    widths = [110 * mm] + [(150 * mm) / max(len(report["columns"]), 1)] * len(report["columns"])
    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle(styles))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


# ================ Отчёт «Отчёт о динамике поставки и монтажа» ================
#
# Повторяет ежедневный отчёт, который заказчик собирал вручную (см.
# Docs/backlog.md): шапка с карточкой объекта, три накопительные кривые,
# две таблицы «план / факт / отклонение» и текстовые блоки.
#
# Что откуда берётся:
#   План СМР      — elements.project_smr_start_date (импорт графика MS Project)
#   План поставки — elements.planned_delivery_date
#   Факт монтажа  — status_history, переход в «Смонтирован»
#   Факт поставки — status_history, переход в «Доставлен»
#
# Кривые строятся НАКОПИТЕЛЬНО по неделям: у заказчика на графике недельная
# сетка (27 июл, 03 авг, 10 авг …), и дневная детализация на полугодовом
# горизонте превратила бы линию в шум.

DYNAMICS_TITLE = "Отчёт о динамике поставки и монтажа"

# Подзаголовок отчёта неизменен и в карточке объекта не хранится (живой
# запрос 2026-07-29): это часть формы самого отчёта, а не свойство объекта.
DYNAMICS_SUBTITLE = "о статусе производства работ и поставке ЖБ изделий на объекте строительства:"


def _week_start(date_str: str) -> str:
    """Понедельник недели, в которую попадает дата. Точка кривой — неделя,
    а не день (см. комментарий выше)."""
    from datetime import date, timedelta

    d = date.fromisoformat(date_str[:10])
    return (d - timedelta(days=d.weekday())).isoformat()


def _week_start_or_none(value) -> Optional[str]:
    """То же, но для дат, пришедших из КАРТОЧКИ ОБЪЕКТА, — контрольных
    сроков и вех. Там лежит свободно правимый JSON настроек, и одно
    непохожее на дату значение роняло весь отчёт пятисоткой
    (`date.fromisoformat('скрыто')`, поймано на обезличенной копии
    2026-08-03). Веха с испорченной датой просто не участвует в сетке
    недель: отчёт про поставку и монтаж, а не про целостность карточки."""
    if not value:
        return None
    try:
        return _week_start(str(value))
    except (ValueError, TypeError):
        return None


def _cumulative(pairs: list, weeks: list) -> list:
    """Накопительный итог по заранее заданной сетке недель. Сетка общая для
    всех кривых — иначе линии на графике стояли бы на разных абсциссах и
    сравнивать их было бы нельзя."""
    by_week: dict = {}
    for week, n in pairs:
        by_week[week] = by_week.get(week, 0) + n
    out, running = [], 0
    for w in weeks:
        running += by_week.get(w, 0)
        out.append(running)
    return out


# Режимы графика «Динамики» (2026-08-14, живой запрос: «поставку и монтаж
# надо разделять между собой, отдельно смотреть поставку, отдельно монтаж»).
# Ряд plan_delivery считался и раньше, но на графике не рисовался вовсе —
# кривая плана поставки была единственной недостающей из четырёх.
# Порядок внутри режима — порядок отрисовки и легенды: сначала план, за ним
# факт той же пары.
DYN_MODE_SERIES = {
    "montage": ["plan_smr", "fact_montage", "forecast_montage"],
    "delivery": ["plan_delivery", "fact_delivery", "forecast_delivery"],
    "both": ["plan_smr", "fact_montage", "forecast_montage",
             "plan_delivery", "fact_delivery", "forecast_delivery"],
}
DYN_MODE_LABELS = {"montage": "Монтаж", "delivery": "Поставка", "both": "Поставка и монтаж"}


def build_dynamics_report(conn, source_file: Optional[str], report_date: Optional[str] = None,
                          element_ids: Optional[list] = None,
                          object_id: Optional[int] = None,
                          week_from: Optional[str] = None,
                          week_to: Optional[str] = None,
                          mode: Optional[str] = None) -> dict:
    from datetime import date

    from app.settings import (
        NOTE_FIELDS, PROJECT_CARD_DEFAULT, get_notes_for_date, get_project_card,
    )

    today = report_date or date.today().isoformat()

    clauses, params = [visible_elements_clause("e")], []
    if source_file:
        clauses.append("e.source_file = ?")
        params.append(source_file)
    if element_ids is not None:
        if not element_ids:
            clauses.append("1=0")
        else:
            clauses.append(f"e.id IN ({','.join('?' * len(element_ids))})")
            params.extend(element_ids)
    where = f"WHERE {' AND '.join(clauses)}"

    режим = mode if mode in DYN_MODE_SERIES else "both"

    total = conn.execute(f"SELECT COUNT(*) AS n FROM elements e {where}", params).fetchone()["n"]

    plan_smr = conn.execute(
        f"SELECT project_smr_start_date AS d, COUNT(*) AS n FROM elements e {where} "
        f"{'AND' if where else 'WHERE'} project_smr_start_date IS NOT NULL GROUP BY d", params).fetchall()
    plan_delivery = conn.execute(
        f"SELECT planned_delivery_date AS d, COUNT(*) AS n FROM elements e {where} "
        f"{'AND' if where else 'WHERE'} planned_delivery_date IS NOT NULL GROUP BY d", params).fetchall()

    # Факт — по ПЕРВОМУ переходу элемента в статус: повторные записи истории
    # (откат и возврат) не должны считаться вторым смонтированным изделием.
    def fact_rows(status: str):
        return conn.execute(
            f"""
            SELECT d, COUNT(*) AS n FROM (
                SELECT date(MIN(sh.changed_at)) AS d
                FROM status_history sh JOIN elements e ON e.id = sh.element_id
                {where} {'AND' if where else 'WHERE'} sh.status = ?
                GROUP BY sh.element_id
            ) GROUP BY d
            """, params + [status]).fetchall()

    fact_montage = fact_rows("installed")
    fact_delivery = fact_rows("delivered")

    # Прогноз — последняя актуализация графика объекта (2026-08-14, живой
    # запрос: «от линии факта должна идти штрихпунктирная линия прогноза»).
    # Двумя рядами, потому что версия несёт две даты: завершение монтажа —
    # это прогноз МОНТАЖА, а начало СМР — срок, к которому изделие обязано
    # быть на площадке, то есть прогноз ПОСТАВКИ (тот же смысл, что у
    # project_smr_start_date в критерии опоздания поставки).
    from app.schedule_versions import cumulative_forecast
    прогноз = cumulative_forecast(conn, object_id) if object_id else {"start": [], "end": []}

    series_raw = {
        "plan_smr": [(_week_start(r["d"]), r["n"]) for r in plan_smr if r["d"]],
        "plan_delivery": [(_week_start(r["d"]), r["n"]) for r in plan_delivery if r["d"]],
        "fact_montage": [(_week_start(r["d"]), r["n"]) for r in fact_montage if r["d"]],
        "fact_delivery": [(_week_start(r["d"]), r["n"]) for r in fact_delivery if r["d"]],
        "forecast_montage": [(_week_start(d), n) for d, n in прогноз["end"]],
        "forecast_delivery": [(_week_start(d), n) for d, n in прогноз["start"]],
    }

    # Карточка и текстовые блоки принадлежат ОБЪЕКТУ (этап D). Без объекта
    # (отчёт по всей системе, source_file не задан и не сопоставлен ни с
    # одним чертежом) берутся пустые: показать здесь карточку «какого-
    # нибудь» объекта хуже, чем не показать никакой — числа в отчёте были
    # бы по одной стройке, а вехи и контрольные даты по другой.
    card = get_project_card(conn, object_id) if object_id else dict(PROJECT_CARD_DEFAULT)
    # Текстовые блоки — НЕ из карточки, а из редакции, действующей на
    # отчётную дату (живой запрос 2026-07-29). Отчёт за прошлую дату должен
    # показывать то, что было актуально тогда, а не сегодняшний текст.
    notes = (get_notes_for_date(conn, object_id, today) if object_id
             else {f: [] for f in NOTE_FIELDS} | {"effective_date": None})
    card = dict(card) | {f: notes[f] for f in NOTE_FIELDS}
    card["notes_effective_date"] = notes["effective_date"]
    weeks = sorted({w for pairs in series_raw.values() for w, _ in pairs} | {_week_start(today)})
    # Вехи и контрольные даты тоже задают правую границу графика — иначе
    # веха «Завершение 3 Захватки» оказалась бы за краем.
    for extra in [card.get("montage_deadline"), card.get("delivery_deadline")] + \
                 [m.get("date") for m in card.get("milestones", [])]:
        неделя = _week_start_or_none(extra)
        if неделя:
            weeks.append(неделя)
    weeks = sorted(set(weeks))

    series = {k: _cumulative(v, weeks) for k, v in series_raw.items()}

    # Кривые ФАКТА обрываются на отчётной дате (живой запрос 2026-07-30):
    # сетка недель тянется вправо до вех и контрольных дат, и накопительный
    # итог держал за отчётной датой горизонтальную полку — она читалась как
    # «факт известен и не растёт», хотя факта там ещё просто нет. Планы,
    # наоборот, идут вперёд до конца сетки — это и есть план.
    # None (а не обрезка списка) — длина всех рядов должна совпадать с
    # длиной weeks: и график, и выгрузка в Excel индексируют их по неделям.
    report_week = _week_start(today)
    cut = weeks.index(report_week) if report_week in weeks else len(weeks) - 1
    for key in ("fact_montage", "fact_delivery"):
        series[key] = [v if i <= cut else None for i, v in enumerate(series[key])]
    # Прогноз — наоборот, обрезается СЛЕВА: до отчётной даты прогнозировать
    # нечего, там уже есть факт, и накопительная кривая прогноза в прошлом
    # читалась бы как второй факт.
    #
    # И ПРИВЯЗЫВАЕТСЯ К ТОЧКЕ ФАКТА (2026-08-14, живой репорт: «почему такая
    # ступенька в текущей дате»). Накопленный прогноз считает и то, что уже
    # должно было быть сделано к отчётной дате: если по прогнозу к сегодня
    # положено 2800, а сделано 500, линия начиналась на 2800 — вертикальным
    # обрывом над концом факта. Обрыв — это и есть отставание, но читался он
    # как ошибка: заказчик просил линию, ИДУЩУЮ ОТ факта.
    #
    # Поэтому кривая сдвигается так, чтобы на отчётной неделе совпасть с
    # фактом, а дальше расти на СВОИ приросты. То есть показывается не «где
    # мы должны были быть», а «когда закончим, если пойдём темпом прогноза»,
    # — это и есть прогноз завершения. Отставание при этом не прячется: оно
    # видно как расстояние по горизонтали между прогнозом и планом.
    for key, факт, сырые in (("forecast_montage", "fact_montage", прогноз["end"]),
                             ("forecast_delivery", "fact_delivery", прогноз["start"])):
        ряд = series[key]
        if not ряд or cut >= len(ряд):
            continue
        основание = series[факт][cut] if cut < len(series[факт]) else None
        сдвиг = (основание or 0) - (ряд[cut] or 0)
        # Справа кривая ОБРЫВАЕТСЯ на своей последней дате — так же, как факт
        # обрывается на отчётной (2026-08-14, живой репорт: «почему монтаж
        # практически прекращается»). Накопительный итог за последней датой
        # версии держал горизонтальную полку, и она читалась как «работы
        # встали», хотя означает ровно обратное: в прогнозе закончились
        # изделия. Линия должна кончаться там, где кончаются данные.
        последняя = _week_start(max(d for d, _ in сырые)) if сырые else None
        предел = weeks.index(последняя) if последняя in weeks else len(ряд) - 1
        series[key] = [None if i < cut or i > предел else (v or 0) + сдвиг
                       for i, v in enumerate(ряд)]

    # Период графика (живой запрос 2026-08-03: «в отчёт динамики добавь
    # интервал дат аналогичный тому что сделан … в правой панели») — это
    # МАСШТАБ ОСИ X, а не пересчёт: ряды накопительные с начала проекта, и
    # окно просто вырезает часть недель, кривая входит в него на своём
    # накопленном уровне. Отчётная дата и таблицы «план/факт/отклонение» от
    # периода не зависят — они «на дату», а не «за интервал».
    #
    # Считается ЗДЕСЬ, на сервере, хотя панель «Статус» то же самое делает
    # в браузере: у формы есть выгрузка в XLSX и PDF, и они обязаны
    # показывать ровно то, что на экране. Обрезать в трёх местах —
    # гарантированно разъехаться.
    weeks_full = [weeks[0], weeks[-1]] if weeks else [None, None]
    lo = _week_start(week_from) if week_from else None
    hi = _week_start(week_to) if week_to else None
    if lo or hi:
        keep = [i for i, w in enumerate(weeks)
                if (lo is None or w >= lo) and (hi is None or w <= hi)]
        weeks = [weeks[i] for i in keep]
        series = {k: [v[i] for i in keep] for k, v in series.items()}

    def upto(pairs, limit_date, only_day=None):
        if only_day:
            return sum(n for d, n in pairs if d == only_day)
        return sum(n for d, n in pairs if d <= limit_date)

    raw_days = {
        "plan_smr": [(r["d"][:10], r["n"]) for r in plan_smr if r["d"]],
        "plan_delivery": [(r["d"][:10], r["n"]) for r in plan_delivery if r["d"]],
        "fact_montage": [(r["d"][:10], r["n"]) for r in fact_montage if r["d"]],
        "fact_delivery": [(r["d"][:10], r["n"]) for r in fact_delivery if r["d"]],
    }

    def block(plan_key, fact_key):
        cum_plan = upto(raw_days[plan_key], today)
        cum_fact = upto(raw_days[fact_key], today)
        day_plan = upto(raw_days[plan_key], today, only_day=today)
        day_fact = upto(raw_days[fact_key], today, only_day=today)
        return {
            "total": total,
            "cumulative": {"plan": cum_plan, "fact": cum_fact, "deviation": cum_fact - cum_plan},
            "day": {"plan": day_plan, "fact": day_fact, "deviation": day_fact - day_plan},
            "percent": round(cum_fact / total * 100) if total else 0,
        }

    return {
        "title": DYNAMICS_TITLE,
        "subtitle": DYNAMICS_SUBTITLE,
        "report_date": today,
        "card": card,
        "weeks": weeks,
        # Границы ВСЕГО срока проекта — ими форма подписывает поля периода и
        # ограничивает выбор: уйти за пределы имеющихся данных нельзя.
        "weeks_full_range": weeks_full,
        "week_from": week_from or None,
        "week_to": week_to or None,
        "series": series,
        "series_labels": {"plan_smr": "Монтаж (план)", "plan_delivery": "Поставка (план)",
                          "fact_delivery": "Поставка (факт)", "fact_montage": "Монтаж (факт)",
                          "forecast_montage": "Монтаж (прогноз)",
                          "forecast_delivery": "Поставка (прогноз)"},
        "forecast_version_id": прогноз.get("version_id"),
        # Какие кривые рисовать (2026-08-14). Ряды считаются ВСЕ и всегда —
        # режим лишь выбирает показываемое, поэтому переключение на экране
        # не требует пересчёта, а выгрузки XLSX/PDF получают тот же список и
        # не расходятся с экраном.
        "mode": режим,
        "series_order": DYN_MODE_SERIES[режим],
        "montage": block("plan_smr", "fact_montage"),
        "delivery": block("plan_delivery", "fact_delivery"),
        # Честная пометка о неполноте плана: на тестовых данных проектные
        # даты есть не у всех изделий, и молча рисовать такую кривую как
        # полный план нельзя (см. разбор в Docs/backlog.md).
        "plan_coverage": {
            "smr": sum(n for _, n in raw_days["plan_smr"]),
            "delivery": sum(n for _, n in raw_days["plan_delivery"]),
            "total": total,
        },
        # Сколько изделий покрывает прогноз. Меньше общего — законно (изделие
        # без привязки к крану, стоянке или этажу в расчёт не встаёт,
        # смонтированные исключаются при пересчёте от факта), но об этом
        # нужно СКАЗАТЬ: иначе кривая, не дорастающая до полного объёма,
        # читается как «монтаж прекратился».
        "forecast_coverage": {"elements": прогноз.get("elements", 0), "total": total},
        # Завершение: когда работы кончатся по прогнозу и насколько это
        # расходится с плановым сроком (2026-08-14, живой запрос). Отсюда и
        # отсечка на графике, и вывод текстом под ним.
        "finish": _finish_summary(raw_days, прогноз, card),
    }


def _finish_summary(raw_days: dict, прогноз: dict, card: dict) -> dict:
    """Плановое и прогнозное завершение по монтажу и по поставке.

    План завершения берётся из КАРТОЧКИ ОБЪЕКТА («окончание монтажа
    изделий» / «окончание поставки»), если он там задан: это директивный
    срок, по которому и отчитываются. Если не задан — последняя дата плана
    в самих данных: она хотя бы честно отвечает «когда план кончается», а
    отсутствие вывода вовсе хуже приблизительного.

    Знак отклонения тот же, что везде в системе: плюс — позже плана, то
    есть опоздание; минус — опережение.
    """
    def позже(даты):
        значения = [d for d, _ in даты if d]
        return max(значения) if значения else None

    def блок(план_даты, срок_из_карточки, прогноз_даты):
        план = _week_start_or_none(срок_из_карточки) and срок_из_карточки[:10] or позже(план_даты)
        план_источник = "карточка объекта" if срок_из_карточки else "последняя дата плана"
        факт_прогноз = позже(прогноз_даты)
        дней = None
        if план and факт_прогноз:
            from datetime import date
            try:
                дней = (date.fromisoformat(факт_прогноз[:10]) - date.fromisoformat(план[:10])).days
            except ValueError:
                дней = None
        return {"plan": план, "plan_source": план_источник,
                "forecast": факт_прогноз, "deviation_days": дней}

    return {
        "montage": блок(raw_days["plan_smr"], card.get("montage_deadline"), прогноз.get("end") or []),
        "delivery": блок(raw_days["plan_delivery"], card.get("delivery_deadline"), прогноз.get("start") or []),
    }


def _dyn_short_date(iso: str) -> str:
    MONTHS = ["янв", "фев", "мар", "апр", "май", "июн", "июл", "авг", "сен", "окт", "ноя", "дек"]
    from datetime import date
    d = date.fromisoformat(iso[:10])
    return f"{d.day:02d} {MONTHS[d.month - 1]}"


def _ru_date_short(iso: Optional[str]) -> str:
    """Дата по-русски; непригодное для разбора — пустая строка, а не
    исключение. Сюда приходят и контрольные сроки из КАРТОЧКИ ОБЪЕКТА
    (свободно правимый JSON настроек), и на одном таком значении падала вся
    выгрузка целиком — тот же случай, что у _week_start_or_none выше."""
    if not iso:
        return ""
    from datetime import date
    try:
        d = date.fromisoformat(str(iso)[:10])
    except (ValueError, TypeError):
        return ""
    return f"{d.day:02d}.{d.month:02d}.{d.year}"


# Порядок по умолчанию — на случай отчёта, собранного до появления режимов
# (в него ключ "series_order" не попадал). Живые отчёты берут порядок из
# самого отчёта: report["series_order"], см. DYN_MODE_SERIES.
DYN_SERIES_ORDER = DYN_MODE_SERIES["both"]
# Пары «план — факт» держатся одного цвета и различаются штрихом (планы
# пунктиром, факты сплошной): так на графике из четырёх кривых видно, что
# синие — это про монтаж, а оранжевые — про поставку, и глазу не нужно
# сверяться с легендой на каждую линию.
# Прогноз — цветом ФАКТА той же пары и штрихпунктиром: он и читается как
# продолжение фактической кривой («от серой линии факта идёт штрихпунктир
# прогноза», формулировка заказчика), а не как третий самостоятельный план.
DYN_SERIES_COLORS = {
    "plan_smr": "#4A86C8", "fact_montage": "#8C99A6", "forecast_montage": "#8C99A6",
    "plan_delivery": "#C2571A", "fact_delivery": "#E8703A", "forecast_delivery": "#E8703A",
}
DYN_SERIES_DASHED = {"plan_smr", "plan_delivery"}
DYN_SERIES_DASHDOT = {"forecast_montage", "forecast_delivery"}


def build_dynamics_report_pdf(report: dict) -> bytes:
    """PDF повторяет экранную вёрстку: шапка, два блока, график, две таблицы,
    открытые вопросы. График рисуется теми же формулами масштабирования, что
    и SVG на экране (см. buildDynamicsChartSvg в app.js) — своей отрисовкой,
    без сторонних библиотек графиков."""
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Flowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.pdf_export import FONT_BOLD, FONT_REGULAR

    card = report["card"]
    weeks = report["weeks"]

    class Chart(Flowable):
        """Свой Flowable: reportlab не умеет графиков, а тащить ради одного
        рисунка стороннюю библиотеку — новый вендоринг, требующий отдельного
        подтверждения (см. CLAUDE.md)."""

        def __init__(self, width, height):
            super().__init__()
            self.width, self.height = width, height

        def draw(self):
            c = self.canv
            L, R, T, B = 34, 8, 26, 34
            series = report["series"]
            # None — «за отчётной датой факта нет» (см. build_dynamics_report).
            порядок = report.get("series_order") or DYN_SERIES_ORDER
            top = max([1] + [v for k in порядок for v in series.get(k, []) if v is not None])
            pow10 = 10 ** (len(str(int(top))) - 1)
            max_y = -(-top // (pow10 / 2)) * (pow10 / 2)
            n = max(len(weeks), 1)

            def X(i):
                return L + (0 if n == 1 else i * (self.width - L - R) / (n - 1))

            def Y(v):
                return B + (v / max_y) * (self.height - T - B)

            c.setFont(FONT_REGULAR, 6)
            for i in range(6):
                v = max_y * i / 5
                y = Y(v)
                c.setStrokeColor(colors.HexColor("#E5E8EC"))
                c.line(L, y, self.width - R, y)
                c.setFillColor(colors.HexColor("#8A94A0"))
                c.drawRightString(L - 3, y - 2, str(int(v)))
            step = 2 if n > 18 else 1
            for i, w in enumerate(weeks):
                if i % step:
                    continue
                c.saveState()
                c.translate(X(i), B - 4)
                c.rotate(45)
                c.setFillColor(colors.HexColor("#8A94A0"))
                c.drawRightString(0, 0, _dyn_short_date(w))
                c.restoreState()

            for key in порядок:
                points = [(i, v) for i, v in enumerate(series.get(key, [])) if v is not None]
                if not any(v for _, v in points):
                    continue
                c.setStrokeColor(colors.HexColor(DYN_SERIES_COLORS[key]))
                c.setLineWidth(1.2)
                # Планы пунктиром, факты сплошной — см. DYN_SERIES_DASHED.
                c.setDash([6, 2, 1, 2] if key in DYN_SERIES_DASHDOT
                          else [3, 2] if key in DYN_SERIES_DASHED else [])
                path = c.beginPath()
                for n, (i, v) in enumerate(points):
                    (path.moveTo if n == 0 else path.lineTo)(X(i), Y(v))
                c.drawPath(path)
            c.setDash([])

            # Вехи за пределами показанного периода не рисуем: индекс недели
            # поджался бы к краю и веха встала бы на чужую неделю. То же
            # правило, что на экране (buildDynamicsChartSvg) — с появлением
            # периода в форме (2026-08-03) окно бывает уже полного срока.
            def in_window(iso):
                # Дата вехи приходит из карточки объекта и может оказаться
                # непригодной для разбора — тогда вехи просто нет, а не 500
                # на всю выгрузку (см. _week_start_or_none).
                неделя = _week_start_or_none(iso)
                return bool(weeks) and неделя is not None and weeks[0] <= неделя <= weeks[-1]

            # Отсечки «завершение по прогнозу» — те же вехи (2026-08-14).
            # Показываются только для тех рядов, что нарисованы: в режиме
            # «только монтаж» отсечка поставки была бы про кривую, которой
            # на графике нет.
            отсечки = []
            for ключ, слово in (("montage", "монтажа"), ("delivery", "поставки")):
                f = (report.get("finish") or {}).get(ключ) or {}
                if not f.get("forecast") or f"forecast_{ключ}" not in порядок:
                    continue
                оба = ("forecast_montage" in порядок) and ("forecast_delivery" in порядок)
                отсечки.append({"label": f"Прогноз завершения {слово}" if оба
                                else "Прогноз завершения", "date": f["forecast"]})
            marks = [m for m in
                     [{"label": "Отчётная дата", "date": report["report_date"]}]
                     + [m for m in card.get("milestones", []) if m.get("date")]
                     + отсечки
                     if in_window(m["date"])]
            c.setFont(FONT_REGULAR, 6)
            for idx, m in enumerate(marks):
                target = m["date"][:10]
                i = max([j for j, w in enumerate(weeks) if w <= target] or [0])
                x = X(i)
                plan = (series.get("plan_smr") or [0])[i] if i < len(series.get("plan_smr", [])) else 0
                y = Y(plan)
                top_y = self.height - 6 - (idx % 2) * 8
                c.setStrokeColor(colors.HexColor("#C0392B"))
                c.setFillColor(colors.HexColor("#C0392B"))
                c.line(x, top_y - 4, x, y + 4)
                p = c.beginPath(); p.moveTo(x - 2.5, y + 6); p.lineTo(x, y + 1); p.lineTo(x + 2.5, y + 6); p.close()
                c.drawPath(p, fill=1, stroke=0)
                text = m["label"] if m["label"] == "Отчётная дата" else f"{m['label']} {_ru_date_short(m['date'])}"
                (c.drawRightString if x > self.width * 0.75 else c.drawCentredString)(x, top_y, text)

            lx = L
            for key in порядок:
                label = report["series_labels"][key]
                c.setStrokeColor(colors.HexColor(DYN_SERIES_COLORS[key]))
                c.setLineWidth(1.4)
                c.setDash([6, 2, 1, 2] if key in DYN_SERIES_DASHDOT
                          else [3, 2] if key in DYN_SERIES_DASHED else [])
                c.line(lx, 6, lx + 12, 6)
                c.setFillColor(colors.HexColor("#4A5460"))
                c.drawString(lx + 15, 4, label)
                lx += 20 + len(label) * 3.4

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm, title=report["title"])
    center = ParagraphStyle("c", fontName=FONT_BOLD, fontSize=12, leading=15, alignment=1)
    sub = ParagraphStyle("s", fontName=FONT_REGULAR, fontSize=8, leading=11, alignment=1)
    subb = ParagraphStyle("sb", fontName=FONT_BOLD, fontSize=8, leading=11, alignment=1)
    small = ParagraphStyle("sm", fontName=FONT_REGULAR, fontSize=7.5, leading=10)
    note = ParagraphStyle("n", fontName=FONT_REGULAR, fontSize=6.5, leading=9,
                          textColor=colors.HexColor("#777777"))

    story = [
        Paragraph(f"Ежедневный отчёт за {_ru_date_short(report['report_date'])}", center),
        Paragraph(pdf_text(report["subtitle"]), sub),
        Paragraph(pdf_text(card.get("title")), subb),
        Spacer(1, 4 * mm),
    ]

    def bullet_box(title, items, head_bg):
        inner = [[Paragraph(f"<b>{pdf_text(title)}</b>", small)]]
        for it in (items or []):
            inner.append([Paragraph(f"•&nbsp;{pdf_text(it)}", small)])
        if not items:
            inner.append([Paragraph("не заполнено", note)])
        t = Table(inner, colWidths=[128 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(head_bg)),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D5D8DC")),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        return t

    story.append(Table([[bullet_box("Ключевые события", card.get("key_events"), "#E8F4EA"),
                         bullet_box("Ключевые задачи", card.get("key_tasks"), "#FDF0E3")]],
                       colWidths=[133 * mm, 133 * mm],
                       style=TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")])))
    story.append(Spacer(1, 4 * mm))
    story.append(Chart(266 * mm, 62 * mm))
    story.append(Spacer(1, 3 * mm))

    # Вывод словами — тот же, что на экране (2026-08-14): когда закончим по
    # прогнозу и на сколько это расходится с планом. Выгрузка обязана
    # повторять экран, иначе распечатанный отчёт говорит меньше исходного.
    вывод = []
    for ключ, что in (("montage", "Монтаж"), ("delivery", "Поставка")):
        f = (report.get("finish") or {}).get(ключ) or {}
        if not f.get("forecast") or f"forecast_{ключ}" not in DYN_MODE_SERIES.get(
                report.get("mode", "both"), []):
            continue
        срок = f"плановый срок {_ru_date_short(f['plan'])}" if f.get("plan") else "плановый срок не задан"
        d = f.get("deviation_days")
        оценка = ""
        if d is not None:
            оценка = (f" — опоздание на {d} дн." if d > 0
                      else f" — опережение на {abs(d)} дн." if d < 0 else " — день в день")
        вывод.append(f"<b>{что}:</b> завершение по прогнозу {_ru_date_short(f['forecast'])}, "
                     f"{срок}{оценка}")
    if вывод:
        story.append(Paragraph("<br/>".join(вывод), ParagraphStyle(
            "finish", parent=small, alignment=1, leading=11)))
        story.append(Spacer(1, 3 * mm))

    def status_table(caption, block, footnote):
        def dev(v):
            return f"+{v}" if v > 0 else str(v)
        rows = [
            [caption, "", "", "", "", "", "", ""],
            ["Всего в проекте", "Накопительно", "", "", f"На {_ru_date_short(report['report_date'])}", "", "", "%"],
            ["", "План", "Факт", "Отклонение", "План", "Факт", "Отклонение", ""],
            [block["total"], block["cumulative"]["plan"], block["cumulative"]["fact"], dev(block["cumulative"]["deviation"]),
             block["day"]["plan"], block["day"]["fact"], dev(block["day"]["deviation"]), f"{block['percent']}%"],
        ]
        t = Table(rows, colWidths=[24 * mm] + [17 * mm] * 6 + [14 * mm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), FONT_REGULAR), ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTNAME", (0, 0), (-1, 2), FONT_BOLD),
            ("SPAN", (0, 0), (-1, 0)), ("SPAN", (1, 1), (3, 1)), ("SPAN", (4, 1), (6, 1)),
            ("SPAN", (0, 1), (0, 2)), ("SPAN", (7, 1), (7, 2)),
            ("BACKGROUND", (0, 0), (-1, 2), colors.HexColor("#EEF2F7")),
            ("GRID", (0, 1), (-1, -1), 0.4, colors.HexColor("#D5D8DC")),
            ("ALIGN", (0, 1), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        # Список флоуаблов, а НЕ KeepTogether: внутри ячейки таблицы
        # KeepTogether заставляет reportlab считать высоту бесконечной и
        # падает с LayoutError «too large on page».
        return [t, Paragraph(footnote, note)] if footnote else [t]

    # Сноску даём, только если срок РАЗОБРАЛСЯ: «окончание монтажа изделий »
    # с пустотой на конце — не подпись, а след ошибки в данных.
    монтаж_до = _ru_date_short(card.get("montage_deadline"))
    поставка_до = _ru_date_short(card.get("delivery_deadline"))
    left = (status_table("Статус монтажа ЖБИ", report["montage"],
                         f"* окончание монтажа изделий {монтаж_до}" if монтаж_до else "")
            + [Spacer(1, 3 * mm)]
            + status_table("Статус поставки ЖБИ", report["delivery"],
                           f"** окончание поставки изделий {поставка_до}" if поставка_до else ""))

    story.append(Table([[left, bullet_box("Открытые вопросы", card.get("open_questions"), "#EEF2F7")]],
                       colWidths=[140 * mm, 126 * mm],
                       style=TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")])))

    cov = report["plan_coverage"]
    if cov["smr"] < cov["total"] or cov["delivery"] < cov["total"]:
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(
            f"Внимание: план СМР задан у {cov['smr']} изделий из {cov['total']}, "
            f"план поставки — у {cov['delivery']}. Кривая плана неполная.",
            ParagraphStyle("w", parent=note, textColor=colors.HexColor("#C0392B"))))

    doc.build(story)
    return buf.getvalue()


def build_dynamics_report_xlsx(report: dict) -> bytes:
    """Excel-версия — таблицы и ряды графика числами. График как рисунок не
    переносим: в Excel его строят из данных, и колонки рядов полезнее
    картинки."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    card = report["card"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Динамика"
    ws.append([f"Ежедневный отчёт за {_ru_date_short(report['report_date'])}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([report["subtitle"]])
    ws.append([card.get("title") or ""])
    ws.append([])

    for caption, block, deadline in (
        ("Статус монтажа ЖБИ", report["montage"], card.get("montage_deadline")),
        ("Статус поставки ЖБИ", report["delivery"], card.get("delivery_deadline")),
    ):
        ws.append([caption])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Всего в проекте", "Накопительно: план", "факт", "отклонение",
                   f"На {_ru_date_short(report['report_date'])}: план", "факт", "отклонение", "%"])
        for i in range(1, 9):
            ws.cell(row=ws.max_row, column=i).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=i).alignment = Alignment(wrap_text=True, horizontal="center")
        ws.append([block["total"], block["cumulative"]["plan"], block["cumulative"]["fact"],
                   block["cumulative"]["deviation"], block["day"]["plan"], block["day"]["fact"],
                   block["day"]["deviation"], block["percent"] / 100])
        ws.cell(row=ws.max_row, column=8).number_format = "0%"
        if deadline:
            ws.append([f"окончание {_ru_date_short(deadline)}"])
        ws.append([])

    # Ряды — те же, что на графике (report["series_order"], см.
    # DYN_MODE_SERIES): выгрузка обязана показывать выбранный режим, а не
    # всё подряд, иначе файл и экран расходятся.
    порядок = report.get("series_order") or DYN_SERIES_ORDER
    ws.append(["Ряды графика (накопительно по неделям)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Неделя"] + [report["series_labels"][k] for k in порядок])
    for i in range(1, len(порядок) + 2):
        ws.cell(row=ws.max_row, column=i).font = Font(bold=True)
    for i, w in enumerate(report["weeks"]):
        ws.append([w] + [report["series"].get(k, [0] * len(report["weeks"]))[i] for k in порядок])

    ws.append([])
    for title, items in (("Ключевые события", card.get("key_events")),
                         ("Ключевые задачи", card.get("key_tasks")),
                         ("Открытые вопросы", card.get("open_questions"))):
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for it in (items or []):
            ws.append([it])
        ws.append([])

    ws.column_dimensions["A"].width = 46
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 16
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
