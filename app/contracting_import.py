"""
Импорт файла "Контрактация" (.xlsx) — см. Docs/backlog.md, "Контрактация
2.0". Реальная структура образца (`Input/Контрактация.xlsx`, лист
"Контрактация"): колонки Покупатель/Поставщик/Договор поставки/
Спецификация/Наименование товара/Кол-во, данные со 2-й строки до первой
полностью пустой строки (в образце — строка 381; дальше в файле идут
черновые заметки автора в тех же колонках, не часть таблицы).

Каждая строка создаёт (или находит существующего) Контрагента (=Поставщик,
без ИНН/КПП/etc — их в файле нет, дозаполняются вручную в справочнике после
импорта), Договор и Спецификацию под ним (парсинг "НОМЕР от ДАТА"), и один
Контракт на эту Спецификацию (find_or_create_contract, app/contracts.py) —
после чего строка становится позицией контракта (тип+марка -> количество).

ОБЪЕКТ импорта задаётся снаружи и обязателен (2026-08-12). До этого импорт
звал find_or_create_agreement без объекта, и загруженный договор оставался с
`agreements.object_id IS NULL` — а объект контракта именно ВЫВОДИТСЯ по
цепочке контракт -> спецификация -> договор (см. schema.sql). Договор без
объекта не принадлежит никому: правила доступа отсеивают его у всех, кроме
администратора сервиса, и загруженная контрактация не видна ни в АРМ, ни в
отчётах, ни в документах контрактации. Дообъектное наследие лечила временная
обработка `app/fill_scope.py` — здесь закрывается сам источник.

Из объекта следует ещё три вещи, каждая — про то, что справочники объектные:
тип по уже загруженным изделиям ищется среди изделий ЭТОГО объекта; марка
позиции заносится в справочник марок объекта (`marks`); и если такая марка
там уже есть в другом написании — берётся СУЩЕСТВУЮЩЕЕ (см. _canonical_mark).

Тип элемента для марки, которой ещё нет ни у одного загруженного элемента:
эвристика по префиксу марки (mark_type_prefixes, app/counterparties.py),
донастраиваемая администратором. Не найдено — позиция создаётся с
element_type=NULL и попадает в сводку импорта как "тип не определён", не
блокирует остальной импорт.
"""

import io
import re
from datetime import date
from typing import Optional

from openpyxl import load_workbook

from app.contracts import find_or_create_contract
from app.counterparties import find_or_create_counterparty, find_or_create_specification

REQUIRED_HEADERS = ["Поставщик", "Договор поставки", "Спецификация", "Наименование товара", "Кол-во"]

_NUMBER_DATE_RE = re.compile(r"^(.*?)\s+от\s+(\d{1,2})\.(\d{1,2})\.(\d{2,4})\s*$", re.IGNORECASE)


class ContractingImportError(Exception):
    def __init__(self, status_code: int, message: str):
        self.status_code = status_code
        self.message = message
        super().__init__(message)


def _fix_short_year(year_str: str) -> Optional[int]:
    """Опечатка вида "206" вместо "2026" (потерян один разряд года) —
    перебор вставки одной цифры в каждую из 4 позиций, из результатов
    оставляем только те, что попадают в правдоподобное окно вокруг
    текущего года. Однозначно разрешается, только если ровно ОДИН
    кандидат попал в окно — иначе (0 или больше 1) считаем неразрешимым,
    не гадаем."""
    if len(year_str) != 3 or not year_str.isdigit():
        return None
    today_year = date.today().year
    lo, hi = today_year - 3, today_year + 6
    candidates = set()
    for pos in range(4):
        for digit in "0123456789":
            cand_str = year_str[:pos] + digit + year_str[pos:]
            if len(cand_str) != 4:
                continue
            cand = int(cand_str)
            if lo <= cand <= hi:
                candidates.add(cand)
    return candidates.pop() if len(candidates) == 1 else None


def parse_number_and_date(raw: str) -> tuple[str, Optional[str], Optional[str]]:
    """Разбирает "НОМЕР от ДД.ММ.ГГ(ГГ)" -> (number, date_iso, warning).
    date_iso — None, если даты в строке не было вовсе (тогда просто номер
    без даты, не ошибка) или год не удалось однозначно разобрать/исправить
    (тогда warning непустой, для сводки импорта, номер всё равно
    возвращается — find_or_create_agreement/specification не требуют
    даты)."""
    raw = raw.strip()
    m = _NUMBER_DATE_RE.match(raw)
    if not m:
        return raw, None, None

    number, day_s, month_s, year_s = m.group(1).strip(), m.group(2), m.group(3), m.group(4)
    warning = None
    if len(year_s) == 2:
        year = 2000 + int(year_s)
    elif len(year_s) == 4:
        year = int(year_s)
    elif len(year_s) == 3:
        fixed = _fix_short_year(year_s)
        if fixed is None:
            return number, None, f"не удалось разобрать год «{year_s}» в дате «{raw}»"
        year = fixed
        warning = f"дата «{raw}» — год исправлен на {year}"
    else:
        return number, None, f"не удалось разобрать год «{year_s}» в дате «{raw}»"

    try:
        parsed_date = date(year, int(month_s), int(day_s))
    except ValueError:
        return number, None, f"некорректная дата «{raw}»"
    return number, parsed_date.isoformat(), warning


def parse_contracting_xlsx(content: bytes) -> list[dict]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise ContractingImportError(422, "Файл повреждён или не является корректным .xlsx")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ContractingImportError(422, "Пустой файл")

    header = [str(h).strip() if h is not None else "" for h in header]
    missing = [h for h in REQUIRED_HEADERS if h not in header]
    if missing:
        raise ContractingImportError(422, f"В файле нет обязательных колонок: {', '.join(missing)}")
    col = {name: idx for idx, name in enumerate(header)}

    def get(row, name):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def is_blank(v) -> bool:
        return v is None or (isinstance(v, str) and not v.strip())

    parsed = []
    incomplete_rows = 0
    row_num = 1
    for row in rows_iter:
        row_num += 1
        values = [get(row, name) for name in REQUIRED_HEADERS]
        if all(is_blank(v) for v in values):
            break  # конец таблицы — дальше черновые заметки, не данные

        supplier, agreement_raw, specification_raw, mark, quantity = values
        if any(is_blank(v) for v in (supplier, agreement_raw, specification_raw, mark)) or quantity is None:
            incomplete_rows += 1
            continue
        try:
            quantity_int = int(quantity)
        except (TypeError, ValueError):
            incomplete_rows += 1
            continue

        parsed.append(
            {
                "row": row_num,
                "supplier": str(supplier).strip(),
                "agreement_raw": str(agreement_raw).strip(),
                "specification_raw": str(specification_raw).strip(),
                "mark": str(mark).strip(),
                "quantity": quantity_int,
            }
        )

    return {"rows": parsed, "incomplete_rows": incomplete_rows}


def _build_mark_lookup(conn, object_id: int) -> dict[str, str]:
    """{марка.lower(): element_type} по загруженным элементам ЭТОГО объекта —
    строится ОДИН раз на весь импорт, не на каждую строку. SQLite не
    приводит кириллицу к одному регистру ни через COLLATE NOCASE, ни
    через lower() (нет ICU-расширения) — регистронезависимое сравнение
    обязано быть на стороне Python, а не в SQL (см. живую проверку:
    lower('15КС1.1') в SQLite остаётся '15КС1.1' как есть).

    Отбор по объекту (2026-08-12): марки нумерует проектировщик в пределах
    здания, одноимённые марки соседних зданий — разные изделия (тот же
    довод, что у ключа справочника `marks`). Без отбора марка соседнего
    объекта молча назначала позиции чужой тип."""
    lookup: dict[str, str] = {}
    for row in conn.execute(
        "SELECT DISTINCT mark, element_type FROM elements WHERE mark IS NOT NULL AND object_id = ?",
        (object_id,),
    ):
        lookup[row["mark"].lower()] = row["element_type"]
    return lookup


def _build_mark_names(conn, object_id: int) -> dict[tuple[Optional[str], str], str]:
    """Написания марок справочника этого объекта: {(тип, марка.lower()):
    марка как в справочнике}. Ключ с типом — потому что владелец марки тип
    (marks: object_id, element_type, name); плюс ключ (None, lower) на
    случай позиции с неопределённым типом — он заполняется, только если во
    всём объекте это написание ОДНО (иначе выбирать за пользователя между
    «К-1» Колонны и «к-1» Панели нечем, и марка остаётся как в файле)."""
    названия: dict[tuple[Optional[str], str], str] = {}
    без_типа: dict[str, set[str]] = {}
    for row in conn.execute(
        "SELECT element_type, name FROM marks WHERE object_id = ?", (object_id,)
    ):
        названия[(row["element_type"], row["name"].lower())] = row["name"]
        без_типа.setdefault(row["name"].lower(), set()).add(row["name"])
    for ключ, варианты in без_типа.items():
        if len(варианты) == 1:
            названия[(None, ключ)] = next(iter(варианты))
    return названия


def _canonical_mark(mark: str, element_type: Optional[str],
                    mark_names: dict[tuple[Optional[str], str], str]) -> str:
    """Написание марки, под которым она уже живёт в справочнике объекта.

    Зачем. Марка позиции контракта — по-прежнему ТЕКСТ (contract_lines.mark),
    и «15кс1.1» из файла рядом с «15КС1.1» из чертежа расщепляет изделие на
    две ветки в остатках контракта, фильтрах и АРМ — ровно та задвоенность,
    ради разбора которой заводился справочник марок. Существующее написание
    здесь ПРИОРИТЕТНЕЕ файла: справочник — то, что пользователь уже сверил.

    Новая марка (в справочнике её нет) пишется как в файле — угадывать
    «правильный» регистр не из чего."""
    return mark_names.get((element_type, mark.lower()), mark_names.get((None, mark.lower()), mark))


_LEADING_DIGITS_RE = re.compile(r"^[0-9]+")


def _resolve_element_type(mark: str, mark_lookup: dict[str, str], prefix_map: dict[str, str]) -> Optional[str]:
    element_type = mark_lookup.get(mark.lower())
    if element_type is not None:
        return element_type
    # Реальные марки часто начинают с числа-позиции ДО буквенного кода
    # (напр. "3Р19.6" — Ригель под номером 3), которое надо отбросить
    # перед сравнением с префиксом — иначе "Р" не совпадёт с началом
    # строки "3Р19.6" (см. Docs/backlog.md, тот же приём уже
    # использовался при анализе образца файла на этапе планирования).
    mark_for_prefix = _LEADING_DIGITS_RE.sub("", mark).lower()
    best_prefix, best_type = None, None
    for prefix, element_type in prefix_map.items():
        # Сравнение БЕЗ УЧЁТА РЕГИСТРА (2026-08-06): марка «15кс1.1» и марка
        # «15КС1.1» — одно и то же изделие, и префикс «Кс» обязан узнавать
        # обе. До этой правки сравнение было точным, и справочник префиксов
        # приходилось держать с регистровыми двойниками («КН» и «Кн», «Кс» и
        # «кс» — они и сейчас есть в сидинге, см. _MARK_TYPE_PREFIX_SEED).
        # Двойники стали безвредны, но больше не нужны.
        ключ = prefix.lower()
        if mark_for_prefix.startswith(ключ) and (best_prefix is None or len(ключ) > len(best_prefix)):
            best_prefix, best_type = ключ, element_type
    return best_type


def resolve_agreement(conn, counterparty_id: int, number: str,
                       agreement_date: Optional[str], object_id: int
                       ) -> tuple[Optional[int], Optional[int], bool]:
    """Договор этого контрагента с этим номером НА ЭТОМ объекте.

    Возвращает (agreement_id, чужой_объект, дозаполнен_объект). Три случая:
      * договора нет            -> заводится с object_id;
      * есть, объект тот же или
        не проставлен вовсе     -> используется, пустой объект дозаполняется
                                   (дообъектное наследие, см. app/fill_scope.py);
      * есть, но объект ЧУЖОЙ   -> (None, id чужого объекта): строка
                                   отклоняется и попадает в сводку.

    Почему не «завести второй договор»: ключ UNIQUE(counterparty_id, number)
    общий на весь сервис, номер договора нельзя переиспользовать на другом
    объекте. Молча вернуть чужой договор (как делала прежняя
    find_or_create_agreement) — хуже всего: позиции файла уехали бы в
    контракт другой стройки. Ручная форма «+ Договор» на тот же случай
    отвечает 400 (см. app/counterparties.py:create_agreement)."""
    row = conn.execute(
        "SELECT id, object_id FROM agreements WHERE counterparty_id = ? AND number = ?",
        (counterparty_id, number),
    ).fetchone()
    if row is None:
        conn.execute(
            "INSERT INTO agreements (counterparty_id, number, agreement_date, object_id) "
            "VALUES (?, ?, ?, ?)",
            (counterparty_id, number, agreement_date, object_id),
        )
        return conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"], None, False
    if row["object_id"] is None:
        conn.execute(
            "UPDATE agreements SET object_id = ?, updated_at = datetime('now') WHERE id = ?",
            (object_id, row["id"]),
        )
        return row["id"], None, True
    if row["object_id"] != object_id:
        return None, row["object_id"], False
    return row["id"], None, False


def import_contracting(conn, parsed: dict, object_id: int) -> dict:
    rows = parsed["rows"]
    prefix_map = {
        r["prefix"]: r["element_type"]
        for r in conn.execute("SELECT prefix, element_type FROM mark_type_prefixes").fetchall()
    }
    mark_lookup = _build_mark_lookup(conn, object_id)
    mark_names = _build_mark_names(conn, object_id)
    counterparty_by_lower = {
        r["short_name"].lower(): r["id"]
        for r in conn.execute("SELECT id, short_name FROM counterparties")
    }

    date_warnings: list[str] = []
    unresolved_type_marks: list[str] = []
    contracts_touched: set[int] = set()
    # Строки, чей договор принадлежит другому объекту: считаем и показываем
    # поимённо — это не «неполная строка», а расхождение файла с базой,
    # которое человек должен увидеть и разобрать (см. resolve_agreement).
    foreign_agreements: dict[str, int] = {}
    foreign_rows = 0
    agreements_object_filled = 0
    marks_created = 0
    # (contract_id, element_type, mark) -> summed quantity ЗА ЭТОТ ЗАПУСК —
    # намеренно не "+=" прямо в БД: повторный запуск того же файла должен
    # выставлять итоговое количество заново, а не бесконечно накапливать
    # его при каждой перезагрузке одного и того же файла (тот же принцип
    # идемпотентности, что у upsert-импорта DXF-элементов).
    line_quantities: dict[tuple[int, Optional[str], str], int] = {}

    for row in rows:
        agreement_number, agreement_date, agr_warning = parse_number_and_date(row["agreement_raw"])
        specification_number, specification_date, spec_warning = parse_number_and_date(row["specification_raw"])
        if agr_warning:
            date_warnings.append(f"Строка {row['row']}, договор «{row['agreement_raw']}»: {agr_warning}")
        if spec_warning:
            date_warnings.append(f"Строка {row['row']}, спецификация «{row['specification_raw']}»: {spec_warning}")

        # Регистронезависимо, как и марки: find_or_create_counterparty
        # сравнивает short_name точным SQL-равенством, и «К-ЖБИ» рядом с
        # «к-жби» завели бы ДВУХ контрагентов (SQLite без ICU не сворачивает
        # кириллицу — см. _build_mark_lookup). Ровно тот же обход уже стоит
        # в импорте истории (_resolve_contract_id, app/history_import.py), и
        # ровно это обещает описание формата (app/import_templates.py).
        counterparty_id = counterparty_by_lower.get(row["supplier"].lower())
        if counterparty_id is None:
            counterparty_id = find_or_create_counterparty(
                conn, full_name=row["supplier"], short_name=row["supplier"]
            )
            counterparty_by_lower[row["supplier"].lower()] = counterparty_id
        agreement_id, чужой, дозаполнен = resolve_agreement(
            conn, counterparty_id, agreement_number, agreement_date, object_id
        )
        if agreement_id is None:
            foreign_rows += 1
            foreign_agreements[f"{row['supplier']}, договор «{agreement_number}»"] = чужой
            continue
        if дозаполнен:
            agreements_object_filled += 1
        specification_id = find_or_create_specification(
            conn, agreement_id, specification_number, specification_date
        )
        contract_id = find_or_create_contract(conn, specification_id)
        contracts_touched.add(contract_id)

        element_type = _resolve_element_type(row["mark"], mark_lookup, prefix_map)
        if element_type is None:
            unresolved_type_marks.append(row["mark"])
        mark = _canonical_mark(row["mark"], element_type, mark_names)
        # Марка попадает в справочник объекта сразу, а не одноразовой
        # релиз-обработкой (_fill_marks_catalog, app/release_tasks.py):
        # позиция контракта — такой же источник марок, как чертёж, и без
        # записи справочника её марки не видно ни в АРМ, ни в фильтрах.
        # Тип обязателен ключом справочника — позицию с неопределённым типом
        # заводить нечем, она ждёт, пока тип донастроят вручную.
        if element_type is not None and (element_type, mark.lower()) not in mark_names:
            marks_created += conn.execute(
                "INSERT OR IGNORE INTO marks (object_id, element_type, name) VALUES (?, ?, ?)",
                (object_id, element_type, mark),
            ).rowcount
            # Тот же файл дальше по строкам обязан узнавать только что
            # заведённое написание — иначе «15кс1.1» после «15КС1.1» в том
            # же файле снова разъедется на две позиции.
            mark_names[(element_type, mark.lower())] = mark
            mark_names.setdefault((None, mark.lower()), mark)

        key = (contract_id, element_type, mark)
        line_quantities[key] = line_quantities.get(key, 0) + row["quantity"]

    inserted_lines = updated_lines = 0
    for (contract_id, element_type, mark), quantity in line_quantities.items():
        existing = conn.execute(
            "SELECT id FROM contract_lines WHERE contract_id = ? AND element_type IS ? AND mark = ?",
            (contract_id, element_type, mark),
        ).fetchone()
        if existing:
            conn.execute("UPDATE contract_lines SET quantity = ? WHERE id = ?", (quantity, existing["id"]))
            updated_lines += 1
        else:
            conn.execute(
                "INSERT INTO contract_lines (contract_id, element_type, mark, quantity) VALUES (?, ?, ?, ?)",
                (contract_id, element_type, mark, quantity),
            )
            inserted_lines += 1

    conn.commit()

    return {
        "rows_processed": len(rows),
        "incomplete_rows_skipped": parsed["incomplete_rows"],
        "contracts_touched": len(contracts_touched),
        "lines_inserted": inserted_lines,
        "lines_updated": updated_lines,
        "marks_created": marks_created,
        "agreements_object_filled": agreements_object_filled,
        "foreign_agreement_rows": foreign_rows,
        "foreign_agreements": sorted(foreign_agreements),
        "unresolved_type_marks": sorted(set(unresolved_type_marks)),
        "date_warnings": date_warnings[:50],
    }
