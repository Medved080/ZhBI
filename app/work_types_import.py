"""
Загрузка справочника видов работ (WBS) из xlsx (Docs/block-accounting.md §4).

Двухфазно, как импорт чертежа и пакетов Revit (решение И3): `analyze`
разбирает файл и считает расхождения, ничего не пишет; `apply` применяет
уже посчитанное по токену.

Формат листа `WBS` (или первого листа книги, если листа с таким именем нет
— старые однолистовые файлы): `Уровень WBS | Идентификатор операции |
Название операции | Единицы измерения | Кодификатор`, плюс необязательные
«Примечание» и «Трек планирования» (2026-09-04) — код группы «Шахматка»
(`app/work_fact.py`, Docs/block-accounting.md §8), расшифровка кодов — на
листе `PlanningTrack` той же книги (`Код | Наименование | Примечание`,
`parse_planning_track`), перезагружаемая вместе со справочником в таблицу
`planning_tracks`. Четыре ловушки исходного файла заказчика (проверено на
реальном файле «WBS МФР типовой», см. документ):

  * у УЗЛА название лежит в колонке «Идентификатор операции», а у
    ОПЕРАЦИИ/ВЕХИ — в «Название операции». Колонка «Уровень WBS» несёт для
    узла НОМЕР УРОВНЯ, а для операции/вехи — слово `оп`/`веха` (тип строки,
    не глубина); операция/веха всегда становится ребёнком ПОСЛЕДНЕГО
    встреченного узла, а не следующего числового уровня;
  * «номер уровня» узла бывает в ДВУХ разных обозначениях, и файл
    выдерживает только одно из них целиком: путь по дереву («1», «1.2»,
    «2.1.3» — глубина считается по числу точек) или голое число абсолютной
    глубины («1», «2», «3», «4» — глубина это само число). В файле «WBS МФР
    типовой» — второй вариант: узел третьего уровня вложенности обозначен
    просто «3», а не «1.1.3». Перепутать критично: посчитать голое число
    точками — все узлы дерева, глубина которых не 1, схлопнутся в одну
    глубину и одноимённые категории из разных веток («Земляные работы»,
    «Металлоконструкции» и т.п.) склеятся по пути в один узел, затерев
    большую часть подчинённых операций. Формат определяется один раз на
    весь лист (`_uses_dotted_levels`): точка хоть где-то в колонке — путь по
    дереву, иначе — голое число;
  * кодификатор МОЖЕТ повторяться и МОЖЕТ отсутствовать — ключом записи
    служит ПУТЬ по дереву (имена всех предков + своё), он и только он
    уникален;
  * пустое название операции — дефект исходника, не повод падать: строка
    остаётся с пустым именем и отдельной строкой в предупреждениях.
"""

import io

from openpyxl import load_workbook

HEADERS = ("Уровень WBS", "Идентификатор операции", "Название операции",
          "Единицы измерения", "Кодификатор")

# Необязательная колонка: встречается не во всех файлах и не у всех строк
# файла, где есть (в «WBS МФР типовой» заполнена у 5 строк из 224) — но
# несёт содержательные пояснения («выбрать нужную технологию»), поэтому
# сохраняется, если колонка в файле есть, и просто не заполняется, если нет.
COL_NOTE = "Примечание"

# Тоже необязательная, тем же приёмом (2026-09-04, живой запрос): код
# «1»..«20» — операция входит в шахматку с визуализацией на модели (имя
# доски — на листе PlanningTrack по этому коду), «0» — линейный трек без
# раскраски блоков, «компл»/«Веха» — вне области этой доработки. У файлов
# без этой колонки (например «WBS МФР типовой») просто не заполняется.
COL_TRACK = "Трек планирования"

# Имена листов нового формата файла (WBS_*.xlsx, 2026-09-04): данные — на
# листе «WBS», расшифровка кодов трека — на «PlanningTrack». У старых
# однолистовых файлов такого имени нет — тогда берётся первый лист книги,
# а расшифровки треков просто не будет (колонки «Трек планирования» у них
# тоже нет).
SHEET_WBS = "WBS"
SHEET_PLANNING_TRACK = "PlanningTrack"

TRACK_HEADERS = ("Код", "Наименование")
TRACK_COL_NOTE = "Примечание"

ROW_NODE = "узел"
ROW_OP = "оп"
ROW_MILESTONE = "веха"

PATH_SEP = " / "

_PENDING = {}
_PENDING_LIMIT = 3


class WorkTypesError(Exception):
    def __init__(self, status_code: int, message: str):
        self.status_code = status_code
        self.message = message
        super().__init__(message)


def _index_by_headers(header_row, required, optional, sheet_label) -> dict:
    values = [str(c.value).strip() if c.value is not None else "" for c in header_row]
    index = {}
    for name in required:
        if name not in values:
            raise WorkTypesError(
                422, "На листе «%s» нет колонки «%s». Ожидаются: %s"
                % (sheet_label, name, ", ".join(required)))
        index[name] = values.index(name)
    for name in optional:
        if name in values:
            index[name] = values.index(name)
    return index


def _header_index(header_row) -> dict:
    return _index_by_headers(header_row, HEADERS, (COL_NOTE, COL_TRACK), SHEET_WBS)


def _code_str(v):
    """Код (трека или строки справочника) — к строке без «.0» у чисел:
    openpyxl отдаёт числовые коды ячейками типа float/int (0, 1, 2…20), и
    только «компл»/«Веха» приходят текстом."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return s or None


def _uses_dotted_levels(sheet, idx) -> bool:
    """Определяет формат колонки «Уровень WBS» по всему листу разом: если
    хоть где-то встретилась точка («2.1.3») — это путь по дереву, и глубина
    везде считается по числу точек; если точек нет ни у одной строки (везде
    голые «1», «2», «3», «4») — число само и есть абсолютная глубина узла.
    Ловушка ровно на этом: у файла «WBS МФР типовой» узел на глубине 3
    обозначен просто «3», а не «1.1.3» — при счёте точек это дало бы
    глубину 1 всем узлам подряд и склеило разные ветки дерева в одну."""
    i = idx["Уровень WBS"]
    for row in sheet.iter_rows(min_row=2):
        v = row[i].value if i < len(row) else None
        cell = str(v).strip() if v is not None else ""
        if cell and cell.lower() not in (ROW_OP, ROW_MILESTONE) and "." in cell:
            return True
    return False


def parse_planning_track(sheet) -> tuple:
    """Разбор листа PlanningTrack: код -> название/примечание (2026-09-04).
    Возвращает (tracks, warnings). Тем же приёмом, что и WBS: «Примечание»
    необязательна, код — строкой без «.0» у чисел (_code_str)."""
    rows_iter = sheet.iter_rows()
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Лист «%s» пуст." % SHEET_PLANNING_TRACK]
    idx = _index_by_headers(header_row, TRACK_HEADERS, (TRACK_COL_NOTE,), SHEET_PLANNING_TRACK)

    warnings = []
    tracks = []
    seen = {}
    for excel_row_no, row in enumerate(rows_iter, start=2):
        def cellv(col):
            i = idx.get(col)
            return row[i].value if i is not None and i < len(row) else None

        code = _code_str(cellv("Код"))
        name = _code_str(cellv("Наименование"))
        note = _code_str(cellv(TRACK_COL_NOTE))
        if code is None and name is None:
            continue  # пустая строка
        if code is None:
            warnings.append("Лист «%s», строка %d: пуст код, строка пропущена"
                            % (SHEET_PLANNING_TRACK, excel_row_no))
            continue
        if not name:
            warnings.append("Лист «%s», строка %d: у кода «%s» нет названия"
                            % (SHEET_PLANNING_TRACK, excel_row_no, code))
        if code in seen:
            warnings.append("Лист «%s»: код «%s» встретился дважды — вторая строка перезатёрла первую"
                            % (SHEET_PLANNING_TRACK, code))
        entry = {"code": code, "name": name or "", "note": note}
        if code in seen:
            tracks[seen[code]] = entry
        else:
            seen[code] = len(tracks)
            tracks.append(entry)
    return tracks, warnings


def parse_xlsx(data: bytes) -> tuple:
    """Разбор книги в плоский список строк дерева с уже посчитанным путём,
    плюс расшифровку кодов трека планирования (лист PlanningTrack, если он
    в книге есть).

    Возвращает (rows, warnings, tracks). `rows` — список словарей: path,
    parent_path, row_kind, code, name, unit, note, track_code, sort_order.
    `tracks` — список словарей code/name/note с листа PlanningTrack (пусто,
    если такого листа в книге нет — старые однолистовые файлы).
    """
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise WorkTypesError(422, "Файл не открылся как xlsx: %s" % e)
    sheet = wb[SHEET_WBS] if SHEET_WBS in wb.sheetnames else wb.worksheets[0]

    tracks, warnings = ([], [])
    if SHEET_PLANNING_TRACK in wb.sheetnames:
        tracks, warnings = parse_planning_track(wb[SHEET_PLANNING_TRACK])

    rows_iter = sheet.iter_rows()
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise WorkTypesError(422, "Лист пуст.")
    idx = _header_index(header_row)
    dotted_levels = _uses_dotted_levels(sheet, idx)

    result = []
    # Стек текущих узлов: [(глубина, путь, порядковый_номер)].
    stack = []
    sort_order = 0
    last_node_path = None

    for excel_row_no, row in enumerate(rows_iter, start=2):
        def cell(col):
            i = idx.get(col)
            if i is None:
                return ""   # необязательная колонка (COL_NOTE), в файле нет
            v = row[i].value if i < len(row) else None
            return str(v).strip() if v is not None else ""

        level_cell = cell("Уровень WBS")
        if not level_cell and not cell("Идентификатор операции") and not cell("Название операции"):
            continue  # пустая строка — не ошибка, просто пропуск

        code = cell("Кодификатор") or None
        unit = cell("Единицы измерения") or None
        note = cell(COL_NOTE) or None
        track_code = cell(COL_TRACK) or None
        kind_word = level_cell.lower()

        if kind_word in (ROW_OP, ROW_MILESTONE):
            row_kind = ROW_OP if kind_word == ROW_OP else ROW_MILESTONE
            name = cell("Название операции")
            if not name:
                warnings.append(
                    "Строка %d: у %s нет названия (колонка «Название операции» пуста)"
                    % (excel_row_no, "операции" if row_kind == ROW_OP else "вехи"))
            if last_node_path is None:
                warnings.append(
                    "Строка %d: %s встретилась раньше первого узла, пропущена"
                    % (excel_row_no, row_kind))
                continue
            if track_code and track_code not in {t["code"] for t in tracks}:
                warnings.append(
                    "Строка %d: код трека «%s» не найден на листе «%s»"
                    % (excel_row_no, track_code, SHEET_PLANNING_TRACK))
            path = last_node_path + PATH_SEP + (name or "(без названия, строка %d)" % excel_row_no)
            sort_order += 1
            result.append({
                "path": path, "parent_path": last_node_path, "row_kind": row_kind,
                "code": code, "name": name, "unit": unit, "note": note,
                "track_code": track_code, "sort_order": sort_order,
            })
            continue

        # Иначе — узел: «Уровень WBS» несёт номер уровня. Формат — либо путь
        # по дереву («1», «1.2», «2.1.3» — глубина по числу точек), либо
        # голое число абсолютной глубины («1», «2», «3», «4») — какой из
        # двух, решено один раз на весь лист в _uses_dotted_levels.
        if not level_cell:
            depth = None
        elif dotted_levels:
            depth = level_cell.count(".") + 1
        else:
            try:
                depth = int(level_cell)
            except ValueError:
                depth = None
        name = cell("Идентификатор операции")
        if not name:
            warnings.append(
                "Строка %d: у узла нет названия (колонка «Идентификатор "
                "операции» пуста), строка пропущена" % excel_row_no)
            continue
        if depth is None:
            warnings.append(
                "Строка %d: не удалось определить уровень узла «%s» (значение "
                "«%s» в колонке «Уровень WBS» не распознано), строка пропущена"
                % (excel_row_no, name, level_cell))
            continue

        while stack and stack[-1][0] >= depth:
            stack.pop()
        parent_path = stack[-1][1] if stack else None
        path = (parent_path + PATH_SEP + name) if parent_path else name
        sort_order += 1
        result.append({
            "path": path, "parent_path": parent_path, "row_kind": ROW_NODE,
            "code": code, "name": name, "unit": unit, "note": note,
            "track_code": None, "sort_order": sort_order,
        })
        stack.append((depth, path))
        last_node_path = path

    seen_paths = {}
    for r in result:
        if r["path"] in seen_paths:
            warnings.append("Путь «%s» встретился дважды — вторая строка перезатёрла первую"
                            % r["path"])
        seen_paths[r["path"]] = r
    return list(seen_paths.values()), warnings, tracks


def _existing(conn, object_id: int) -> dict:
    return {
        row["path"]: dict(row)
        for row in conn.execute(
            "SELECT id, path, row_kind, code, name, unit, note, retired_at FROM work_types "
            "WHERE object_id = ?", (object_id,))
    }


def analyze(conn, object_id: int, data: bytes) -> dict:
    """Фаза 1. В БД не пишет ничего."""
    parsed, warnings, tracks = parse_xlsx(data)
    have = _existing(conn, object_id)

    parsed_paths = {r["path"] for r in parsed}
    new_rows = [r for r in parsed if r["path"] not in have]
    existing_active = [r for r in parsed if r["path"] in have and not have[r["path"]]["retired_at"]]
    reviving = [r for r in parsed if r["path"] in have and have[r["path"]]["retired_at"]]
    retiring = [p for p, row in have.items() if p not in parsed_paths and not row["retired_at"]]

    return {
        "object_id": object_id,
        "total_rows": len(parsed),
        "new": [{"путь": r["path"], "тип": r["row_kind"], "единица": r["unit"]}
                for r in new_rows],
        "reviving": [{"путь": r["path"]} for r in reviving],
        "retiring": sorted(retiring),
        "unchanged": len(existing_active) - len(reviving),
        "tracks": [{"код": t["code"], "название": t["name"]} for t in tracks],
        "warnings": warnings,
        "_parsed": parsed,
        "_tracks": tracks,
    }


def apply(conn, object_id: int, analysis: dict) -> dict:
    """Фаза 2: применяет уже посчитанное. Сопоставление — ПО ПУТИ.
    Пропавшие пути помечаются `retired_at`, а не удаляются — иначе
    потерялась бы простановленная по ним история статусов."""
    parsed = analysis["_parsed"]
    tracks = analysis.get("_tracks", [])
    have = _existing(conn, object_id)
    parsed_paths = {r["path"] for r in parsed}

    path_to_id = {p: row["id"] for p, row in have.items()}
    added = 0
    revived = 0
    for r in parsed:
        existing = have.get(r["path"])
        parent_id = path_to_id.get(r["parent_path"]) if r["parent_path"] else None
        if existing is None:
            cur = conn.execute(
                "INSERT INTO work_types (object_id, parent_id, path, row_kind, code, "
                "name, unit, note, planning_track_code, sort_order, retired_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,NULL)",
                (object_id, parent_id, r["path"], r["row_kind"], r["code"],
                 r["name"], r["unit"], r["note"], r["track_code"], r["sort_order"]),
            )
            path_to_id[r["path"]] = cur.lastrowid
            added += 1
        else:
            if existing["retired_at"]:
                revived += 1
            conn.execute(
                "UPDATE work_types SET parent_id = ?, code = ?, name = ?, unit = ?, note = ?, "
                "planning_track_code = ?, sort_order = ?, retired_at = NULL WHERE id = ?",
                (parent_id, r["code"], r["name"], r["unit"], r["note"], r["track_code"],
                 r["sort_order"], existing["id"]),
            )

    retired = 0
    for path, row in have.items():
        if path not in parsed_paths and not row["retired_at"]:
            conn.execute(
                "UPDATE work_types SET retired_at = datetime('now') WHERE id = ?",
                (row["id"],),
            )
            retired += 1

    # Справочник треков (planning_tracks) — маленький, без своей истории:
    # переустанавливается целиком по объекту, а не сверяется построчно, как
    # work_types (см. модуль docstring и schema.sql).
    have_track_codes = {
        row["code"] for row in conn.execute(
            "SELECT code FROM planning_tracks WHERE object_id = ?", (object_id,))
    }
    parsed_track_codes = {t["code"] for t in tracks}
    for code in have_track_codes - parsed_track_codes:
        conn.execute(
            "DELETE FROM planning_tracks WHERE object_id = ? AND code = ?", (object_id, code))
    for t in tracks:
        conn.execute(
            "INSERT INTO planning_tracks (object_id, code, name, note) VALUES (?,?,?,?) "
            "ON CONFLICT (object_id, code) DO UPDATE SET name = excluded.name, note = excluded.note",
            (object_id, t["code"], t["name"], t["note"]),
        )

    conn.commit()
    return {"added": added, "revived": revived, "retired": retired, "total": len(parsed),
            "tracks": len(tracks)}


def remember_pending(analysis: dict) -> str:
    import uuid
    token = uuid.uuid4().hex
    _PENDING[token] = analysis
    while len(_PENDING) > _PENDING_LIMIT:
        _PENDING.pop(next(iter(_PENDING)))
    return token


def get_pending(token: str) -> dict:
    analysis = _PENDING.get(token)
    if analysis is None:
        raise WorkTypesError(
            410, "Результат разбора уже недоступен (сервер перезапускался или "
            "разбор устарел). Загрузите файл заново.")
    return analysis


def forget_pending(token: str) -> None:
    _PENDING.pop(token, None)
