"""
Импорт графика MS Project ("Прогноз СМР.xlsx", см. Docs/backlog.md,
"Контрактация 2.0", п.5/6). Реальная структура образца: колонки
Тип и Подтип/Начало/Окончание/Кран/Захватка/Стоянка/Этаж — каждая строка
описывает блок работ для комбинации Кран+Стоянка+Этаж+Тип+Подтип. Всем
элементам, физически попадающим в этот блок, проставляются два новых
поля: project_delivery_date ("Дата завершения СМР", из "Окончание" —
конец СМР по этому блоку) и project_smr_start_date (из "Начало" — начало
СМР по этому блоку; к этому моменту элемент должен физически быть на
площадке, см. критерий опоздания в app/static/app.js,
computeDeliveryLateStatus).

"Захватка" УЧАСТВУЕТ в сопоставлении с 2026-08-14. До этой даты она
читалась и отбрасывалась — считалось, что стоянка однозначно определяет
захватку. На реальном графике это неверно: в файле
«grafik_ms_project (москвич 14.08.2026).xlsx» стоянки 4, 5, 10, 11, 16 и 17
у КАЖДОГО из трёх кранов относятся к двум разным захваткам. Беды это ещё не
принесло (ключ со стоянкой и этажом всё равно оставался уникальным), но
следующий график развёз бы даты по соседней захватке молча. Строка без
разбираемого номера захватки не отбрасывается — сопоставление тогда идёт
без неё, как раньше: старые файлы должны грузиться.

"Тип и Подтип" — свободный текст ("Колонна нижняя", "Панелей лифтовой
шахты до отм. +15.000" — родительный падеж от "Панель"), не совпадает
буквально со словарём element_type/subtype — используется ЗАКРЫТЫЙ список
алиасов (TYPE_SUBTYPE_ALIASES), а не парсер общего вида: строка, которой
нет в списке, попадает в сводку как unmatched, не угадывается.

Кран/Стоянка сопоставляются элементу ЧЕРЕЗ его собственные
zone_crane_id/zone_stance_id (см. app/db.py) — зоны с одинаковым именем
повторяются в каждом source_file, сравнение "в лоб" по имени зоны было бы
неоднозначным без привязки к конкретному элементу.
"""

import io
import json

from app import activity
import re
from datetime import date, datetime
from typing import Optional

from openpyxl import load_workbook

# Колонка с видом работ называется по-разному в разных выгрузках MS Project:
# «Тип и Подтип» в образце «Прогноз СМР.xlsx» и «Название задачи» в выгрузке
# 6-го уровня структуры (реальный файл актуализированного графика,
# 2026-08-14). Содержимое одно и то же, поэтому принимаются оба имени — а не
# «любая первая колонка»: перепутанный файл должен отвергаться, а не
# разбираться наугад.
WORK_KIND_HEADERS = ["Тип и Подтип", "Название задачи"]
REQUIRED_HEADERS = ["Начало", "Окончание", "Кран", "Захватка", "Стоянка", "Этаж"]

# Полный список из образца "Прогноз СМР.xlsx" (18 уникальных строк) —
# элевационный суффикс "до отм. +X" у Панели намеренно ОТБРОШЕН при
# сопоставлении с подтипом: он избыточен с колонкой Этаж (проверено на
# примере — 3 строки на разных этажах с одним и тем же suffix-набором
# субтипа). Строка, которой здесь нет, попадает в unmatched_type_subtype
# сводки импорта — не гадаем.
TYPE_SUBTYPE_ALIASES: dict[str, tuple[str, str]] = {
    "Колонна нижняя": ("Колонна", "нижняя"),
    "Колонна верхняя": ("Колонна", "верхняя"),
    "Колонна средняя нижний ярус": ("Колонна", "средняя нижний ярус"),
    "Колонна средняя верхний ярус": ("Колонна", "средняя верхний ярус"),
    "Ригель периметральный": ("Ригель", "периметральный"),
    "Ригель на отм. +15.000": ("Ригель", "на отм. +15.000"),
    "Ригель на отм. +25.800": ("Ригель", "на отм. +25.800"),
    "Ригель на отм. +34.700": ("Ригель", "на отм. +34.700"),
    "Ригель на отм. +39.200": ("Ригель", "на отм. +39.200"),
    "Плита перекрытия на отм. +15.000": ("Плита перекрытия", "на отм. +15.000"),
    "Плита перекрытия на отм. +25.800": ("Плита перекрытия", "на отм. +25.800"),
    "Плита перекрытия на отм. +34.700": ("Плита перекрытия", "на отм. +34.700"),
    "Плита перекрытия на отм. +39.200": ("Плита перекрытия", "на отм. +39.200"),
    "Панелей лифтовой шахты до отм. +15.000": ("Панель", "ЛифтоваяШахта"),
    "Панелей лифтовой шахты до отм. +25.800": ("Панель", "ЛифтоваяШахта"),
    "Панелей лифтовой шахты до отм. +34.700": ("Панель", "ЛифтоваяШахта"),
    "Панелей шахты подъемника до отм. +15.000": ("Панель", "ШахтаПодъемника"),
    "Панелей шахты подъемника до отм. +25.800": ("Панель", "ШахтаПодъемника"),
    "Панелей шахты подъемника до отм. +34.700": ("Панель", "ШахтаПодъемника"),
}

_CRANE_NUMBER_RE = re.compile(r"(\d+)")
_TRAILING_NUMBER_RE = re.compile(r"(\d+)\s*$")
# "Этаж" в реальном образце — "1 этаж", число ПЕРЕД словом, а не в конце
# строки (в отличие от "Стоянка 01"), поэтому здесь отдельный шаблон:
# _TRAILING_NUMBER_RE на таком значении не находил ничего и молча уводил
# в skipped_rows ВСЕ строки файла (живой репорт, см. Docs/backlog.md).
_ANY_NUMBER_RE = re.compile(r"(\d+)")


def _manual_fields(raw) -> set:
    """elements.manual_fields — JSON-список имён полей, правленных руками
    (см. PATCH /elements/{id}/fields). Битое значение трактуем как «правок
    нет»: импорт не должен падать из-за одной испорченной строки."""
    if not raw:
        return set()
    try:
        value = json.loads(raw)
    except (ValueError, TypeError):
        return set()
    return set(value) if isinstance(value, list) else set()


class ScheduleImportError(Exception):
    def __init__(self, status_code: int, message: str):
        self.status_code = status_code
        self.message = message
        super().__init__(message)


def _normalize_type_subtype(raw: str) -> str:
    # Запятая отбрасывается вместе со схлопыванием пробелов: MS Project
    # экспортирует одну и ту же работу и как "Колонна средняя, нижний ярус",
    # и как "Колонна средняя нижний ярус" (реальные два образца) — это
    # НЕ угадывание типа, а нормализация записи перед точным поиском по
    # закрытому списку алиасов.
    return re.sub(r"\s+", " ", raw.replace(",", " ").strip())


def _extract_number(raw, pattern: re.Pattern) -> Optional[int]:
    if raw is None:
        return None
    m = pattern.search(str(raw).strip())
    return int(m.group(1)) if m else None


# В реальном образце "Начало"/"Окончание" — ТЕКСТ вида "Ср 01.07.26"
# (день недели + дата, двузначный год), не datetime-значение — так
# MS Project экспортирует даты задач по умолчанию. День недели —
# избыточная информация, отбрасывается регэкспом перед разбором.
_WEEKDAY_PREFIX_RE = re.compile(r"^[А-Яа-я]{2}\s+")


def _to_date_iso(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _WEEKDAY_PREFIX_RE.sub("", str(value).strip())
    for fmt in ("%d.%m.%y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_schedule_xlsx(content: bytes) -> dict:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise ScheduleImportError(422, "Файл повреждён или не является корректным .xlsx")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ScheduleImportError(422, "Пустой файл")

    header = [str(h).strip() if h is not None else "" for h in header]
    missing = [h for h in REQUIRED_HEADERS if h not in header]
    вид_работ = next((h for h in WORK_KIND_HEADERS if h in header), None)
    if вид_работ is None:
        missing = [" или ".join(f"«{h}»" for h in WORK_KIND_HEADERS)] + missing
    if missing:
        raise ScheduleImportError(422, f"В файле нет обязательных колонок: {', '.join(missing)}")
    col = {name: idx for idx, name in enumerate(header)}

    def get(row, name):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    parsed = []
    skipped_rows = []
    row_num = 1
    for row in rows_iter:
        row_num += 1
        type_subtype_raw = get(row, вид_работ)
        if type_subtype_raw is None or not str(type_subtype_raw).strip():
            continue  # пустая строка — пропуск, не ошибка (не конец таблицы, просто разделитель)

        normalized = _normalize_type_subtype(str(type_subtype_raw))
        element_type, subtype = TYPE_SUBTYPE_ALIASES.get(normalized, (None, None))

        start_date = _to_date_iso(get(row, "Начало"))
        end_date = _to_date_iso(get(row, "Окончание"))
        crane_number = _extract_number(get(row, "Кран"), _CRANE_NUMBER_RE)
        stance_number = _extract_number(get(row, "Стоянка"), _TRAILING_NUMBER_RE)
        floor_number = _extract_number(get(row, "Этаж"), _ANY_NUMBER_RE)
        zakhvatka_number = _extract_number(get(row, "Захватка"), _TRAILING_NUMBER_RE)

        problems = []
        if element_type is None:
            problems.append(f"тип/подтип «{normalized}» не распознан")
        if end_date is None:
            problems.append("не удалось разобрать дату «Окончание»")
        if crane_number is None:
            problems.append("не удалось разобрать номер крана")
        if stance_number is None:
            problems.append("не удалось разобрать номер стоянки")
        if floor_number is None:
            problems.append("не удалось разобрать номер этажа")

        if problems:
            skipped_rows.append(f"Строка {row_num}: {'; '.join(problems)}")
            continue

        parsed.append(
            {
                "row": row_num,
                "type_subtype_raw": normalized,
                "element_type": element_type,
                "subtype": subtype,
                "project_smr_start_date": start_date,
                "project_delivery_date": end_date,
                "crane_number": crane_number,
                "stance_number": stance_number,
                "floor": floor_number,
                "zakhvatka_number": zakhvatka_number,
            }
        )

    return {"rows": parsed, "skipped_rows": skipped_rows}


def import_schedule(conn, parsed: dict, user=None, request_id: str = None,
                    object_id: Optional[int] = None, kind: str = "baseline",
                    source_file: Optional[str] = None, note: Optional[str] = None) -> dict:
    """Загрузка графика.

    object_id — объект, к которому относится файл (2026-08-14). До этой даты
    сопоставление шло по ВСЕЙ базе: пока здание одно — работает, со вторым
    даты уехали бы в чужой дом молча. Без объекта (None) поведение прежнее,
    и это оставлено только ради вызовов из старых скриптов.

    kind — 'baseline' (базовый, директивный) или 'current' (актуализированный
    прогноз). Оба сохраняются ВЕРСИЕЙ (schedule_versions), которая копится, —
    по ним видно, как менялся прогноз. Разница одна и важная: базовый, кроме
    того, проставляет даты в сами изделия (project_smr_start_date /
    project_delivery_date), потому что на этих полях держится вся остальная
    система — фильтры, подписи, отчёты, аналитическая справка. Актуализация
    полей изделия НЕ трогает: она отвечает на вопрос «насколько отстаём», а
    не «когда надо».
    """
    if kind not in ("baseline", "current"):
        raise ScheduleImportError(422, "Неизвестный вид графика")
    if kind == "current" and object_id is None:
        # Версия обязана принадлежать объекту (внешний ключ), и «график
        # неизвестно чего» — не то, что стоит уметь заводить.
        raise ScheduleImportError(422, "Для актуализированного графика нужно выбрать объект")

    rows = parsed["rows"]
    matched_elements_total = 0
    unmatched_blocks: list[str] = []
    touched_element_ids: set[int] = set()
    # Даты версии: element_id -> (начало, окончание). Пишутся ВСЕГДА, и для
    # базового тоже — иначе базовую версию не с чем было бы сравнивать после
    # ручной правки полей изделия.
    version_dates: dict[int, tuple] = {}
    # Обе даты СМР админ может поправить руками в форме элемента — ровно
    # потому, что для части блоков строки в графике нет вовсе (живой разбор:
    # 18 «Ригелей периметральных» остались без дат навсегда). Такое поле
    # помечено в elements.manual_fields, и импорт его НЕ перезаписывает —
    # иначе ручная правка исчезала бы молча при следующей загрузке графика.
    # Считаем такие случаи и показываем в сводке: не применить значение
    # молча — так же плохо, как молча затереть.
    manual_kept: dict[str, int] = {}

    объектное = "AND e.object_id = ?" if object_id is not None else ""
    for row in rows:
        candidates = conn.execute(
            f"""
            SELECT e.id, e.manual_fields, zc.name AS crane_name, zs.name AS stance_name,
                   zz.name AS zakhvatka_name
            FROM elements e
            LEFT JOIN zones zc ON zc.id = e.zone_crane_id AND e.zone_crane_status = 'matched'
            LEFT JOIN zones zs ON zs.id = e.zone_stance_id AND e.zone_stance_status = 'matched'
            LEFT JOIN zones zz ON zz.id = e.zone_zakhvatka_id AND e.zone_zakhvatka_status = 'matched'
            WHERE e.floor = ? AND e.element_type = ? AND e.subtype IS ? {объектное}
            """,
            (row["floor"], row["element_type"], row["subtype"])
            + ((object_id,) if object_id is not None else ()),
        ).fetchall()

        def подходит(c):
            if _extract_number(c["crane_name"], _CRANE_NUMBER_RE) != row["crane_number"]:
                return False
            if _extract_number(c["stance_name"], _TRAILING_NUMBER_RE) != row["stance_number"]:
                return False
            # Захватка сверяется, только если она разобрана в строке файла:
            # старые выгрузки без внятной колонки должны грузиться, как
            # грузились (см. заголовок модуля).
            if row.get("zakhvatka_number") is not None:
                if _extract_number(c["zakhvatka_name"], _TRAILING_NUMBER_RE) != row["zakhvatka_number"]:
                    return False
            return True

        matched = [c for c in candidates if подходит(c)]

        if not matched:
            unmatched_blocks.append(
                f"Захватка {row.get('zakhvatka_number') or '—'}/Кран {row['crane_number']}/"
                f"Стоянка {row['stance_number']}/Этаж {row['floor']}/"
                f"{row['type_subtype_raw']} — ни один элемент не найден"
            )
            continue

        for c in matched:
            version_dates[c["id"]] = (row["project_smr_start_date"], row["project_delivery_date"])

        # Даты в САМИ изделия проставляет только базовый график: поля
        # изделия — это директивные сроки, «когда надо». Актуализированный
        # живёт версией и полей не касается (см. docstring функции).
        if kind == "baseline":
            # Поля обновляются ПООТДЕЛЬНОСТИ: одну дату могли править руками, а
            # вторую нет, и общий UPDATE обеими колонками затёр бы правку.
            updates: dict[str, list[tuple]] = {"project_delivery_date": [], "project_smr_start_date": []}
            for c in matched:
                manual = _manual_fields(c["manual_fields"])
                for field in updates:
                    if field in manual:
                        manual_kept[field] = manual_kept.get(field, 0) + 1
                        continue
                    updates[field].append((row[field], c["id"]))
                    touched_element_ids.add(c["id"])

            for field, payload in updates.items():
                if payload:
                    conn.executemany(
                        f"UPDATE elements SET {field} = ?, updated_at = datetime('now') WHERE id = ?",
                        payload,
                    )
        matched_elements_total += len(matched)

    # Событие на каждое изделие, которому реально проставили даты
    # (2026-08-03): импорт графика — единственный способ, которым эти даты
    # появляются массово, и без поэлементной записи их изменение не видно
    # ни в истории изменений изделия, ни в фильтре «Изменения». Сводка
    # операции пишется вызывающим и связана общим request_id.
    if touched_element_ids:
        for снимок in conn.execute(
            f"SELECT id, element_type, subtype, mark, project_smr_start_date, project_delivery_date "
            f"FROM elements WHERE id IN ({','.join('?' * len(touched_element_ids))})",
            tuple(touched_element_ids),
        ).fetchall():
            activity.log(
                "schedule_import", user=user, entity_type="element", entity_id=снимок["id"],
                element_type=снимок["element_type"], subtype=снимок["subtype"], mark=снимок["mark"],
                new_value=f"Дата начала СМР: {снимок['project_smr_start_date'] or '—'}; "
                          f"Дата завершения СМР: {снимок['project_delivery_date'] or '—'}",
                request_id=request_id,
            )

    version_id = None
    if object_id is not None and version_dates:
        version_id = save_version(
            conn, object_id, kind, version_dates,
            source_file=source_file, user=user, note=note,
        )

    conn.commit()

    return {
        "rows_processed": len(rows),
        "rows_skipped": len(parsed["skipped_rows"]),
        "skipped_rows": parsed["skipped_rows"][:50],
        "elements_updated": len(touched_element_ids),
        "elements_in_version": len(version_dates),
        "unmatched_blocks": unmatched_blocks,
        "manual_kept": manual_kept,
        "kind": kind,
        "version_id": version_id,
    }


def save_version(conn, object_id: int, kind: str, dates: dict,
                 source_file: Optional[str] = None, user=None,
                 note: Optional[str] = None, origin: str = "import") -> int:
    """Сохранить набор дат версией графика. Возвращает id версии.

    Базовая версия у объекта ОДНА (уникальный индекс в схеме): повторная
    загрузка базового графика заменяет её целиком, а не копит вторую. Это
    решение пользователя: базовый график — то, с чем сравнивают, и двух
    оснований для сравнения быть не может. Актуализированные, наоборот,
    копятся — по ним и видно, как менялся прогноз.
    """
    title = ("Базовый график" if kind == "baseline"
             else f"Актуализация от {_today_ru()}")
    if kind == "baseline":
        conn.execute("DELETE FROM schedule_versions WHERE object_id = ? AND kind = 'baseline'",
                     (object_id,))
    cur = conn.execute(
        "INSERT INTO schedule_versions (object_id, kind, title, source_file, origin, loaded_by, note) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (object_id, kind, title, source_file, origin,
         user["id"] if user is not None else None, note),
    )
    version_id = cur.lastrowid
    conn.executemany(
        "INSERT INTO schedule_version_dates (version_id, element_id, smr_start_date, smr_end_date) "
        "VALUES (?, ?, ?, ?)",
        [(version_id, eid, начало, окончание) for eid, (начало, окончание) in dates.items()],
    )
    return version_id


def _today_ru() -> str:
    return date.today().strftime("%d.%m.%Y")
