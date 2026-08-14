"""
Расчёт графика СМР внутри системы (2026-08-14).

Откуда взялось. До сих пор график считался СНАРУЖИ: заказчик выгружал из
системы ведомость объёмов, дообогащал её в Excel-обработке Power Query
двумя своими величинами (темп монтажа и порядок работ), гнал три листа в
MS Project и запускал там макрос. Обратно в систему приезжал результат —
файл с датами. Механика расчёта нигде не была описана: движком служил сам
MS Project.

Механика восстановлена по присланным файлам («!_Обработка данных из
WEB.xlsx» + «grafik_ms_project (москвич 14.08.2026).xlsx», разбор —
Docs/requirements-2026-08-14.md) и проверена на них же: из 663 работ 245
совпали день-в-день, остальные разошлись ровно на сутки — расхождение
объясняется тем, как MS Project округляет дробные длительности до
отображаемых дат.

Алгоритм целиком:

1. длительность работы = количество изделий / темп (в сутках, дробная);
2. внутри одного фронта (кран + стоянка + этаж) виды работ идут в порядке
   `schedule_work_kinds.order_no` — колонна нижняя, ригель периметральный,
   колонна средняя, ригель, плита и так вверх по отметкам;
3. фронты одного крана идут в порядке `schedule_flow.order_no`, встык, без
   простоев;
4. краны работают независимо и параллельно, каждый со своей даты старта;
5. календарь СЕМИДНЕВНЫЙ — выходных нет (в графике заказчика есть и
   субботы, и воскресенья).

Пересчёт от факта (требование того же совещания: «запланировано 10 колонн,
смонтировал 20 — остаток пересчитать»): изделия, уже имеющие фактическую
дату монтажа, из количества фронта ВЫЧИТАЮТСЯ, и остаток раскладывается
дальше. Смонтированный фронт целиком просто исчезает из очереди.

Результат сохраняется версией графика (`origin = 'calc'`) — той же
таблицей, что и загруженный извне, и дальше живёт как обычная актуализация:
участвует в отклонении и в кривой прогноза.
"""

import io
import re
import sqlite3
from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel

from app import activity
from app.access import assert_object_feature
from app.auth import get_current_user
from app.db import get_connection
from app.models import Status
from app.schedule_import import save_version

router = APIRouter(prefix="/schedule-calc", tags=["schedule"])

# Темп, применяемый к виду работ, для которого его не задали. Не ноль и не
# «пропустить»: пропуск молча выкинул бы изделия из графика, а ноль дал бы
# бесконечную длительность. Единица честно означает «сутки на изделие» —
# заметно медленно, и в сводке расчёта такие виды работ перечислены прямо.
DEFAULT_RATE = 1.0


def _work_kinds(conn: sqlite3.Connection, object_id: int) -> dict:
    """(тип, подтип) → {"rate", "order"}. Подтип NULL и пустая строка — одно
    и то же: в модели у части изделий подтипа нет вовсе."""
    out = {}
    for r in conn.execute(
        "SELECT element_type, subtype, rate_per_day, order_no FROM schedule_work_kinds "
        "WHERE object_id = ?", (object_id,)
    ):
        out[(r["element_type"], r["subtype"] or None)] = {
            "rate": r["rate_per_day"], "order": r["order_no"],
        }
    return out


def _flow(conn: sqlite3.Connection, object_id: int) -> dict:
    """(кран, стоянка, этаж) → порядковый номер фронта."""
    return {
        (r["crane_name"], r["stance_name"], r["floor"]): r["order_no"]
        for r in conn.execute(
            "SELECT crane_name, stance_name, floor, order_no FROM schedule_flow WHERE object_id = ?",
            (object_id,),
        )
    }


def build_wbs(conn: sqlite3.Connection, object_id: int, skip_installed: bool = True) -> list:
    """Ведомость фронтов работ: по одной строке на
    (кран, стоянка, этаж, тип, подтип) с количеством изделий и их id.

    Это ровно тот лист `01_WBS`, который заказчик собирал Power Query, — с
    той разницей, что количество берётся прямо из модели, а не из выгрузки.

    skip_installed — пересчёт от факта: смонтированные изделия из количества
    исключаются (см. заголовок модуля).
    """
    условие = "" if not skip_installed else f"AND e.current_status <> '{Status.INSTALLED.value}'"
    rows = conn.execute(
        f"""
        SELECT e.id, e.element_type, e.subtype, e.floor,
               zc.name AS crane, zs.name AS stance, zz.name AS zakhvatka
        FROM elements e
        LEFT JOIN zones zc ON zc.id = e.zone_crane_id AND e.zone_crane_status = 'matched'
        LEFT JOIN zones zs ON zs.id = e.zone_stance_id AND e.zone_stance_status = 'matched'
        LEFT JOIN zones zz ON zz.id = e.zone_zakhvatka_id AND e.zone_zakhvatka_status = 'matched'
        WHERE e.object_id = ? AND e.is_current = 1 {условие}
        """,
        (object_id,),
    ).fetchall()

    фронты = {}
    for r in rows:
        if not r["crane"] or not r["stance"] or r["floor"] is None:
            continue  # изделие без привязки к зонам в график не встаёт — его негде разместить
        ключ = (r["crane"], r["stance"], r["floor"], r["element_type"], r["subtype"] or None)
        фронты.setdefault(ключ, []).append(r["id"])
    return [
        {"crane": k[0], "stance": k[1], "floor": k[2], "element_type": k[3], "subtype": k[4],
         "element_ids": ids, "quantity": len(ids)}
        for k, ids in фронты.items()
    ]


def calculate(conn: sqlite3.Connection, object_id: int, start_date: str,
              skip_installed: bool = True) -> dict:
    """Посчитать даты. Возвращает {"dates": {element_id: (начало, конец)},
    "warnings": [...], "fronts": N} — сохранение отдельным шагом, чтобы
    расчёт можно было показать до записи."""
    виды = _work_kinds(conn, object_id)
    поток = _flow(conn, object_id)
    if not виды:
        raise ScheduleCalcError("Не заданы темпы и порядок монтажа — считать нечем")
    if not поток:
        raise ScheduleCalcError("Не задан поток (очередь стоянок и этажей у кранов)")

    wbs = build_wbs(conn, object_id, skip_installed=skip_installed)
    if not wbs:
        raise ScheduleCalcError("Нет изделий с привязкой к крану, стоянке и этажу")

    предупреждения = []
    без_темпа = sorted({
        f"{w['element_type']} {w['subtype'] or ''}".strip() for w in wbs
        if (виды.get((w["element_type"], w["subtype"])) or {}).get("rate") in (None, 0)
    })
    if без_темпа:
        предупреждения.append(
            f"Темп не задан ({DEFAULT_RATE} изд./сутки принят по умолчанию): "
            + ", ".join(без_темпа[:10]) + ("…" if len(без_темпа) > 10 else "")
        )
    вне_потока = sorted({
        f"{w['crane']} · {w['stance']} · этаж {w['floor']}" for w in wbs
        if (w["crane"], w["stance"], w["floor"]) not in поток
    })
    if вне_потока:
        предупреждения.append(
            "Фронты вне потока (встали в конец очереди своего крана): "
            + ", ".join(вне_потока[:10]) + ("…" if len(вне_потока) > 10 else "")
        )

    по_кранам = {}
    for w in wbs:
        по_кранам.setdefault(w["crane"], []).append(w)

    начало_проекта = date.fromisoformat(start_date[:10])
    даты = {}
    for кран, работы in по_кранам.items():
        # Фронты в порядке потока; не найденные в потоке — в конец (и об этом
        # сказано в предупреждениях выше), внутри — по порядку видов работ.
        def ключ_сортировки(w):
            фронт = поток.get((w["crane"], w["stance"], w["floor"]), 10 ** 6)
            вид = (виды.get((w["element_type"], w["subtype"])) or {}).get("order")
            return (фронт, вид if вид is not None else 10 ** 6,
                    w["element_type"], w["subtype"] or "")

        смещение = 0.0
        for w in sorted(работы, key=ключ_сортировки):
            темп = (виды.get((w["element_type"], w["subtype"])) or {}).get("rate") or DEFAULT_RATE
            длительность = w["quantity"] / темп
            начало = начало_проекта + timedelta(days=смещение)
            конец = начало_проекта + timedelta(days=смещение + длительность)
            смещение += длительность
            # Все изделия фронта получают даты САМОГО ФРОНТА, а не свои
            # собственные внутри него: строка графика описывает блок работ,
            # и дробить её на изделия — придумывать точность, которой в
            # исходных данных нет (ровно так же устроен импорт из MS Project).
            # date + timedelta с дробными сутками отбрасывает часы — то есть
            # день ОТОБРАЖАЕТСЯ округлённым вниз, а накопление смещения
            # остаётся дробным (смещение — float). Ровно так же ведёт себя
            # MS Project, чем и объясняются расхождения в сутки при сверке.
            for eid in w["element_ids"]:
                даты[eid] = (начало.isoformat(), конец.isoformat())

    return {"dates": даты, "warnings": предупреждения, "fronts": len(wbs), "elements": len(даты)}


class ScheduleCalcError(Exception):
    def __init__(self, message: str):
        super().__init__(message)
        self.message = message


# ---------------------------------------------------------------- эндпоинты

# ---------------------------------------- загрузка исходных данных файлом
#
# Заказчик ведёт эти три величины в своей Excel-обработке (Power Query,
# «!_Обработка данных из WEB.xlsx») и присылает её целиком. Набивать полтора
# десятка темпов и две сотни строк потока руками, когда файл уже есть, —
# работа ради работы, поэтому форма умеет читать его напрямую.
#
# Колонки ищутся ПО ЗАГОЛОВКАМ, а не по номерам: на листе `hide` полезные
# данные начинаются с четвёртой колонки, а первая занята другим списком, и
# любой сдвиг при правке файла молча увёл бы разбор не туда.
#
# Читаются четыре листа, каждый по отдельности и каждый необязателен:
#   `hide`           — темп и порядок монтажа (основной источник);
#   `02_Technology`  — порядок монтажа (если на `hide` его нет);
#   `01_WBS`         — темп (запасной источник: там он проставлен построчно);
#   `03_Flow`        — поток: кран, стоянка, этаж, порядок.
# Файл без единого узнанного листа отвергается: молча загрузить «ничего» —
# худший исход, человек будет думать, что данные приняты.

WORK_KIND_SHEETS = ("hide", "02_Technology", "01_WBS")
FLOW_SHEET = "03_Flow"


def _find_header(ws, нужные: set, предел: int = 20):
    """Строка заголовков и карта «имя колонки → СПИСОК индексов».

    Именно список: на листе `hide` заголовок «Элемент» встречается ДВАЖДЫ —
    в первой колонке лежит посторонний алфавитный список, а рабочие данные
    начинаются с четвёртой. Взять первое попавшееся вхождение значит увести
    темпы и порядок к чужим видам работ (поймано при проверке на файле
    заказчика: «Колонна верхняя» получала темп «Колонны нижней»).

    Возвращает (номер строки, карта) или (None, None).
    """
    for номер, строка in enumerate(ws.iter_rows(values_only=True), start=1):
        if номер > предел:
            break
        карта = {}
        for i, значение in enumerate(строка):
            if значение is None:
                continue
            имя = re.sub(r"\s+", " ", str(значение)).strip()
            if имя:
                карта.setdefault(имя, []).append(i)
        if нужные <= set(карта):
            return номер, карта
    return None, None


def _column_left_of(карта: dict, имя: str, границы: list) -> Optional[int]:
    """Колонка `имя`, ближайшая СЛЕВА к своим данным.

    Из нескольких одноимённых колонок берётся последняя перед первой из
    колонок-«границ» (Темп, Порядок): в таблице подпись стоит слева от
    чисел, к которым относится.
    """
    индексы = карта.get(имя) or []
    if not индексы:
        return None
    предел = min(границы) if границы else None
    if предел is None:
        return индексы[0]
    слева = [i for i in индексы if i < предел]
    return max(слева) if слева else индексы[0]


def _rows_after(ws, номер_заголовка: int):
    for номер, строка in enumerate(ws.iter_rows(values_only=True), start=1):
        if номер > номер_заголовка:
            yield строка


def parse_inputs_xlsx(content: bytes) -> dict:
    """Разобрать обработку заказчика. Ничего не пишет в базу."""
    from openpyxl import load_workbook

    from app.schedule_import import TYPE_SUBTYPE_ALIASES, _normalize_type_subtype

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise ScheduleCalcError("Файл повреждён или не является корректным .xlsx")

    предупреждения = []
    не_узнаны = set()
    виды = {}     # (тип, подтип) -> {"rate", "order"}
    листы = []

    def вид_работ(значение):
        """Текст вида работ → (тип, подтип) по тому же закрытому списку, что
        и импорт графика: нераспознанное не угадывается, а перечисляется."""
        имя = _normalize_type_subtype(str(значение))
        пара = TYPE_SUBTYPE_ALIASES.get(имя)
        if пара is None:
            не_узнаны.add(имя)
        return пара

    for имя_листа in WORK_KIND_SHEETS:
        if имя_листа not in wb.sheetnames:
            continue
        ws = wb[имя_листа]
        # На «hide» и «02_Technology» колонка называется «Элемент», на
        # «01_WBS» — тоже; различаются наборы соседних колонок.
        номер, карта = _find_header(ws, {"Элемент"})
        if номер is None:
            continue
        кол_темп = (карта.get("Темп") or [None])[0]
        имя_порядка = next((k for k in ("Порядок монтажа", "Порядок") if k in карта), None)
        кол_порядок = карта[имя_порядка][0] if имя_порядка else None
        if кол_темп is None and кол_порядок is None:
            continue
        # Колонка с названием вида работ — та, что относится к этим числам
        # (ближайшая слева), а не первая одноимённая на листе.
        кол_вид = _column_left_of(карта, "Элемент",
                                  [i for i in (кол_темп, кол_порядок) if i is not None])
        if кол_вид is None:
            continue
        прочитано = 0
        for строка in _rows_after(ws, номер):
            сырое = строка[кол_вид] if кол_вид < len(строка) else None
            if сырое is None or not str(сырое).strip():
                continue
            пара = вид_работ(сырое)
            if пара is None:
                continue
            запись = виды.setdefault(пара, {"rate": None, "order": None})
            if кол_темп is not None and кол_темп < len(строка):
                темп = строка[кол_темп]
                if isinstance(темп, (int, float)) and темп > 0 and запись["rate"] is None:
                    запись["rate"] = float(темп)
            if кол_порядок is not None and кол_порядок < len(строка):
                порядок = строка[кол_порядок]
                if isinstance(порядок, (int, float)) and запись["order"] is None:
                    запись["order"] = int(порядок)
            прочитано += 1
        if прочитано:
            листы.append(f"«{имя_листа}» — строк {прочитано}")

    поток = []
    if FLOW_SHEET in wb.sheetnames:
        ws = wb[FLOW_SHEET]
        номер, карта = _find_header(ws, {"Кран", "Стоянка", "Этаж", "Порядок"})
        if номер is not None:
            for строка in _rows_after(ws, номер):
                def значение(имя):
                    i = карта[имя][0]
                    return строка[i] if i < len(строка) else None
                кран, стоянка, этаж, порядок = (значение("Кран"), значение("Стоянка"),
                                                значение("Этаж"), значение("Порядок"))
                if not кран or not стоянка or этаж is None:
                    continue
                try:
                    поток.append({
                        "crane_name": str(кран).strip(),
                        "stance_name": str(стоянка).strip(),
                        "floor": int(str(этаж).strip()),
                        "order_no": int(порядок) if isinstance(порядок, (int, float)) else None,
                    })
                except (TypeError, ValueError):
                    continue
            if поток:
                листы.append(f"«{FLOW_SHEET}» — фронтов {len(поток)}")

    if not виды and not поток:
        raise ScheduleCalcError(
            "В файле не нашлось ни темпов, ни порядка, ни потока. Ожидаются листы "
            "«hide» или «02_Technology» (колонки «Элемент», «Темп», «Порядок монтажа») "
            "и «03_Flow» (колонки «Кран», «Стоянка», «Этаж», «Порядок»).")
    if не_узнаны:
        предупреждения.append(
            "Не распознаны виды работ (строки пропущены): "
            + ", ".join(sorted(не_узнаны)[:10])
            + (f" и ещё {len(не_узнаны) - 10}" if len(не_узнаны) > 10 else ""))

    return {
        "work_kinds": [
            {"element_type": t, "subtype": st,
             "rate_per_day": v["rate"], "order_no": v["order"]}
            for (t, st), v in виды.items()
        ],
        "flow": поток,
        "sheets": листы,
        "warnings": предупреждения,
    }


@router.post("/inputs/parse")
def parse_inputs(object_id: int = Form(...), file: UploadFile = File(...),
                 user: sqlite3.Row = Depends(get_current_user)):
    """Разобрать файл обработки и вернуть значения В ФОРМУ, не записывая.

    Не сохраняем сразу намеренно: файл заказчика описывает его расчёт, а не
    обязательно то, что должно лежать в системе, — человек должен увидеть
    числа рядом с количествами из модели и нажать «Сохранить».
    """
    conn = get_connection()
    try:
        assert_object_feature(conn, user, object_id, "schedule", "write")
    finally:
        conn.close()
    from app.upload_limits import read_upload_limited
    content = read_upload_limited(file.file)
    try:
        return parse_inputs_xlsx(content)
    except ScheduleCalcError as e:
        raise HTTPException(status_code=422, detail=e.message)


@router.get("/inputs")
def get_inputs(object_id: int, user: sqlite3.Row = Depends(get_current_user)):
    """Исходные данные расчёта: виды работ (темп и порядок) и поток.

    Вместе с ними отдаются РЕАЛЬНЫЕ виды работ и фронты объекта, взятые из
    модели: без этого таблицу пришлось бы набивать вручную, гадая, какие
    сочетания вообще есть. Строка, которой в модели нет, помечается —
    зона могла исчезнуть при переимпорте чертежа, и молча пропасть из
    настройки такая строка не должна.
    """
    conn = get_connection()
    try:
        assert_object_feature(conn, user, object_id, "schedule", "read")
        виды = _work_kinds(conn, object_id)
        поток = _flow(conn, object_id)
        wbs = build_wbs(conn, object_id, skip_installed=False)

        в_модели_виды = {}
        в_модели_фронты = {}
        for w in wbs:
            в_модели_виды[(w["element_type"], w["subtype"])] = \
                в_модели_виды.get((w["element_type"], w["subtype"]), 0) + w["quantity"]
            ключ = (w["crane"], w["stance"], w["floor"])
            в_модели_фронты[ключ] = в_модели_фронты.get(ключ, 0) + w["quantity"]

        from app.reports import natural_key
        строки_видов = []
        for ключ in sorted(set(виды) | set(в_модели_виды),
                           key=lambda k: (natural_key(k[0]), natural_key(k[1] or ""))):
            запись = виды.get(ключ) or {}
            строки_видов.append({
                "element_type": ключ[0], "subtype": ключ[1],
                "rate_per_day": запись.get("rate"), "order_no": запись.get("order"),
                "quantity": в_модели_виды.get(ключ, 0),
                "in_model": ключ in в_модели_виды,
            })
        строки_потока = []
        for ключ in sorted(set(поток) | set(в_модели_фронты),
                           key=lambda k: (natural_key(k[0]), natural_key(k[1]), k[2])):
            строки_потока.append({
                "crane_name": ключ[0], "stance_name": ключ[1], "floor": ключ[2],
                "order_no": поток.get(ключ),
                "quantity": в_модели_фронты.get(ключ, 0),
                "in_model": ключ in в_модели_фронты,
            })
        return {"work_kinds": строки_видов, "flow": строки_потока}
    finally:
        conn.close()


class WorkKindIn(BaseModel):
    element_type: str
    subtype: Optional[str] = None
    rate_per_day: Optional[float] = None
    order_no: Optional[int] = None


class FlowIn(BaseModel):
    crane_name: str
    stance_name: str
    floor: int
    order_no: Optional[int] = None


class InputsIn(BaseModel):
    object_id: int
    work_kinds: list[WorkKindIn]
    flow: list[FlowIn]


@router.put("/inputs")
def put_inputs(body: InputsIn, user: sqlite3.Row = Depends(get_current_user)):
    """Замена обеих таблиц целиком. Целиком, а не построчно: форма правит их
    как единый лист (так же они и жили в Excel у заказчика), и построчное
    сохранение потребовало бы отдельно решать судьбу удалённых строк."""
    conn = get_connection()
    try:
        assert_object_feature(conn, user, body.object_id, "schedule", "write")
        conn.execute("DELETE FROM schedule_work_kinds WHERE object_id = ?", (body.object_id,))
        conn.executemany(
            "INSERT INTO schedule_work_kinds (object_id, element_type, subtype, rate_per_day, order_no) "
            "VALUES (?, ?, ?, ?, ?)",
            [(body.object_id, w.element_type, w.subtype or None, w.rate_per_day, w.order_no)
             for w in body.work_kinds],
        )
        conn.execute("DELETE FROM schedule_flow WHERE object_id = ?", (body.object_id,))
        conn.executemany(
            "INSERT INTO schedule_flow (object_id, crane_name, stance_name, floor, order_no) "
            "VALUES (?, ?, ?, ?, ?)",
            [(body.object_id, f.crane_name, f.stance_name, f.floor, f.order_no or 0)
             for f in body.flow],
        )
        conn.commit()
        activity.log("schedule_inputs_save", user=user, entity_type="object", entity_id=body.object_id,
                     new_value=f"видов работ {len(body.work_kinds)}, фронтов {len(body.flow)}")
        return {"ok": True, "work_kinds": len(body.work_kinds), "flow": len(body.flow)}
    finally:
        conn.close()


class CalcIn(BaseModel):
    object_id: int
    # Дата старта работ. Обязательна и задаётся человеком: вывести её из
    # данных нельзя — «сегодня» дало бы график, начинающийся задним числом
    # на объекте, где монтаж ещё не начинался.
    start_date: str
    skip_installed: bool = True
    # save=False — «показать, что получится», без записи версии.
    save: bool = True
    note: Optional[str] = None


@router.post("")
def run_calc(body: CalcIn, user: sqlite3.Row = Depends(get_current_user)):
    conn = get_connection()
    try:
        assert_object_feature(conn, user, body.object_id, "schedule", "write")
        try:
            итог = calculate(conn, body.object_id, body.start_date, body.skip_installed)
        except ScheduleCalcError as e:
            raise HTTPException(status_code=422, detail=e.message)
        version_id = None
        if body.save:
            version_id = save_version(
                conn, body.object_id, "current", итог["dates"],
                source_file=None, user=user, origin="calc",
                note=body.note or f"Расчёт системы от {body.start_date}",
            )
            conn.commit()
            activity.log("schedule_calc", user=user, entity_type="object", entity_id=body.object_id,
                         new_value=f"расчёт графика: фронтов {итог['fronts']}, "
                                   f"изделий {итог['elements']}, старт {body.start_date}")
        return {
            "fronts": итог["fronts"], "elements": итог["elements"],
            "warnings": итог["warnings"], "version_id": version_id,
            # Границы посчитанного — чтобы форма сразу показала, во что
            # уложился график, не выкачивая девять тысяч дат.
            "first_date": min((d[0] for d in итог["dates"].values()), default=None),
            "last_date": max((d[1] for d in итог["dates"].values()), default=None),
        }
    finally:
        conn.close()
