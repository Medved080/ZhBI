"""
Массовая правка ИСТОРИИ СТАТУСОВ через Excel (2026-08-01, живой запрос;
формат переделан 2026-08-03 по живому репорту).

Второй режим той же формы, что и правка реквизитов
(app/element_bulk_edit.py): выгрузить -> поправить снаружи -> загрузить ->
отметить флажками, что применить. Отдаёт и принимает те же структуры
(columns/elements/changes/rejected), поэтому табличный экран подтверждения
переиспользуется целиком, без второй реализации.

**Строка на ЗАПИСЬ истории, а не на элемент.** Первая версия формата была
матрицей: строка на элемент, колонка на каждый статус, в ячейке дата. Живой
репорт закрыл её: в матрице не видно ни автора изменения, ни времени
установки статуса — а это ровно то, ради чего историю и открывают. Матрица
физически не могла их показать: у элемента семь статусов и по автору с
временем на каждый, то есть три колонки на статус вместо одной, и вдобавок
она не умела представить ПОВТОР статуса (откат на «Запланирован» после
инцидента) — такие элементы приходилось запирать от правки целиком.

Теперь строка = запись `status_history` со всеми своими колонками: статус
(ОДНОЙ колонкой), момент установки, кто изменил, комментарий, снимок
контракта. Ключ строки — `id` записи; строка с пустым номером и заполненным
UID элемента это НОВАЯ запись.

**Удалить запись файлом нельзя** — та же причина, по которой нельзя
очистить контракт в правке реквизитов: удаление события это отдельное
осознанное действие (оно есть в карточке элемента), а пустая ячейка или
стёртая строка в Excel слишком дёшевы. Строка, удалённая из файла, просто
не рассматривается.

**Автоматически другие записи не меняются** (прямое указание пользователя):
поправили дату «Отгружен» — меняется только она. Единственное исключение —
общий с импортом истории сдвиг самой ранней записи «Запланирован»
(_shift_planned_before_first_event): она датирована МОМЕНТОМ ИМПОРТА
чертежа, и событие, проставленное задним числом, оказалось бы раньше неё,
то есть правка отменила бы сама себя.
"""

import io
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app import activity, impersonation
from app.contracts import recompute_status_and_actual_date, sync_element_contract
from app.element_bulk_edit import _contract_catalog
from app.element_fields import EXCEL_DATETIME_FORMAT, ru_date_text, to_excel_date
from app.history_import import (
    _shift_planned_before_first_event as shift_planned_before_first_event,
    normalize_changed_at,
)
from app.models import STATUS_LABELS_RU, STATUS_ORDER

SHEET_DATA = "История статусов"
SHEET_STATUSES = "Справочник статусов"
SHEET_USERS = "Пользователи"
SHEET_OBJECTS = "Объекты"

KEY_COLUMN = "record_id"
UID_COLUMN = "element_uid"

# Порядок жизненного цикла — он же порядок в справочнике и он же основание
# для проверки последовательности. Один список, а не три: разойдясь, они
# дали бы проверку, не соответствующую тому, что человек видит в файле.
STATUS_KEYS = [s.value for s in STATUS_ORDER]
STATUS_TITLES = {s.value: STATUS_LABELS_RU[s] for s in STATUS_ORDER}
STATUS_BY_TITLE = {v: k for k, v in STATUS_TITLES.items()}

# Время суток для НОВОЙ записи, если в ячейке только дата без времени.
# Полдень, а не полночь: записи, созданные импортом чертежа, несут реальное
# время, и событие в 00:00 того же дня оказалось бы раньше них (та же
# причина, что в массовой контрактации).
NEW_RECORD_TIME = "12:00:00"

# (ключ, подпись, правимая ли). Порядок тот, в котором читают: сначала «что
# это и о чём», потом правимое, потом справочное.
COLUMNS = [
    (KEY_COLUMN, "№ записи (не менять)", False),
    (UID_COLUMN, "UID элемента (не менять)", False),
    ("object_name", "Объект", False),
    ("element_type", "Тип элемента", False),
    ("subtype", "Подтип", False),
    ("mark", "Марка", False),
    ("address", "Адрес по осям", False),
    ("current_status", "Текущий статус элемента", False),
    ("status", "Статус", True),
    ("changed_at", "Дата и время установки", True),
    ("changed_by", "Кто изменил", True),
    ("changed_by_login", "Учётная запись (справочно)", False),
    ("comment", "Комментарий", True),
    ("contract_name", "Контракт на момент записи (справочно)", False),
]

EDITABLE = [key for key, _, editable in COLUMNS if editable]

# Подписи полей для сводки расхождений — те же, что заголовки колонок:
# разойдясь, они не дали бы связать чип «Статус (12)» с колонкой файла.
FIELD_LABELS = {key: label for key, label, _ in COLUMNS}
FIELD_LABELS["__new_record__"] = "Новая запись"


def columns_spec() -> list:
    return [{"key": k, "label": l, "editable": e} for k, l, e in COLUMNS]


def _load(conn, element_ids: Optional[set] = None) -> tuple[dict, list]:
    """Элементы и все записи истории — двумя запросами, а не запросом истории
    на каждый элемент: на 9422 строках это тот же N+1, что уже стоил проекту
    2,7 секунды на другой форме.

    `element_ids` — отбор схемы при ВЫГРУЗКЕ («Учитывать фильтр»); сверка
    загруженного файла его не передаёт никогда, иначе строка про элемент вне
    текущего фильтра была бы отвергнута как «нет такого UID». Отсев в Python,
    а не `id IN (...)`: тысячи значений не помещаются в лимит переменных
    SQLite.
    """
    elements = {
        r["id"]: r for r in conn.execute(
            """
            SELECT e.id, e.element_uid, e.element_type, e.subtype, e.mark, e.address,
                   e.current_status, o.name AS object_name
            FROM elements e LEFT JOIN objects o ON o.id = e.object_id
            WHERE e.is_current = 1 AND e.element_uid IS NOT NULL
            """
        )
        if element_ids is None or r["id"] in element_ids
    }
    records = conn.execute(
        """
        SELECT h.id, h.element_id, h.status, h.changed_at, h.changed_by,
               h.changed_by_user_id, h.comment, h.contract_id,
               u.domain_login AS changed_by_login
        FROM status_history h
        LEFT JOIN users u ON u.id = h.changed_by_user_id
        ORDER BY h.element_id, h.changed_at, h.id
        """
    ).fetchall()
    return elements, [r for r in records if r["element_id"] in elements]


def _record_values(record, element, contract_names: dict) -> dict:
    """Значения одной строки файла (и одной строки экрана подтверждения —
    ОДНА функция на оба, как в правке реквизитов: разойдясь, экран перестал
    бы показывать элемент так, как он выглядит в файле)."""
    return {
        KEY_COLUMN: record["id"],
        UID_COLUMN: element["element_uid"],
        "object_name": element["object_name"],
        "element_type": element["element_type"],
        "subtype": element["subtype"],
        "mark": element["mark"],
        "address": element["address"],
        # Статусы — ПОДПИСЬЮ, как в интерфейсе, а не кодом: человек,
        # открывший файл, читает «Смонтирован», а не «installed». Обратно
        # разбирается только колонка «Статус» (STATUS_BY_TITLE).
        "current_status": STATUS_TITLES.get(element["current_status"], element["current_status"]),
        "status": STATUS_TITLES.get(record["status"], record["status"]),
        "changed_at": record["changed_at"],
        "changed_by": record["changed_by"],
        "changed_by_login": record["changed_by_login"],
        "comment": record["comment"],
        "contract_name": contract_names.get(record["contract_id"]),
    }


def _sort_key(values: dict) -> tuple:
    return (values["object_name"] or "", values["element_type"] or "",
            values["mark"] or "", values[UID_COLUMN] or "",
            str(values["changed_at"] or ""), values[KEY_COLUMN] or 0)


def build_status_workbook(conn, element_ids: Optional[set] = None) -> Workbook:
    elements, records = _load(conn, element_ids)
    contract_names = {c["id"]: c["name"] for c in _contract_catalog(conn)}
    rows = sorted(
        (_record_values(r, elements[r["element_id"]], contract_names) for r in records),
        key=_sort_key,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA
    ws.append([label for _, label, _ in COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Момент — НАСТОЯЩЕЙ датой Excel с русским форматом, а не текстом
    # «2026-07-28 12:00:00» (та же конвенция, что у дат в правке реквизитов):
    # файл загружается обратно, и normalize_changed_at принимает и datetime
    # от openpyxl, и строку. Заодно колонка остаётся сортируемой как время,
    # а не как строка.
    момент_кол = [i + 1 for i, (key, _, _) in enumerate(COLUMNS) if key == "changed_at"][0]
    # Номер строки считаем САМИ: ws.max_row в openpyxl это max() по всем
    # ячейкам листа, O(n) на каждое обращение — на 9,4 тыс. строк обращение
    # в цикле стоило бы десятки секунд (живой репорт по выгрузке реквизитов).
    for номер, values in enumerate(rows, start=2):
        ws.append([to_excel_date(values[key]) if key == "changed_at" else values[key]
                   for key, _, _ in COLUMNS])
        ws.cell(row=номер, column=момент_кол).number_format = EXCEL_DATETIME_FORMAT
    _widen(ws)

    # ---- листы справочников ----
    ws_s = wb.create_sheet(SHEET_STATUSES)
    ws_s.append(["Статус", "№ в жизненном цикле"])
    for n, key in enumerate(STATUS_KEYS, start=1):
        ws_s.append([STATUS_TITLES[key], n])
    ws_s.append([])
    ws_s.append(["Строка = одна запись истории. Пустой «№ записи» + заполненный UID = новая запись."])
    ws_s.append(["Удалить запись файлом нельзя — это делается в карточке элемента."])
    ws_s.append(["Даты последних записей статусов не должны убывать по этому порядку."])
    _widen(ws_s)

    users = conn.execute(
        "SELECT last_name, first_name, patronymic, domain_login FROM users ORDER BY last_name, first_name"
    ).fetchall()
    ws_u = wb.create_sheet(SHEET_USERS)
    ws_u.append(["Кто изменил (это значение выбирается в колонке «Кто изменил»)", "Учётная запись"])
    for u in users:
        ws_u.append([_display_name(u), u["domain_login"]])
    _widen(ws_u)

    ws_o = wb.create_sheet(SHEET_OBJECTS)
    ws_o.append(["Объект", "Описание"])
    for r in conn.execute("SELECT name, COALESCE(description,'') AS description FROM objects ORDER BY name"):
        ws_o.append([r["name"], r["description"]])
    _widen(ws_o)

    _add_dropdowns(ws, len(rows), len(users))
    return wb


def _display_name(user) -> str:
    """ФИО одной строкой. Та же склейка, что у format_display_name
    (app/auth.py) — она и пишется в status_history.changed_by, поэтому
    значение из выпадающего списка обязано совпадать с ней символ в символ."""
    return " ".join(p for p in (user["last_name"], user["first_name"], user["patronymic"]) if p)


def _col_letter(key: str) -> str:
    return get_column_letter([k for k, _, _ in COLUMNS].index(key) + 1)


def _add_dropdowns(ws, n_rows: int, n_users: int) -> None:
    """Выпадающие списки в правимых колонках со списочными значениями —
    статус и автор. Валидация Excel при этом НЕ гарантия (вставка через
    буфер её обходит), поэтому те же значения проверяет сервер: список —
    удобство, проверка — обязанность."""
    last = max(n_rows + 1, 2)
    specs = [
        ("status", f"'{SHEET_STATUSES}'!$A$2:$A${len(STATUS_KEYS) + 1}", True),
        ("changed_by", f"'{SHEET_USERS}'!$A$2:$A${n_users + 1}", n_users > 0),
    ]
    for key, formula, present in specs:
        if not present:
            continue
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Значение должно быть выбрано из списка на листе справочника."
        dv.errorTitle = "Недопустимое значение"
        ws.add_data_validation(dv)
        col = _col_letter(key)
        dv.add(f"{col}2:{col}{last}")


def _widen(ws) -> None:
    for i, column in enumerate(ws.iter_cols(), start=1):
        width = max((len(str(c.value)) for c in column if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 12), 45)


# ---------------------------------------------------------------- разбор

def _read_sheet(file_bytes: bytes) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if SHEET_DATA not in wb.sheetnames:
        raise ValueError(f"В файле нет листа «{SHEET_DATA}». "
                         f"Загружайте тот файл, который выгрузила система в режиме «История статусов».")
    ws = wb[SHEET_DATA]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError(f"Лист «{SHEET_DATA}» пуст")
    label_to_key = {label: key for key, label, _ in COLUMNS}
    index = {}
    for i, cell in enumerate(header):
        key = label_to_key.get(str(cell).strip() if cell is not None else "")
        if key:
            index[key] = i
    for обязательная in (KEY_COLUMN, UID_COLUMN):
        if обязательная not in index:
            подпись = dict((k, l) for k, l, _ in COLUMNS)[обязательная]
            raise ValueError(f"В файле нет колонки «{подпись}» — без неё строки не с чем сопоставить.")
    out = []
    for n, raw in enumerate(rows, start=2):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        out.append((n, {key: (raw[i] if i < len(raw) else None) for key, i in index.items()}))
    return out


def _text(raw) -> Optional[str]:
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None


def _coerce_status(raw) -> Optional[str]:
    """Ячейка -> код статуса. Принимается и русская подпись (так выгружается),
    и сам код — файл могли собрать из другой выгрузки."""
    text = _text(raw)
    if text is None:
        return None
    if text in STATUS_BY_TITLE:
        return STATUS_BY_TITLE[text]
    if text in STATUS_TITLES:
        return text
    raise ValueError(f"Неизвестный статус «{text}». Допустимые: "
                     + ", ".join(STATUS_TITLES[k] for k in STATUS_KEYS))


def _coerce_moment(raw) -> Optional[str]:
    """Ячейка -> 'ГГГГ-ММ-ДД ЧЧ:ММ:СС'. Принимает и настоящую дату Excel, и
    текст: openpyxl отдаёт первое, если ячейка так отформатирована, и второе,
    если её набрали руками.

    Ровная полночь заменяется на полдень (NEW_RECORD_TIME). Отличить «ввёл
    дату без времени» от «ввёл 00:00:00» нечем — ячейка с датой приходит из
    openpyxl тем же `datetime` с нулевым временем; а полночь в этой предметной
    области не значит ничего, кроме «время не указали», и при этом ставит
    событие раньше всех записей того же дня.
    """
    if raw is None or (isinstance(raw, str) and not str(raw).strip()):
        return None
    stamp = normalize_changed_at(raw)
    if not stamp:
        raise ValueError(f"Не удалось разобрать дату «{raw}» — ожидается ДД.ММ.ГГГГ ЧЧ:ММ:СС "
                         f"или ячейка в формате даты")
    if stamp.endswith(" 00:00:00"):
        stamp = stamp[:11] + NEW_RECORD_TIME
    return stamp


def analyze(conn, file_bytes: bytes) -> dict:
    parsed = _read_sheet(file_bytes)
    elements, records = _load(conn)
    by_uid = {e["element_uid"]: e for e in elements.values()}
    by_record = {r["id"]: r for r in records}
    history_by_element: dict = {}
    for r in records:
        history_by_element.setdefault(r["element_id"], []).append(r)
    contract_names = {c["id"]: c["name"] for c in _contract_catalog(conn)}

    rejected = []
    предложения: dict = {}   # element_id -> список правок/новых записей

    seen_records = set()
    for line_no, values in parsed:
        отказ = None
        uid = _text(values.get(UID_COLUMN))
        raw_id = values.get(KEY_COLUMN)
        element = by_uid.get(uid) if uid else None
        if element is None:
            отказ = ("Пустой UID элемента — строку не с чем сопоставить" if not uid
                     else "Элемент с таким UID не найден среди актуальных")
        record = None
        if отказ is None and _text(raw_id) is not None:
            try:
                record_id = int(float(str(raw_id).strip()))
            except (TypeError, ValueError):
                record_id = None
            record = by_record.get(record_id) if record_id is not None else None
            if record is None:
                отказ = f"Запись истории № {raw_id} не найдена — её могли удалить после выгрузки"
            elif record_id in seen_records:
                отказ = f"Запись № {record_id} встречается в файле дважды — какую строку применять, неизвестно"
            elif record["element_id"] != element["id"]:
                # UID и номер записи разошлись: строку собрали из двух разных
                # (сортировкой, копированием). Угадывать, что из них верно,
                # нельзя — правка ушла бы не тому элементу.
                отказ = (f"Запись № {record_id} принадлежит другому элементу — "
                         f"UID в строке не совпадает с номером записи")
            else:
                seen_records.add(record_id)
        if отказ:
            rejected.append({"line": line_no, "uid": uid, "reason": отказ})
            continue

        новые = {}
        for key in EDITABLE:
            if key not in values:
                continue
            try:
                if key == "status":
                    новые[key] = _coerce_status(values[key])
                elif key == "changed_at":
                    новые[key] = _coerce_moment(values[key])
                else:
                    новые[key] = _text(values[key])
            except ValueError as exc:
                отказ = str(exc)
                break
        if отказ:
            rejected.append({"line": line_no, "uid": uid, "reason": отказ})
            continue

        if record is None:
            # Новая запись: без статуса и момента она бессмысленна — событие
            # обязано отвечать «что» и «когда».
            если_нет = [FIELD_LABELS[k] for k in ("status", "changed_at") if not новые.get(k)]
            if если_нет:
                rejected.append({"line": line_no, "uid": uid,
                                 "reason": "Новая запись без обязательных полей: " + ", ".join(если_нет)})
                continue
        предложения.setdefault(element["id"], []).append(
            {"line": line_no, "record": record, "values": новые})

    changes, rows_out = [], []
    for element_id, items in предложения.items():
        element = elements[element_id]
        существующие = history_by_element.get(element_id, [])
        item_changes, order_error, повтор = _element_changes(element, существующие, items)
        if not item_changes:
            continue
        if order_error and not повтор:
            # Прежнее поведение матрицы: нарушенный порядок — отказ. Оставлено
            # там, где порядок однозначен, то есть у элемента нет повторов
            # статуса.
            for line in sorted({c["line"] for c in item_changes}):
                rejected.append({"line": line, "uid": element["element_uid"], "reason": order_error})
            continue
        for c in item_changes:
            if order_error:
                # У элемента есть повтор статуса (откат после инцидента) —
                # «порядок» тут неоднозначен по существу, поэтому
                # предупреждаем, а решает человек флажком.
                c["warning"] = order_error
            changes.append(c)
        # Только те строки, где правка реально есть: гнать на экран строку,
        # в которой ничего не изменилось, значит показать её как затронутую.
        затронутые = {c["row_id"] for c in item_changes}
        rows_out.extend(r for r in _row_snapshots(element, items, contract_names)
                        if r["row_id"] in затронутые)

    return {
        "rows_read": len(parsed),
        "elements_touched": len({c["element_id"] for c in changes}),
        "columns": columns_spec(),
        "elements": rows_out,
        "changes": changes,
        "rejected": rejected,
    }


def _row_id(record, line_no: int) -> str:
    """Идентификатор СТРОКИ экрана подтверждения. Строка здесь — запись
    истории, а не элемент (у одного элемента их бывает семь), поэтому
    element_id в этой роли больше не годится."""
    return f"h{record['id']}" if record is not None else f"new{line_no}"


def _row_snapshots(element, items, contract_names: dict) -> list:
    out = []
    for item in items:
        record, новые = item["record"], item["values"]
        if record is not None:
            values = _record_values(record, element, contract_names)
        else:
            # Новая запись: показываем ровно то, что предложено файлом —
            # «было» у неё нет.
            values = {key: None for key, _, _ in COLUMNS}
            values.update({
                UID_COLUMN: element["element_uid"],
                "object_name": element["object_name"],
                "element_type": element["element_type"],
                "subtype": element["subtype"],
                "mark": element["mark"],
                "address": element["address"],
                "current_status": STATUS_TITLES.get(element["current_status"], element["current_status"]),
                "status": STATUS_TITLES.get(новые.get("status"), новые.get("status")),
                "changed_at": новые.get("changed_at"),
                "changed_by": новые.get("changed_by"),
                "comment": новые.get("comment"),
            })
        out.append({"row_id": _row_id(record, item["line"]), "element_id": element["id"],
                    "uid": element["element_uid"], "values": values})
    return out


def _element_changes(element, существующие, items):
    """Правки одного элемента + ошибка порядка статусов, если она есть.

    Порядок считается по ИТОГОВОМУ состоянию — то, что уже в базе, плюс то,
    что пришло файлом. Проверять только файл было бы дырой: «Смонтирован»
    раньше уже существующего «Отгружен» — ровно то нарушение, о котором
    просил сообщать пользователь, и в файле при этом одна строка.
    """
    changes = []
    итог = {r["id"]: {"status": r["status"], "changed_at": r["changed_at"]} for r in существующие}
    for item in items:
        record, новые = item["record"], item["values"]
        row_id = _row_id(record, item["line"])
        общее = {
            "row_id": row_id, "element_id": element["id"], "uid": element["element_uid"],
            "line": item["line"], "mark": element["mark"], "element_type": element["element_type"],
        }
        if record is None:
            итог[row_id] = {"status": новые["status"], "changed_at": новые["changed_at"]}
            # Момент — по-русски: строку собирает сервер, и разбирать её
            # обратно на клиенте ради формата даты было бы хуже, чем собрать
            # сразу так, как её прочитают.
            подпись = (f"{STATUS_TITLES[новые['status']]} · {ru_date_text(новые['changed_at'])}"
                       + (f" · {новые['changed_by']}" if новые.get("changed_by") else ""))
            changes.append({
                **общее, "field": "__new_record__", "column": KEY_COLUMN,
                "field_label": FIELD_LABELS["__new_record__"],
                "was": None, "now": подпись, "is_new": True,
                # Значения новой записи едут вместе с правкой: применение
                # НЕ перечитывает файл (иначе применилось бы не то, что
                # показали на экране).
                "record_values": новые,
            })
            continue
        for key in EDITABLE:
            if key not in новые:
                continue
            было, стало = record[key], новые[key]
            if стало is None or стало == было:
                # Пустая ячейка — «не трогать», а не «очистить»: она слишком
                # дёшева, её ставят случайно. Очистка комментария делается в
                # карточке элемента.
                continue
            changes.append({
                **общее, "field": key, "column": key, "field_label": FIELD_LABELS[key],
                "was": STATUS_TITLES.get(было, было) if key == "status" else было,
                "now": STATUS_TITLES.get(стало, стало) if key == "status" else стало,
            })
            итог.setdefault(record["id"], {})[key] = стало

    порядок = _sequence_error(итог)
    статусы = [v["status"] for v in итог.values()]
    повтор = len(статусы) != len(set(статусы))
    return changes, порядок, повтор


def _sequence_error(итог: dict) -> Optional[str]:
    """Даты ПОСЛЕДНИХ записей каждого статуса не должны убывать вдоль
    жизненного цикла. Пропуски допустимы: элемент мог не проходить
    «Контрактацию» явно — сравниваются только встретившиеся статусы."""
    последняя: dict = {}
    for v in итог.values():
        момент = v.get("changed_at")
        if not момент:
            continue
        if v["status"] not in последняя or момент > последняя[v["status"]]:
            последняя[v["status"]] = момент
    filled = [(k, последняя[k]) for k in STATUS_KEYS if k in последняя]
    for i in range(1, len(filled)):
        (prev_key, prev_date), (key, date) = filled[i - 1], filled[i]
        if date < prev_date:
            return (f"«{STATUS_TITLES[key]}» ({date}) раньше, чем "
                    f"«{STATUS_TITLES[prev_key]}» ({prev_date}) — нарушен порядок статусов")
    return None


# ---------------------------------------------------------------- запись

def apply_changes(conn, selections: list, user_name: str, user_id: Optional[int]) -> dict:
    """Применяет отмеченные правки истории.

    На вход приходит то же, что вернул analyze, но отфильтрованное флажками —
    заново файл не читается: перечитывание между показом и применением
    означало бы, что применить могли не то, что показали.
    """
    разрешено = set(EDITABLE) | {"__new_record__"}
    неизвестные = {str(sel.get("field")) for sel in selections} - разрешено
    if неизвестные:
        raise ValueError("Недопустимые поля для правки: " + ", ".join(sorted(неизвестные)))

    по_строкам: dict = {}
    for sel in selections:
        по_строкам.setdefault(str(sel.get("row_id")), []).append(sel)

    applied_elements, inserted, updated, skipped = set(), 0, 0, []
    подробности: dict = {}
    for row_id, items in по_строкам.items():
        element_id = int(items[0]["element_id"])
        element = conn.execute(
            "SELECT id, element_type, subtype, mark FROM elements WHERE id = ?", (element_id,)
        ).fetchone()
        if element is None:
            skipped.append({"element_id": element_id, "reason": "Элемент исчез между сверкой и применением"})
            continue

        новая = next((s for s in items if s["field"] == "__new_record__"), None)
        if новая is not None:
            v = новая.get("record_values") or {}
            статус, момент = v.get("status"), v.get("changed_at")
            if статус not in STATUS_TITLES or not момент:
                skipped.append({"element_id": element_id,
                                "reason": "Новая запись без статуса или без момента установки"})
                continue
            автор = v.get("changed_by") or user_name
            conn.execute(
                "INSERT INTO status_history (element_id, status, changed_at, changed_by, "
                "changed_by_user_id, comment) VALUES (?, ?, ?, ?, ?, ?)",
                (element_id, статус, момент, автор, _user_id_for(conn, автор),
                 v.get("comment") or "Массовая правка истории через Excel"),
            )
            inserted += 1
            подробности.setdefault(element_id, []).append(
                f"добавлено «{STATUS_TITLES[статус]}» на {момент}")
            applied_elements.add(element_id)
            continue

        record_id = int(row_id[1:]) if row_id.startswith("h") else None
        record = conn.execute(
            "SELECT * FROM status_history WHERE id = ? AND element_id = ?", (record_id, element_id)
        ).fetchone() if record_id else None
        if record is None:
            skipped.append({"element_id": element_id,
                            "reason": f"Запись истории № {record_id} исчезла между сверкой и применением"})
            continue

        поля, тексты = {}, []
        for sel in items:
            field, стало = sel["field"], sel.get("now")
            if field == "status":
                стало = STATUS_BY_TITLE.get(стало, стало)
                if стало not in STATUS_TITLES:
                    skipped.append({"element_id": element_id, "reason": f"Неизвестный статус «{sel.get('now')}»"})
                    continue
            if field == "changed_at":
                стало = normalize_changed_at(стало)
                if not стало:
                    skipped.append({"element_id": element_id, "reason": "Не удалось разобрать момент установки"})
                    continue
            поля[field] = стало
            тексты.append(f"{FIELD_LABELS[field]}: {sel.get('was') or '—'} → {sel.get('now')}")
            if field == "changed_by":
                # ФИО и учётная запись — одна пара: оставить прежний
                # changed_by_user_id при новом ФИО значило бы, что событие
                # подписано одним человеком, а ссылается на другого.
                поля["changed_by_user_id"] = _user_id_for(conn, стало)
        if not поля:
            continue
        присвоения = ", ".join(f"{f} = :{f}" for f in поля)
        conn.execute(f"UPDATE status_history SET {присвоения} WHERE id = :id",
                     {**поля, "id": record["id"]})
        updated += 1
        подробности.setdefault(element_id, []).extend(тексты)
        applied_elements.add(element_id)

    for element_id in applied_elements:
        # Запись «Запланирован» датирована МОМЕНТОМ ИМПОРТА чертежа, а не
        # реальным планированием: событие, проставленное задним числом,
        # оказывается раньше неё, и элемент по правилу «текущий статус =
        # последняя по changed_at» откатился бы в «Запланирован» — правка
        # отменила бы сама себя. Тот же сдвиг, что делает импорт истории;
        # вторая реализация разошлась бы с ней.
        shift_planned_before_first_event(conn, element_id)
        # Текущий статус и фактическая дата — производные от истории, их
        # обязательно пересчитать: правка меняет, какая запись самая поздняя.
        # Контракт следом — инвариант «Запланирован ⇒ контракт пуст».
        effective, _ = recompute_status_and_actual_date(conn, element_id)
        sync_element_contract(conn, element_id, effective)
        row = conn.execute(
            "SELECT element_type, subtype, mark FROM elements WHERE id = ?", (element_id,)
        ).fetchone()
        activity.log(
            "status_bulk_edit", user_name=impersonation.plain_name(user_name), user_id=user_id,
            entity_type="element", entity_id=element_id,
            element_type=row["element_type"], subtype=row["subtype"], mark=row["mark"],
            new_value="; ".join(подробности.get(element_id, []))[:500],
            details={"source": "xlsx", "effective_status": effective},
        )

    conn.commit()
    return {"elements_updated": len(applied_elements), "records_inserted": inserted,
            "records_updated": updated, "skipped": skipped}


def _user_id_for(conn, display_name: Optional[str]) -> Optional[int]:
    """Учётная запись по ФИО. None — такого пользователя нет: ФИО в истории
    это СНИМОК («кто изменил тогда»), и человек мог быть удалён; запрещать
    из-за этого правку нельзя, но и оставлять ссылку на чужую учётную запись
    тоже."""
    if not display_name:
        return None
    for u in conn.execute("SELECT id, last_name, first_name, patronymic FROM users"):
        if _display_name(u) == display_name:
            return u["id"]
    return None
