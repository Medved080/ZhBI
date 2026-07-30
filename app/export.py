from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.contracts import build_document_label
from app.db import visible_elements_clause
from app.models import STATUS_LABELS_RU

ELEMENT_COLUMNS = [
    ("id", "ID"),
    ("dxf_handle", "DXF handle"),
    ("layer", "Слой"),
    ("element_type", "Тип"),
    ("subtype", "Подтип"),
    ("mark", "Марка"),
    ("mark_source", "Источник марки"),
    ("x", "X, мм"),
    ("y", "Y, мм"),
    ("z", "Z, мм"),
    ("elevation_mm", "Отметка, мм"),
    ("address", "Адрес по осям"),
    ("axis_status", "Статус адресации"),
    ("axis_number", "Числовая ось"),
    ("axis_letter", "Буквенная ось"),
    ("nearest_axis_number", "Ближайшая числовая ось"),
    ("nearest_axis_letter", "Ближайшая буквенная ось"),
    ("offset_x_mm", "Смещение X, мм"),
    ("offset_y_mm", "Смещение Y, мм"),
    # "Контрактация 2.0" (см. Docs/backlog.md) — четыре независимые шкалы
    # дат поставки, простые колонки elements, доступны в обоих экспортах
    # (снимок на дату и полная история) без отдельной логики.
    ("planned_delivery_date", "Плановая дата поставки"),
    ("project_delivery_date", "Дата завершения СМР"),
    ("project_smr_start_date", "Начало СМР"),
    ("actual_delivery_date", "Фактическая дата поставки"),
]

# Привязка к зонам (см. Docs/backlog.md, "Разбор структурированных имён
# слоёв DWG/DXF...") — только id+status хранятся в elements, реальное
# название зоны резолвится отдельным запросом к zones (см. _zone_names),
# тем же способом, что и карточка элемента на фронтенде (zoneBindingText
# в app/static/index.html) — сюда попадает не сырой id, а имя зоны.
ZONE_COLUMNS = [
    ("zone_zakhvatka_id", "zone_zakhvatka_status", "Захватка"),
    ("zone_crane_id", "zone_crane_status", "Кран"),
    ("zone_stance_id", "zone_stance_status", "Стоянка"),
]

ZONE_STATUS_LABELS_RU = {
    "unmatched": "не определено", "needs_review": "требует проверки", "not_applicable": "неприменимо",
}


def _zone_names(conn) -> dict:
    return {r["id"]: r["name"] for r in conn.execute("SELECT id, name FROM zones").fetchall()}


# Реквизиты контракта — ТРИ отдельные колонки (Поставщик / Договор /
# Спецификация) вместо одной склеенной "Контракт" (живой запрос
# пользователя, 2026-07-28): в выгрузке по ним нужно фильтровать и
# сводить сводные таблицы, а из строки вида
# "Контрагент/Договор № 5 от .../Спецификация № 2 от ..." Excel этого не
# умеет. Формат каждого документа — общий с интерфейсом
# (build_document_label, app/contracts.py), не своя копия.
CONTRACT_COLUMNS = ["Поставщик", "Договор (номер и дата)", "Спецификация (номер и дата)"]

_EMPTY_CONTRACT_CELLS = ["", "", ""]


def _contract_labels(conn) -> dict:
    # "Контрактация 2.0" (см. Docs/backlog.md) — contracts.supplier убран,
    # контрагент резолвится через цепочку specification->agreement->
    # counterparty, та же схема, что app/contracts.py:_specification_chain.
    # Тема контракта (contracts.theme) в выгрузку не попадает — она часть
    # СКЛЕЕННОГО наименования (build_contract_name), а по отдельным
    # реквизитам ей места нет.
    return {
        r["id"]: [
            r["counterparty_short_name"],
            build_document_label(r["agreement_number"], r["agreement_date"]),
            build_document_label(r["specification_number"], r["specification_date"]),
        ]
        for r in conn.execute(
            """
            SELECT co.id AS id,
                   c.short_name AS counterparty_short_name,
                   a.number AS agreement_number, a.agreement_date AS agreement_date,
                   s.number AS specification_number, s.specification_date AS specification_date
            FROM contracts co
            JOIN specifications s ON s.id = co.specification_id
            JOIN agreements a ON a.id = s.agreement_id
            JOIN counterparties c ON c.id = a.counterparty_id
            """
        ).fetchall()
    }


def _zone_cell(row, id_field, status_field, zone_names) -> str:
    status = row[status_field]
    if status == "matched":
        zid = row[id_field]
        name = zone_names.get(zid)
        return name if name else (f"зона #{zid}" if zid else "")
    return ZONE_STATUS_LABELS_RU.get(status, status or "")


def _contract_cells(contract_id, contract_labels) -> list:
    """Три ячейки (Поставщик / Договор / Спецификация) на один contract_id.
    Контракт, которого уже нет в справочнике (удалён после того, как был
    записан в историю), не молчит пустотой — попадает в колонку Поставщик
    как "#id", как и раньше в склеенной колонке."""
    if not contract_id:
        return list(_EMPTY_CONTRACT_CELLS)
    return contract_labels.get(contract_id, [f"#{contract_id}", "", ""])


def _autosize(ws):
    for i, col in enumerate(ws.columns, start=1):
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 60)


def _day_bounds(date_from, date_to):
    lo = f"{date_from} 00:00:00" if date_from else None
    hi = f"{date_to} 23:59:59" if date_to else None
    return lo, hi


def _element_ids_clause(element_ids, column="id"):
    """element_ids — None (без ограничения, экспорт "все элементы") или
    список id, уже отфильтрованных на клиенте текущим состоянием фильтров
    (см. Docs/backlog.md, "Экспорт с учётом фильтра") — фильтры сложные и
    целиком живут в app.js (passesPlacementFilters), пересчитывать их на
    бэкенде было бы дублированием логики; вместо этого клиент присылает
    готовый список id. Пустой список — легитимный результат (фильтр не
    оставил ни одного элемента), не ошибка — возвращает заведомо ложное
    условие, а не пытается собрать "IN ()" (невалидный SQL)."""
    if element_ids is None:
        return None, []
    if not element_ids:
        return "1=0", []
    placeholders = ",".join("?" * len(element_ids))
    return f"{column} IN ({placeholders})", list(element_ids)


def build_snapshot_xlsx(conn, source_file, date, element_ids=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Статус на дату"

    zone_names = _zone_names(conn)
    contract_labels = _contract_labels(conn)

    header = (
        [label for _, label in ELEMENT_COLUMNS]
        + [label for _, _, label in ZONE_COLUMNS]
        + CONTRACT_COLUMNS
        + ["Статус", "Статус изменён", "Кто изменил"]
    )
    ws.append(header)

    clauses = [visible_elements_clause()]
    params = []
    if source_file:
        clauses.append("source_file = ?")
        params.append(source_file)
    ids_clause, ids_params = _element_ids_clause(element_ids)
    if ids_clause:
        clauses.append(ids_clause)
        params.extend(ids_params)
    where = f"WHERE {' AND '.join(clauses)}"
    elements = conn.execute(f"SELECT * FROM elements {where} ORDER BY id", params).fetchall()

    for el in elements:
        # Всегда через status_history (не el["current_status"]/["updated_at"]
        # напрямую даже без date) — только там есть changed_by (ФИО того, кто
        # менял статус, живой запрос пользователя), а без date нужна именно
        # САМАЯ ПОЗДНЯЯ запись — тот же инвариант, что у current_status
        # (см. recompute_status_and_actual_date, app/contracts.py).
        query = "SELECT status, changed_at, changed_by FROM status_history WHERE element_id = ?"
        query_params = [el["id"]]
        if date:
            query += " AND changed_at <= ?"
            query_params.append(f"{date} 23:59:59")
        query += " ORDER BY changed_at DESC LIMIT 1"
        row = conn.execute(query, query_params).fetchone()
        status = row["status"] if row else None
        changed_at = row["changed_at"] if row else None
        changed_by = row["changed_by"] if row else None

        values = [el[key] for key, _ in ELEMENT_COLUMNS]
        values.extend(_zone_cell(el, id_field, status_field, zone_names) for id_field, status_field, _ in ZONE_COLUMNS)
        values.extend(_contract_cells(el["contract_id"], contract_labels))
        values.append(STATUS_LABELS_RU.get(status, status) if status else "(нет данных на эту дату)")
        values.append(changed_at or "")
        values.append(changed_by or "")
        ws.append(values)

    _autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_history_xlsx(conn, source_file, date_from, date_to, element_ids=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "История статусов"

    zone_names = _zone_names(conn)
    contract_labels = _contract_labels(conn)

    header = (
        [label for _, label in ELEMENT_COLUMNS]
        + [label for _, _, label in ZONE_COLUMNS]
        + ["Статус", "Изменено", "Кто изменил", "Комментарий"]
        # "на момент изменения" в заголовке сохранено намеренно (было и у
        # прежней склеенной колонки): здесь реквизиты берутся из
        # status_history.contract_id, то есть отражают, что было ТОГДА, а не
        # текущую привязку элемента.
        + [f"{label} на момент изменения" for label in CONTRACT_COLUMNS]
    )
    ws.append(header)

    lo, hi = _day_bounds(date_from, date_to)
    clauses = []
    params = []
    if source_file:
        clauses.append("e.source_file = ?")
        params.append(source_file)
    if lo:
        clauses.append("sh.changed_at >= ?")
        params.append(lo)
    if hi:
        clauses.append("sh.changed_at <= ?")
        params.append(hi)
    ids_clause, ids_params = _element_ids_clause(element_ids, column="e.id")
    if ids_clause:
        clauses.append(ids_clause)
        params.extend(ids_params)
    # ОСОЗНАННОЕ отличие от снимка "Статус на дату" выше: тут
    # visible_elements_clause НЕ применяется. Выгрузка истории — архив
    # произошедшего и средство переноса прогресса на другой сервер; элемент,
    # исчезнувший из актуального чертежа (is_current=0), свою историю
    # сохраняет, и молча выбросить её из архива хуже, чем показать строку по
    # элементу, которого на схеме уже нет.
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""

    rows = conn.execute(
        f"""
        SELECT e.*, sh.status as event_status, sh.changed_at, sh.changed_by, sh.comment,
               sh.contract_id as event_contract_id
        FROM status_history sh
        JOIN elements e ON e.id = sh.element_id
        {where}
        ORDER BY e.id, sh.changed_at
        """,
        params,
    ).fetchall()

    for r in rows:
        values = [r[key] for key, _ in ELEMENT_COLUMNS]
        values.extend(_zone_cell(r, id_field, status_field, zone_names) for id_field, status_field, _ in ZONE_COLUMNS)
        values.append(STATUS_LABELS_RU.get(r["event_status"], r["event_status"]))
        values.append(r["changed_at"])
        values.append(r["changed_by"] or "")
        values.append(r["comment"] or "")
        # Контракт НА МОМЕНТ этого события (sh.contract_id), а не текущий
        # контракт элемента — история должна отражать, что было тогда,
        # а не что сейчас (см. Docs/backlog.md).
        values.extend(_contract_cells(r["event_contract_id"], contract_labels))
        ws.append(values)

    _autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
