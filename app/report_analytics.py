"""
Отчёт «Аналитическая справка» (живой запрос 2026-08-11, прототип согласован
в той же сессии).

Вопрос, на который он отвечает: **что сейчас мешает стройке и чем это
кончится**. Остальные отчёты показывают состояние («сколько смонтировано»,
«чем закрыта стоянка», «как шла динамика»); этот — разрыв между тем, что
стройке нужно в ближайшее время, и тем, что под это есть.

Шесть разделов, каждый отвечает на свой вопрос:

  0. Резюме        — показатели и выводы; ниже — блоки «События, задачи,
     вопросы» на дату (те же, что в ежедневном отчёте, app/settings.py:
     своего хранилища справка не заводит).
  1.1 Обеспечение ближайших этапов — этапы, где СМР уже идёт или начнётся
     в горизонте: закрыта ли их потребность контрактами.
  1.2 Общий прогресс контрактации — то же по объекту целиком, по типам.
  2.1 Фронт работ  — текущий незакрытый ярус каждой стоянки и следующий.
  2.2 Критический путь — чего именно не хватает на текущих ярусах.
  2.3 Динамика     — накопительные кривые за весь срок.

Инварианты (решения пользователя 2026-08-11, менять только с ним):

  **Потребность** — изделия МОДЕЛИ, разложенные по `project_smr_start_date`
      (дата начала СМР: к ней изделие обязано быть на площадке;
      `project_delivery_date` означает ЗАВЕРШЕНИЕ СМР и к поставке
      отношения не имеет — см. app/schedule_import.py).

  **Ярус** — отметка яруса стоянки (`elements.zone_stance_level_id` →
      `zone_levels.elevation_mm`), а не этаж и не подтип: логика «нельзя
      ставить верхний, пока не смонтирован нижний» — про физическую
      отметку. Этап СМР — кран + стоянка + ЭТАЖ: тем же блоком размечен
      график MS Project, откуда берутся даты.

  **Зачёт контрактации — по маркам.** Изделие считается закрытым, только
      если законтрактована ЕГО марка: сотня плит другой марки потребность
      не закрывает. Колонка «по типу» рядом — справочная; расхождение между
      ними и есть «контракт есть, но не тот».

  **«На площадке»** — заполненная `actual_delivery_date`, а не статус:
      статусы «В производстве» и «Отгружено» на этой стройке не ведутся,
      входящих данных по ним нет.

  **Очередь завода.** Срока изготовления в днях у заказчика нет — есть темп
      завода (app/capacity.py). Весь непоставленный объём завода по типу
      выстраивается в очередь по датам потребности и списывается темпом
      «шт./календарный день» со дня справки. Изделия с плановой датой
      поставки показывают ПЛАН (это обещание завода, а не наша оценка), но
      мощность занимают и сдвигают тех, кто за ними. Дефицит без контракта
      идёт отдельной очередью со средним темпом по типу; если темпа нет ни
      у кого — «сроки неизвестны», и вывод «успевает / не успевает» не
      делается вовсе.

  Оценка честно оптимистичная: заказы этих же заводов на ДРУГИЕ стройки
  системе неизвестны и в очередь не попадают. Это сказано в самом отчёте, а
  не только здесь.

Область — объект ЦЕЛИКОМ: фильтр схемы справка не учитывает (решение
пользователя). Данные, как и у остальных отчётов, считает СЕРВЕР — экран,
XLSX и PDF берут результат одной функции и разойтись не могут.
"""

import math
import sqlite3
from datetime import date, datetime, timedelta
from typing import Optional

from app.capacity import CapacityBook
from app.models import Status
from app.reports import natural_key, pdf_text
from app.settings import get_notes_for_date

TITLE = "Аналитическая справка"

# Горизонт «ближайшего времени» — выбирается в форме, умолчание месяц.
DEFAULT_HORIZON_DAYS = 30
HORIZONS = [
    {"days": 14, "label": "2 недели"},
    {"days": 30, "label": "1 месяц"},
    {"days": 42, "label": "6 недель"},
    {"days": 60, "label": "2 месяца"},
    {"days": 90, "label": "квартал"},
]

# Смонтировано: изделие уже стоит в конструкции. «Принято» сюда же — это
# состояние ПОСЛЕ монтажа, и считать его несмонтированным значило бы
# показывать закрытые ярусы вечно незакрытыми.
DONE_STATUSES = (Status.INSTALLED.value, Status.ACCEPTED.value)

NO_LEVEL = "Ярус не определён"
UNMAPPED_LABEL = "Не размечено"

# Сколько строк критического пути отдавать. Ограничение есть, потому что при
# полностью незаконтрактованном объекте сюда попадёт весь перечень марок; о
# срезке отчёт сообщает явно — молчаливое усечение читалось бы как «это всё».
CRITICAL_ROWS_LIMIT = 300


# Колонки таблиц — ОДИН список на экран, Excel и PDF: три копии заголовков
# разошлись бы на первой же правке. kind задаёт выравнивание на экране и
# формат ячейки в Excel («date» кладётся настоящей датой, а не текстом).
STAGE_COLUMNS = [
    {"key": "crane", "label": "Кран", "kind": "text"},
    {"key": "stance", "label": "Стоянка", "kind": "text"},
    {"key": "floor", "label": "Этаж", "kind": "num"},
    {"key": "elevation", "label": "Отметка", "kind": "text"},
    {"key": "start", "label": "Старт СМР", "kind": "date"},
    {"key": "days_left", "label": "Дней до старта", "kind": "num"},
    {"key": "element_type", "label": "Тип", "kind": "text"},
    {"key": "mark", "label": "Марка", "kind": "text"},
    {"key": "need", "label": "Потребность", "kind": "num"},
    {"key": "contracted", "label": "Законтрактовано по марке", "kind": "num"},
    {"key": "deficit", "label": "Дефицит", "kind": "num"},
    # «Ждём поставки» отделяет бумажный дефицит от физического: изделия
    # позиции могут уже стоять в конструкции, а контракт под них так и не
    # оформлен — это разные проблемы и разные адресаты.
    {"key": "awaiting", "label": "Ждём поставки", "kind": "num"},
    {"key": "contracted_by_type", "label": "Законтрактовано по типу", "kind": "num"},
    {"key": "counterparty", "label": "Завод", "kind": "text"},
    {"key": "ready", "label": "Расчётная готовность", "kind": "date"},
    {"key": "verdict", "label": "Оценка", "kind": "verdict"},
]

PROGRESS_COLUMNS = [
    {"key": "element_type", "label": "Тип", "kind": "text"},
    {"key": "need", "label": "Потребность (модель)", "kind": "num"},
    {"key": "credited", "label": "Законтрактовано (зачёт по маркам)", "kind": "num"},
    {"key": "deficit", "label": "Дефицит", "kind": "num"},
    {"key": "percent", "label": "%", "kind": "num"},
    {"key": "contracted_total", "label": "Законтрактовано всего по типу", "kind": "num"},
    {"key": "assigned", "label": "Привязано к изделиям", "kind": "num"},
    {"key": "not_credited", "label": "Не в зачёт (марка не та)", "kind": "num"},
]

FRONT_COLUMNS = [
    {"key": "crane", "label": "Кран", "kind": "text"},
    {"key": "stance", "label": "Стоянка", "kind": "text"},
    {"key": "elevation", "label": "Ярус", "kind": "text"},
    {"key": "need", "label": "Нужно", "kind": "num"},
    {"key": "installed", "label": "Смонтировано", "kind": "num"},
    {"key": "on_site", "label": "На площадке", "kind": "num"},
    {"key": "contracted_not_delivered", "label": "Законтрактовано, не поставлено", "kind": "num"},
    {"key": "no_contract", "label": "Нет контракта", "kind": "num"},
    {"key": "state", "label": "Состояние", "kind": "verdict"},
]

CRITICAL_COLUMNS = [
    {"key": "crane", "label": "Кран", "kind": "text"},
    {"key": "stance", "label": "Стоянка", "kind": "text"},
    {"key": "elevation", "label": "Ярус", "kind": "text"},
    {"key": "element_type", "label": "Тип", "kind": "text"},
    {"key": "mark", "label": "Марка", "kind": "text"},
    {"key": "missing", "label": "Не хватает", "kind": "num"},
    {"key": "no_contract", "label": "Из них без контракта", "kind": "num"},
    {"key": "counterparty", "label": "Завод", "kind": "text"},
    {"key": "need_date", "label": "Нужно к дате", "kind": "date"},
    {"key": "plan_date", "label": "Плановая поставка", "kind": "date"},
    {"key": "ready", "label": "Расчётная готовность", "kind": "date"},
    {"key": "verdict", "label": "Оценка", "kind": "verdict"},
]

SERIES_LABELS = {
    "need": "Потребность", "contracted": "Законтрактовано",
    "delivered": "Поставлено", "installed": "Смонтировано",
}

# Приписка про природу оценки. Уходит и на экран, и в оба файла: файл живёт
# своей жизнью, и предупреждение, оставшееся только на экране, бесполезно.
DISCLAIMER = (
    "Расчётная готовность считается по очереди завода (темп из карточки контрагента) "
    "и учитывает только заказы ЭТОЙ стройки: заказы тех же заводов на другие объекты "
    "системе неизвестны, поэтому оценка оптимистичная."
)


def _parse(значение) -> Optional[date]:
    """Даты в базе лежат текстом 'ГГГГ-ММ-ДД' и сравниваются как текст (см.
    Docs/DECISIONS.md); здесь они нужны объектами — считать сроки."""
    if not значение:
        return None
    try:
        return datetime.strptime(str(значение)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _iso(d: Optional[date]) -> Optional[str]:
    return d.isoformat() if d else None


def _elevation_label(mm: Optional[int]) -> str:
    """5000 → «+5.000», 0 → «±0.000», None → «Ярус не определён». Формат тот
    же, что в подтипах и на чертеже, — по нему ярус и опознают."""
    if mm is None:
        return NO_LEVEL
    if mm == 0:
        return "±0.000"
    знак = "+" if mm > 0 else "−"
    целые, остаток = divmod(abs(mm), 1000)
    return f"{знак}{целые}.{остаток:03d}"


def _zone_label(prefix: str, number, name) -> str:
    """«Кран 2» / «Стоянка 07». Номер первичен (имя менялось между версиями
    чертежа), имя — запасной вариант для зоны без номера."""
    if number is not None:
        return f"{prefix} {number}"
    return name or f"{prefix} не определён"


# ---------------------------------------------------------------- очередь


class _Queue:
    """Очередь завода по одному типу изделий.

    Списывается с даты справки темпом per_day (шт./календарный день).
    Позиция i (считая с нуля) готова через (i+1)/per_day дней — вверх до
    целого дня: половину изделия на площадку не привезут.
    """

    def __init__(self, per_day: Optional[float], start: date):
        self.per_day = per_day
        self.start = start
        self.выпущено = 0.0

    def take(self) -> Optional[date]:
        if not self.per_day:
            return None      # темпа нет — «сроки неизвестны», а не «ноль дней»
        self.выпущено += 1.0
        return self.start + timedelta(days=math.ceil(self.выпущено / self.per_day))


def _build_queues(изделия: list, book: CapacityBook, today: date) -> dict:
    """id изделия → расчётная дата готовности (или None, если темп неизвестен).

    Очередь одна на (завод, тип). Изделия без контракта идут отдельной
    очередью того же типа со средним темпом по объекту.

    Темп завода при разных переопределениях в его контрактах берётся
    НАИБОЛЬШИЙ: очередь моделирует мощность ЗАВОДА, а переопределение
    описывает темп одного документа — завод в целом не может быть медленнее
    самого быстрого своего контракта.
    """
    темпы: dict = {}
    for e in изделия:
        if e["actual_delivery_date"] or e["current_status"] in DONE_STATUSES:
            continue          # уже на площадке или в конструкции: мощность завода оно не ест
        cp_id = book.counterparty_of_contract.get(e["contract_id"]) if e["contract_id"] else None
        ключ = (cp_id, e["element_type"])
        if cp_id is None:
            темпы.setdefault(ключ, book.average.get(e["element_type"]))
            continue
        темп = book.for_contract(e["contract_id"], e["element_type"])
        прежний = темпы.get(ключ)
        if темп and (прежний is None or темп > прежний):
            темпы[ключ] = темп
        темпы.setdefault(ключ, None)

    очереди = {ключ: _Queue(темп, today) for ключ, темп in темпы.items()}

    # Порядок очереди — по дате потребности: раньше нужно, раньше делают.
    # Изделия без даты уходят в хвост (иначе они заняли бы мощность впереди
    # тех, у кого срок известен, — то есть неразмеченность двигала бы сроки).
    def ключ_сортировки(e):
        нужно = _parse(e["project_smr_start_date"])
        return (0, нужно, e["id"]) if нужно else (1, date.max, e["id"])

    готовность = {}
    ждут = (e for e in изделия
            if not e["actual_delivery_date"] and e["current_status"] not in DONE_STATUSES)
    for e in sorted(ждут, key=ключ_сортировки):
        cp_id = book.counterparty_of_contract.get(e["contract_id"]) if e["contract_id"] else None
        готовность[e["id"]] = очереди[(cp_id, e["element_type"])].take()
    return готовность


# ------------------------------------------------------- зачёт контрактов


def _allocate(потребности: list, запас: dict) -> dict:
    """Разложить законтрактованные количества по потребностям.

    `потребности` — список (ключ_запаса, дата_потребности, порядковый_ключ,
    сколько_нужно, id_строки); `запас` — ключ → сколько выкуплено.
    Возвращает id_строки → сколько ей досталось.

    Раскладка ПО ДАТАМ, а не поровну: контракт закрывает сначала то, что
    нужно раньше. Иначе ранний этап показывал бы дефицит при том, что
    изделия для него давно куплены, — и справка гнала бы тревогу туда, где
    её нет.
    """
    остаток = dict(запас)
    выдано = {}
    for ключ, дата, порядок, нужно, id_строки in sorted(
        потребности, key=lambda p: (p[1] or date.max, p[2])
    ):
        свободно = остаток.get(ключ, 0)
        доля = min(нужно, свободно)
        if доля:
            остаток[ключ] = свободно - доля
        выдано[id_строки] = доля
    return выдано


# ------------------------------------------------------------ сам отчёт


def build_analytics_report(conn: sqlite3.Connection, object_id: int,
                           report_date: Optional[str] = None,
                           horizon_days: Optional[int] = None) -> dict:
    today = _parse(report_date) or date.today()
    горизонт = horizon_days or DEFAULT_HORIZON_DAYS
    if горизонт not in [h["days"] for h in HORIZONS]:
        горизонт = DEFAULT_HORIZON_DAYS
    конец_горизонта = today + timedelta(days=горизонт)

    изделия = [dict(r) for r in conn.execute(
        """
        SELECT e.id, e.element_type, e.mark, e.floor, e.contract_id, e.current_status,
               e.project_smr_start_date, e.planned_delivery_date, e.actual_delivery_date,
               e.zone_crane_id, e.zone_stance_id, e.zone_stance_level_id,
               zc.number AS crane_number, zc.name AS crane_name,
               zs.number AS stance_number, zs.name AS stance_name,
               zl.elevation_mm AS level_mm
        FROM elements e
        LEFT JOIN zones zc ON zc.id = e.zone_crane_id
        LEFT JOIN zones zs ON zs.id = e.zone_stance_id
        LEFT JOIN zone_levels zl ON zl.id = e.zone_stance_level_id
        WHERE e.object_id = ? AND e.is_current = 1
        """,
        (object_id,),
    ).fetchall()]

    # Позиции контрактов ЭТОГО объекта: объект выводится по цепочке контракт
    # → спецификация → договор (своего поля у контракта нет, см. schema.sql).
    # Архивные не участвуют — отработанный документ ничего не обеспечивает.
    законтрактовано_марка: dict = {}
    законтрактовано_тип: dict = {}
    for p in conn.execute(
        """
        SELECT cl.element_type, cl.mark, cl.quantity
        FROM contract_lines cl
        JOIN contracts co ON co.id = cl.contract_id
        JOIN specifications s ON s.id = co.specification_id
        JOIN agreements a ON a.id = s.agreement_id
        WHERE a.object_id = ? AND co.is_archived = 0
        """,
        (object_id,),
    ):
        тип = p["element_type"] or ""
        марка = (p["mark"] or "").strip()
        законтрактовано_тип[тип] = законтрактовано_тип.get(тип, 0) + p["quantity"]
        if марка:
            ключ = (тип, марка)
            законтрактовано_марка[ключ] = законтрактовано_марка.get(ключ, 0) + p["quantity"]

    book = CapacityBook(conn, object_id)
    готовность = _build_queues(изделия, book, today)

    # ---------- разметка изделий по этапам и ярусам ----------
    неразмечено_дата = [e for e in изделия if not e["project_smr_start_date"]]
    неразмечено_ярус = [e for e in изделия if e["zone_stance_level_id"] is None]

    этапы: dict = {}
    for e in изделия:
        ключ = (e["zone_crane_id"], e["zone_stance_id"], e["floor"])
        этап = этапы.setdefault(ключ, {
            "crane": _zone_label("Кран", e["crane_number"], e["crane_name"]),
            "stance": _zone_label("Стоянка", e["stance_number"], e["stance_name"]),
            "floor": e["floor"],
            "elevations": set(),
            "start": None,
            "elements": [],
        })
        этап["elements"].append(e)
        if e["level_mm"] is not None:
            этап["elevations"].add(e["level_mm"])
        нужно = _parse(e["project_smr_start_date"])
        if нужно and (этап["start"] is None or нужно < этап["start"]):
            этап["start"] = нужно

    # ---------- 1.1 обеспечение ближайших этапов ----------
    #
    # Раскладка контрактов по этапам считается ПО ВСЕМ этапам, а не только по
    # тем, что попали в горизонт: этапы, начавшиеся раньше, уже израсходовали
    # свою часть выкупленного, и не вычесть её значило бы посчитать одни и те
    # же изделия дважды.
    потребности_марка, потребности_тип, строки_этапов = [], [], {}
    for ключ_этапа, этап in этапы.items():
        по_позициям: dict = {}
        for e in этап["elements"]:
            позиция = (e["element_type"] or "", (e["mark"] or "").strip())
            по_позициям.setdefault(позиция, []).append(e)
        for (тип, марка), список in по_позициям.items():
            id_строки = (ключ_этапа, тип, марка)
            строки_этапов[id_строки] = {"stage": ключ_этапа, "element_type": тип,
                                        "mark": марка or None, "elements": список}
            порядок = (этап["start"] or date.max, str(ключ_этапа), тип, марка)
            if марка:
                потребности_марка.append(((тип, марка), этап["start"], порядок, len(список), id_строки))
            потребности_тип.append((тип, этап["start"], порядок, len(список), id_строки))

    выдано_марка = _allocate(потребности_марка, законтрактовано_марка)
    выдано_тип = _allocate(потребности_тип, законтрактовано_тип)

    def _готовность_строки(список: list) -> tuple:
        """Дата, к которой закроется вся позиция, и чем она получена.

        План поставки главнее расчёта (решение пользователя): где он есть,
        показывается он. Неизвестный темп делает неизвестной всю позицию —
        «часть посчитали, часть нет» было бы хуже молчания."""
        дата, источник, неизвестно = None, None, False
        for e in список:
            if e["actual_delivery_date"] or e["current_status"] in DONE_STATUSES:
                continue
            план = _parse(e["planned_delivery_date"])
            своя, свой_источник = (план, "plan") if план else (готовность.get(e["id"]), "queue")
            if своя is None:
                неизвестно = True
                continue
            if дата is None or своя > дата:
                дата, источник = своя, свой_источник
        if неизвестно:
            return None, "unknown"
        return дата, источник

    def _вердикт(дефицит: int, готово: Optional[date], источник: Optional[str],
                 старт: Optional[date], ждём: Optional[int] = None) -> dict:
        if дефицит <= 0:
            return {"code": "closed", "label": "закрыт"}
        # Изделия уже на площадке (или в конструкции), а контрактом позиция не
        # закрыта — это расхождение УЧЁТА, а не срыв стройки, и путать его с
        # «нечего монтировать» нельзя: иначе справка поднимает тревогу там,
        # где работа идёт своим ходом.
        if ждём == 0:
            return {"code": "paper", "label": "поставлено, не закрыто контрактом"}
        if старт is not None and старт <= today:
            return {"code": "started", "label": "СМР идёт, дефицит"}
        if источник == "unknown" or готово is None:
            return {"code": "unknown", "label": "сроки неизвестны"}
        if старт is None:
            return {"code": "no_date", "label": "нет даты СМР"}
        if готово <= старт:
            return {"code": "in_time", "label": "успеть можно"}
        return {"code": "late", "label": f"не успевает, +{(готово - старт).days} дн."}

    строки_горизонта = []
    for id_строки, строка in строки_этапов.items():
        этап = этапы[строка["stage"]]
        if этап["start"] is None or этап["start"] > конец_горизонта:
            continue
        # Полностью смонтированный этап из справки уходит: вопрос «чем его
        # обеспечить» для него уже решён стройкой.
        if all(e["current_status"] in DONE_STATUSES for e in этап["elements"]):
            continue
        нужно = len(строка["elements"])
        закрыто = выдано_марка.get(id_строки, 0) if строка["mark"] else выдано_тип.get(id_строки, 0)
        дефицит = max(нужно - закрыто, 0)
        ждём = sum(1 for e in строка["elements"]
                   if not e["actual_delivery_date"] and e["current_status"] not in DONE_STATUSES)
        готово, источник = _готовность_строки(строка["elements"])
        завод = None
        for e in строка["elements"]:
            завод = book.counterparty_for_contract(e["contract_id"])
            if завод:
                break
        строки_горизонта.append({
            "crane": этап["crane"], "stance": этап["stance"], "floor": этап["floor"],
            "elevation": _elevation_label(min(этап["elevations"]) if этап["elevations"] else None),
            "start": _iso(этап["start"]),
            "days_left": (этап["start"] - today).days if этап["start"] else None,
            "element_type": строка["element_type"] or None,
            "mark": строка["mark"],
            "need": нужно,
            "contracted": закрыто,
            "deficit": дефицит,
            "awaiting": ждём,
            "contracted_by_type": выдано_тип.get(id_строки, 0),
            "ready": _iso(готово),
            "ready_source": источник,
            "counterparty": завод,
            "verdict": _вердикт(дефицит, готово, источник, этап["start"], ждём),
        })
    строки_горизонта.sort(key=lambda r: (
        r["start"] or "9999", natural_key(r["crane"]), natural_key(r["stance"]),
        r["floor"] if r["floor"] is not None else 99, r["element_type"] or "", r["mark"] or ""))

    итог_горизонта = {
        "need": sum(r["need"] for r in строки_горизонта),
        "contracted": sum(r["contracted"] for r in строки_горизонта),
        "deficit": sum(r["deficit"] for r in строки_горизонта),
    }

    # ---------- 1.2 общий прогресс контрактации ----------
    по_типам: dict = {}
    for e in изделия:
        тип = e["element_type"] or ""
        свод = по_типам.setdefault(тип, {"element_type": тип, "need": 0, "credited": 0,
                                         "assigned": 0, "marks": {}})
        свод["need"] += 1
        свод["marks"][(e["mark"] or "").strip()] = свод["marks"].get((e["mark"] or "").strip(), 0) + 1
        if e["contract_id"]:
            свод["assigned"] += 1
    прогресс = []
    for тип, свод in по_типам.items():
        # Зачёт по марке: изделия закрыты ровно настолько, насколько
        # законтрактована ИХ марка. Изделия без марки в зачёт не идут — им
        # нечего сопоставить (и это видно в дефиците, а не спрятано).
        зачтено = sum(min(сколько, законтрактовано_марка.get((тип, марка), 0))
                      for марка, сколько in свод["marks"].items() if марка)
        всего_по_типу = законтрактовано_тип.get(тип, 0)
        прогресс.append({
            "element_type": тип or None,
            "need": свод["need"],
            "credited": зачтено,
            "deficit": max(свод["need"] - зачтено, 0),
            "percent": round(100 * зачтено / свод["need"]) if свод["need"] else 0,
            "contracted_total": всего_по_типу,
            "assigned": свод["assigned"],
            "not_credited": max(всего_по_типу - зачтено, 0),
        })
    прогресс.sort(key=lambda r: -r["need"])
    итог_прогресса = {
        "need": sum(r["need"] for r in прогресс),
        "credited": sum(r["credited"] for r in прогресс),
        "deficit": sum(r["deficit"] for r in прогресс),
        "contracted_total": sum(r["contracted_total"] for r in прогресс),
        "assigned": sum(r["assigned"] for r in прогресс),
        "not_credited": sum(r["not_credited"] for r in прогресс),
    }
    итог_прогресса["percent"] = (round(100 * итог_прогресса["credited"] / итог_прогресса["need"])
                                 if итог_прогресса["need"] else 0)

    # ---------- 2.1 фронт работ по стоянкам ----------
    стоянки: dict = {}
    for e in изделия:
        if e["zone_stance_level_id"] is None:
            continue     # без яруса изделие в логику «нижний → верхний» не встраивается
        ключ = (e["zone_crane_id"], e["zone_stance_id"])
        стоянка = стоянки.setdefault(ключ, {
            "crane": _zone_label("Кран", e["crane_number"], e["crane_name"]),
            "stance": _zone_label("Стоянка", e["stance_number"], e["stance_name"]),
            "levels": {},
        })
        стоянка["levels"].setdefault(e["level_mm"], []).append(e)

    def _разложение(список: list) -> dict:
        """Четыре колонки в СУММЕ дают потребность яруса — иначе таблицу
        нельзя прочитать. Поэтому смонтированное считается один раз и
        дальше не разбирается: изделие в конструкции, вопрос «поставлено ли
        оно» для него закрыт (а в накопленных данных встречаются
        смонтированные без фактической даты поставки — дата заполняется
        импортом, статус ставится руками)."""
        смонтировано = [e for e in список if e["current_status"] in DONE_STATUSES]
        осталось = [e for e in список if e["current_status"] not in DONE_STATUSES]
        на_площадке = sum(1 for e in осталось if e["actual_delivery_date"])
        ждём = [e for e in осталось if not e["actual_delivery_date"]]
        по_контракту = sum(1 for e in ждём if e["contract_id"])
        return {"need": len(список), "installed": len(смонтировано), "on_site": на_площадке,
                "contracted_not_delivered": по_контракту,
                "no_contract": len(ждём) - по_контракту}

    фронт = []
    for (crane_id, stance_id), стоянка in стоянки.items():
        уровни = sorted(стоянка["levels"].items(), key=lambda kv: (kv[0] is None, kv[0]))
        текущий = None
        for индекс, (отметка, список) in enumerate(уровни):
            if any(e["current_status"] not in DONE_STATUSES for e in список):
                текущий = индекс
                break
        if текущий is None:
            continue     # стоянка закрыта целиком — на критическом пути её нет
        for сдвиг, роль in ((0, "current"), (1, "next")):
            индекс = текущий + сдвиг
            if индекс >= len(уровни):
                continue
            отметка, список = уровни[индекс]
            строка = {"crane": стоянка["crane"], "stance": стоянка["stance"],
                      "elevation": _elevation_label(отметка), "role": роль,
                      "crane_id": crane_id, "stance_id": stance_id, "level_mm": отметка}
            строка.update(_разложение(список))
            дефицит_яруса = строка["need"] - строка["installed"]
            if роль == "next":
                строка["state"] = {"code": "waiting",
                                   "label": f"ждёт закрытия {_elevation_label(уровни[текущий][0])}"}
            elif строка["no_contract"]:
                строка["state"] = {"code": "no_contract", "label": "не обеспечен"}
            elif строка["contracted_not_delivered"]:
                строка["state"] = {"code": "waiting_delivery",
                                   "label": f"ждёт поставки {строка['contracted_not_delivered']} шт."}
            else:
                строка["state"] = {"code": "on_site", "label": f"хвост монтажа {дефицит_яруса} шт."}
            фронт.append(строка)
    фронт.sort(key=lambda r: (natural_key(r["crane"]), natural_key(r["stance"]),
                              r["level_mm"] if r["level_mm"] is not None else 10 ** 9))

    # ---------- 2.2 чего не хватает на критическом пути ----------
    текущие_ярусы = {(r["crane_id"], r["stance_id"], r["level_mm"]) for r in фронт if r["role"] == "current"}
    критические: dict = {}
    for e in изделия:
        ключ_яруса = (e["zone_crane_id"], e["zone_stance_id"], e["level_mm"])
        # На критическом пути — только то, чего физически нет на площадке и
        # что ещё не стоит в конструкции.
        if (ключ_яруса not in текущие_ярусы or e["actual_delivery_date"]
                or e["current_status"] in DONE_STATUSES):
            continue
        ключ = ключ_яруса + ((e["element_type"] or ""), (e["mark"] or "").strip())
        строка = критические.setdefault(ключ, {
            "crane": _zone_label("Кран", e["crane_number"], e["crane_name"]),
            "stance": _zone_label("Стоянка", e["stance_number"], e["stance_name"]),
            "elevation": _elevation_label(e["level_mm"]),
            "element_type": e["element_type"], "mark": e["mark"],
            "missing": 0, "with_contract": 0, "elements": [],
        })
        строка["missing"] += 1
        строка["elements"].append(e)
        if e["contract_id"]:
            строка["with_contract"] += 1

    критический_путь = []
    for строка in критические.values():
        нужно_к = min((_parse(e["project_smr_start_date"]) for e in строка["elements"]
                       if e["project_smr_start_date"]), default=None)
        план = min((_parse(e["planned_delivery_date"]) for e in строка["elements"]
                    if e["planned_delivery_date"]), default=None)
        готово, источник = _готовность_строки(строка["elements"])
        завод = next((book.counterparty_for_contract(e["contract_id"])
                      for e in строка["elements"] if e["contract_id"]), None)
        критический_путь.append({
            "crane": строка["crane"], "stance": строка["stance"], "elevation": строка["elevation"],
            "element_type": строка["element_type"], "mark": строка["mark"],
            "missing": строка["missing"],
            "with_contract": строка["with_contract"],
            "no_contract": строка["missing"] - строка["with_contract"],
            "counterparty": завод,
            "need_date": _iso(нужно_к),
            "plan_date": _iso(план),
            "ready": _iso(готово),
            "ready_source": источник,
            "verdict": _вердикт(строка["missing"], готово, источник, нужно_к),
        })
    критический_путь.sort(key=lambda r: (r["need_date"] or "9999", -r["missing"]))
    срезано = max(len(критический_путь) - CRITICAL_ROWS_LIMIT, 0)
    критический_путь = критический_путь[:CRITICAL_ROWS_LIMIT]

    # ---------- 2.3 динамика ----------
    динамика = _build_dynamics(conn, object_id, изделия, законтрактовано_марка, today)

    # ---------- 0. резюме ----------
    смонтировано_всего = sum(1 for e in изделия if e["current_status"] in DONE_STATUSES)
    этапов_без_контракта = sum(1 for r in строки_горизонта
                               if r["contracted"] == 0 and r["deficit"] > 0 and r["awaiting"] > 0)
    заблокировано = sum(1 for r in фронт if r["role"] == "current" and r["state"]["code"] != "on_site")
    показатели = [
        {"key": "contracting", "value": f"{итог_прогресса['percent']} %",
         "label": "законтрактовано от потребности",
         "hint": f"{итог_прогресса['credited']} из {итог_прогресса['need']} шт."},
        {"key": "stages_no_contract", "value": этапов_без_контракта,
         "label": "позиций в горизонте без контракта",
         "tone": "bad" if этапов_без_контракта else "ok"},
        {"key": "deficit_horizon", "value": итог_горизонта["deficit"],
         "label": f"дефицит на горизонт {горизонт} дн., шт.",
         "tone": "bad" if итог_горизонта["deficit"] else "ok"},
        {"key": "blocked", "value": заблокировано,
         "label": "фронтов не обеспечено",
         "tone": "bad" if заблокировано else "ok"},
        {"key": "installed", "value": f"{round(100 * смонтировано_всего / len(изделия)) if изделия else 0} %",
         "label": "смонтировано", "hint": f"{смонтировано_всего} шт."},
    ]

    выводы = _build_conclusions(строки_горизонта, фронт, критический_путь, прогресс,
                                неразмечено_дата, неразмечено_ярус, book, today)

    объект = conn.execute("SELECT name FROM objects WHERE id = ?", (object_id,)).fetchone()
    заметки = get_notes_for_date(conn, object_id, today.isoformat())

    return {
        "title": TITLE,
        "object_id": object_id,
        "object_name": объект["name"] if объект else None,
        "report_date": today.isoformat(),
        "horizon_days": горизонт,
        "horizon_end": конец_горизонта.isoformat(),
        "horizons": HORIZONS,
        "disclaimer": DISCLAIMER,
        "tiles": показатели,
        "conclusions": выводы,
        "notes": заметки,
        "stages": {"columns": STAGE_COLUMNS, "rows": строки_горизонта, "total": итог_горизонта},
        "progress": {"columns": PROGRESS_COLUMNS, "rows": прогресс, "total": итог_прогресса},
        "front": {"columns": FRONT_COLUMNS, "rows": фронт},
        "critical": {"columns": CRITICAL_COLUMNS, "rows": критический_путь, "truncated": срезано},
        "dynamics": динамика,
        "unmapped": {"no_smr_date": len(неразмечено_дата), "no_level": len(неразмечено_ярус)},
        # Заводы без заполненного норматива — причина всех «сроки неизвестны»
        # в отчёте; без этого списка непонятно, где чинить.
        "capacity_gaps": _capacity_gaps(изделия, book),
    }


def _capacity_gaps(изделия: list, book: CapacityBook) -> list:
    """Пары «завод + тип», по которым есть непоставленные изделия, а темпа
    нет. Ровно они превращают строки отчёта в «сроки неизвестны»."""
    дыры = {}
    for e in изделия:
        if e["actual_delivery_date"] or e["current_status"] in DONE_STATUSES:
            continue
        тип = e["element_type"] or ""
        if e["contract_id"]:
            if book.for_contract(e["contract_id"], тип):
                continue
            завод = book.counterparty_for_contract(e["contract_id"]) or "—"
        else:
            if book.average.get(тип):
                continue
            завод = None      # дефицит без контракта: считается средним по типу
        дыры[(завод, тип)] = дыры.get((завод, тип), 0) + 1
    return [{"counterparty": завод, "element_type": тип or None, "elements": сколько}
            for (завод, тип), сколько in sorted(дыры.items(), key=lambda kv: -kv[1])]


def _week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _build_dynamics(conn, object_id: int, изделия: list, законтрактовано_марка: dict,
                    today: date) -> dict:
    """Четыре накопительные кривые по неделям: потребность, контрактация,
    поставка, монтаж. Сжатая версия «Графика контрактации и поставки» —
    одна картинка на объект; подробная раскладка по маркам живёт там.

    Контрактация раскладывается по дате СПЕЦИФИКАЦИИ (своей даты у контракта
    нет и намеренно не будет, см. app/report_contracting.py), монтаж — по
    первому переходу изделия в «Смонтирован»/«Принят» в истории статусов.
    """
    точки = {"need": {}, "contracted": {}, "delivered": {}, "installed": {}}

    def плюс(шкала: str, d: Optional[date], сколько: int = 1):
        if d is None or not сколько:
            return
        неделя = _week_start(d)
        точки[шкала][неделя] = точки[шкала].get(неделя, 0) + сколько

    for e in изделия:
        плюс("need", _parse(e["project_smr_start_date"]))
        плюс("delivered", _parse(e["actual_delivery_date"]))

    for r in conn.execute(
        """
        SELECT cl.quantity AS quantity, s.specification_date AS spec_date
        FROM contract_lines cl
        JOIN contracts co ON co.id = cl.contract_id
        JOIN specifications s ON s.id = co.specification_id
        JOIN agreements a ON a.id = s.agreement_id
        WHERE a.object_id = ? AND co.is_archived = 0
        """,
        (object_id,),
    ):
        плюс("contracted", _parse(r["spec_date"]), r["quantity"])

    for r in conn.execute(
        f"""
        SELECT MIN(sh.changed_at) AS at
        FROM status_history sh
        JOIN elements e ON e.id = sh.element_id
        WHERE e.object_id = ? AND e.is_current = 1
          AND sh.status IN ({','.join('?' * len(DONE_STATUSES))})
        GROUP BY sh.element_id
        """,
        (object_id, *DONE_STATUSES),
    ):
        плюс("installed", _parse(r["at"]))

    все = [d for шкала in точки.values() for d in шкала]
    if not все:
        return {"weeks": [], "series": {k: [] for k in точки}, "today_index": None}
    недели, текущая = [], min(все)
    предел = max(max(все), _week_start(today))
    while текущая <= предел:
        недели.append(текущая)
        текущая += timedelta(days=7)

    ряды = {}
    for шкала, значения in точки.items():
        накопление, ряд = 0, []
        for неделя in недели:
            накопление += значения.get(неделя, 0)
            ряд.append(накопление)
        ряды[шкала] = ряд
    неделя_справки = _week_start(today)
    return {
        "weeks": [w.isoformat() for w in недели],
        "series": ряды,
        "today_index": недели.index(неделя_справки) if неделя_справки in недели else None,
    }


def _build_conclusions(строки_горизонта: list, фронт: list, критический_путь: list,
                       прогресс: list, неразмечено_дата: list, неразмечено_ярус: list,
                       book: CapacityBook, today: date) -> list:
    """Выводы считаются по правилам, а не пишутся человеком: рядом в справке
    есть блоки «События, задачи, вопросы», которые ведёт человек, и смешивать
    два разных по природе текста в одном списке нельзя — было бы не понять,
    что здесь наблюдение системы, а что мнение."""
    выводы = []

    for r in sorted(строки_горизонта, key=lambda r: (r["verdict"]["code"] != "started", -r["deficit"])):
        # «paper» в выводы не идёт: изделия на месте, стройке ничего не
        # мешает — это работа для того, кто ведёт контракты, и она видна в
        # таблице 1.1.
        if r["verdict"]["code"] in ("closed", "no_date", "paper"):
            continue
        адрес = f"{r['crane']} / {r['stance']}" + (f" / этаж {r['floor']}" if r["floor"] is not None else "")
        марка = f" {r['mark']}" if r["mark"] else ""
        когда = f"СМР с {r['start'][8:10]}.{r['start'][5:7]}" if r["start"] else "дата СМР не задана"
        хвост = ""
        if r["verdict"]["code"] == "unknown":
            хвост = " Срок изготовления посчитать не по чему: производительность завода не задана."
        elif r["ready"]:
            хвост = f" Расчётная готовность — {r['ready'][8:10]}.{r['ready'][5:7]}.{r['ready'][:4]}."
        выводы.append({
            "severity": "critical" if r["verdict"]["code"] in ("started", "late") else "warning",
            "text": (f"{адрес}: {когда}. {r['element_type'] or 'Тип не определён'}{марка} — "
                     f"не хватает {r['deficit']} шт."
                     f"{' Контракта нет.' if r['contracted'] == 0 else ''}{хвост}"),
        })
        if len(выводы) >= 6:
            break

    заблокированные = [r for r in фронт if r["role"] == "next"]
    for r in заблокированные[:2]:
        предыдущий = next((f for f in фронт
                           if f["role"] == "current" and f["crane"] == r["crane"]
                           and f["stance"] == r["stance"]), None)
        if not предыдущий:
            continue
        осталось = предыдущий["need"] - предыдущий["installed"]
        if осталось <= 0:
            continue
        выводы.append({
            "severity": "critical",
            "text": (f"{r['crane']} / {r['stance']}: ярус {r['elevation']} не начать — на ярусе "
                     f"{предыдущий['elevation']} не смонтировано {осталось} шт., из них "
                     f"{предыдущий['contracted_not_delivered'] + предыдущий['no_contract']} "
                     f"ещё не поставлено."),
        })

    for r in прогресс:
        if r["percent"] < 60 and r["deficit"] > 0:
            выводы.append({
                "severity": "warning",
                "text": (f"{r['element_type'] or 'Тип не определён'}: контрактация {r['percent']} % — "
                         f"не закрыто {r['deficit']} шт. из {r['need']}."),
            })
    if all(r["percent"] >= 95 for r in прогресс) and прогресс:
        выводы.append({"severity": "ok", "text": "Контрактация закрыта по всем типам изделий."})
    elif not any(в["severity"] == "critical" for в in выводы):
        закрытые = [r["element_type"] for r in прогресс if r["percent"] >= 95]
        if закрытые:
            выводы.append({"severity": "ok",
                           "text": "Контрактация закрыта: " + ", ".join(x or "—" for x in закрытые) + "."})

    if неразмечено_дата or неразмечено_ярус:
        выводы.append({
            "severity": "data",
            "text": (f"Не размечено: без даты начала СМР — {len(неразмечено_дата)} шт., "
                     f"без привязки к ярусу — {len(неразмечено_ярус)} шт. "
                     f"В расчёт горизонта и критического пути они не входят."),
        })
    if not book.average:
        выводы.append({
            "severity": "data",
            "text": ("Производительность заводов не заполнена ни по одному типу — сроки изготовления "
                     "не рассчитываются. Заполняется в карточке контрагента, закладка "
                     "«Производительность»."),
        })
    return выводы


# ---------- выгрузка того же отчёта в файлы ----------
#
# Обе функции получают УЖЕ ПОСТРОЕННЫЙ отчёт, а не строят его заново: иначе
# числа на экране, в Excel и в PDF со временем разошлись бы (общее правило
# всех отчётов проекта).


def _cell_value(row: dict, column: dict):
    """Значение колонки в виде, годном и для Excel, и для PDF. «verdict» —
    словарь {code, label}: в файл уходит подпись, код нужен только экрану
    (он им красит)."""
    значение = row.get(column["key"])
    if column["kind"] == "verdict":
        return (значение or {}).get("label") or ""
    return значение


def _subtitle(report: dict) -> str:
    return (f"{report.get('object_name') or 'Объект не задан'} · на "
            f"{_ru(report['report_date'])} · горизонт {report['horizon_days']} дн. "
            f"(до {_ru(report['horizon_end'])}) · по объекту целиком, без фильтра схемы")


def _ru(iso: Optional[str]) -> str:
    d = _parse(iso)
    return d.strftime("%d.%m.%Y") if d else "—"


def build_analytics_report_xlsx(report: dict) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    жирный = Font(bold=True)
    шапка = PatternFill("solid", fgColor="F2F4F7")

    def лист(name: str, columns: list, rows: list, total: Optional[dict] = None,
             total_label: str = "Итого"):
        ws = wb.create_sheet(name)
        ws.append([c["label"] for c in columns])
        for ячейка in ws[1]:
            ячейка.font = жирный
            ячейка.fill = шапка
            ячейка.alignment = Alignment(horizontal="center", wrap_text=True)
        for row in rows:
            строка = []
            for c in columns:
                значение = _cell_value(row, c)
                # Дата кладётся НАСТОЯЩЕЙ датой: по текстовой не отсортировать
                # и не отфильтровать, а этот отчёт в Excel именно крутят.
                строка.append(_parse(значение) if c["kind"] == "date" and значение else значение)
            ws.append(строка)
            for i, c in enumerate(columns, start=1):
                if c["kind"] == "date":
                    ws.cell(row=ws.max_row, column=i).number_format = "DD.MM.YYYY"
        if total:
            ws.append([total_label] + [
                total.get(c["key"]) if c["key"] in total else None for c in columns[1:]])
            for ячейка in ws[ws.max_row]:
                ячейка.font = жирный
        for i, c in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(11, min(len(c["label"]) + 2, 28))
        ws.freeze_panes = "A2"
        return ws

    # Первым листом — резюме: с него отчёт и читают.
    ws = wb.active
    ws.title = "Резюме"
    ws.append([report["title"]])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([_subtitle(report)])
    ws.append([])
    for плитка in report["tiles"]:
        ws.append([плитка["label"], плитка["value"], плитка.get("hint") or ""])
    ws.append([])
    ws.append(["Выводы"])
    ws.cell(row=ws.max_row, column=1).font = жирный
    for вывод in report["conclusions"]:
        ws.append([вывод["text"]])
    заметки = report.get("notes") or {}
    for ключ, подпись in (("key_events", "Ключевые события"), ("key_tasks", "Ключевые задачи"),
                          ("open_questions", "Открытые вопросы")):
        пункты = заметки.get(ключ) or []
        if not пункты:
            continue
        ws.append([])
        ws.append([подпись])
        ws.cell(row=ws.max_row, column=1).font = жирный
        for пункт in пункты:
            ws.append([пункт])
    ws.append([])
    ws.append([report["disclaimer"]])
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30

    лист("Ближайшие этапы", report["stages"]["columns"], report["stages"]["rows"],
         report["stages"]["total"], "Итого по горизонту")
    лист("Контрактация", report["progress"]["columns"], report["progress"]["rows"],
         report["progress"]["total"])
    лист("Фронт работ", report["front"]["columns"], report["front"]["rows"])
    лист("Критический путь", report["critical"]["columns"], report["critical"]["rows"])

    динамика = report.get("dynamics") or {}
    if динамика.get("weeks"):
        ws = wb.create_sheet("Динамика")
        ws.append(["Неделя"] + [SERIES_LABELS[k] for k in SERIES_LABELS])
        for ячейка in ws[1]:
            ячейка.font = жирный
        for i, неделя in enumerate(динамика["weeks"]):
            ws.append([_parse(неделя)] + [динамика["series"][k][i] for k in SERIES_LABELS])
            ws.cell(row=ws.max_row, column=1).number_format = "DD.MM.YYYY"
        ws.column_dimensions["A"].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_analytics_report_pdf(report: dict) -> bytes:
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.pdf_export import FONT_BOLD, FONT_REGULAR

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=8 * mm, rightMargin=8 * mm, topMargin=8 * mm, bottomMargin=8 * mm,
        title=report["title"],
    )
    ширина = doc.width
    заголовок = ParagraphStyle("t", fontName=FONT_BOLD, fontSize=14, leading=18)
    раздел = ParagraphStyle("h2", fontName=FONT_BOLD, fontSize=10, leading=13,
                            spaceBefore=4 * mm, spaceAfter=1.5 * mm)
    мелкий = ParagraphStyle("s", fontName=FONT_REGULAR, fontSize=8, leading=11,
                            textColor=colors.HexColor("#555555"))
    обычный = ParagraphStyle("p", fontName=FONT_REGULAR, fontSize=8.5, leading=11)
    ячейка = ParagraphStyle("c", fontName=FONT_REGULAR, fontSize=6.5, leading=8)
    шапка = ParagraphStyle("h", fontName=FONT_BOLD, fontSize=6.5, leading=8, alignment=1)

    ТОН = {"critical": "#c0392b", "warning": "#a86500", "ok": "#1e7e34", "data": "#6b7280"}

    story = [Paragraph(pdf_text(report["title"]), заголовок),
             Paragraph(pdf_text(_subtitle(report)), мелкий), Spacer(1, 3 * mm)]

    # Показатели — одной строкой-таблицей: они и на экране читаются как ряд
    # плиток, а не как список.
    плитки = [[Paragraph(f"<b>{pdf_text(t['value'])}</b><br/>{pdf_text(t['label'])}"
                         + (f"<br/>{pdf_text(t['hint'])}" if t.get("hint") else ""), обычный)
               for t in report["tiles"]]]
    таблица = Table(плитки, colWidths=[ширина / len(report["tiles"])] * len(report["tiles"]))
    таблица.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#dde1e6")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dde1e6")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"), ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story += [таблица, Spacer(1, 3 * mm), Paragraph("Выводы", раздел)]
    for вывод in report["conclusions"]:
        story.append(Paragraph(
            f'<font color="{ТОН.get(вывод["severity"], "#000000")}">■</font> {pdf_text(вывод["text"])}',
            обычный))

    заметки = report.get("notes") or {}
    for ключ, подпись in (("key_events", "Ключевые события"), ("key_tasks", "Ключевые задачи"),
                          ("open_questions", "Открытые вопросы")):
        пункты = заметки.get(ключ) or []
        if not пункты:
            continue
        story.append(Paragraph(подпись, раздел))
        for пункт in пункты:
            story.append(Paragraph("— " + pdf_text(пункт), обычный))

    def секция(title: str, columns: list, rows: list, note: str = "",
               total: Optional[dict] = None, total_label: str = "Итого"):
        story.append(Paragraph(pdf_text(title), раздел))
        if note:
            story.append(Paragraph(pdf_text(note), мелкий))
        if not rows:
            story.append(Paragraph("Нет данных", мелкий))
            return
        данные = [[Paragraph(pdf_text(c["label"]), шапка) for c in columns]]
        for row in rows:
            данные.append([
                Paragraph(pdf_text(_ru(_cell_value(row, c)) if c["kind"] == "date"
                                   else _cell_value(row, c)), ячейка)
                for c in columns
            ])
        if total:
            данные.append([Paragraph(f"<b>{pdf_text(total_label)}</b>", ячейка)] + [
                Paragraph(f"<b>{pdf_text(total.get(c['key']))}</b>" if c["key"] in total else "",
                          ячейка) for c in columns[1:]])
        t = Table(данные, colWidths=[ширина / len(columns)] * len(columns), repeatRows=1)
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#dde1e6")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2f4f7")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 1.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ]))
        story.append(t)

    # В PDF идут только строки с дефицитом: документ читают как справку, а
    # закрытые позиции в нём — это сотни строк, ради которых теряется
    # смысл. Сколько их скрыто, сказано явно; полный перечень есть в XLSX.
    дефицитные = [r for r in report["stages"]["rows"] if r["deficit"] > 0]
    скрыто = len(report["stages"]["rows"]) - len(дефицитные)
    секция("1.1. Обеспечение ближайших этапов СМР", report["stages"]["columns"], дефицитные,
           f"Показаны позиции с дефицитом; закрытых позиций скрыто: {скрыто}."
           if скрыто else "", report["stages"]["total"], "Итого по горизонту")
    секция("1.2. Общий прогресс контрактации", report["progress"]["columns"],
           report["progress"]["rows"], "", report["progress"]["total"])
    секция("2.1. Фронт работ по стоянкам", report["front"]["columns"], report["front"]["rows"],
           "Текущий незакрытый ярус каждой стоянки и следующий за ним.")
    секция("2.2. Чего не хватает на критическом пути", report["critical"]["columns"],
           report["critical"]["rows"],
           f"Показано {CRITICAL_ROWS_LIMIT} строк из {CRITICAL_ROWS_LIMIT + report['critical']['truncated']}."
           if report["critical"]["truncated"] else "")

    неразмечено = report.get("unmapped") or {}
    if неразмечено.get("no_smr_date") or неразмечено.get("no_level"):
        story.append(Paragraph(
            f"Не размечено: без даты начала СМР — {неразмечено.get('no_smr_date', 0)} шт., "
            f"без привязки к ярусу — {неразмечено.get('no_level', 0)} шт. "
            f"В расчёт горизонта и критического пути они не входят.", мелкий))
    story += [Spacer(1, 2 * mm), Paragraph(pdf_text(report["disclaimer"]), мелкий)]

    doc.build(story)
    return buf.getvalue()
