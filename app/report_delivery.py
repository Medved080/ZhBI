"""
Отчёт «График поставки» (живой запрос 2026-08-02).

Смысл: все ЖБ изделия должны быть на объекте К ДАТЕ НАЧАЛА СМР
(`elements.project_smr_start_date`, заполняется импортом графика MS
Project) — значит из этой даты раскладывается календарная ПОТРЕБНОСТЬ в
поставке. По горизонтали — календарь (день / неделя / месяц), по
вертикали — иерархия группировки, СОСТАВ И ПОРЯДОК КОТОРОЙ ЗАДАЁТ
ПОЛЬЗОВАТЕЛЬ (одному нужен разрез по поставщикам, другому — по захваткам
и этажам; фиксированная иерархия, как в отчёте «Статусы», здесь не
годится).

В ячейке ТРИ числа (живой запрос того же дня) — три независимые шкалы
дат одного и того же изделия:
  потребность — `project_smr_start_date`, «к этой дате изделие обязано
                быть на площадке» (проектный график);
  план        — `planned_delivery_date`, на какую дату поставку реально
                запланировали (контракт/логистика);
  факт        — `actual_delivery_date`, когда поставили на самом деле.
Изделие с потребностью на 5 августа, запланированное на 7-е и
поставленное 9-го, даёт по одной единице в ТРИ РАЗНЫЕ колонки — ровно так
и видно, где расхождение. Складывать шкалы в одну колонку значило бы
прятать то, ради чего отчёт и нужен.

Изделия вне выбранного периода не отбрасываются, а попадают в служебные
колонки «Ранее» / «Позже» (решение пользователя): иначе итог справа
перестал бы сходиться с реальным объёмом группы, и сужение периода молча
меняло бы общее число изделий. Третья служебная колонка — «Без даты»:
проектная дата на реальном файле есть не у всех изделий, и молча их
терять нельзя (та же честность, что у `plan_coverage` в «Динамике»).
Факт в «Без даты» не попадает никогда: изделие без фактической даты
просто не поставлено, а не «поставлено неизвестно когда».

Разбор ячейки по маркам (`build_delivery_cell_detail`) считается ОТДЕЛЬНЫМ
запросом при наведении мыши, а не кладётся в тело отчёта: на реальном
файле это тысячи троек «строка × колонка × марка», и в общий ответ они не
помещаются. Ячейка адресуется не подписью строки, а сырыми значениями
уровней группировки (`gkey` каждого узла) — подписи генерируются (имя
контракта собирается из реквизитов) и ключом быть не могут.

Данные, как и у остальных отчётов, считает СЕРВЕР — экран, XLSX и PDF
берут результат одной и той же функции и разойтись не могут.
"""

from datetime import date, timedelta
from typing import Optional

from app.db import visible_elements_clause
from app.reports import NO_FLOOR, NO_ZAKHVATKA, _floor_label, _item_label, natural_key, pdf_text

TITLE = "График поставки ЖБИ"
ROOT_LABEL = "Группировка"
TOTAL_LABEL = "Итого"

NO_COUNTERPARTY = "Контрагент не определён"
NO_CONTRACT = "Контракт не назначен"
NO_STANCE = "Стоянка не определена"
NO_STANCE_IN_CRANE = "вне стоянок"
NO_MARK = "без марки"

# Три шкалы одной ячейки. Порядок задаёт и порядок чисел на экране, и
# порядок колонок в Excel — он один на все три представления.
SCALES = [
    {"key": "need", "label": "Потребность", "column": "project_smr_start_date",
     "hint": "к дате начала СМР"},
    {"key": "plan", "label": "План", "column": "planned_delivery_date",
     "hint": "плановая дата поставки"},
    {"key": "fact", "label": "Факт", "column": "actual_delivery_date",
     "hint": "фактическая поставка"},
]
SCALE_KEYS = [s["key"] for s in SCALES]
NEED, PLAN, FACT = 0, 1, 2

# Порядок в списке — порядок по умолчанию в форме выбора; сам ПОРЯДОК
# группировки задаёт пользователь и присылает его в group_by.
GROUPS = [
    {"key": "counterparty", "label": "Контрагент"},
    {"key": "contract", "label": "Контракт"},
    {"key": "zakhvatka", "label": "Захватка"},
    {"key": "stance", "label": "Стоянка"},
    {"key": "floor", "label": "Этаж"},
    {"key": "type", "label": "Тип/подтип элемента"},
    # Марка — уточнение типа, поэтому идёт следом за ним. Именно по марке
    # изделия взаимозаменяемы, и именно она нужна, когда график читают
    # «чего конкретно не хватает» (2026-08-02, живой запрос).
    {"key": "mark", "label": "Марка"},
]
GROUP_KEYS = [g["key"] for g in GROUPS]
GROUP_LABELS = {g["key"]: g["label"] for g in GROUPS}
DEFAULT_GROUPS = ["counterparty", "contract", "type"]

STEPS = ("day", "week", "month")
STEP_LABELS = {"day": "День", "week": "Неделя", "month": "Месяц"}

# Потолок числа календарных колонок. Не «чтобы красиво»: 400 колонок ×
# сотни строк — это и вес ответа, и намертво повисшая таблица в браузере.
# Упереться в него можно только явно попросив дневной шаг на многолетнем
# периоде — тогда честнее сказать об этом, чем молча подменить шаг.
MAX_PERIOD_COLUMNS = 400

# Границы автоподбора шага, когда клиент шаг не прислал (первое открытие).
AUTO_DAY_MAX = 45
AUTO_WEEK_MAX = 200

MONTHS_SHORT = ["янв", "фев", "мар", "апр", "май", "июн",
                "июл", "авг", "сен", "окт", "ноя", "дек"]

BEFORE_KEY, AFTER_KEY, NONE_KEY = "before", "after", "none"

# Статус изделия, которое физически лежит на площадке и ещё не пущено в
# дело, — единственный кандидат на перемещение с другой стоянки.
AVAILABLE_STATUS = "delivered"


# ---------- перекрыта ли потребность (живой запрос 2026-08-02) ----------
#
# Потребность изделия перекрыта, если оно окажется (или уже оказалось) на
# площадке НЕ ПОЗЖЕ своей даты начала СМР. Считается по САМОМУ изделию, а
# не сравнением чисел в ячейке: поставка неделей раньше срока перекрывает
# потребность точно так же, как поставка день в день, а сравнение
# «потребность против плана В ТОЙ ЖЕ КОЛОНКЕ» объявило бы её непокрытой.
#
# Факт ВАЖНЕЕ плана: если изделие уже приехало, судим по тому, когда оно
# приехало на самом деле. План в срок при опоздавшем факте потребность НЕ
# перекрывает — изделия к сроку на площадке не было, и красить такую
# ячейку белым значило бы прятать настоящий срыв.
def _covered(need_d: Optional[date], plan_d: Optional[date], fact_d: Optional[date]) -> bool:
    if need_d is None:
        return True     # сравнивать не с чем — это колонка «Без даты»
    if fact_d is not None:
        return fact_d <= need_d
    if plan_d is not None:
        return plan_d <= need_d
    return False        # ни плана, ни факта — потребность ничем не закрыта


# Та же формула в SQL — для разбора ячейки по маркам. substr(...,1,10):
# даты текстовые, у части записей исторически со временем.
_COVERED_SQL = """
    (CASE WHEN e.project_smr_start_date IS NULL THEN 1
          WHEN e.actual_delivery_date IS NOT NULL
               THEN substr(e.actual_delivery_date,1,10) <= substr(e.project_smr_start_date,1,10)
          WHEN e.planned_delivery_date IS NOT NULL
               THEN substr(e.planned_delivery_date,1,10) <= substr(e.project_smr_start_date,1,10)
          ELSE 0 END)
"""


# ---------- календарная сетка ----------

def _parse(value: Optional[str]) -> Optional[date]:
    try:
        return date.fromisoformat(value[:10]) if value else None
    except (ValueError, TypeError):
        return None


def _bucket_start(d: date, step: str) -> date:
    """Начало периода, в который попадает дата. Ключ колонки — ISO-дата
    этого начала (а не «2026-W32»/«2026-08»): так ключи сортируются как
    обычные даты и одинаково устроены при любом шаге."""
    if step == "week":
        return d - timedelta(days=d.weekday())
    if step == "month":
        return d.replace(day=1)
    return d


def _bucket_next(d: date, step: str) -> date:
    if step == "week":
        return d + timedelta(days=7)
    if step == "month":
        return date(d.year + 1, 1, 1) if d.month == 12 else date(d.year, d.month + 1, 1)
    return d + timedelta(days=1)


def _period_label(d: date, step: str) -> str:
    if step == "month":
        return f"{MONTHS_SHORT[d.month - 1]} {d.year}"
    if step == "week":
        end = d + timedelta(days=6)
        return f"{d.day:02d}.{d.month:02d}–{end.day:02d}.{end.month:02d}"
    return f"{d.day:02d}.{d.month:02d}"


def _ru(d) -> str:
    d = d if isinstance(d, date) else _parse(d)
    return f"{d.day:02d}.{d.month:02d}.{d.year}" if d else ""


def _build_columns(start: date, end: date, step: str) -> list:
    """Календарь строится СПЛОШНЫМ, включая дни без единой поставки: это
    график, а не список событий — пустой день в середине недели такой же
    результат, как заполненный."""
    cols, cur = [], _bucket_start(start, step)
    while cur <= end:
        cols.append({
            "key": cur.isoformat(),
            "label": _period_label(cur, step),
            "kind": "period",
            "date": cur.isoformat(),
            # Выходной подсвечивается только при дневном шаге — у недели и
            # месяца понятия «выходной» нет.
            "weekend": step == "day" and cur.weekday() >= 5,
        })
        cur = _bucket_next(cur, step)
    return cols


def column_bounds(column: str, start: date, end: date, step: str):
    """Границы дат колонки — (lo, hi), любая может быть None (открыта).
    Возвращает False для колонки «Без даты»: у неё границ нет вовсе, там
    условие «дата не задана». Общая для отчёта и для разбора ячейки."""
    if column == NONE_KEY:
        return False
    if column == BEFORE_KEY:
        return (None, start - timedelta(days=1))
    if column == AFTER_KEY:
        return (end + timedelta(days=1), None)
    d = _parse(column)
    if d is None:
        return None
    return (max(d, start), min(_bucket_next(d, step) - timedelta(days=1), end))


# ---------- разбор параметров ----------

def _normalize_groups(group_by: Optional[list]) -> list:
    """Чужие ключи молча отбрасываются, дубли схлопываются, пустой список
    заменяется значением по умолчанию: параметр приходит с клиента, и
    строить SQL/дерево по непроверенному списку нельзя."""
    if not group_by:
        return list(DEFAULT_GROUPS)
    seen, out = set(), []
    for key in group_by:
        if key in GROUP_KEYS and key not in seen:
            seen.add(key)
            out.append(key)
    return out or list(DEFAULT_GROUPS)


def _stance_label(stance: Optional[str], crane: Optional[str]) -> str:
    """Стоянка неизвестна, но кран известен — «Кран 2 · вне стоянок», а не
    общее «не определено»: тот же смысл, что у псевдо-значения
    `no-stance:<craneId>` в фильтрах схемы (см. CLAUDE.md)."""
    if stance:
        return f"{crane} · {stance}" if crane else stance
    if crane:
        return f"{crane} · {NO_STANCE_IN_CRANE}"
    return NO_STANCE


def _site_address(crane: Optional[str], stance: Optional[str], floor) -> str:
    """Физический адрес на площадке — «Кран · Стоянка · Этаж» (решение
    пользователя). Именно по нему изделие ищут ногами, и он не зависит от
    того, какая группировка выбрана в отчёте."""
    return " · ".join([_stance_label(stance, crane), _floor_label(floor)])


_NO_VALUE_LABELS = {NO_COUNTERPARTY, NO_CONTRACT, NO_ZAKHVATKA, NO_STANCE, NO_FLOOR, NO_MARK}


def _sort_key(label: str):
    # «не определено» — всегда в конец группы, при любом уровне.
    return (1 if label in _NO_VALUE_LABELS else 0, natural_key(label))


# Один и тот же набор JOIN-ов нужен и отчёту, и разбору ячейки: если их
# развести, условия группировки в одном месте начнут ссылаться на алиасы,
# которых нет в другом.
_JOINS = """
        FROM elements e
        LEFT JOIN zones zz ON zz.id = e.zone_zakhvatka_id
        LEFT JOIN zones zs ON zs.id = e.zone_stance_id
        LEFT JOIN zones zc ON zc.id = e.zone_crane_id
        LEFT JOIN contracts c ON c.id = e.contract_id
        LEFT JOIN specifications sp ON sp.id = c.specification_id
        LEFT JOIN agreements ag ON ag.id = sp.agreement_id
        LEFT JOIN counterparties cp ON cp.id = ag.counterparty_id
"""


def _base_filters(source_file: Optional[str], element_ids: Optional[list]):
    clauses, params = [visible_elements_clause("e")], []
    if source_file:
        clauses.append("e.source_file = ?")
        params.append(source_file)
    if element_ids is not None:
        if not element_ids:
            clauses.append("1=0")
        else:
            clauses.append(f"e.id IN ({','.join('?' * len(element_ids))})")
            params.extend(element_ids)
    return clauses, params


# ---------- сам отчёт ----------

def build_delivery_schedule_report(
    conn,
    source_file: Optional[str] = None,
    element_ids: Optional[list] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    step: Optional[str] = None,
    group_by: Optional[list] = None,
) -> dict:
    from app.contracts import build_contract_name

    groups = _normalize_groups(group_by)
    clauses, params = _base_filters(source_file, element_ids)
    where = f"WHERE {' AND '.join(clauses)}"

    # Один запрос со ВСЕМИ разрезами сразу, а не по запросу на выбранную
    # группировку: набор строк тот же (группировка лишь схлопывает их
    # сильнее или слабее), а так смена порядка уровней не переписывает SQL.
    rows = conn.execute(
        f"""
        SELECT cp.id AS cp_id, cp.short_name AS cp_name,
               ag.number AS ag_number, ag.agreement_date AS ag_date,
               sp.number AS sp_number, sp.specification_date AS sp_date,
               c.id AS contract_id, c.theme AS contract_theme,
               e.zone_zakhvatka_id AS zakhvatka_id, zz.name AS zakhvatka,
               e.zone_stance_id AS stance_id, zs.name AS stance,
               e.zone_crane_id AS crane_id, zc.name AS crane,
               e.floor AS floor, e.element_type AS element_type, e.subtype AS subtype,
               e.mark AS mark,
               e.project_smr_start_date AS need_date,
               e.planned_delivery_date AS plan_date,
               e.actual_delivery_date AS fact_date,
               COUNT(*) AS n
        {_JOINS}
        {where}
        GROUP BY cp.id, ag.number, ag.agreement_date, sp.number, sp.specification_date,
                 c.id, c.theme, e.zone_zakhvatka_id, e.zone_stance_id, e.zone_crane_id,
                 e.floor, e.element_type, e.subtype, e.mark,
                 e.project_smr_start_date, e.planned_delivery_date, e.actual_delivery_date
        """,
        params,
    ).fetchall()

    known = [d for r in rows for d in
             (_parse(r["need_date"]), _parse(r["plan_date"]), _parse(r["fact_date"])) if d]

    start = _parse(date_from) or (min(known) if known else date.today())
    end = _parse(date_to) or (max(known) if known else date.today())
    if end < start:
        start, end = end, start

    span = (end - start).days + 1
    if step not in STEPS:
        # Первое открытие: шаг подбирается по ширине периода. Полгода по
        # дням — это 180 колонок, читать нечего; переключатель шага рядом.
        step = "day" if span <= AUTO_DAY_MAX else ("week" if span <= AUTO_WEEK_MAX else "month")

    period_columns = _build_columns(start, end, step)
    if len(period_columns) > MAX_PERIOD_COLUMNS:
        raise ValueError(
            f"Период даёт {len(period_columns)} колонок при шаге «{STEP_LABELS[step]}» — "
            f"это больше {MAX_PERIOD_COLUMNS}. Сузьте период или укрупните шаг."
        )

    columns = (
        [{"key": BEFORE_KEY, "label": "Ранее", "kind": "edge",
          "title": f"Раньше {_ru(start)}"}]
        + period_columns
        + [{"key": AFTER_KEY, "label": "Позже", "kind": "edge",
            "title": f"Позже {_ru(end)}"},
           {"key": NONE_KEY, "label": "Без даты", "kind": "edge",
            "title": "Потребность — без даты начала СМР, план — без плановой даты "
                     "поставки. Факт сюда не попадает: изделие без фактической "
                     "даты просто не поставлено."}]
    )
    valid_keys = {c["key"] for c in columns}

    def column_of(value: Optional[str]) -> Optional[str]:
        d = _parse(value)
        if d is None:
            return None
        if d < start:
            return BEFORE_KEY
        if d > end:
            return AFTER_KEY
        key = _bucket_start(d, step).isoformat()
        # Страховка: начало периода первой колонки может лежать ЛЕВЕЕ start
        # (неделя/месяц, в который попадает start) — такая дата всё равно
        # относится к первой колонке, а не «Ранее».
        return key if key in valid_keys else BEFORE_KEY

    # Дерево: узел = {values: {ключ колонки: [потребность, план, факт]}}.
    # Разрежённо (нулевые ячейки не хранятся) — при 180 колонках и сотнях
    # строк плотная матрица была бы в основном нулями.
    def new_node(label, level, gkey=None):
        # gaps — сколько изделий этой ячейки НЕ перекрыты (см. _covered).
        # Отдельной картой, а не четвёртым числом в values: три числа
        # ячейки — это шкалы, а непокрытая потребность их подсветка, и в
        # Excel/PDF она тоже оформление, а не ещё одна колонка.
        return {"label": label, "level": level, "gkey": gkey,
                "values": {}, "total": [0, 0, 0],
                "gaps": {}, "gap_total": 0, "children": {}}

    root = new_node(TOTAL_LABEL, -1)

    def add(node, col, idx, n):
        # col=None бывает только у факта (изделие не поставлено) и у плана
        # без плановой даты — тогда не попадает ни в ячейку, ни в итог
        # справа: иначе «Итого» по факту равнялся бы общему числу изделий,
        # а не числу поставленных.
        if col is None:
            return
        cell = node["values"].setdefault(col, [0, 0, 0])
        cell[idx] += n
        node["total"][idx] += n

    def add_gap(node, col, n):
        node["gaps"][col] = node["gaps"].get(col, 0) + n
        node["gap_total"] += n

    def label_and_key(key, r):
        """Подпись уровня и его СЫРОЕ значение. Сырое нужно для разбора
        ячейки при наведении: подписи генерируются (имя контракта — из
        реквизитов) и обратно в условие отбора не разбираются."""
        if key == "counterparty":
            return (r["cp_name"] or NO_COUNTERPARTY), r["cp_id"]
        if key == "contract":
            if not r["contract_id"]:
                return NO_CONTRACT, None
            return build_contract_name(
                r["cp_name"] or NO_COUNTERPARTY, r["ag_number"] or "", r["ag_date"],
                r["sp_number"] or "", r["sp_date"], r["contract_theme"]), r["contract_id"]
        if key == "zakhvatka":
            return (r["zakhvatka"] or NO_ZAKHVATKA), r["zakhvatka_id"]
        if key == "stance":
            return _stance_label(r["stance"], r["crane"]), [r["stance_id"], r["crane_id"]]
        if key == "floor":
            return _floor_label(r["floor"]), r["floor"]
        if key == "mark":
            return (r["mark"] or NO_MARK), r["mark"]
        return _item_label(r["element_type"], r["subtype"]), [r["element_type"], r["subtype"]]

    need_with_date = 0
    plan_with_date = 0
    total_elements = 0
    for r in rows:
        n = r["n"]
        total_elements += n
        cols_by_scale = [column_of(r["need_date"]), column_of(r["plan_date"]),
                         column_of(r["fact_date"])]
        if cols_by_scale[NEED] is None:
            cols_by_scale[NEED] = NONE_KEY  # изделие без проектной даты
        else:
            need_with_date += n
        if cols_by_scale[PLAN] is None:
            cols_by_scale[PLAN] = NONE_KEY  # плановая дата поставки не назначена
        else:
            plan_with_date += n
        # Факт без даты не значит «поставлено неизвестно когда» — значит «не
        # поставлено»; такой элемент в факт не попадает вовсе.

        # Непокрытая потребность считается по САМОМУ изделию и ложится в
        # колонку его ПОТРЕБНОСТИ: подсвечивается тот срок, к которому
        # изделия не будет, а не тот, когда его привезут.
        gap = 0 if _covered(_parse(r["need_date"]), _parse(r["plan_date"]),
                            _parse(r["fact_date"])) else n

        node = root
        for idx, col in enumerate(cols_by_scale):
            add(node, col, idx, n)
        if gap:
            add_gap(node, cols_by_scale[NEED], gap)
        for level, key in enumerate(groups):
            label, gkey = label_and_key(key, r)
            node = node["children"].setdefault(label, new_node(label, level, gkey))
            for idx, col in enumerate(cols_by_scale):
                add(node, col, idx, n)
            if gap:
                add_gap(node, cols_by_scale[NEED], gap)

    def finish(node) -> dict:
        children = [finish(node["children"][k]) for k in sorted(node["children"], key=_sort_key)]
        return {"label": node["label"], "level": node["level"], "gkey": node["gkey"],
                "values": node["values"], "total": node["total"],
                "gaps": node["gaps"], "gap_total": node["gap_total"], "children": children}

    tree = finish(root)

    report = {
        "title": TITLE,
        "root_label": ROOT_LABEL,
        "total_label": TOTAL_LABEL,
        "scales": SCALES,
        "step": step,
        "step_label": STEP_LABELS[step],
        "date_from": start.isoformat(),
        "date_to": end.isoformat(),
        "group_by": groups,
        "group_labels": [GROUP_LABELS[k] for k in groups],
        "columns": columns,
        "rows": tree["children"],
        "total": {"label": TOTAL_LABEL, "values": tree["values"], "total": tree["total"],
                  "gaps": tree["gaps"], "gap_total": tree["gap_total"]},
        # Та же честная пометка, что в «Динамике»: проектная дата задана не
        # у всех изделий, и график по части объёма внешне неотличим от
        # полного.
        "plan_coverage": {"need": need_with_date, "plan": plan_with_date, "total": total_elements},
    }
    # Подпись и предупреждение о неполноте плана считаются ЗДЕСЬ и уходят
    # готовым текстом: их показывают экран, Excel и PDF, и три копии одной
    # фразы разошлись бы на первой же правке формулировки.
    report["subtitle"] = subtitle(report)
    report["warning"] = coverage_warning(report)
    return report


def flatten(report: dict) -> list:
    out = []

    def walk(node):
        out.append(node)
        for child in node.get("children", []):
            walk(child)

    for row in report["rows"]:
        walk(row)
    return out


def coverage_warning(report: dict) -> str:
    cov = report["plan_coverage"]
    parts = []
    if cov["need"] < cov["total"]:
        parts.append(f"дата начала СМР задана у {cov['need']} изделий из {cov['total']}")
    if cov["plan"] < cov["total"]:
        parts.append(f"плановая дата поставки — у {cov['plan']}")
    if not parts:
        return ""
    return ("Внимание: " + ", ".join(parts) +
            " — остальные показаны в колонке «Без даты». График неполный.")


def subtitle(report: dict) -> str:
    return (f"Период {_ru(report['date_from'])} — {_ru(report['date_to'])}, "
            f"шаг «{report['step_label']}». Группировка: {' → '.join(report['group_labels'])}. "
            f"В ячейке: " + " / ".join(s["label"].lower() for s in SCALES) + ".")


# ---------- разбор одной ячейки по маркам (подсказка при наведении) ----------

def _group_conditions(groups: list, path: list):
    """Условия отбора по УРОВНЯМ группировки из сырых значений (`gkey`).
    Везде `IS`, а не `=`: значение уровня сплошь и рядом NULL («контрагент
    не определён», «этаж не определён»), и обычное сравнение вернуло бы
    NULL вместо истины."""
    clauses, params = [], []
    for key, value in zip(groups, path or []):
        if key == "counterparty":
            clauses.append("cp.id IS ?")
            params.append(value)
        elif key == "contract":
            clauses.append("e.contract_id IS ?")
            params.append(value)
        elif key == "zakhvatka":
            clauses.append("e.zone_zakhvatka_id IS ?")
            params.append(value)
        elif key == "stance":
            stance_id, crane_id = (value or [None, None])[:2]
            clauses.append("e.zone_stance_id IS ? AND e.zone_crane_id IS ?")
            params.extend([stance_id, crane_id])
        elif key == "floor":
            clauses.append("e.floor IS ?")
            params.append(value)
        elif key == "mark":
            clauses.append("e.mark IS ?")
            params.append(value)
        elif key == "type":
            element_type, subtype = (value or [None, None])[:2]
            clauses.append("e.element_type IS ? AND e.subtype IS ?")
            params.extend([element_type, subtype])
    return clauses, params


def _scale_in_column(column_field: str, bounds, column: str):
    """Фрагмент SQL «дата этой шкалы попадает в колонку» + его параметры.
    substr(...,1,10) — даты в этих колонках текстовые, у части записей
    исторически со временем; сравнивать надо ровно день."""
    if bounds is False:  # колонка «Без даты»
        return f"{column_field} IS NULL", []
    lo, hi = bounds
    parts, params = [f"{column_field} IS NOT NULL"], []
    if lo is not None:
        parts.append(f"substr({column_field},1,10) >= ?")
        params.append(lo.isoformat())
    if hi is not None:
        parts.append(f"substr({column_field},1,10) <= ?")
        params.append(hi.isoformat())
    return "(" + " AND ".join(parts) + ")", params


def build_delivery_cell_detail(
    conn,
    source_file: Optional[str] = None,
    element_ids: Optional[list] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    step: Optional[str] = None,
    group_by: Optional[list] = None,
    path: Optional[list] = None,
    column: Optional[str] = None,
) -> dict:
    """Что стоит за одной ячейкой: разбивка по маркам и, если чего-то не
    хватает, откуда это можно взять.

    «Не перекрыто» = изделия, которые требуются к дате колонки в ЭТОЙ
    группировке и к своему сроку на площадке не окажутся (см. `_covered`) —
    та же величина, которой подсвечена сама ячейка. «Можно взять» = изделия той же марки
    со статусом «Доставлен» (лежат на площадке и ещё не смонтированы),
    находящиеся ВНЕ этой ячейки — их адрес даётся физически, «Кран ·
    Стоянка · Этаж» (решение пользователя), а не подписью строки отчёта:
    ногами изделие ищут именно так, и от выбранной группировки это не
    зависит. У кого собственный срок СМР раньше даты колонки, помечается
    отдельно — забрать такое изделие значит сорвать соседний участок."""
    groups = _normalize_groups(group_by)
    start, end = _parse(date_from), _parse(date_to)
    if start is None or end is None:
        raise ValueError("Не задан период отчёта")
    if step not in STEPS:
        step = "day"
    bounds = column_bounds(column, start, end, step)
    if bounds is None:
        raise ValueError("Неизвестная колонка отчёта")

    base_clauses, base_params = _base_filters(source_file, element_ids)
    group_clauses, group_params = _group_conditions(groups, path)

    need_sql, need_params = _scale_in_column("e.project_smr_start_date", bounds, column)
    plan_sql, plan_params = _scale_in_column("e.planned_delivery_date", bounds, column)
    # Факт в колонку «Без даты» не попадает (см. модуль выше) — там условие
    # заведомо ложно, а не «фактической даты нет».
    if bounds is False:
        fact_sql, fact_params = "0", []
    else:
        fact_sql, fact_params = _scale_in_column("e.actual_delivery_date", bounds, column)

    in_cell = f"({need_sql} OR {plan_sql} OR {fact_sql})"
    in_cell_params = need_params + plan_params + fact_params

    where_cell = " AND ".join(base_clauses + group_clauses + [in_cell])
    marks = conn.execute(
        f"""
        SELECT e.mark AS mark,
               SUM(CASE WHEN {need_sql} THEN 1 ELSE 0 END) AS need,
               SUM(CASE WHEN {plan_sql} THEN 1 ELSE 0 END) AS plan,
               SUM(CASE WHEN {fact_sql} THEN 1 ELSE 0 END) AS fact,
               SUM(CASE WHEN {need_sql} AND e.actual_delivery_date IS NOT NULL
                        THEN 1 ELSE 0 END) AS need_delivered,
               SUM(CASE WHEN {need_sql} AND NOT {_COVERED_SQL}
                        THEN 1 ELSE 0 END) AS uncovered
        {_JOINS}
        WHERE {where_cell}
        GROUP BY e.mark
        """,
        need_params + plan_params + fact_params + need_params + need_params
        + base_params + group_params + in_cell_params,
    ).fetchall()

    # Общая потребность в этих марках на ту же дату ПО ВСЕМУ отбору отчёта
    # (без условий группировки) — «сколько таких изделий вообще нужно к
    # этой дате», ради чего подсказка и заводилась.
    total_need = {
        r["mark"]: r["n"] for r in conn.execute(
            f"SELECT e.mark AS mark, COUNT(*) AS n {_JOINS} "
            f"WHERE {' AND '.join(base_clauses + [need_sql])} GROUP BY e.mark",
            base_params + need_params).fetchall()
    }

    # Источники ищем ровно для того, что не перекрыто: именно эти изделия и
    # надо чем-то закрыть к сроку.
    deficits = {r["mark"]: r["uncovered"] for r in marks if r["uncovered"] > 0}
    sources: dict = {}
    if deficits:
        keys = list(deficits)
        placeholders = ",".join("?" * len(keys))
        # NOT(...) — исключение самой ячейки: изделия, которые уже посчитаны
        # в «поставлено», не могут быть одновременно и источником.
        not_cell = f"NOT ({' AND '.join(group_clauses + [in_cell])})" if group_clauses else f"NOT {in_cell}"
        rows = conn.execute(
            f"""
            SELECT e.mark AS mark, zc.name AS crane, zs.name AS stance, e.floor AS floor,
                   COUNT(*) AS n, MIN(e.project_smr_start_date) AS earliest_need
            {_JOINS}
            WHERE {' AND '.join(base_clauses)} AND e.current_status = ?
              AND e.mark IN ({placeholders}) AND {not_cell}
            GROUP BY e.mark, zc.name, zs.name, e.floor
            ORDER BY n DESC
            """,
            base_params + [AVAILABLE_STATUS] + keys + group_params + in_cell_params,
        ).fetchall()
        limit_date = bounds[1] if bounds is not False and bounds[1] else end
        for r in rows:
            earliest = _parse(r["earliest_need"])
            sources.setdefault(r["mark"], []).append({
                "where": _site_address(r["crane"], r["stance"], r["floor"]),
                "count": r["n"],
                "earliest_need": r["earliest_need"][:10] if r["earliest_need"] else None,
                # Срок самого источника не позже даты колонки — забирать
                # такое изделие значит сорвать чужой участок.
                "urgent": bool(earliest and limit_date and earliest <= limit_date),
            })

    # Сначала то, чего не хватает больше всего: подсказка обрезается по
    # высоте экрана, и в неё должно попасть самое важное, а не то, чья
    # марка раньше по алфавиту.
    def mark_order(r):
        return (-r["uncovered"], -r["need"], natural_key(r["mark"] or NO_MARK))

    out_marks = []
    for r in sorted(marks, key=mark_order):
        mark = r["mark"]
        out_marks.append({
            "mark": mark or NO_MARK,
            "need": r["need"], "plan": r["plan"], "fact": r["fact"],
            "delivered": r["need_delivered"],
            "deficit": r["uncovered"],
            "total_need": total_need.get(mark, r["need"]),
            "sources": sources.get(mark, []),
        })

    col_label = ("без даты" if column == NONE_KEY
                 else "ранее" if column == BEFORE_KEY
                 else "позже" if column == AFTER_KEY
                 else (_ru(bounds[0]) if bounds[0] == bounds[1]
                       else f"{_ru(bounds[0])} — {_ru(bounds[1])}"))
    return {"column": column, "column_label": col_label, "marks": out_marks,
            "available_status_label": "доставлено, не смонтировано"}


# ---------- выгрузка ----------
#
# Обе функции получают УЖЕ ПОСТРОЕННЫЙ отчёт, а не строят его заново —
# иначе числа на экране, в Excel и в PDF со временем разошлись бы.

def build_delivery_schedule_xlsx(report: dict) -> bytes:
    """В Excel каждая шкала — СВОЯ колонка под общей объединённой шапкой
    даты, а не «12/8/5» одной строкой: в файле с числами работают формулами
    и сводными, и текст сделал бы это невозможным."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "График поставки"

    thin = Side(style="thin", color="D5D8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EEF2F7")
    edge_fill = PatternFill("solid", fgColor="F5F0E6")
    # Непокрытая потребность — заливкой, а не отдельной колонкой: это
    # подсветка тех же трёх чисел, а не четвёртая шкала.
    gap_fill = PatternFill("solid", fgColor="FBE3E0")

    ws.append([report["title"]])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([report["subtitle"]])
    warn = report["warning"]
    if warn:
        ws.append([warn])
        ws.cell(row=ws.max_row, column=1).font = Font(color="C0392B")

    columns = report["columns"]
    width = len(SCALES)
    # +2, а не append([]) + max_row: пустой append в openpyxl не двигает
    # max_row, и разделительная строка молча пропадала бы.
    head_row = ws.max_row + 2
    ws.cell(row=head_row, column=1, value=report["root_label"])
    ws.merge_cells(start_row=head_row, start_column=1, end_row=head_row + 1, end_column=1)

    def head_block(col_index: int, label: str):
        ws.cell(row=head_row, column=col_index, value=label)
        ws.merge_cells(start_row=head_row, start_column=col_index,
                       end_row=head_row, end_column=col_index + width - 1)
        for k, scale in enumerate(SCALES):
            ws.cell(row=head_row + 1, column=col_index + k, value=scale["label"].lower())

    for i, col in enumerate(columns):
        head_block(2 + i * width, col["label"])
    total_col = 2 + len(columns) * width
    head_block(total_col, report["total_label"])
    last_col = total_col + width - 1

    for row in (head_row, head_row + 1):
        for i in range(1, last_col + 1):
            cell = ws.cell(row=row, column=i)
            cell.font = Font(bold=True)
            cell.fill = head_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if i > 1 else "left",
                                       vertical="center", wrap_text=True)

    edge_cols = set()
    for i, col in enumerate(columns):
        if col["kind"] == "edge":
            edge_cols |= {2 + i * width + k for k in range(width)}

    def write_row(label, values, total, indent, gaps=None, gap_total=0):
        ws.append([label])
        r = ws.max_row
        # Отступ вложенности — свойством ячейки, а не пробелами в тексте:
        # текст остаётся пригодным для фильтров и формул (как в «Статусах»).
        ws.cell(row=r, column=1).alignment = Alignment(indent=indent)
        gap_cols = set()
        for i, col in enumerate(columns):
            trio = values.get(col["key"]) or [0] * width
            for k in range(width):
                ws.cell(row=r, column=2 + i * width + k, value=trio[k] or None)
            if (gaps or {}).get(col["key"]):
                gap_cols |= {2 + i * width + k for k in range(width)}
        for k in range(width):
            ws.cell(row=r, column=total_col + k, value=total[k] or None)
        if gap_total:
            gap_cols |= {total_col + k for k in range(width)}
        for i in range(1, last_col + 1):
            ws.cell(row=r, column=i).border = border
            if i in gap_cols:
                ws.cell(row=r, column=i).fill = gap_fill
            elif i in edge_cols:
                ws.cell(row=r, column=i).fill = edge_fill
        return r

    for node in flatten(report):
        r = write_row(node["label"], node["values"], node["total"], node["level"] * 2,
                      node.get("gaps"), node.get("gap_total", 0))
        # Группировка строк «+/−» — тем же приёмом, что в отчёте «Статусы».
        if node["level"] > 0:
            ws.row_dimensions[r].outlineLevel = min(node["level"], 7)
        if node["level"] < 2:
            for i in range(1, last_col + 1):
                ws.cell(row=r, column=i).font = Font(bold=True)

    total = report["total"]
    r = write_row(total["label"], total["values"], total["total"], 0,
                  total.get("gaps"), total.get("gap_total", 0))
    for i in range(1, last_col + 1):
        ws.cell(row=r, column=i).font = Font(bold=True)
        ws.cell(row=r, column=i).fill = head_fill

    ws.column_dimensions["A"].width = 46
    for i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.freeze_panes = ws.cell(row=head_row + 2, column=2)
    ws.sheet_properties.outlinePr.summaryBelow = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Сколько календарных колонок помещается на альбомном A4 читаемым кеглем.
PDF_COLUMNS_PER_BLOCK = 12


def build_delivery_schedule_pdf(report: dict) -> bytes:
    """Календарь режется на блоки по PDF_COLUMNS_PER_BLOCK колонок, каждый
    блок — своя таблица с повтором колонки группировки и итога. Иначе 180
    дневных колонок пришлось бы либо сжать до нечитаемых, либо обрезать
    молча — а обрезанный отчёт выглядит как полный."""
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.pdf_export import FONT_BOLD, FONT_REGULAR

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
        title=report["title"],
    )
    title_style = ParagraphStyle("t", fontName=FONT_BOLD, fontSize=13, leading=17)
    sub_style = ParagraphStyle("s", fontName=FONT_REGULAR, fontSize=8, leading=11,
                               textColor=colors.HexColor("#666666"))
    warn_style = ParagraphStyle("w", parent=sub_style, textColor=colors.HexColor("#C0392B"))

    story = [Paragraph(pdf_text(report["title"]), title_style),
             Paragraph(pdf_text(report["subtitle"]), sub_style)]
    warn = report["warning"]
    if warn:
        story.append(Paragraph(pdf_text(warn), warn_style))
    story.append(Spacer(1, 4 * mm))

    nodes = flatten(report) + [{"label": report["total"]["label"], "level": -1,
                                "values": report["total"]["values"],
                                "total": report["total"]["total"],
                                "gaps": report["total"].get("gaps", {}),
                                "gap_total": report["total"].get("gap_total", 0)}]
    columns = report["columns"]

    def cell(trio):
        if not trio or not any(trio):
            return ""
        return "/".join(str(v or 0) for v in trio)

    for offset in range(0, len(columns), PDF_COLUMNS_PER_BLOCK):
        block = columns[offset:offset + PDF_COLUMNS_PER_BLOCK]
        data = [[report["root_label"]] + [c["label"] for c in block] + [report["total_label"]]]
        styles = [
            ("FONTNAME", (0, 0), (-1, -1), FONT_REGULAR),
            ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2F7")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5D8DC")),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (1, 0), (-1, -1), 1), ("RIGHTPADDING", (1, 0), (-1, -1), 1),
        ]
        for node in nodes:
            indent = " " * (max(node["level"], 0) * 3)
            data.append([indent + node["label"]]
                        + [cell(node["values"].get(c["key"])) for c in block]
                        + [cell(node["total"])])
            # Непокрытая потребность — та же подсветка, что на экране и в
            # Excel: без неё PDF показывал бы ровно те же числа без главного.
            gaps = node.get("gaps") or {}
            for j, c in enumerate(block):
                if gaps.get(c["key"]):
                    styles.append(("BACKGROUND", (j + 1, len(data) - 1), (j + 1, len(data) - 1),
                                   colors.HexColor("#FBE3E0")))
            if node.get("gap_total"):
                styles.append(("BACKGROUND", (len(block) + 1, len(data) - 1),
                               (len(block) + 1, len(data) - 1), colors.HexColor("#FBE3E0")))
            if node["level"] < 2:
                styles.append(("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1), FONT_BOLD))
            if node["level"] < 0:
                styles.append(("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1),
                               colors.HexColor("#EEF2F7")))
        widths = [66 * mm] + [(211 * mm) / (len(block) + 1)] * (len(block) + 1)
        table = Table(data, colWidths=widths, repeatRows=1)
        table.setStyle(TableStyle(styles))
        story.append(table)
        story.append(Spacer(1, 5 * mm))

    doc.build(story)
    return buf.getvalue()
