"""
Массовая правка КОНТРАКТАЦИИ через Excel (2026-08-10, живой запрос).

Третий режим той же формы, что правка реквизитов
(app/element_bulk_edit.py) и истории статусов (app/status_bulk_edit.py):
выгрузить -> поправить снаружи -> загрузить -> отметить флажками, что
применить. Отдаёт и принимает те же структуры
(columns/elements/changes/rejected), поэтому табличный экран подтверждения
переиспользуется целиком, без второй реализации.

**Строка = ПОЗИЦИЯ контракта** (`contract_lines`), а рядом в той же строке
— реквизиты всех её владельцев по цепочке контракт -> спецификация ->
договор -> контрагент. Реестром, а не деревом: в Excel правят таблицу, и
единственный способ показать реквизиты владельцев рядом с позицией — их
повторение в каждой строке. Отсюда же главная особенность разбора, которой
нет у двух прежних режимов: **одна сущность-владелец встречается в десятках
строк**, и если строки предлагают ей РАЗНЫЕ значения, применить нельзя ни
одно — такое расхождение отклоняется (`_owner_conflicts`), а не решается
«кто последний, тот и прав».

Правится всё, кроме порождаемых значений: позиция (тип, марка,
количество), тема контракта, номер и дата спецификации, номер и дата
договора, все реквизиты контрагента. Наименование контракта не правится —
оно целиком генерируется из цепочки (build_contract_name), и колонка в
файле справочная.

**Новые позиции заводятся файлом** (решение пользователя 2026-08-10, по
образцу отдельного инструмента «Импорт контрактации из XLS»): строка с
пустым «№ позиции» и заполненным «№ контракта» — новая позиция этого
контракта. А вот новые контракты, спецификации, договоры и контрагенты
файлом НЕ заводятся: для этого есть импорт контрактации, который умеет
разбирать «НОМЕР от ДАТА» и заводить всю цепочку. Здесь же неизвестный
номер контракта — почти всегда опечатка, и заводить по ней документ
означало бы мусор в справочнике.

**Удалить позицию файлом нельзя** — как и запись истории, и контракт у
элемента в двух прежних режимах: удаление это отдельное осознанное
действие (оно есть в форме контракта), а стёртая строка в Excel слишком
дёшева. Строка, удалённая из файла, просто не рассматривается; пустая
ячейка означает «не трогать», а не «очистить».

Ключ строки — id позиции. У него есть известная слабость: правка контракта
через форму переписывает его позиции целиком (DELETE + INSERT, см.
app/contracts.update_contract), то есть id позиций после такой правки
другие. Поэтому строка сверяется НЕ ТОЛЬКО по id: у найденной позиции
обязан совпасть и номер контракта, иначе строка отклоняется — иначе правка
ушла бы в чужую позицию, чей id случайно совпал.
"""

import io
import sqlite3
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app import activity, contract_guard, impersonation
from app.contracts import build_contract_name
from app.element_fields import EXCEL_DATE_FORMAT, to_excel_date
from app.models import ZHBI_ELEMENT_TYPES

SHEET_DATA = "Контрактация"
SHEET_CONTRACTS = "Контракты"
SHEET_TYPES = "Типы элементов"
SHEET_OBJECTS = "Объекты"

KEY_COLUMN = "line_id"
CONTRACT_KEY_COLUMN = "contract_id"

# (ключ, подпись, правимая ли). Порядок тот, в котором читают: сначала
# ключи и «что это», потом сама позиция, потом её владельцы сверху вниз по
# цепочке — контракт, спецификация, договор, контрагент.
COLUMNS = [
    (KEY_COLUMN, "№ позиции (не менять)", False),
    (CONTRACT_KEY_COLUMN, "№ контракта (не менять)", False),
    ("contract_name", "Контракт (справочно)", False),
    ("object_name", "Объект договора (справочно)", False),
    ("element_type", "Тип элемента", True),
    ("mark", "Марка", True),
    ("quantity", "Количество", True),
    ("theme", "Тема контракта", True),
    ("spec_number", "Номер спецификации", True),
    ("spec_date", "Дата спецификации", True),
    ("agr_number", "Номер договора", True),
    ("agr_date", "Дата договора", True),
    ("cp_short_name", "Контрагент (краткое наименование)", True),
    ("cp_full_name", "Контрагент (полное наименование)", True),
    ("cp_inn", "ИНН", True),
    ("cp_kpp", "КПП", True),
    ("cp_ogrn", "ОГРН", True),
    ("cp_legal_address", "Юридический адрес", True),
    ("cp_contact_person", "Контактное лицо", True),
    ("cp_contact_phone", "Телефон", True),
    ("cp_code", "Код контрагента", True),
]

EDITABLE = [key for key, _, editable in COLUMNS if editable]

FIELD_LABELS = {key: label for key, label, _ in COLUMNS}
FIELD_LABELS["__new_line__"] = "Новая позиция"

# Куда пишется каждое правимое поле: (таблица, колонка БД). Одна таблица —
# один владелец, и по ней же считается, какой сущности принадлежит правка
# (см. _owner_key): без этого две строки одной спецификации с разными
# номерами применились бы обе, последняя поверх первой.
FIELD_TARGETS = {
    "element_type": ("contract_lines", "element_type"),
    "mark": ("contract_lines", "mark"),
    "quantity": ("contract_lines", "quantity"),
    "theme": ("contracts", "theme"),
    "spec_number": ("specifications", "number"),
    "spec_date": ("specifications", "specification_date"),
    "agr_number": ("agreements", "number"),
    "agr_date": ("agreements", "agreement_date"),
    "cp_short_name": ("counterparties", "short_name"),
    "cp_full_name": ("counterparties", "full_name"),
    "cp_inn": ("counterparties", "inn"),
    "cp_kpp": ("counterparties", "kpp"),
    "cp_ogrn": ("counterparties", "ogrn"),
    "cp_legal_address": ("counterparties", "legal_address"),
    "cp_contact_person": ("counterparties", "contact_person"),
    "cp_contact_phone": ("counterparties", "contact_phone"),
    "cp_code": ("counterparties", "code"),
}

# Правимые сущности и их человеческие имена — для текстов отказов и
# журнала. Именительный падеж и кавычки по месту употребления: склонять
# названия сущностей в шаблонах строк («одной и той же контрагента»)
# получается только уродливо.
OWNER_LABELS = {
    "contract_lines": "Позиция",
    "contracts": "Контракт",
    "specifications": "Спецификация",
    "agreements": "Договор",
    "counterparties": "Контрагент",
}

DATE_FIELDS = {"spec_date", "agr_date"}
INT_FIELDS = {"quantity"}

# Поля, обязательные для НОВОЙ позиции: без них она бессмысленна — позиция
# отвечает «сколько чего законтрактовано», и марка вдобавок участвует в
# ключе позиции (idx_contract_lines_unique: контракт + тип + марка).
NEW_LINE_REQUIRED = ("mark", "quantity")


def columns_spec() -> list:
    return [{"key": k, "label": l, "editable": e} for k, l, e in COLUMNS]


# ---------------------------------------------------------------- выгрузка

def _rows(conn) -> list:
    """Все позиции всех контрактов ОДНИМ запросом с JOIN на всю цепочку
    владельцев. Именно одним: отдельный запрос за контрагентом на каждую
    позицию — тот же N+1, что уже стоил проекту 2,7 секунды на другой
    форме.

    Контракт БЕЗ позиций тоже попадает в файл — строкой с пустым «№
    позиции» (LEFT JOIN): иначе в только что заведённый контракт нельзя
    было бы добавить первую позицию файлом, ради чего новые позиции и
    заводятся."""
    return conn.execute(
        """
        SELECT l.id AS line_id, l.element_type, l.mark, l.quantity,
               co.id AS contract_id, co.theme,
               s.id AS specification_id, s.number AS spec_number, s.specification_date AS spec_date,
               a.id AS agreement_id, a.number AS agr_number, a.agreement_date AS agr_date,
               cp.id AS counterparty_id, cp.short_name AS cp_short_name, cp.full_name AS cp_full_name,
               cp.inn AS cp_inn, cp.kpp AS cp_kpp, cp.ogrn AS cp_ogrn,
               cp.legal_address AS cp_legal_address, cp.contact_person AS cp_contact_person,
               cp.contact_phone AS cp_contact_phone, cp.code AS cp_code,
               o.name AS object_name
        FROM contracts co
        JOIN specifications s ON s.id = co.specification_id
        JOIN agreements a ON a.id = s.agreement_id
        JOIN counterparties cp ON cp.id = a.counterparty_id
        LEFT JOIN objects o ON o.id = a.object_id
        LEFT JOIN contract_lines l ON l.contract_id = co.id
        ORDER BY cp.short_name, a.number, s.number, co.id, l.element_type, l.mark, l.id
        """
    ).fetchall()


def _display_values(row) -> dict:
    """Значения одной строки по всем колонкам COLUMNS.

    Одна функция на выгрузку в XLS и на экран подтверждения — тот же приём,
    что в двух прежних режимах: своя сборка на экране разошлась бы с файлом
    при первой же правке состава колонок."""
    values = {key: (row[key] if key in row.keys() else None) for key, _, _ in COLUMNS}
    values["contract_name"] = build_contract_name(
        row["cp_short_name"], row["agr_number"], row["agr_date"],
        row["spec_number"], row["spec_date"], row["theme"],
    )
    return values


def build_contracting_workbook(conn) -> Workbook:
    """Снимок контрактации на текущий момент: лист данных + листы
    справочников, с выпадающим списком в колонке «Тип элемента».

    Отбора по фильтру схемы здесь нет и быть не может: строка файла — не
    элемент, а позиция контракта, и «элементы, прошедшие фильтр» к ней
    отношения не имеют. Форма в этом режиме галочку «Учитывать фильтр»
    прячет."""
    rows = _rows(conn)
    # Типы — из константы, а не из справочника подтипов: тот стал объектным
    # и у нового объекта пуст (см. app/element_bulk_edit.build_export_workbook).
    types = list(ZHBI_ELEMENT_TYPES)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA
    ws.append([label for _, label, _ in COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    столбцы_дат = [i + 1 for i, (key, _, _) in enumerate(COLUMNS) if key in DATE_FIELDS]
    # Номер строки считаем САМИ, а не через ws.max_row: в openpyxl это не
    # счётчик, а max() по всем ячейкам листа — O(n) на каждое обращение, и
    # на выгрузке реквизитов такое обращение в цикле стоило 88 секунд.
    for номер, row in enumerate(rows, start=2):
        values = _display_values(row)
        ws.append([to_excel_date(values[key]) if key in DATE_FIELDS else values[key]
                   for key, _, _ in COLUMNS])
        for i in столбцы_дат:
            ws.cell(row=номер, column=i).number_format = EXCEL_DATE_FORMAT

    # ---- листы справочников ----
    ws_c = wb.create_sheet(SHEET_CONTRACTS)
    ws_c.append(["№ контракта", "Наименование", "Контрагент", "Договор", "Спецификация", "Тема"])
    seen = set()
    for r in rows:
        if r["contract_id"] in seen:
            continue
        seen.add(r["contract_id"])
        v = _display_values(r)
        ws_c.append([r["contract_id"], v["contract_name"], r["cp_short_name"],
                     r["agr_number"], r["spec_number"], r["theme"]])
    ws_c.append([])
    ws_c.append(["Строка = одна позиция контракта. Пустой «№ позиции» + заполненный "
                 "«№ контракта» = новая позиция этого контракта."])
    ws_c.append(["Новые контракты, спецификации, договоры и контрагенты этим файлом не "
                 "заводятся — для этого есть «Импорт контрактации из XLS»."])
    ws_c.append(["Удалить позицию файлом нельзя — это делается в форме контракта."])

    ws_t = wb.create_sheet(SHEET_TYPES)
    ws_t.append(["Тип элемента"])
    for t in types:
        ws_t.append([t])

    ws_o = wb.create_sheet(SHEET_OBJECTS)
    ws_o.append(["Объект", "Описание"])
    for r in conn.execute("SELECT name, COALESCE(description,'') AS description FROM objects ORDER BY name"):
        ws_o.append([r["name"], r["description"]])

    _add_dropdowns(ws, len(rows), types)
    for sheet in (ws, ws_c, ws_t, ws_o):
        _widen(sheet)
    return wb


def _col_letter(key: str) -> str:
    return get_column_letter([k for k, _, _ in COLUMNS].index(key) + 1)


def _add_dropdowns(ws, n_rows: int, types) -> None:
    """Выпадающий список типа элемента. Валидация Excel при этом НЕ гарантия
    (вставка через буфер её обходит), поэтому то же значение проверяет
    сервер: список — удобство, проверка — обязанность."""
    if not types:
        return
    last = max(n_rows + 1, 2)
    dv = DataValidation(type="list", formula1=f"'{SHEET_TYPES}'!$A$2:$A${len(types) + 1}",
                        allow_blank=True)
    dv.error = "Значение должно быть выбрано из списка на листе справочника."
    dv.errorTitle = "Недопустимое значение"
    ws.add_data_validation(dv)
    col = _col_letter("element_type")
    dv.add(f"{col}2:{col}{last}")


def _widen(ws) -> None:
    for i, column in enumerate(ws.iter_cols(), start=1):
        width = max((len(str(c.value)) for c in column if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 45)


# ---------------------------------------------------------------- разбор

def _read_sheet(file_bytes: bytes) -> list:
    """Строки листа данных словарями по КЛЮЧАМ колонок. Сопоставление
    колонок — по ПОДПИСИ, а не по порядку: пользователь имеет полное право
    вставить свою колонку с пометками или скрыть ненужные."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if SHEET_DATA not in wb.sheetnames:
        raise ValueError(f"В файле нет листа «{SHEET_DATA}». "
                         f"Загружайте тот файл, который выгрузила система в режиме «Контрактация».")
    ws = wb[SHEET_DATA]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError(f"Лист «{SHEET_DATA}» пуст")

    label_to_key = {label: key for key, label, _ in COLUMNS}
    index = {}
    for i, cell in enumerate(header):
        key = label_to_key.get(str(cell).strip() if cell is not None else "")
        if key:
            index[key] = i
    for обязательная in (KEY_COLUMN, CONTRACT_KEY_COLUMN):
        if обязательная not in index:
            raise ValueError(f"В файле нет колонки «{FIELD_LABELS[обязательная]}» — "
                             f"без неё строки не с чем сопоставить.")

    out = []
    for n, raw in enumerate(rows, start=2):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        out.append((n, {key: (raw[i] if i < len(raw) else None) for key, i in index.items()}))
    return out


def _text(raw) -> Optional[str]:
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None


def _int_id(raw) -> Optional[int]:
    """Ячейка -> целый идентификатор. Excel отдаёт числа float'ом (1503.0),
    поэтому через float, а не прямым int()."""
    text = _text(raw)
    if text is None:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _coerce(field: str, raw):
    """Ячейка -> значение для БД. Пустая ячейка это «не трогать» (её
    обрабатывает вызывающий), сюда она приходит только непустой."""
    if field in INT_FIELDS:
        if isinstance(raw, float) and raw.is_integer():
            значение = int(raw)
        else:
            try:
                значение = int(str(raw).strip())
            except (TypeError, ValueError):
                raise ValueError(f"«{FIELD_LABELS[field]}» должно быть целым числом, получено «{raw}»")
        if значение < 0:
            raise ValueError(f"«{FIELD_LABELS[field]}» не может быть отрицательным")
        return значение
    if field in DATE_FIELDS:
        if hasattr(raw, "strftime"):   # date/datetime от openpyxl
            return raw.strftime("%Y-%m-%d")
        text = str(raw).strip()
        try:
            datetime.strptime(text, "%Y-%m-%d")
        except ValueError:
            raise ValueError(f"«{FIELD_LABELS[field]}»: ожидается дата в виде ГГГГ-ММ-ДД "
                             f"или ячейка в формате даты, получено «{text}»")
        return text
    return str(raw).strip()


def _owner_key(row, field: str) -> tuple:
    """Какой СУЩНОСТИ принадлежит правка этого поля: (таблица, id). Позиции
    контракта у новой строки ещё нет — у неё ключ по номеру строки файла,
    чтобы две новые позиции не считались одной и той же сущностью."""
    table, _ = FIELD_TARGETS[field]
    id_by_table = {
        "contract_lines": row["line_id"],
        "contracts": row["contract_id"],
        "specifications": row["specification_id"],
        "agreements": row["agreement_id"],
        "counterparties": row["counterparty_id"],
    }
    return table, id_by_table[table]


def _row_id(line_id: Optional[int], line_no: int) -> str:
    """Идентификатор СТРОКИ экрана подтверждения: позиция или, у новой, —
    номер строки файла."""
    return f"l{line_id}" if line_id is not None else f"new{line_no}"


def analyze(conn, file_bytes: bytes) -> dict:
    """Сверяет файл с базой. Ничего не пишет."""
    parsed = _read_sheet(file_bytes)
    rows = _rows(conn)
    by_line = {r["line_id"]: r for r in rows if r["line_id"] is not None}
    by_contract = {}
    for r in rows:
        by_contract.setdefault(r["contract_id"], r)
    types = set(ZHBI_ELEMENT_TYPES)
    существующие_позиции = {
        (r["contract_id"], r["element_type"], r["mark"]) for r in rows if r["line_id"] is not None
    }

    changes, rejected, rows_out = [], [], []
    seen_lines = set()
    правки_владельцев = {}   # (таблица, id) -> {поле: (значение, [номера строк])}

    for line_no, values in parsed:
        line_id = _int_id(values.get(KEY_COLUMN))
        contract_id = _int_id(values.get(CONTRACT_KEY_COLUMN))
        row = by_line.get(line_id) if line_id is not None else None

        отказ = None
        if line_id is not None and row is None:
            отказ = (f"Позиция № {line_id} не найдена — её могли удалить после выгрузки. "
                     f"Правка контракта через форму заводит его позиции заново, с новыми номерами.")
        elif row is not None and contract_id is not None and row["contract_id"] != contract_id:
            # Строку собрали из двух разных (сортировкой, копированием).
            # Угадывать, что верно, нельзя — правка ушла бы в чужой контракт.
            отказ = (f"Позиция № {line_id} принадлежит контракту № {row['contract_id']}, "
                     f"а в строке указан № {contract_id}")
        elif line_id is not None and line_id in seen_lines:
            отказ = f"Позиция № {line_id} встречается в файле дважды — какую строку применять, неизвестно"
        elif line_id is None:
            # Новая позиция: контракт обязан существовать (новые документы
            # этим файлом не заводятся, см. модуль docstring).
            if contract_id is None:
                отказ = "Пустой «№ позиции» и пустой «№ контракта» — строку не с чем сопоставить"
            elif contract_id not in by_contract:
                отказ = (f"Контракт № {contract_id} не найден. Новые контракты этим файлом не "
                         f"заводятся — воспользуйтесь «Импорт контрактации из XLS».")
        if отказ:
            rejected.append({"line": line_no, "reason": отказ})
            continue
        if line_id is not None:
            seen_lines.add(line_id)

        # Значения ячеек -> значения полей. Пустая ячейка — «не трогать».
        новые, ошибка = {}, None
        for field in EDITABLE:
            if field not in values:
                continue
            raw = values[field]
            if raw is None or (isinstance(raw, str) and not str(raw).strip()):
                continue
            try:
                новые[field] = _coerce(field, raw)
            except ValueError as exc:
                ошибка = str(exc)
                break
        if ошибка is None and новые.get("element_type") and новые["element_type"] not in types:
            ошибка = (f"Тип элемента «{новые['element_type']}» не из справочника. "
                      f"Выбирайте значение из списка на листе «{SHEET_TYPES}».")
        if ошибка:
            rejected.append({"line": line_no, "reason": ошибка})
            continue

        if row is None:
            # --- новая позиция
            нет = [FIELD_LABELS[k] for k in NEW_LINE_REQUIRED if новые.get(k) is None]
            if нет:
                rejected.append({"line": line_no,
                                 "reason": "Новая позиция без обязательных полей: " + ", ".join(нет)})
                continue
            ключ = (contract_id, новые.get("element_type"), новые.get("mark"))
            if ключ in существующие_позиции:
                rejected.append({
                    "line": line_no,
                    "reason": f"У контракта № {contract_id} уже есть позиция с этим типом и маркой — "
                              f"позиция ключуется парой «тип + марка». Поправьте существующую.",
                })
                continue
            существующие_позиции.add(ключ)
            образец = by_contract[contract_id]
            подпись = " · ".join(str(новые[k]) for k in ("element_type", "mark", "quantity")
                                 if новые.get(k) is not None)
            общее = {"row_id": _row_id(None, line_no), "line": line_no,
                     "contract_id": contract_id, "line_id": None,
                     "mark": новые.get("mark"), "element_type": новые.get("element_type")}
            changes.append({
                **общее, "field": "__new_line__", "column": KEY_COLUMN,
                "field_label": FIELD_LABELS["__new_line__"],
                "was": None, "now": подпись, "is_new": True,
                # Значения новой позиции едут вместе с правкой: применение
                # НЕ перечитывает файл (иначе применилось бы не то, что
                # показали на экране).
                "line_values": {k: новые.get(k) for k in ("element_type", "mark", "quantity")},
            })
            снимок = {key: None for key, _, _ in COLUMNS}
            снимок.update({
                CONTRACT_KEY_COLUMN: contract_id,
                "contract_name": _display_values(образец)["contract_name"],
                "object_name": образец["object_name"],
                "element_type": новые.get("element_type"),
                "mark": новые.get("mark"),
                "quantity": новые.get("quantity"),
            })
            rows_out.append({"row_id": общее["row_id"], "line_id": None,
                             "contract_id": contract_id, "values": снимок})
            # Реквизиты владельцев в строке новой позиции не правятся: у неё
            # они справочные (взяты из контракта), и принимать их как правку
            # значило бы менять документ «между делом», заводя позицию.
            continue

        # --- существующая позиция
        текущие = _display_values(row)
        общее = {"row_id": _row_id(row["line_id"], line_no), "line": line_no,
                 "line_id": row["line_id"], "contract_id": row["contract_id"],
                 "mark": row["mark"], "element_type": row["element_type"]}
        строчные = []
        for field, стало in новые.items():
            было = текущие[field]
            if стало == было:
                continue
            строчные.append({
                **общее, "field": field, "column": field, "field_label": FIELD_LABELS[field],
                "was": было, "now": стало,
                # Кому принадлежит правка — нужно и при применении (какую
                # таблицу писать), и при сверке противоречий.
                "owner_table": FIELD_TARGETS[field][0],
                "owner_id": _owner_key(row, field)[1],
            })
            ключ = _owner_key(row, field)
            правки_владельцев.setdefault(ключ, {}).setdefault(field, []).append((стало, line_no))
        if строчные:
            changes.extend(строчные)
            rows_out.append({"row_id": общее["row_id"], "line_id": row["line_id"],
                             "contract_id": row["contract_id"], "values": текущие})

    # Противоречия между строками одного владельца (см. модуль docstring) и
    # столкновения номеров документов — обе проверки снимают уже собранные
    # правки, поэтому идут после основного прохода.
    changes, конфликты = _owner_conflicts(changes, правки_владельцев)
    rejected.extend(конфликты)
    changes, дубли = _document_number_conflicts(conn, changes)
    rejected.extend(дубли)
    changes = _dedupe_owner_changes(changes)

    # На экран — только строки, где правка реально осталась.
    затронутые = {c["row_id"] for c in changes}
    rows_out = [r for r in rows_out if r["row_id"] in затронутые]
    return {
        "rows_read": len(parsed),
        "elements_touched": len({c["row_id"] for c in changes}),
        "columns": columns_spec(),
        "elements": rows_out,
        "changes": changes,
        "rejected": rejected,
    }


def _dedupe_owner_changes(changes: list) -> list:
    """Правка ДОКУМЕНТА (контракта, спецификации, договора, контрагента)
    приходит из всех его строк сразу — у спецификации их столько, сколько
    во всех её контрактах позиций. Записывается она один раз, поэтому и на
    экране показывается один раз, у первой своей строки: иначе смена одного
    номера спецификации выглядела бы как двадцать правок, из которых
    девятнадцать ничего не делают, а снятие флажка с одной ничего бы не
    меняло. Правки самих позиций не схлопываются — они и правда разные."""
    видели, out = set(), []
    for c in changes:
        table = c.get("owner_table")
        if table is None or table == "contract_lines":
            out.append(c)
            continue
        ключ = (table, c["owner_id"], c["field"])
        if ключ in видели:
            continue
        видели.add(ключ)
        out.append(c)
    return out


def _owner_conflicts(changes: list, правки_владельцев: dict) -> tuple[list, list]:
    """Одна сущность-владелец повторяется в десятках строк файла. Если
    строки предлагают ей РАЗНЫЕ значения одного поля, применить нельзя ни
    одно: «кто последний, тот и прав» здесь означало бы, что результат
    зависит от порядка строк в файле."""
    плохие = set()
    rejected = []
    for (table, owner_id), поля in правки_владельцев.items():
        for field, предложения in поля.items():
            значения = {значение for значение, _ in предложения}
            if len(значения) < 2:
                continue
            плохие.add((table, owner_id, field))
            строки = sorted({line for _, line in предложения})
            rejected.append({
                "line": строки[0],
                "reason": (f"«{FIELD_LABELS[field]}»: строки {', '.join(str(n) for n in строки)} "
                           f"задают одной и той же записи «{OWNER_LABELS[table]}» разные значения "
                           f"({', '.join(sorted(str(v) for v in значения))}). Применить нельзя ни "
                           f"одно: одна запись повторяется в файле много раз."),
            })
    if not плохие:
        return changes, rejected
    оставшиеся = [c for c in changes
                  if (c.get("owner_table"), c.get("owner_id"), c["field"]) not in плохие]
    return оставшиеся, rejected


def _document_number_conflicts(conn, changes: list) -> tuple[list, list]:
    """Номер договора уникален у контрагента, номер спецификации — в
    договоре (уникальные индексы в схеме). Проверяем ЗАРАНЕЕ, а не ловим
    IntegrityError при записи: применение идёт пачкой, и упавшая на середине
    транзакция оставила бы пользователя без внятного ответа, что именно не
    легло."""
    rejected, плохие = [], set()
    # Одна и та же правка документа приходит в десятках строк файла (у
    # спецификации столько строк, сколько во всех её контрактах позиций).
    # Проверяем и отклоняем ПО СУЩНОСТИ, иначе один занятый номер дал бы
    # двадцать одинаковых строк отказа.
    сказано = set()
    for c in changes:
        if c["field"] == "agr_number":
            занято = conn.execute(
                "SELECT a.id FROM agreements a JOIN agreements own ON own.id = ? "
                "WHERE a.counterparty_id = own.counterparty_id AND a.number = ? AND a.id <> own.id",
                (c["owner_id"], c["now"]),
            ).fetchone()
            чей = "контрагента"
        elif c["field"] == "spec_number":
            занято = conn.execute(
                "SELECT s.id FROM specifications s JOIN specifications own ON own.id = ? "
                "WHERE s.agreement_id = own.agreement_id AND s.number = ? AND s.id <> own.id",
                (c["owner_id"], c["now"]),
            ).fetchone()
            чей = "договора"
        else:
            continue
        if занято:
            плохие.add(id(c))
            ключ = (c["field"], c["owner_id"], c["now"])
            if ключ not in сказано:
                сказано.add(ключ)
                rejected.append({
                    "line": c["line"],
                    "reason": f"«{FIELD_LABELS[c['field']]}» = «{c['now']}» уже занят у этого {чей}",
                })
    if not плохие:
        return changes, rejected
    return [c for c in changes if id(c) not in плохие], rejected


# ---------------------------------------------------------------- запись

def apply_changes(conn, selections: list, user_name: str, user_id: Optional[int]) -> dict:
    """Применяет ОТМЕЧЕННЫЕ пользователем правки.

    На вход приходит то же, что вернул analyze, но отфильтрованное
    флажками — заново файл не читается: перечитывание между показом и
    применением означало бы, что применить могли не то, что показали.
    """
    разрешено = set(EDITABLE) | {"__new_line__"}
    неизвестные = {str(sel.get("field")) for sel in selections} - разрешено
    if неизвестные:
        raise ValueError("Недопустимые поля для правки: " + ", ".join(sorted(неизвестные)))

    lines_inserted, lines_updated, skipped = 0, 0, []
    # Правки владельцев собираются по СУЩНОСТИ, а не по строке файла: одна
    # спецификация приходит в десятках строк, и записывать её столько же раз
    # значило бы столько же записей в журнале об одном и том же.
    по_сущностям: dict = {}

    for sel in selections:
        field = sel["field"]
        if field == "__new_line__":
            contract_id = int(sel["contract_id"])
            v = sel.get("line_values") or {}
            if not conn.execute("SELECT id FROM contracts WHERE id = ?", (contract_id,)).fetchone():
                skipped.append({"reason": f"Контракт № {contract_id} исчез между сверкой и применением"})
                continue
            try:
                conn.execute(
                    "INSERT INTO contract_lines (contract_id, element_type, mark, quantity) "
                    "VALUES (?, ?, ?, ?)",
                    (contract_id, v.get("element_type"), v.get("mark"), int(v.get("quantity") or 0)),
                )
            except sqlite3.IntegrityError:
                # Позицию с той же парой «тип + марка» успели завести между
                # сверкой и применением — уникальный индекс её и ловит.
                skipped.append({"reason": f"У контракта № {contract_id} уже есть позиция "
                                          f"«{v.get('element_type') or '—'} / {v.get('mark') or '—'}»"})
                continue
            lines_inserted += 1
            activity.log(
                "contracting_bulk_edit", user_name=impersonation.plain_name(user_name), user_id=user_id,
                entity_type="contract", entity_id=contract_id,
                element_type=v.get("element_type"), mark=v.get("mark"),
                new_value=f"позиция добавлена: {v.get('element_type') or '—'} / "
                          f"{v.get('mark') or '—'} — {v.get('quantity')}",
                details={"source": "xlsx"},
            )
            continue

        table, column = FIELD_TARGETS[field]
        owner_id = sel.get("owner_id")
        if table == "contract_lines":
            owner_id = sel.get("line_id")
        if owner_id is None:
            skipped.append({"reason": f"Правка «{FIELD_LABELS[field]}» пришла без владельца"})
            continue
        по_сущностям.setdefault((table, int(owner_id)), []).append(sel)

    for (table, owner_id), items in по_сущностям.items():
        существует = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (owner_id,)).fetchone()
        if существует is None:
            skipped.append({"reason": f"«{OWNER_LABELS[table]}» № {owner_id} исчезла(-ёл) "
                                      f"между сверкой и применением"})
            continue
        поля, тексты = {}, []
        for sel in items:
            _, column = FIELD_TARGETS[sel["field"]]
            поля[column] = sel.get("now")
            тексты.append(f"{FIELD_LABELS[sel['field']]}: {sel.get('was') if sel.get('was') is not None else '—'}"
                          f" → {sel.get('now')}")
        присвоения = ", ".join(f"{c} = :{c}" for c in поля)
        # updated_at есть у всех правимых таблиц, кроме contract_lines —
        # у позиции своих отметок времени нет, её время живёт у контракта.
        if table != "contract_lines":
            присвоения += ", updated_at = datetime('now')"
        # Правка ПОЗИЦИИ может выдернуть основание из-под уже привязанных
        # изделий — уменьшением количества, сменой марки или типа (2026-08-14,
        # см. app/contract_guard.py). Снимаем покрытие контракта до правки,
        # чтобы сверить после: предсказывать результат по отдельным полям
        # значило бы завести вторую формулу остатка рядом с настоящей.
        покрытие_до = None
        contract_id_позиции = None
        if table == "contract_lines":
            contract_id_позиции = int(items[0]["contract_id"])
            покрытие_до = contract_guard.coverage_state(conn, contract_id_позиции)
        try:
            conn.execute(f"UPDATE {table} SET {присвоения} WHERE id = :id", {**поля, "id": owner_id})
        except sqlite3.IntegrityError as exc:
            skipped.append({"reason": f"«{OWNER_LABELS[table]}» № {owner_id}: {exc}"})
            continue
        if покрытие_до is not None:
            беда = contract_guard.regressions(
                покрытие_до, contract_guard.coverage_state(conn, contract_id_позиции))
            if беда:
                # Возврат прежних значений, а не откат всей транзакции:
                # остальные строки файла ни при чём, и обрывать их из-за
                # одной позиции незачем. Прежние значения — в `существует`,
                # прочитанном до правки.
                conn.execute(
                    f"UPDATE {table} SET {', '.join(f'{c} = :{c}' for c in поля)} WHERE id = :id",
                    {**{c: существует[c] for c in поля}, "id": owner_id},
                )
                skipped.append({"reason": f"Позиция контракта № {contract_id_позиции}: "
                                          + "; ".join(беда)
                                          + ". Сначала переназначьте изделия на другой "
                                            "контракт или снимите с них привязку."})
                continue
        if table == "contract_lines":
            lines_updated += 1
        activity.log(
            "contracting_bulk_edit", user_name=impersonation.plain_name(user_name), user_id=user_id,
            entity_type={"contract_lines": "contract", "contracts": "contract",
                         "specifications": "specification", "agreements": "agreement",
                         "counterparties": "counterparty"}[table],
            entity_id=int(items[0]["contract_id"]) if table in ("contract_lines", "contracts") else owner_id,
            # Тип и марка — только у позиции: в журнале они колонки отбора, и
            # заполнять их у правки контрагента значило бы приписать реквизиту
            # справочника изделие, которого он не касается.
            element_type=items[0].get("element_type") if table == "contract_lines" else None,
            mark=items[0].get("mark") if table == "contract_lines" else None,
            new_value="; ".join(тексты)[:500],
            details={"source": "xlsx", "таблица": table, "запись": owner_id},
        )

    conn.commit()
    return {
        "lines_inserted": lines_inserted,
        "lines_updated": lines_updated,
        "entities_updated": len(по_сущностям),
        "skipped": skipped,
    }
