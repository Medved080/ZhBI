"""
Отчёт «Статус комплектации» (живой запрос 2026-08-03, по образцу заказчика
«Статус комплектации.xlsx»).

Смысл: ПЛОСКИЙ перечень «что, где и по какому контракту» — строка на
КАЖДОЕ ИЗДЕЛИЕ, а не на группу одинаковых (живой запрос 2026-08-03:
«не группируй, этот отчёт по индивидуальным позициям элементов»). Это не
сводка по статусам (для неё есть «Статусы») и не календарь (для него
«График поставки»), а рабочий список комплектации, который заказчик до сих
пор собирал руками: по нему видно, чем закрыта конкретная стоянка
конкретного крана и на какие даты по ней есть план, факт и потребность.

Раз строка — отдельная позиция, у каждой есть **GUID** (`element_uid`) —
тот же ключ, которым изделие адресуется в массовой правке через Excel
(`app/element_bulk_edit.py`). Он и делает отчёт пригодным для сверки с
внешними перечнями: марка у изделий повторяется десятками, а GUID нет.
Колонка «Кол-во» осталась (она есть в образце заказчика) и всегда равна
единице — сумма по ней даёт число позиций.

Дерева здесь нет НАМЕРЕННО: образец — плоская таблица, которую в Excel
крутят сводными и фильтрами, а любая наша группировка этому мешала бы
(колонка со склеенным заголовком в сводную таблицу не входит). По той же
причине реквизиты контракта разложены на три колонки — «Завод»,
«Договор», «Спецификация», — а не склеены в имя контракта; формат каждого
документа общий с интерфейсом и XLS-экспортом (`build_document_label`).

Три даты — те же три независимые шкалы, что и в «Графике поставки»
(см. app/report_delivery.py):
  Плановая дата поставки    — `planned_delivery_date` (контракт/логистика);
  Фактическая дата поставки — `actual_delivery_date` (переход в «Доставлено»);
  Требуемая дата поставки   — «Начало СМР (прогноз)» из АКТУАЛИЗИРОВАННОГО
      графика: к этому числу изделие обязано быть на площадке, из него и
      раскладывается потребность. С 2026-08-30 (живой запрос) это прогноз
      последней актуализации объекта, а НЕ директивное поле
      `project_smr_start_date`: поле хранит то, что обещали изначально, а
      снабжение работает по текущему графику. Откуда берётся дата и почему
      версия ищется по объекту изделия — `FORECAST_JOIN` в
      app/schedule_versions.py. Не `smr_end_date` — та означает ЗАВЕРШЕНИЕ
      СМР (см. app/schedule_import.py) и к поставке отношения не имеет.
      Прогноза на изделие нет — ячейка ПУСТАЯ (решение пользователя
      2026-08-30): подстановка директивной даты смешала бы в одной колонке
      два разных срока, и понять, какой из них перед тобой, стало бы
      нельзя. Молчать об этом тоже нельзя — сколько позиций осталось без
      прогноза и почему, говорит строка над таблицей (`warning`).

Кран и стоянка выводятся НОМЕРОМ (`zones.number`), как в образце, а не
именем: имя формата «Стоянка 03» менялось между версиями чертежа, а по
номеру в Excel сортируют и фильтруют. Имя остаётся запасным вариантом на
случай зоны без номера.

Данные, как и у остальных отчётов, считает СЕРВЕР — экран, XLSX и PDF
берут результат одной и той же функции и разойтись не могут.
"""

from typing import Optional

from app.contracts import build_document_label
from app.db import visible_elements_clause
from app.models import STATUS_LABELS_RU, STATUS_ORDER, Status
from app.reports import natural_key, pdf_text
from app.schedule_versions import FORECAST_JOIN, FORECAST_START

TITLE = "Статус комплектации"

# Колонки — ровно те и в том порядке, что в образце заказчика. kind
# определяет и выравнивание на экране, и формат ячейки в Excel:
# «date» кладётся НАСТОЯЩЕЙ датой с числовым форматом, а не текстом.
COLUMNS = [
    {"key": "crane", "label": "Кран", "kind": "num"},
    {"key": "stance", "label": "Стоянка", "kind": "num"},
    {"key": "element_type", "label": "Тип", "kind": "text"},
    {"key": "subtype", "label": "Подтип", "kind": "text"},
    {"key": "mark", "label": "Маркировка изделий", "kind": "text"},
    {"key": "count", "label": "Кол-во", "kind": "num"},
    # Текущий статус — сразу за изделием, а не в конце строки: дальше идут
    # реквизиты контракта и даты, и «в каком состоянии изделие» между ними
    # потерялось бы. Ячейка красится цветом статуса ИЗ НАСТРОЕК системы
    # (таблица status_colors), а не своей палитрой — на схеме, в легенде и в
    # этом отчёте один и тот же статус обязан быть одного цвета.
    {"key": "status", "label": "Статус", "kind": "status"},
    {"key": "counterparty", "label": "Завод", "kind": "text"},
    {"key": "agreement", "label": "Договор", "kind": "text"},
    {"key": "specification", "label": "Спецификация", "kind": "text"},
    {"key": "plan_date", "label": "Плановая дата поставки", "kind": "date"},
    {"key": "fact_date", "label": "Фактическая дата поставки", "kind": "date"},
    {"key": "need_date", "label": "Требуемая дата поставки", "kind": "date"},
    # GUID — последней колонкой: читают отчёт по левым колонкам («кран,
    # стоянка, марка»), а GUID нужен при сверке с внешним перечнем и
    # копировании, и в начале строки он только отодвигал бы смысл вправо.
    {"key": "guid", "label": "GUID", "kind": "text"},
]

TOTAL_LABEL = "Итого"

# Порядок сортировки строк. Дата сортируется как ISO-текст — тот же приём,
# что в истории статусов (см. CLAUDE.md): формат хранения одинаковый, и
# разбирать её в объект ради сравнения незачем.
SORT_KEYS = ["crane", "stance", "element_type", "subtype", "mark",
             # Не подпись статуса, а его номер в технологическом порядке
             # (STATUS_ORDER): по алфавиту «В производстве» встало бы раньше
             # «Запланирован», и одинаковые изделия шли бы вперемешку.
             "status_order", "counterparty", "agreement", "specification",
             "plan_date", "fact_date", "need_date",
             # Последним — GUID: одинаковых по всем реквизитам позиций теперь
             # много (группировки нет), и без него их взаимный порядок
             # зависел бы от того, как база вернула строки.
             "guid"]


# Цвет на случай статуса, которого нет в настройках (в таблицу его не
# записали, а в данных он есть) — серый, как «неизвестно», а не пустая
# ячейка: пустая читалась бы как «статуса нет вовсе».
DEFAULT_STATUS_COLOR = "#9aa0a6"


def status_label(code: Optional[str]) -> Optional[str]:
    """Русская подпись статуса — та же, что на схеме и в остальных отчётах
    (STATUS_LABELS_RU). Неизвестный код показывается как есть: молча
    подменять его пустотой значило бы прятать расхождение в данных."""
    if not code:
        return None
    try:
        return STATUS_LABELS_RU[Status(code)]
    except ValueError:
        return code


def status_index(code: Optional[str]) -> Optional[int]:
    """Номер статуса в технологическом порядке (STATUS_ORDER) — ключ
    сортировки. Неизвестный статус уходит в конец."""
    for i, status in enumerate(STATUS_ORDER):
        if status.value == code:
            return i
    return len(STATUS_ORDER)


def _sort_key(value):
    """Пусто — всегда в конец, остальное «по-человечески» (Стоянка 2 раньше
    Стоянки 10, см. natural_key). Номера зон приходят числами, марки и
    реквизиты — текстом; общий ключ приводит и то, и другое к одному виду."""
    text = "" if value is None else str(value)
    return (1, []) if not text else (0, natural_key(text))


def build_completion_report(conn, source_file: Optional[str],
                            element_ids: Optional[list] = None,
                            object_id: Optional[int] = None) -> dict:
    """element_ids — необязательное сужение до конкретных элементов (тот же
    приём, что у остальных отчётов и XLS-экспорта: фильтры схемы живут на
    клиенте, сервер получает готовый список id).

    object_id — АКТИВНЫЙ объект (живой запрос 2026-08-30: «отчёт только по
    активному объекту, чужие в нём не нужны»). Отбор по чертежу его уже
    почти обеспечивал, но держался на одном признаке: стоило клиенту
    прислать пустой `source_file` (объект без чертежа оставляет в
    `state.sourceFile` файл ПРЕДЫДУЩЕГО объекта), и администратор сервиса
    получал перечень по всей базе. Условие по объекту ставится ВДОБАВОК к
    чертежу, а не вместо: два независимых сужения на один отчёт дешевле
    одного разбирательства, почему в перечне чужие изделия.

    Для ЭТОГО отчёта сужение — основной режим работы: галочка «Учитывать
    текущий фильтр схемы» у него включена по умолчанию (живой запрос),
    потому что перечень комплектации читают по конкретной захватке или
    стоянке, а не по всей стройке разом."""
    clauses, params = [visible_elements_clause("e")], []
    if object_id is not None:
        clauses.append("e.object_id = ?")
        params.append(object_id)
    if source_file:
        clauses.append("e.source_file = ?")
        params.append(source_file)
    if element_ids is not None:
        if not element_ids:
            clauses.append("1=0")
        else:
            clauses.append(f"e.id IN ({','.join('?' * len(element_ids))})")
            params.extend(element_ids)
    where = f"WHERE {' AND '.join(clauses)}"

    rows = conn.execute(
        f"""
        SELECT zc.number AS crane_number, zc.name AS crane_name,
               zs.number AS stance_number, zs.name AS stance_name,
               e.element_type AS element_type, e.subtype AS subtype, e.mark AS mark,
               cp.short_name AS cp_name,
               ag.number AS ag_number, ag.agreement_date AS ag_date,
               sp.number AS sp_number, sp.specification_date AS sp_date,
               e.planned_delivery_date AS plan_date,
               e.actual_delivery_date AS fact_date,
               {FORECAST_START} AS need_date,
               e.element_uid AS guid,
               e.current_status AS status
        FROM elements e
        {FORECAST_JOIN}
        LEFT JOIN zones zc ON zc.id = e.zone_crane_id
        LEFT JOIN zones zs ON zs.id = e.zone_stance_id
        LEFT JOIN contracts c ON c.id = e.contract_id
        LEFT JOIN specifications sp ON sp.id = c.specification_id
        LEFT JOIN agreements ag ON ag.id = sp.agreement_id
        LEFT JOIN counterparties cp ON cp.id = ag.counterparty_id
        {where}
        """,
        params,
    ).fetchall()

    # Цвета берутся ИЗ НАСТРОЕК (таблица status_colors), а не из своей
    # палитры: один и тот же статус на схеме, в легенде и в отчёте обязан
    # быть одного цвета, иначе отчёт спорит с экраном.
    status_colors = {
        row["status"]: row["color"]
        for row in conn.execute("SELECT status, color FROM status_colors").fetchall()
    }

    def zone_value(number, name):
        # Номер — то, что в образце; имя — запасной вариант для зоны без
        # номера (такие бывают у чертежей старого формата имён слоёв).
        return number if number is not None else (name or None)

    def document(number, date_str):
        # Тот же формат «НОМЕР от ДД.ММ.ГГГГ», что в карточке контракта и в
        # XLS-экспорте: одна функция на все места (app/contracts.py).
        return build_document_label(number, date_str) if number else None

    out = [{
        "crane": zone_value(r["crane_number"], r["crane_name"]),
        "stance": zone_value(r["stance_number"], r["stance_name"]),
        "element_type": r["element_type"],
        "subtype": r["subtype"] or None,
        "mark": r["mark"] or None,
        # Всегда 1: строка отчёта — одно изделие. Колонка оставлена, потому
        # что она есть в образце заказчика, и сумма по ней в итоге даёт
        # число позиций.
        "count": 1,
        # Три поля на одну колонку: подпись (её видит человек), цвет из
        # настроек системы (им красится ячейка на экране, в Excel и в PDF) и
        # номер в технологическом порядке — по нему строки сортируются.
        "status": status_label(r["status"]),
        "status_color": status_colors.get(r["status"]) or DEFAULT_STATUS_COLOR,
        "status_order": status_index(r["status"]),
        "counterparty": r["cp_name"] or None,
        "agreement": document(r["ag_number"], r["ag_date"]),
        "specification": document(r["sp_number"], r["sp_date"]),
        "plan_date": r["plan_date"] or None,
        "fact_date": r["fact_date"] or None,
        "need_date": r["need_date"] or None,
        "guid": r["guid"] or None,
    } for r in rows]
    out.sort(key=lambda row: tuple(_sort_key(row[k]) for k in SORT_KEYS))

    report = {
        "title": TITLE,
        "columns": COLUMNS,
        "rows": out,
        # Строка = изделие, поэтому число одно (2026-08-03, после отказа от
        # группировки): «строк» и «изделий» теперь означали бы одно и то же.
        "total": {"label": TOTAL_LABEL, "count": len(out)},
    }
    # Предупреждение считается ЗДЕСЬ и уходит готовым текстом — его
    # показывают и экран, и PDF, а две копии одной фразы разошлись бы на
    # первой же правке формулировки (тот же приём, что в сводной).
    report["warning"] = forecast_warning(report)
    return report


def forecast_warning(report: dict) -> str:
    """Честная пометка о пустых «Требуемых датах» — та же, что у сводной
    (coverage_warning в app/report_pivot.py) и по той же причине: колонка
    берётся из актуализированного графика, и у изделия, не попавшего в
    последнюю актуализацию, она пустая. Без этой строки пустота читалась бы
    как сбой выгрузки, а не как «прогноза на это изделие нет».

    Причины перечислены те же, что в forecast_gap (app/schedule_versions.py),
    — сюда они попадают текстом, а не числами: отчёт о комплектации не место
    для разбора графика, для него есть форма «График СМР».
    """
    всего = len(report["rows"])
    без = sum(1 for r in report["rows"] if not r["need_date"])
    if not всего or not без:
        return ""
    return (f"Требуемая дата поставки — «Начало СМР (прогноз)» из последней "
            f"актуализации графика СМР; она заполнена у {всего - без} позиций из "
            f"{всего}. У остальных {без} прогноза нет (по объекту не загружен "
            f"актуализированный график, изделие уже смонтировано либо не "
            f"привязано к крану, стоянке и этажу) — у них ячейка пустая.")


# ---------- выгрузка того же отчёта в файлы ----------
#
# Обе функции получают УЖЕ ПОСТРОЕННЫЙ отчёт, а не строят его заново —
# иначе числа на экране, в Excel и в PDF со временем разошлись бы.

def _rgb(color: Optional[str]):
    """'#22c55e' → (34, 197, 94). Цвет приходит из настроек, где его правит
    человек, — непохожее на #RRGGBB не разбирается, а игнорируется: отчёт
    из-за цвета падать не должен."""
    if not isinstance(color, str):
        return None
    text = color.strip().lstrip("#")
    if len(text) != 6:
        return None
    try:
        return tuple(int(text[i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return None


def _is_light(color: Optional[str]) -> bool:
    """Светлая ли подложка — по воспринимаемой яркости, а не по среднему:
    зелёный кажется куда светлее синего той же величины. По ней выбирается
    чёрный или белый текст поверх заливки."""
    rgb = _rgb(color)
    if not rgb:
        return True
    r, g, b = rgb
    return (0.299 * r + 0.587 * g + 0.114 * b) > 150


def _xlsx_color(color: Optional[str]) -> Optional[str]:
    rgb = _rgb(color)
    return "%02X%02X%02X" % rgb if rgb else None


def build_completion_report_xlsx(report: dict) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from app.element_fields import EXCEL_DATE_FORMAT, to_excel_date

    wb = Workbook()
    ws = wb.active
    ws.title = "Комплектация"

    thin = Side(style="thin", color="D5D8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EEF2F7")

    columns = report["columns"]
    header = [c["label"] for c in columns]
    ws.append(header)
    header_row = 1
    for i in range(1, len(header) + 1):
        cell = ws.cell(row=header_row, column=i)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Предупреждение о пустых «Требуемых датах» — ПРИМЕЧАНИЕМ к заголовку
    # колонки, а не строкой над таблицей (как в сводной). Лист перечня
    # намеренно начинается с шапки в первой строке: по нему строят сводные
    # в Excel, у него автофильтр и закреплённая шапка, и текстовая строка
    # сверху сдвинула бы всё это — вместе с уже настроенными у заказчика
    # диапазонами.
    if report.get("warning"):
        from openpyxl.comments import Comment
        need_col = [c["key"] for c in columns].index("need_date") + 1
        примечание = Comment(report["warning"], "ЖБИ-трекер")
        примечание.width, примечание.height = 420, 110
        ws.cell(row=header_row, column=need_col).comment = примечание

    # Номер строки ведём СВОИМ счётчиком: `ws.max_row` в openpyxl — не
    # счётчик, а максимум по всем ячейкам листа, и обращение к нему в цикле
    # по строкам даёт квадратичный рост (см. CLAUDE.md; на выгрузке
    # реквизитов это стоило 88 секунд вместо 1,7). Строк здесь тысячи.
    row = header_row
    for data in report["rows"]:
        # Даты — НАСТОЯЩИМИ датами с числовым форматом, а не текстом: по
        # такой колонке в Excel сортируют и строят сводные (см. CLAUDE.md).
        ws.append([to_excel_date(data[c["key"]]) if c["kind"] == "date" else data[c["key"]]
                   for c in columns])
        row += 1
        for i, c in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=i)
            cell.border = border
            if c["kind"] == "date":
                cell.number_format = EXCEL_DATE_FORMAT
            elif c["kind"] == "status":
                # Заливка — тем же цветом, что на схеме (настройки системы).
                # Цвет текста подбирается под яркость подложки: жёлтый
                # «Контрактация» с белым текстом не читается вовсе.
                заливка = _xlsx_color(data.get("status_color"))
                if заливка:
                    cell.fill = PatternFill("solid", fgColor=заливка)
                    cell.font = Font(color="000000" if _is_light(data["status_color"]) else "FFFFFF")
                cell.alignment = Alignment(horizontal="center")

    total = report["total"]
    # Итог — только в колонке «Кол-во»: складывать номера стоянок или даты
    # бессмысленно, а пустая ячейка в остальных колонках это и показывает.
    ws.append([total["label"]] + [""] * (len(columns) - 1))
    row += 1
    ws.cell(row=row, column=[c["key"] for c in columns].index("count") + 1).value = total["count"]
    for i in range(1, len(header) + 1):
        cell = ws.cell(row=row, column=i)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.border = border

    widths = {"crane": 8, "stance": 10, "element_type": 18, "subtype": 26,
              "mark": 20, "count": 9, "counterparty": 22, "agreement": 22,
              "specification": 22, "plan_date": 16, "fact_date": 18, "need_date": 16,
              "status": 16, "guid": 34}
    for i, c in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c["key"], 16)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    # Автофильтр по шапке: перечень комплектации в Excel именно фильтруют —
    # ради этого он и плоский.
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(header))}{row - 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_completion_report_pdf(report: dict, subtitle: str = "") -> bytes:
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.element_fields import ru_date_text
    from app.pdf_export import FONT_BOLD, FONT_REGULAR

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
        title=report["title"],
    )
    title_style = ParagraphStyle("t", fontName=FONT_BOLD, fontSize=14, leading=18)
    sub_style = ParagraphStyle("s", fontName=FONT_REGULAR, fontSize=9, leading=12,
                               textColor=colors.HexColor("#666666"))
    warn_style = ParagraphStyle("w", parent=sub_style, textColor=colors.HexColor("#C0392B"))

    story = [Paragraph(pdf_text(report["title"]), title_style)]
    if subtitle:
        story.append(Paragraph(pdf_text(subtitle), sub_style))
    # Тот же текст, что на экране: пустые «Требуемые даты» объясняются и в
    # файле — распечатку читают отдельно от экрана (см. forecast_warning).
    if report.get("warning"):
        story.append(Paragraph(pdf_text(report["warning"]), warn_style))
    story.append(Spacer(1, 5 * mm))

    columns = report["columns"]
    # Длинные значения (подтип, марка, реквизиты договора) переносятся
    # внутри ячейки: без Paragraph reportlab обрезал бы их по ширине
    # колонки молча. pdf_text — обязательное экранирование пользовательского
    # текста (см. app/reports.py).
    cell_style = ParagraphStyle("c", fontName=FONT_REGULAR, fontSize=6.5, leading=8)
    head_style = ParagraphStyle("h", fontName=FONT_BOLD, fontSize=6.5, leading=8, alignment=1)
    # GUID — 32 символа подряд без пробелов: обычным кеглем он либо не влез
    # бы в колонку, либо потребовал переноса ПО СИМВОЛАМ, а тот удвоил бы
    # высоту каждой строки таблицы (то есть и число страниц). Меньший кегль
    # оставляет его одной строкой.
    guid_style = ParagraphStyle("g", fontName=FONT_REGULAR, fontSize=5, leading=7)

    def cell(value, kind, key=None):
        if value is None or value == "":
            return ""
        if kind == "date":
            return ru_date_text(value)
        if kind == "num":
            return str(value)
        return Paragraph(pdf_text(value), guid_style if key == "guid" else cell_style)

    data = [[Paragraph(pdf_text(c["label"]), head_style) for c in columns]]
    # Заливка ячейки статуса задаётся стилем таблицы по КООРДИНАТАМ, а не
    # свойством ячейки: у reportlab фон — это свойство таблицы, а не текста.
    status_col = next((i for i, c in enumerate(columns) if c["kind"] == "status"), None)
    status_fills = []
    # Подряд идущие строки одного статуса красятся ОДНОЙ командой на весь
    # диапазон, а не построчно. Отчёт отсортирован в том числе по статусу,
    # поэтому таких пробегов немного; построчная заливка добавляла к сборке
    # PDF пять секунд на девяти тысячах строк (замер на обезличенной копии).
    пробег = None   # {"цвет": str, "с": int, "по": int}

    def закрыть_пробег():
        if not пробег or status_col is None:
            return
        rgb = _rgb(пробег["цвет"])
        if not rgb:
            return
        r, g, b = rgb
        начало, конец = (status_col, пробег["с"]), (status_col, пробег["по"])
        status_fills.append(("BACKGROUND", начало, конец, colors.Color(r / 255, g / 255, b / 255)))
        # Тёмная подложка — белый текст: иначе тёмно-синий «Отгружен»
        # чёрными буквами не читается (та же поправка, что в Excel).
        if not _is_light(пробег["цвет"]):
            status_fills.append(("TEXTCOLOR", начало, конец, colors.white))

    for row in report["rows"]:
        data.append([cell(row[c["key"]], c["kind"], c["key"]) for c in columns])
        цвет = row.get("status_color")
        строка = len(data) - 1
        if пробег and пробег["цвет"] == цвет:
            пробег["по"] = строка
        else:
            закрыть_пробег()
            пробег = {"цвет": цвет, "с": строка, "по": строка} if цвет else None
    закрыть_пробег()

    total = report["total"]
    data.append([total["label"] if c["key"] == "crane"
                 else (total["count"] if c["key"] == "count" else "")
                 for c in columns])

    # Доли ширины полосы набора: под текстовые колонки её нужно больше, чем
    # под номер крана и количество.
    # Доли подобраны не на глаз: каждая колонка шире самого длинного
    # НЕРАЗРЫВНОГО слова, которое в неё попадает (заголовок или значение) —
    # иначе reportlab молча вывел бы его за границу ячейки.
    shares = {"crane": 0.5, "stance": 0.75, "element_type": 1.1, "subtype": 1.35,
              "mark": 1.1, "count": 0.6, "counterparty": 1.15, "agreement": 1.1,
              "specification": 1.15, "plan_date": 0.95, "fact_date": 1.05, "need_date": 0.95,
              "status": 1.2,
              # GUID — 32 символа без единого пробела: перенести его негде,
              # поэтому колонка широкая, а кегль в ней меньше (см. guid_style).
              "guid": 2.25}
    band = 278 * mm
    total_share = sum(shares.get(c["key"], 1.0) for c in columns)
    widths = [band * shares.get(c["key"], 1.0) / total_share for c in columns]

    last = len(data) - 1
    table = Table(data, colWidths=widths, repeatRows=1)
    # Выравнивание — по КЛЮЧУ колонки, а не по её номеру: номера сдвинулись
    # уже один раз (добавился GUID), и жёсткие индексы центрировали бы
    # чужую колонку молча.
    centered = [i for i, c in enumerate(columns)
                if c["kind"] in ("num", "date", "status") or c["key"] == "guid"]
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), FONT_REGULAR),
        ("FONTSIZE", (0, 0), (-1, -1), 6.5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2F7")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5D8DC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        *[("ALIGN", (i, 1), (i, -1), "CENTER") for i in centered],
        *status_fills,
        ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("FONTNAME", (0, last), (-1, last), FONT_BOLD),
        ("BACKGROUND", (0, last), (-1, last), colors.HexColor("#EEF2F7")),
    ]))
    story.append(table)
    doc.build(story)
    return buf.getvalue()
