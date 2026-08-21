"""
Выгрузка диаграммы Ганта графика СМР в XLSX и PDF (2026-08-22).

Оба файла собираются из ОДНОГО дерева — того же `gantt_tree`, что рисует
экран (`app/schedule_versions.py`). Считать их порознь значило бы завести
третий и четвёртый источник тех же чисел, и однажды они разошлись бы: ровно
эту причину описывает шапка `app/reports.py`.

Выгружается ВСЁ дерево, а не текущая глубина показа (решение пользователя):
файл живёт своей жизнью, и «в нём нет стоянок, потому что я их свернул» —
не то свойство, которое стоит объяснять получателю письма. В XLSX глубина
восстанавливается своими средствами Excel: у строк выставлен уровень
группировки, и дерево сворачивается кнопками «1 2 3 4» слева.

Оформление у обоих файлов одно и то же и повторяет экран: две полосы на
этап (план — синяя, прогноз — оранжевая), чем выше уровень группировки, тем
полоса толще и темнее.
"""

import io
from datetime import date, timedelta
from typing import Optional

# Уровни дерева: подпись и оттенки полос. Порядок совпадает с
# schedule_versions.GANTT_LEVELS — это он и есть, только с оформлением.
LEVELS = {
    "crane": {"label": "Кран", "plan": "1F4E79", "fc": "9C3D0E", "pdf_h": 5.0},
    "stance": {"label": "Стоянка", "plan": "2E6DA8", "fc": "C2571A", "pdf_h": 4.2},
    "floor": {"label": "Этаж", "plan": "4A86C8", "fc": "E8703A", "pdf_h": 3.4},
    "item": {"label": "Тип + подтип", "plan": "7BA9DA", "fc": "F0956C", "pdf_h": 2.8},
}
LEVEL_ORDER = ["crane", "stance", "floor", "item"]

MONTHS_RU = ["янв", "фев", "мар", "апр", "май", "июн",
             "июл", "авг", "сен", "окт", "ноя", "дек"]


def flatten(nodes: list, depth: int = 0) -> list:
    """Дерево в плоский список [(узел, глубина)] — сверху вниз, в том же
    порядке, в котором строки идут на экране."""
    out = []
    for n in nodes:
        out.append((n, depth))
        out.extend(flatten(n["children"], depth + 1))
    return out


def _d(iso: Optional[str]) -> Optional[date]:
    if not iso:
        return None
    try:
        return date.fromisoformat(str(iso)[:10])
    except (ValueError, TypeError):
        return None


def _span(node: dict, ряд: str):
    """Начало и конец полосы. Заполнена одна дата из двух — полоса
    вырождается в точку, но существует: у изделия бывает известен только
    один срок, и выбросить строку значило бы потерять то, что известно."""
    a = _d(node["plan_start"] if ряд == "plan" else node["forecast_start"])
    b = _d(node["plan_end"] if ряд == "plan" else node["forecast_end"])
    return (a or b), (b or a)


def _ru(d) -> str:
    """Дата по-русски — для PDF и подписей, где нужен готовый текст.
    В ЯЧЕЙКИ XLSX уходит настоящая дата с форматом (см. build_gantt_xlsx):
    текст «01.10.2026» лишил бы колонку сортировки по дате."""
    d = d if isinstance(d, date) else _d(d)
    return d.strftime("%d.%m.%Y") if d else ""


def _subtitle(data: dict, object_name: str) -> str:
    части = [f"Объект: {object_name}"]
    if data.get("version_id"):
        момент = str(data.get("loaded_at") or "")
        части.append(f"прогноз: «{data.get('version_title') or 'без названия'}»"
                     + (f" от {_ru(момент)} {момент[11:16]}".rstrip() if момент else ""))
    else:
        части.append("версий графика нет — показаны только директивные сроки")
    части.append(f"изделий на диаграмме: {data.get('elements', 0)}")
    if data.get("installed"):
        части.append(f"смонтировано: {data['installed']}")
    return "; ".join(части)


def _footnotes(data: dict) -> list:
    """То же, что строка под диаграммой на экране: о чём файл молчит и
    почему. Без неё половинчатое дерево читается как потерянные данные."""
    строки = []
    if data.get("undated"):
        строки.append(f"Не показано {data['undated']} изделий: у них нет ни директивных "
                      "дат СМР, ни прогноза.")
    if data.get("no_forecast"):
        строки.append(f"У {data['no_forecast']} изделий в этой версии нет прогноза — "
                      "у них заполнена только плановая строка.")
    return строки


# ============================================================ XLSX

def build_gantt_xlsx(data: dict, object_name: str) -> bytes:
    """Таблица и рисованная диаграмма на одном листе.

    Слева — колонки с числами и датами, справа — сетка по неделям с заливкой
    ячеек. На каждый этап ДВЕ строки, план и прогноз: одна строка вместила бы
    только одну полосу, а закрашивать обе в одной ячейке нечем — Excel не
    умеет двух цветов в одной заливке.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from app.element_fields import EXCEL_DATE_FORMAT

    узлы = flatten(data["nodes"])
    wb = Workbook()
    ws = wb.active
    ws.title = "Диаграмма Ганта"
    # Итог СВЕРХУ группы: строка крана идёт ПЕРЕД своими стоянками, и без
    # этого Excel искал бы итог под ними и группировал не то.
    ws.sheet_properties.outlinePr.summaryBelow = False

    ЗАГОЛОВКИ = ["Этап СМР", "Уровень", "Строка", "Изделий", "Смонтировано",
                 "% монтажа", "Начало", "Завершение", "Отклонение, дн."]
    ПЕРВАЯ_НЕДЕЛЯ = len(ЗАГОЛОВКИ) + 1   # колонка J

    # --- шапка листа
    ws.cell(row=1, column=1, value="График СМР — диаграмма Ганта").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=_subtitle(data, object_name)).font = Font(size=9, color="666666")

    # --- сетка недель: от понедельника самой ранней даты до последней
    начало = _d(data.get("min_date"))
    конец = _d(data.get("max_date"))
    недели = []
    if начало and конец:
        t = начало - timedelta(days=начало.weekday())
        while t <= конец:
            недели.append(t)
            t += timedelta(days=7)

    ШАПКА1, ШАПКА2, ПЕРВАЯ = 4, 5, 6
    тонкая = Side(style="thin", color="D5D8DC")
    for i, имя in enumerate(ЗАГОЛОВКИ):
        c = ws.cell(row=ШАПКА1, column=i + 1, value=имя)
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=ШАПКА1, start_column=i + 1, end_row=ШАПКА2, end_column=i + 1)
    # Месяцы над неделями объединённой ячейкой, под ними — дата понедельника.
    начало_месяца = 0
    for i, w in enumerate(недели):
        c = ws.cell(row=ШАПКА2, column=ПЕРВАЯ_НЕДЕЛЯ + i, value=f"{w.day:02d}.{w.month:02d}")
        c.font = Font(size=7, color="666666")
        c.alignment = Alignment(textRotation=90, horizontal="center")
        последняя = i == len(недели) - 1
        if последняя or недели[i + 1].month != w.month:
            m = ws.cell(row=ШАПКА1, column=ПЕРВАЯ_НЕДЕЛЯ + начало_месяца,
                        value=f"{MONTHS_RU[w.month - 1]} {w.year % 100:02d}")
            m.font = Font(bold=True, size=9)
            m.alignment = Alignment(horizontal="center")
            if i > начало_месяца:
                ws.merge_cells(start_row=ШАПКА1, start_column=ПЕРВАЯ_НЕДЕЛЯ + начало_месяца,
                               end_row=ШАПКА1, end_column=ПЕРВАЯ_НЕДЕЛЯ + i)
            начало_месяца = i + 1

    заливки = {(ур, ряд): PatternFill("solid", fgColor=LEVELS[ур][ряд])
               for ур in LEVEL_ORDER for ряд in ("plan", "fc")}

    строка = ПЕРВАЯ
    for узел, глубина in узлы:
        ур = узел["level"]
        for ряд, подпись in (("plan", "План"), ("fc", "Прогноз")):
            a, b = _span(узел, ряд)
            # Отступ пробелами: уровень группировки Excel показывает кнопками
            # слева, но при развёрнутом дереве вложенность по ним не читается.
            ws.cell(row=строка, column=1, value="    " * глубина + узел["label"])
            ws.cell(row=строка, column=2, value=LEVELS[ур]["label"])
            ws.cell(row=строка, column=3, value=подпись)
            if ряд == "plan":
                ws.cell(row=строка, column=4, value=узел["quantity"])
                ws.cell(row=строка, column=5, value=узел["installed"] or None)
                if узел["installed"]:
                    ws.cell(row=строка, column=6, value=узел["fact_pct"] / 100).number_format = "0 %"
            # В ячейке — настоящая дата с форматом, а не текст: текст лишил
            # бы колонку сортировки и любых вычислений по срокам (то же
            # решение, что в app/export.py::_format_dates).
            for колонка, значение in ((7, a), (8, b)):
                c = ws.cell(row=строка, column=колонка, value=значение)
                if значение:
                    c.number_format = EXCEL_DATE_FORMAT
            if ряд == "fc" and узел["deviation_end"] is not None:
                ws.cell(row=строка, column=9, value=узел["deviation_end"])
            for c in range(1, ЗАГОЛОВКИ.__len__() + 1):
                ws.cell(row=строка, column=c).font = Font(
                    size=9, bold=глубина == 0, color="000000")

            if a and b:
                for i, w in enumerate(недели):
                    # Неделя закрашивается, если пересекается с работой хотя бы
                    # одним днём: работа короче недели иначе исчезала бы с
                    # диаграммы совсем, а таких тут большинство.
                    if w + timedelta(days=6) >= a and w <= b:
                        ws.cell(row=строка, column=ПЕРВАЯ_НЕДЕЛЯ + i).fill = заливки[(ур, ряд)]
            # Уровень группировки: у крана 0 (он сам итог), у его стоянок 1 и
            # так вниз. Обе строки этапа получают один уровень — иначе
            # «Прогноз» сворачивался бы отдельно от своего «Плана».
            ws.row_dimensions[строка].outline_level = глубина
            строка += 1

    for сноска in _footnotes(data):
        ws.cell(row=строка + 1, column=1, value=сноска).font = Font(size=9, color="666666")
        строка += 1

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 9
    for имя in ("D", "E", "F", "I"):
        ws.column_dimensions[имя].width = 12
    for имя in ("G", "H"):
        ws.column_dimensions[имя].width = 12
    for i in range(len(недели)):
        ws.column_dimensions[get_column_letter(ПЕРВАЯ_НЕДЕЛЯ + i)].width = 3.2
    ws.freeze_panes = ws.cell(row=ПЕРВАЯ, column=ПЕРВАЯ_НЕДЕЛЯ)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================= PDF

# Шаги подписи чисел месяца — тот же ряд, что на экране (ganttDayStep):
# первый шаг, при котором соседние подписи расходятся.
PDF_DAY_STEPS = [1, 2, 5, 10]


def build_gantt_pdf(data: dict, object_name: str) -> bytes:
    """Диаграмма во всю ширину альбомного A3, с переносом на новые страницы.

    Рисуется своим кодом по канве, как и график «Динамики» (см.
    build_dynamics_report_pdf): reportlab не умеет диаграмм Ганта, а тащить
    ради одной ради стороннюю библиотеку — новый вендоринг, требующий
    отдельного подтверждения (CLAUDE.md).
    """
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.pdfgen import canvas as rl_canvas

    from app.pdf_export import FONT_BOLD, FONT_REGULAR

    W, H = landscape(A3)
    M = 24                 # поля страницы
    NAME_W = 240           # колонка названий
    ROW_H = 11
    HEAD_H = 34            # шапка страницы (заголовок + подзаголовок)
    SCALE_H = 22           # шкала месяцев и чисел
    FOOT_H = 26

    узлы = flatten(data["nodes"])
    начало = _d(data.get("min_date"))
    конец = _d(data.get("max_date"))

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=(W, H), pageCompression=1)
    c.setTitle("График СМР — диаграмма Ганта")

    if not узлы or not начало or not конец:
        c.setFont(FONT_BOLD, 14)
        c.drawString(M, H - M - 14, "График СМР — диаграмма Ганта")
        c.setFont(FONT_REGULAR, 9)
        c.drawString(M, H - M - 34, "Рисовать нечего: ни у одного изделия объекта нет "
                                    "ни директивных дат СМР, ни прогноза.")
        c.showPage()
        c.save()
        return buf.getvalue()

    # Домен — от начала месяца самой ранней даты до начала месяца за самой
    # поздней: полоса, упирающаяся в край сетки, читается как обрезанная.
    ОТ = date(начало.year, начало.month, 1)
    ДО = date(конец.year + конец.month // 12, конец.month % 12 + 1, 1)
    дней = max(1, (ДО - ОТ).days)
    X0 = M + NAME_W
    CHART_W = W - M - X0
    px = CHART_W / дней

    def X(d: date) -> float:
        return X0 + (d - ОТ).days * px

    шаг = next((s for s in PDF_DAY_STEPS if s * px >= 13), 0)

    сноски = _footnotes(data)
    строк_на_стр = int((H - M - HEAD_H - SCALE_H - M - FOOT_H) // ROW_H)
    страниц = max(1, -(-len(узлы) // строк_на_стр))

    def шапка(номер: int):
        c.setFont(FONT_BOLD, 13)
        c.drawString(M, H - M - 11, "График СМР — диаграмма Ганта")
        c.setFont(FONT_REGULAR, 8)
        c.setFillColor(HexColor("#666666"))
        c.drawString(M, H - M - 24, _subtitle(data, object_name))
        c.drawRightString(W - M, H - M - 24, f"стр. {номер} из {страниц}")
        c.setFillColor(HexColor("#000000"))

        верх = H - M - HEAD_H
        # Шкала: месяцы, при достаточном масштабе — числа месяца, засечки.
        c.setFont(FONT_BOLD, 7)
        m = ОТ
        while m < ДО:
            x = X(m)
            c.setFillColor(HexColor("#666666"))
            c.drawString(x + 2, верх - 8, f"{MONTHS_RU[m.month - 1]} {m.year % 100:02d}")
            m = date(m.year + m.month // 12, m.month % 12 + 1, 1)
        if шаг:
            c.setFont(FONT_REGULAR, 6)
            d = ОТ
            while d < ДО:
                в_месяце = ((date(d.year + d.month // 12, d.month % 12 + 1, 1)
                             - timedelta(days=1)).day)
                if (d.day - 1) % шаг == 0 and в_месяце - d.day + 1 >= шаг:
                    x = X(d)
                    c.setFillColor(HexColor("#8A94A0"))
                    c.drawString(x + 1, верх - 18, str(d.day))
                    c.setStrokeColor(HexColor("#D5D8DC"))
                    c.line(x, верх - 20, x, верх - 22)
                d += timedelta(days=1)
        c.setFillColor(HexColor("#000000"))
        c.setFont(FONT_BOLD, 7)
        c.drawString(M, верх - 18, "Этап СМР")
        c.drawRightString(X0 - 46, верх - 18, "изд.")
        c.drawRightString(X0 - 8, верх - 18, "монтаж")

        низ = M + FOOT_H
        c.setStrokeColor(HexColor("#D5D8DC"))
        c.setLineWidth(0.4)
        c.line(M, верх - SCALE_H, W - M, верх - SCALE_H)
        c.line(X0 - 6, верх - SCALE_H, X0 - 6, низ)
        # Вертикали месяцев на всю высоту строк — по ним читаются сроки.
        m = ОТ
        while m < ДО:
            c.line(X(m), верх - SCALE_H, X(m), низ)
            m = date(m.year + m.month // 12, m.month % 12 + 1, 1)
        сегодня = date.today()
        if ОТ <= сегодня < ДО:
            c.setStrokeColor(HexColor("#C0392B"))
            c.setLineWidth(0.7)
            c.line(X(сегодня), верх - SCALE_H, X(сегодня), низ)
        return верх - SCALE_H

    def подвал():
        c.setFont(FONT_REGULAR, 7)
        c.setFillColor(HexColor("#666666"))
        y = M + FOOT_H - 10
        c.drawString(M, y, "Синяя полоса — план (директивные даты), оранжевая — прогноз; "
                           "чем выше уровень группировки, тем полоса толще и темнее. "
                           "Красная вертикаль — сегодняшний день.")
        for i, s in enumerate(сноски[:2]):
            c.drawString(M, y - 9 * (i + 1), s)
        c.setFillColor(HexColor("#000000"))

    for стр in range(страниц):
        верх = шапка(стр + 1)
        for i, (узел, глубина) in enumerate(узлы[стр * строк_на_стр:(стр + 1) * строк_на_стр]):
            y = верх - i * ROW_H
            ур = узел["level"]
            c.setFont(FONT_BOLD if глубина == 0 else FONT_REGULAR, 6.5)
            c.setFillColor(HexColor("#000000"))
            подпись = узел["label"]
            # Обрезаем ПО ШИРИНЕ, а не по числу знаков: подписи разной длины,
            # и «Плита перекрытия на отм. …» иначе то влезала бы, то нет.
            доступно = NAME_W - глубина * 8 - 88
            while подпись and c.stringWidth(подпись, FONT_REGULAR, 6.5) > доступно:
                подпись = подпись[:-1]
            c.drawString(M + глубина * 8, y - 7, подпись)
            c.setFont(FONT_REGULAR, 6.5)
            c.setFillColor(HexColor("#666666"))
            c.drawRightString(X0 - 46, y - 7, str(узел["quantity"]))
            if узел["installed"]:
                процент = узел["fact_pct"] if узел["fact_pct"] >= 1 else "<1"
                c.drawRightString(X0 - 8, y - 7, f"{процент} %")

            for ряд, отступ in (("plan", 2.0), ("fc", 6.0)):
                a, b = _span(узел, ряд)
                if not a:
                    continue
                h = LEVELS[ур]["pdf_h"]
                x1, x2 = X(a), X(max(b, a))
                c.setFillColor(HexColor("#" + LEVELS[ур][ряд]))
                # Работа длиной в день даёт нулевую ширину — рисуем 0,8 пункта:
                # это по-прежнему точка на шкале, но точка нарисованная.
                c.rect(x1, y - отступ - h, max(0.8, x2 - x1), h, stroke=0, fill=1)
        подвал()
        c.showPage()

    c.save()
    return buf.getvalue()
