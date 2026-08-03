"""
Массовая правка реквизитов элементов через Excel (2026-08-01, живой запрос).

Круг: выгрузить снимок всех объектов ОДНИМ файлом -> пользователь правит в
Excel -> загрузить обратно -> система показывает список расхождений ->
пользователь флажками выбирает, что применять.

Три решения, объясняющие форму этого модуля.

1. **Ключ — element_uid, а не id и не (source_file, dxf_handle).** uid для
   того и заведён (этап 1): он переживает перерисовку чертежа заказчиком, а
   handle обнулялся дважды из шести переходов между версиями. Файл,
   выгруженный сегодня, обязан лечь и после переимпорта чертежа.

2. **История статусов в файл НЕ входит** (прямое указание пользователя).
   Статус и фактическая дата меняются диалогом со своей датой и автором —
   правка их «между делом» в таблице лишила бы события даты.

3. **Справочники — отдельными листами, а в правимых колонках выпадающие
   списки.** Валидация Excel при этом НЕ гарантия: она не мешает вставить
   значение через copy/paste, поэтому всё то же самое проверяется на
   сервере. Список — удобство, проверка — обязанность.

Контракт правится здесь по своим правилам (см. CONTRACT_* ниже): его можно
проставить или исправить, но нельзя очистить и нельзя тронуть у элемента в
статусе «Запланирован» — снятие контракта это событие (откат статуса), а не
правка реквизита.
"""

import io
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app import activity
from app.contracts import (
    apply_status_change,
    build_contract_name,
    recompute_status_and_actual_date,
    sync_element_contract,
)
# Приватное имя по месту объявления, но задача у него общая: не дать записи
# «Запланирован» от импорта чертежа перекрыть событие, датированное задним
# числом. Вторая реализация того же сдвига неминуемо разошлась бы с первой.
from app.history_import import _shift_planned_before_first_event as shift_planned_before_first_event
from app.models import STATUS_LABELS_RU, STATUS_ORDER
from app.element_fields import (
    EXCEL_DATE_FORMAT,
    to_excel_date,
    DATE_FIELDS,
    EDITABLE_FIELDS,
    FIELD_LABELS,
    FieldError,
    check_subtype,
    coerce_field,
    contract_mismatch,
    write_fields,
)

SHEET_DATA = "Элементы"
SHEET_CONTRACTS = "Контракты"
SHEET_TYPES = "Типы элементов"
SHEET_SUBTYPES = "Подтипы"
SHEET_OBJECTS = "Объекты"

KEY_COLUMN = "element_uid"

# Колонки листа «Элементы»: (ключ, подпись, правимая ли).
# Порядок ровно тот, в котором их читает человек: сначала «что это и где»,
# потом правимое, потом справочное. Ключ — первым и подписан явно, чтобы
# его не приняли за мусор и не удалили.
COLUMNS = [
    (KEY_COLUMN, "UID (не менять)", False),
    ("object_name", "Объект", False),
    ("current_status", "Статус", False),
    ("element_type", FIELD_LABELS["element_type"], True),
    ("subtype", FIELD_LABELS["subtype"], True),
    ("mark", FIELD_LABELS["mark"], True),
    ("elevation_mm", FIELD_LABELS["elevation_mm"], True),
    ("floor", FIELD_LABELS["floor"], True),
    ("address", FIELD_LABELS["address"], True),
    ("planned_delivery_date", FIELD_LABELS["planned_delivery_date"], True),
    ("project_smr_start_date", FIELD_LABELS["project_smr_start_date"], True),
    ("project_delivery_date", FIELD_LABELS["project_delivery_date"], True),
    ("contract_name", FIELD_LABELS["contract_id"], True),
    ("counterparty", "Контрагент (справочно)", False),
    ("agreement", "Договор (справочно)", False),
    ("specification", "Спецификация (справочно)", False),
    ("actual_delivery_date", "Фактическая дата поставки", False),
    ("zone_zakhvatka", "Захватка", False),
    ("zone_crane", "Кран", False),
    ("zone_stance", "Стоянка", False),
    ("source_file", "Чертёж", False),
    ("dxf_handle", "Handle в DXF", False),
]

CONTRACT_COLUMN = "contract_name"

# Код статуса -> подпись, как в интерфейсе.
STATUS_LABELS_RU_BY_VALUE = {s.value: STATUS_LABELS_RU[s] for s in STATUS_ORDER}


def _contract_catalog(conn) -> list:
    """Контракты с расшифровкой цепочки. Наименование строит
    build_contract_name — то же, что видно в интерфейсе и в других
    выгрузках; своя склейка здесь разошлась бы с ними при первой же правке
    формата."""
    rows = conn.execute(
        """
        SELECT co.id, co.theme,
               s.number AS spec_number, s.specification_date,
               a.number AS agr_number, a.agreement_date,
               cp.short_name AS counterparty
        FROM contracts co
        JOIN specifications s ON s.id = co.specification_id
        JOIN agreements a ON a.id = s.agreement_id
        JOIN counterparties cp ON cp.id = a.counterparty_id
        ORDER BY cp.short_name, a.number, s.number
        """
    ).fetchall()
    out = []
    for r in rows:
        out.append({
            "id": r["id"],
            "name": build_contract_name(
                r["counterparty"], r["agr_number"], r["agreement_date"],
                r["spec_number"], r["specification_date"], r["theme"],
            ),
            "counterparty": r["counterparty"],
            "agreement": r["agr_number"],
            "specification": r["spec_number"],
            "theme": r["theme"],
        })
    return out


def _element_rows(conn) -> list:
    """Все актуальные элементы всех объектов — одним запросом с JOIN на
    зоны и объект. Именно одним: на 9422 строках отдельный запрос за
    названием зоны на каждую строку — это тот самый N+1, который уже стоил
    проекту 2,7 секунды на открытии окна массовой смены статуса."""
    return conn.execute(
        """
        SELECT e.*, o.name AS object_name,
               zz.name AS zone_zakhvatka, zc.name AS zone_crane, zs.name AS zone_stance
        FROM elements e
        LEFT JOIN objects o ON o.id = e.object_id
        LEFT JOIN zones zz ON zz.id = e.zone_zakhvatka_id
        LEFT JOIN zones zc ON zc.id = e.zone_crane_id
        LEFT JOIN zones zs ON zs.id = e.zone_stance_id
        WHERE e.is_current = 1
        ORDER BY o.name, e.element_type, e.mark, e.id
        """
    ).fetchall()


# Колонки, которые в Excel должны быть НАСТОЯЩИМИ датами с русским
# форматом (живой запрос 2026-08-03): три правимые даты плюс справочная
# фактическая. `actual_delivery_date` тоже дата, хотя и не правится.
_DATE_COLUMNS = {"planned_delivery_date", "project_smr_start_date",
                 "project_delivery_date", "actual_delivery_date"}


def display_values(row, contract_by_id: dict) -> dict:
    """Значения одной строки по всем колонкам COLUMNS.

    Одна функция на выгрузку в XLS и на экран подтверждения правок (живой
    запрос 2026-08-01: «выводи элементы также как они выводятся при
    сохранении во внешние XLS»). Своя сборка на экране разошлась бы с
    файлом при первой же правке состава колонок — а именно их совпадение и
    есть смысл требования.
    """
    contract = contract_by_id.get(row["contract_id"])
    derived = {
        "object_name": row["object_name"],
        "contract_name": contract["name"] if contract else None,
        "counterparty": contract["counterparty"] if contract else None,
        "agreement": contract["agreement"] if contract else None,
        "specification": contract["specification"] if contract else None,
        "zone_zakhvatka": row["zone_zakhvatka"],
        "zone_crane": row["zone_crane"],
        "zone_stance": row["zone_stance"],
    }
    # Статус — ПОДПИСЬЮ, как в интерфейсе, а не кодом (живой запрос
    # 2026-08-01). Колонка справочная и обратно не разбирается.
    derived["current_status"] = STATUS_LABELS_RU_BY_VALUE.get(
        row["current_status"], row["current_status"])
    return {key: (derived[key] if key in derived else row[key]) for key, _, _ in COLUMNS}


def build_export_workbook(conn) -> Workbook:
    """Снимок на текущий момент: лист данных + четыре листа справочников,
    с выпадающими списками в правимых колонках."""
    contracts = _contract_catalog(conn)
    contract_names = [c["name"] for c in contracts]
    types = [r["element_type"] for r in conn.execute(
        "SELECT DISTINCT element_type FROM allowed_subtypes ORDER BY element_type")]
    subtypes = conn.execute(
        "SELECT element_type, subtype FROM allowed_subtypes ORDER BY element_type, subtype").fetchall()
    objects = conn.execute("SELECT name, COALESCE(description,'') AS description FROM objects ORDER BY name").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA

    ws.append([label for _, label, _ in COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    by_id = {c["id"]: c for c in contracts}
    # Даты — настоящими датами Excel с русским форматом (живой запрос
    # 2026-08-03), а не текстом «2026-10-01». Текст «01.10.2026» тут
    # недопустим отдельно: этот же файл загружается обратно, и coerce_field
    # принимает либо дату от openpyxl, либо ISO-строку.
    столбцы_дат = [i + 1 for i, (key, _, _) in enumerate(COLUMNS) if key in _DATE_COLUMNS]
    # Номер строки считаем САМИ, а не через ws.max_row (живой репорт
    # 2026-08-03: «не формирует файл»). `max_row` в openpyxl — не счётчик, а
    # max() по всем ячейкам листа, то есть O(n) на каждое обращение; четыре
    # обращения на строку при 9422 строках давали 88 СЕКУНД на сборку файла,
    # и выгрузка выглядела как зависшая. Тот же файл со счётчиком строится
    # за 0,7 с.
    элементы = _element_rows(conn)
    for номер, row in enumerate(элементы, start=2):   # 1-я строка — заголовки
        values = display_values(row, by_id)
        ws.append([to_excel_date(values[key]) if key in _DATE_COLUMNS else values[key]
                   for key, _, _ in COLUMNS])
        for i in столбцы_дат:
            ws.cell(row=номер, column=i).number_format = EXCEL_DATE_FORMAT

    # ---- листы справочников ----
    ws_c = wb.create_sheet(SHEET_CONTRACTS)
    ws_c.append(["Наименование (это значение выбирается в колонке «Контракт»)",
                 "Контрагент", "Договор", "Спецификация", "Тема"])
    for c in contracts:
        ws_c.append([c["name"], c["counterparty"], c["agreement"], c["specification"], c["theme"]])

    ws_t = wb.create_sheet(SHEET_TYPES)
    ws_t.append(["Тип элемента"])
    for t in types:
        ws_t.append([t])

    ws_s = wb.create_sheet(SHEET_SUBTYPES)
    ws_s.append(["Тип элемента", "Подтип"])
    for r in subtypes:
        ws_s.append([r["element_type"], r["subtype"]])

    ws_o = wb.create_sheet(SHEET_OBJECTS)
    ws_o.append(["Объект", "Описание"])
    for r in objects:
        ws_o.append([r["name"], r["description"]])

    # len(элементы), а не повторный _element_rows(conn): тот же запрос на
    # 9422 строки второй раз — ради одного числа.
    _add_dropdowns(ws, len(элементы), contract_names, types, subtypes)
    _widen(ws)
    for sheet in (ws_c, ws_t, ws_s, ws_o):
        _widen(sheet)
    return wb


def _col_letter(key: str) -> str:
    return get_column_letter([k for k, _, _ in COLUMNS].index(key) + 1)


def _add_dropdowns(ws, n_rows: int, contract_names, types, subtypes) -> None:
    """Выпадающие списки ссылками на диапазоны справочных листов.

    Подтип — ПЛОСКИЙ список всех допустимых значений, а не зависящий от
    типа. Каскад в Excel требует INDIRECT и именованных диапазонов, а
    русские подписи с пробелами и точками («на отм. +15.000») в имена
    диапазонов не годятся — формула вышла бы хрупкой. Пару тип+подтип
    проверяет сервер при загрузке и говорит внятно, что не так; это
    надёжнее, чем формула, которая может молча перестать работать.
    """
    last = max(n_rows + 1, 2)
    specs = [
        ("element_type", f"'{SHEET_TYPES}'!$A$2:$A${len(types) + 1}", bool(types)),
        ("subtype", f"'{SHEET_SUBTYPES}'!$B$2:$B${len(subtypes) + 1}", bool(subtypes)),
        (CONTRACT_COLUMN, f"'{SHEET_CONTRACTS}'!$A$2:$A${len(contract_names) + 1}", bool(contract_names)),
    ]
    for key, formula, present in specs:
        if not present:
            continue
        # allow_blank: пустая ячейка законна — «значение не задано».
        # showErrorMessage=False сознательно НЕ ставим: пусть Excel
        # предупреждает о значении вне списка. Но запретить он его не может
        # (вставка через буфер валидацию обходит), поэтому серверная
        # проверка остаётся обязательной.
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Значение должно быть выбрано из списка на листе справочника."
        dv.errorTitle = "Недопустимое значение"
        ws.add_data_validation(dv)
        col = _col_letter(key)
        dv.add(f"{col}2:{col}{last}")


def _widen(ws) -> None:
    """Ширина колонок по самому длинному значению, но не больше 45 символов
    — иначе колонка «Наименование контракта» растягивает лист так, что
    остальное уезжает за экран."""
    for i, column in enumerate(ws.iter_cols(), start=1):
        width = max((len(str(c.value)) for c in column if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 45)


# ---------------------------------------------------------------- разбор

def _read_sheet(file_bytes: bytes) -> list:
    """Строки листа «Элементы» словарями по КЛЮЧАМ колонок (не по подписям):
    подпись может быть переведена или переименована, а ключ — контракт
    между выгрузкой и загрузкой.

    Сопоставление колонок — по ПОДПИСИ из COLUMNS, а не по порядку:
    пользователь имеет полное право вставить свою колонку с пометками или
    скрыть ненужные, и файл от этого не должен становиться нечитаемым.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if SHEET_DATA not in wb.sheetnames:
        raise ValueError(f"В файле нет листа «{SHEET_DATA}». "
                         f"Загружайте тот файл, который выгрузила система.")
    ws = wb[SHEET_DATA]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("Лист «Элементы» пуст")

    label_to_key = {label: key for key, label, _ in COLUMNS}
    index = {}
    for i, cell in enumerate(header):
        key = label_to_key.get(str(cell).strip() if cell is not None else "")
        if key:
            index[key] = i
    if KEY_COLUMN not in index:
        raise ValueError(f"В файле нет колонки «{dict((k, l) for k, l, _ in COLUMNS)[KEY_COLUMN]}» — "
                         f"без неё строки не с чем сопоставить.")

    out = []
    for n, raw in enumerate(rows, start=2):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        out.append((n, {key: (raw[i] if i < len(raw) else None) for key, i in index.items()}))
    return out


def analyze(conn, file_bytes: bytes) -> dict:
    """Сверяет файл с базой. Ничего не пишет.

    Возвращает список расхождений построчно — каждое расхождение это
    (элемент, поле, было, стало), чтобы пользователь мог отметить флажками
    отдельные поля отдельных элементов, а не файл целиком.
    """
    parsed = _read_sheet(file_bytes)
    catalog = _contract_catalog(conn)
    contracts = {c["name"]: c for c in catalog}
    by_id = {c["id"]: c for c in catalog}
    # Те же строки с теми же JOIN, что уходят в выгрузку: экран подтверждения
    # обязан показывать элемент ровно так, как он выглядит в файле.
    elements = {r["element_uid"]: r for r in _element_rows(conn) if r["element_uid"]}

    changes, rejected = [], []
    seen = set()
    for line_no, values in parsed:
        uid = values.get(KEY_COLUMN)
        uid = str(uid).strip() if uid is not None else ""
        if not uid:
            rejected.append({"line": line_no, "reason": "Пустой UID — строку не с чем сопоставить"})
            continue
        if uid in seen:
            rejected.append({"line": line_no, "uid": uid,
                             "reason": "UID повторяется в файле — какую из строк применять, неизвестно"})
            continue
        seen.add(uid)
        row = elements.get(uid)
        if row is None:
            rejected.append({"line": line_no, "uid": uid,
                             "reason": "Элемент с таким UID не найден среди актуальных"})
            continue

        item_changes, item_rejected = _diff_row(conn, row, values, contracts, line_no)
        changes.extend(item_changes)
        rejected.extend(item_rejected)

    # Строки затронутых элементов — для табличного экрана подтверждения.
    # Только затронутые: гнать все 9422 строки ради десятка правок значило бы
    # переслать полтора мегабайта, чтобы показать десять ячеек.
    touched = {c["element_id"] for c in changes}
    rows_out = [
        {"element_id": row["id"], "uid": row["element_uid"],
         "values": display_values(row, by_id)}
        for row in elements.values() if row["id"] in touched
    ]
    return {
        "rows_read": len(parsed),
        "elements_touched": len(touched),
        "columns": [{"key": k, "label": l, "editable": e} for k, l, e in COLUMNS],
        "elements": rows_out,
        "changes": changes,
        "rejected": rejected,
    }


def _diff_row(conn, row, values: dict, contracts: dict, line_no: int) -> tuple[list, list]:
    changes, rejected = [], []

    def describe(field, was, now, **extra):
        return {
            "element_id": row["id"], "uid": row["element_uid"], "line": line_no,
            "mark": row["mark"], "element_type": row["element_type"],
            "field": field, "field_label": FIELD_LABELS.get(field, field),
            # Колонка ЭКРАНА, к которой относится правка: у контракта поле
            # называется contract_id, а колонка — contract_name, и без этой
            # пары табличный экран не знал бы, куда её приткнуть.
            "column": CONTRACT_COLUMN if field == "contract_id" else field,
            "was": was, "now": now, **extra,
        }

    # --- обычные поля
    proposed = {}
    for field in EDITABLE_FIELDS:
        if field not in values:
            continue
        try:
            new = coerce_field(field, values[field])
        except FieldError as exc:
            rejected.append({"line": line_no, "uid": row["element_uid"], "reason": str(exc)})
            continue
        if new != row[field]:
            proposed[field] = new

    # Пара тип+подтип проверяется на ИТОГОВОМ сочетании, а не на каждом
    # поле по отдельности: сменить и тип, и подтип одной строкой — законно,
    # и промежуточное сочетание «новый тип + старый подтип» проверять
    # бессмысленно, оно нигде не окажется записанным.
    if proposed:
        final_type = proposed.get("element_type", row["element_type"])
        final_subtype = proposed.get("subtype", row["subtype"])
        err = check_subtype(conn, final_type, final_subtype)
        if err and ("element_type" in proposed or "subtype" in proposed):
            rejected.append({"line": line_no, "uid": row["element_uid"], "reason": err})
            proposed.pop("element_type", None)
            proposed.pop("subtype", None)

        warn = None
        if "mark" in proposed or "element_type" in proposed:
            warn = contract_mismatch(
                conn, row["contract_id"],
                proposed.get("element_type", row["element_type"]),
                proposed.get("mark", row["mark"]),
            )
        for field, new in proposed.items():
            changes.append(describe(field, row[field], new,
                                    warning=warn if field in ("mark", "element_type") else None))

    # --- контракт: свои правила
    if CONTRACT_COLUMN in values:
        rej, change = _diff_contract(row, values[CONTRACT_COLUMN], contracts, line_no, describe)
        if rej:
            rejected.append(rej)
        if change:
            changes.append(change)
    return changes, rejected


def _diff_contract(row, raw, contracts: dict, line_no: int, describe):
    """Правила правки контракта файлом (см. Docs/TZ.md, раздел 5):
    проставить или исправить можно, очистить — нельзя, у «Запланировано» —
    нельзя вовсе.

    Очистка запрещена не из строгости: контракт снимается ОТКАТОМ статуса,
    и это событие с датой и автором. Пустая ячейка в Excel слишком дёшева
    для такого — её ставят случайно, а восстановить потом нечем.
    """
    text = str(raw).strip() if raw is not None else ""
    current_name = None
    for name, c in contracts.items():
        if c["id"] == row["contract_id"]:
            current_name = name
            break
    if text == (current_name or ""):
        return None, None
    if not text:
        return {
            "line": line_no, "uid": row["element_uid"],
            "reason": "Очистить контракт файлом нельзя — он снимается откатом статуса "
                      "на «Запланирован» (это событие, а не правка реквизита)",
        }, None
    if text not in contracts:
        return {
            "line": line_no, "uid": row["element_uid"],
            "reason": f"Контракт «{text}» не найден. Выбирайте значение из списка "
                      f"на листе «{SHEET_CONTRACTS}».",
        }, None
    # У «Запланирован» контракт обязан быть пуст — но это не повод
    # отказывать: назначение контракта И ЕСТЬ переход в «Контрактацию»
    # (живой запрос 2026-08-01). Поэтому такая правка не отклоняется, а
    # помечается needs_contracting: при применении элементу добавится
    # запись истории «Контрактация» на дату, которую спросят у
    # пользователя. Без даты применить нельзя — статус это событие, а у
    # события должно быть когда.
    return None, describe(
        "contract_id", current_name, text,
        contract_id=contracts[text]["id"], is_contract=True,
        needs_contracting=(row["current_status"] == "planned"),
    )


# ---------------------------------------------------------------- запись

def apply_changes(conn, selections: list, user_name: str, user_id: Optional[int],
                  contracting_date: Optional[str] = None) -> dict:
    """Применяет ОТМЕЧЕННЫЕ пользователем изменения.

    На вход приходит то же, что вернул analyze, но отфильтрованное
    флажками — заново файл не читается. Так пользователь применяет ровно
    то, что видел на экране: перечитывание файла между показом и
    применением означало бы, что применить могли не то, что показали.
    """
    # Имя поля приходит из клиентского словаря и попадает в `SET {поле} = ...`
    # (app/element_fields.apply_field_changes). Инъекции тут не было —
    # несуществующая колонка роняла запрос на sqlite3.Row до commit, — но
    # записать можно было ЛЮБУЮ реально существующую колонку `elements`:
    # `object_id` (увести элемент на чужой объект), `current_status` и
    # `contract_id` в обход истории, `is_current`, `element_uid`,
    # `manual_fields`. То есть обходились инварианты «статус меняется только
    # событием» и «объект выводится из элемента, а не принимается
    # параметром» (аудит безопасности 2026-08-03).
    #
    # Проверка та же и по тому же списку, что у одиночной правки
    # (app/main.py, _ELEMENT_EDITABLE_FIELDS) — иначе «в форме нельзя, а
    # через файл прошло», ровно та асимметрия, ради устранения которой
    # проверки полей и выносили в app/element_fields.py.
    разрешено = set(EDITABLE_FIELDS) | {"contract_id"}
    неизвестные = {str(sel.get("field")) for sel in selections} - разрешено
    if неизвестные:
        raise ValueError("Недопустимые поля для правки: " + ", ".join(sorted(неизвестные)))

    by_element: dict = {}
    for sel in selections:
        by_element.setdefault(int(sel["element_id"]), []).append(sel)

    applied, skipped = 0, []
    for element_id, items in by_element.items():
        row = conn.execute("SELECT * FROM elements WHERE id = ?", (element_id,)).fetchone()
        if row is None:
            skipped.append({"element_id": element_id, "reason": "Элемент исчез между сверкой и применением"})
            continue

        values, contract_id = {}, None
        for sel in items:
            field = sel["field"]
            if field == "contract_id":
                contract_id = sel.get("contract_id")
                continue
            try:
                values[field] = coerce_field(field, sel.get("now"))
            except FieldError as exc:
                skipped.append({"element_id": element_id, "reason": str(exc)})

        if values:
            err = check_subtype(
                conn,
                values.get("element_type", row["element_type"]),
                values.get("subtype", row["subtype"]),
            )
            if err:
                skipped.append({"element_id": element_id, "reason": err})
                values.pop("element_type", None)
                values.pop("subtype", None)

        changed, manual = ({}, None)
        if values:
            changed, manual = write_fields(conn, element_id, row, values)

        if contract_id is not None:
            if row["current_status"] == "planned":
                # Назначение контракта запланированному элементу — это
                # переход в «Контрактацию». Идём ЧЕРЕЗ apply_status_change,
                # а не UPDATE: только так появятся запись истории, снимок
                # контракта, журнал действий и проверка остатка позиции
                # контракта. Своя вставка в status_history разошлась бы с
                # обычной сменой статуса при первой же правке правил.
                if not contracting_date:
                    skipped.append({"element_id": element_id,
                                    "reason": "Не указана дата статуса «Контрактация»"})
                    continue
                apply_status_change(
                    conn, element_id, "contracting", True, int(contract_id),
                    contracting_date, "Массовая правка через Excel", user_name, user_id,
                )
                # Запись «Запланирован» датирована МОМЕНТОМ ИМПОРТА чертежа,
                # а не реальным планированием. Дата контрактации задним
                # числом оказывается раньше неё, и элемент по правилу
                # «текущий статус = последняя по changed_at» откатился бы
                # обратно в «Запланирован» — правка отменила бы сама себя.
                # Тот же сдвиг, что делает импорт истории; вторая
                # реализация разошлась бы с ней (см. Docs/backlog.md).
                if shift_planned_before_first_event(conn, element_id):
                    recompute_status_and_actual_date(conn, element_id)
                    sync_element_contract(conn, element_id, "contracting", True, int(contract_id))
                changed = {**changed, "contract_id": (row["contract_id"], int(contract_id)),
                           "current_status": ("planned", "contracting")}
            else:
                conn.execute(
                    "UPDATE elements SET contract_id = ?, updated_at = datetime('now') WHERE id = ?",
                    (int(contract_id), element_id),
                )
                changed = {**changed, "contract_id": (row["contract_id"], int(contract_id))}

        if changed:
            applied += 1
            activity.log(
                "element_bulk_edit", user_name=user_name, user_id=user_id,
                entity_type="element", entity_id=element_id,
                element_type=row["element_type"], subtype=row["subtype"], mark=row["mark"],
                old_value="; ".join(f"{f}: {was}" for f, (was, _) in changed.items())[:500],
                new_value="; ".join(f"{f}: {now}" for f, (_, now) in changed.items())[:500],
                details={"manual_fields": manual, "source": "xlsx"},
            )

    conn.commit()
    return {"elements_updated": applied, "skipped": skipped}
