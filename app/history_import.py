"""
Импорт истории статусов из .xlsx (обратная операция к export.build_history_xlsx)
— переносит прогресс поставки/монтажа на другой сервер, где тот же чертёж уже
загружен (тот же source_file, те же dxf_handle). Тот же механизм используется
для аварийного восстановления статусов после потери БД (см. Docs/backlog.md,
2026-07-28) — по заранее сохранённой выгрузке "Статус на дату"
(export.build_snapshot_xlsx).

Сопоставление элементов — по (source_file, dxf_handle), как и при обычном
upsert элементов. Элементы, которых нет в целевой БД, пропускаются и
попадают в сводку как unmatched — заказчик должен сначала загрузить тот же
DXF на новом сервере.

Три режима:

* "sync" (по умолчанию в интерфейсе, «Скорректировать даты») — строка файла
  СОПОСТАВЛЯЕТСЯ с уже существующей записью истории и ОБНОВЛЯЕТ её дату,
  а не добавляет вторую. Это и есть обычный обратный круг «выгрузили →
  поправили даты в Excel → загрузили»: у "merge" изменённая дата не
  совпадала с существующей и давала ВТОРУЮ запись того же статуса
  (задвоение, живой репорт 2026-07-30).
* "merge" — история дополняется, дубли пропускаются: если у элемента уже
  есть запись с тем же статусом в ту же ДАТУ, строка из файла
  пропускается (см. Docs/backlog.md п.3). Оставлен как есть.
* "replace" — существующая история сопоставленных элементов удаляется
  перед импортом.

Идентичность записи истории в режиме "sync" — (элемент, статус) + номер
повторения: у status_history нет своего идентификатора в выгрузке, поэтому
строки файла и записи БД для одной пары (элемент, статус) сортируются по
дате и сопоставляются попарно по порядку. Лишние строки файла
добавляются, лишние записи БД остаются нетронутыми (удаление — это
"replace", а не коррекция) и попадают в сводку.

Порядок жизненного цикла: запись "Запланирован", созданная импортом
чертежа, датирована МОМЕНТОМ ИМПОРТА, поэтому перенос реальной даты
поставки/монтажа назад делает её ПОЗЖЕ фактического события — а текущим
статусом элемента считается последняя по changed_at запись, то есть
правка молча отменяла бы сама себя (в живом файле — 88 элементов из 90
отредактированных). Поэтому после импорта САМАЯ РАННЯЯ запись
"Запланирован" элемента сдвигается на минуту раньше самого раннего
события другого статуса, если оказалась позже него
(_shift_planned_before_first_event). Сдвигается только самая ранняя —
запись "Запланирован" от РУЧНОГО откката статуса лежит позже неё и
остаётся на месте.

Принимает ДВА разных формата листа с одинаковой сутью строки "элемент
получил такой-то статус в такой-то момент": "История статусов"
(build_history_xlsx, колонка "Изменено", по записи на КАЖДОЕ событие) и
"Статус на дату" (build_snapshot_xlsx, колонка "Статус изменён", ОДНА
запись на элемент — его статус на момент выгрузки). Для восстановления
после потери БД это ничем не хуже: строка снимка становится единственной
записью истории элемента, current_status после импорта — то, что было в
снимке (см. import_history ниже).
"""

import io
from collections import defaultdict
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from app import activity, contract_guard
from app.contracting_import import parse_number_and_date, resolve_agreement
from app.db import touch_elements
from app.contracts import (
    find_or_create_contract,
    adopt_contract_from_history,
    recompute_status_and_actual_date,
)
from app.counterparties import (
    find_or_create_agreement,
    find_or_create_counterparty,
    find_or_create_specification,
)
from app.models import STATUS_LABELS_RU, Status

STATUS_LABEL_TO_VALUE = {label: status.value for status, label in STATUS_LABELS_RU.items()}
PLANNED_STATUS = Status.PLANNED.value

# Формат, в котором даты живут в status_history (его же пишет
# apply_status_change и наша выгрузка) — единственный, в котором строковое
# сравнение changed_at совпадает с хронологическим: на нём держатся и
# ORDER BY changed_at (текущий статус элемента), и попарное сопоставление
# в режиме "sync" ниже.
MOMENT_FORMAT = "%Y-%m-%d %H:%M:%S"

# Что принимаем в колонке даты. Excel отдаёт отредактированную вручную
# ячейку уже объектом datetime (в живом файле 2026-07-30 такими пришли
# ровно 90 правок из 9589 строк), но пользователь может и набрать текст —
# в т.ч. в привычном ДД.ММ.ГГГГ, как во всём остальном интерфейсе.
_MOMENT_INPUT_FORMATS = (
    MOMENT_FORMAT,
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%d.%m.%Y %H:%M:%S",
    "%d.%m.%Y %H:%M",
    "%d.%m.%Y",
)

REQUIRED_HEADERS = ["DXF handle", "Статус"]
CHANGED_AT_HEADER_CANDIDATES = ["Изменено", "Статус изменён"]

# Реквизиты контракта — три колонки, которые выгрузка отдаёт с 2026-07-29
# (см. app/export.py, CONTRACT_COLUMNS). На листе "История статусов" у тех
# же колонок есть суффикс "на момент изменения" — принимаем оба варианта,
# как уже сделано для даты изменения выше.
CONTRACT_HEADER_CANDIDATES = {
    "supplier": ["Поставщик", "Поставщик на момент изменения"],
    "agreement": ["Договор (номер и дата)", "Договор (номер и дата) на момент изменения"],
    "specification": ["Спецификация (номер и дата)", "Спецификация (номер и дата) на момент изменения"],
    # Старый формат выгрузки (до 2026-07-29) — одна склеенная колонка
    # "Контракт" вида "Контрагент/Договор № от .../Спецификация № от ...".
    # Разбирать её обратно не пытаемся: тема контракта в скобках и слэши
    # внутри номеров документов (реальный пример: "2/09.04-ПОСТ") делают
    # разбор неоднозначным. Такие файлы импортируются как раньше, без
    # реквизитов — об этом сообщается в сводке.
    "legacy": ["Контракт", "Контракт на момент изменения"],
}


class HistoryImportError(Exception):
    def __init__(self, status_code: int, message: str):
        self.status_code = status_code
        self.message = message
        super().__init__(message)


def parse_moment(value: str):
    """Строка из status_history → datetime, или None, если формат чужой.

    Отдельно от normalize_changed_at: тут разбираются значения, УЖЕ лежащие
    в БД (их пишем мы сами), и «не разобралось» — повод не трогать запись,
    а не повод отвергнуть строку файла.
    """
    if not value:
        return None
    for fmt in _MOMENT_INPUT_FORMATS:
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None


def normalize_changed_at(value):
    """Ячейка даты/времени из xlsx → строка 'ГГГГ-ММ-ДД ЧЧ:ММ:СС' (None —
    ячейка пустая или формат нераспознаваем).

    Приводить к одному виду обязательно: даты в status_history сравниваются
    и сортируются КАК ТЕКСТ (см. MOMENT_FORMAT), поэтому записанные как есть
    'ГГГГ-ММ-ДД' (без времени) или '15.07.2026' встали бы в хронологию не на
    своё место — и текущий статус элемента посчитался бы по неверной
    последней записи.
    """
    if value is None:
        return None
    # datetime — подкласс date, порядок проверок важен
    if isinstance(value, datetime):
        return value.strftime(MOMENT_FORMAT)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).strftime(MOMENT_FORMAT)
    moment = parse_moment(str(value))
    return moment.strftime(MOMENT_FORMAT) if moment else None


def parse_history_xlsx(content: bytes):
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HistoryImportError(422, "Файл повреждён или не является корректным .xlsx")

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise HistoryImportError(422, "Пустой файл")

    header = [str(h).strip() if h is not None else "" for h in header]
    missing = [h for h in REQUIRED_HEADERS if h not in header]
    if missing:
        raise HistoryImportError(
            422, f"В файле нет обязательных колонок: {', '.join(missing)}"
        )
    changed_at_header = next((h for h in CHANGED_AT_HEADER_CANDIDATES if h in header), None)
    if changed_at_header is None:
        raise HistoryImportError(
            422,
            "В файле нет колонки с датой/временем статуса "
            f"({' или '.join(CHANGED_AT_HEADER_CANDIDATES)})",
        )
    col = {name: idx for idx, name in enumerate(header)}

    def get(row, name):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # Какие из колонок реквизитов реально есть в этом файле
    contract_headers = {
        key: next((h for h in candidates if h in header), None)
        for key, candidates in CONTRACT_HEADER_CANDIDATES.items()
    }
    has_contract_columns = all(contract_headers[k] for k in ("supplier", "agreement", "specification"))

    def text(row, header_name):
        if not header_name:
            return None
        value = get(row, header_name)
        if value is None:
            return None
        value = str(value).strip()
        return value or None

    parsed = []
    invalid_dates = []
    for row in rows:
        dxf_handle = get(row, "DXF handle")
        status_label = get(row, "Статус")
        raw_changed_at = get(row, changed_at_header)
        if not dxf_handle or not status_label or not raw_changed_at:
            continue
        status = STATUS_LABEL_TO_VALUE.get(str(status_label).strip())
        if status is None:
            continue
        changed_at = normalize_changed_at(raw_changed_at)
        if changed_at is None:
            # Строку с непонятной датой пропускаем НЕ молча: записать её как
            # есть значило бы сломать хронологию элемента (см.
            # normalize_changed_at), а промолчать — оставить пользователя
            # думать, что правка применилась.
            invalid_dates.append(f"{dxf_handle}: «{raw_changed_at}»")
            continue
        parsed.append(
            {
                "dxf_handle": str(dxf_handle),
                "status": status,
                "changed_at": changed_at,
                "changed_by": get(row, "Кто изменил") or None,
                "comment": get(row, "Комментарий") or None,
                "supplier_raw": text(row, contract_headers["supplier"]) if has_contract_columns else None,
                "agreement_raw": text(row, contract_headers["agreement"]) if has_contract_columns else None,
                "specification_raw": text(row, contract_headers["specification"]) if has_contract_columns else None,
            }
        )

    return {
        "rows": parsed,
        "has_contract_columns": has_contract_columns,
        "has_legacy_contract_column": bool(contract_headers["legacy"]),
        "invalid_dates": len(invalid_dates),
        "invalid_date_examples": invalid_dates[:20],
    }


def _object_for_source_file(conn, source_file: str):
    """Объект чертежа, из которого идёт импорт. None — чертёж ни к одному
    объекту не привязан (дообъектное наследие).

    Спрашивать объект у человека здесь не нужно, в отличие от импорта
    контрактации: файл истории привязан к КОНКРЕТНОМУ чертежу (сопоставление
    идёт по (source_file, dxf_handle)), а у чертежа объект уже известен."""
    row = conn.execute(
        "SELECT object_id FROM object_drawings WHERE source_file = ? LIMIT 1", (source_file,)
    ).fetchone()
    return row["object_id"] if row else None


def _resolve_contract_id(conn, row, cache, warnings, counterparty_by_lower,
                         object_id=None, object_warnings=None):
    """Реквизиты строки → contract_id, с созданием недостающих звеньев
    цепочки Контрагент→Договор→Спецификация→Контракт (согласовано с
    пользователем: создавать на лету, а не отвергать строку).

    Переиспользует ровно те же find_or_create_*, что и импорт контрактации
    (app/contracting_import.py), и тот же parse_number_and_date — формат
    "НОМЕР от ДД.ММ.ГГГГ" разбирается одной и той же функцией, которой он
    и собирался при выгрузке (build_document_label, app/contracts.py).

    Кэш по тройке (поставщик, договор, спецификация) — в выгрузке тысячи
    строк на десяток контрактов, без него на каждую строку шли бы четыре
    SELECT.
    """
    supplier = row.get("supplier_raw")
    agreement_raw = row.get("agreement_raw")
    specification_raw = row.get("specification_raw")
    if not supplier or not agreement_raw or not specification_raw:
        return None  # обычная ситуация: у элемента просто нет контракта

    key = (supplier, agreement_raw, specification_raw)
    if key in cache:
        return cache[key]

    agreement_number, agreement_date, agr_warning = parse_number_and_date(agreement_raw)
    specification_number, specification_date, spec_warning = parse_number_and_date(specification_raw)
    if agr_warning:
        warnings.append(f"Договор «{agreement_raw}»: {agr_warning}")
    if spec_warning:
        warnings.append(f"Спецификация «{specification_raw}»: {spec_warning}")

    # Регистронезависимое сопоставление кириллицы — ТОЛЬКО на стороне
    # Python: SQLite без ICU не приводит кириллицу к одному регистру ни
    # через COLLATE NOCASE, ни через lower() (живой баг на марках
    # "15КС1.1"/"15кс1.1", см. Docs/backlog.md, "Контрактация 2.0").
    # find_or_create_counterparty сравнивает short_name точным SQL-равенством,
    # поэтому "Партнер" и "партнер" в отредактированном вручную файле
    # создали бы ДВУХ контрагентов. Сначала ищем сами, без учёта регистра.
    existing_id = counterparty_by_lower.get(supplier.lower())
    if existing_id is not None:
        counterparty_id = existing_id
    else:
        counterparty_id = find_or_create_counterparty(conn, full_name=supplier, short_name=supplier)
        counterparty_by_lower[supplier.lower()] = counterparty_id

    # Договор заводится НА ОБЪЕКТ чертежа (2026-08-12). До этого здесь
    # звалась find_or_create_agreement без объекта — та же дыра, что
    # чинилась в импорте контрактации: договор с `object_id IS NULL` не
    # принадлежит ни одной стройке, и восстановленная история ссылалась на
    # контракт, невидимый всем, кроме администратора сервиса. Обе проверки —
    # одной функцией (resolve_agreement, app/contracting_import.py): вторая
    # реализация того же правила разошлась бы с первой молча.
    if object_id is None:
        # Чертёж без объекта — дообъектное наследие: заводим как раньше,
        # без объекта, и говорим об этом в сводке, а не молчим.
        agreement_id = find_or_create_agreement(conn, counterparty_id, agreement_number, agreement_date)
    else:
        agreement_id, чужой_объект, _ = resolve_agreement(
            conn, counterparty_id, agreement_number, agreement_date, object_id
        )
        if agreement_id is None:
            (object_warnings if object_warnings is not None else warnings).append(
                f"Договор «{agreement_raw}» ({supplier}) заведён на другой объект "
                f"(№{чужой_объект}) — строки по нему остались без контракта"
            )
            cache[key] = None
            return None
    specification_id = find_or_create_specification(
        conn, agreement_id, specification_number, specification_date
    )
    contract_id = find_or_create_contract(conn, specification_id)
    cache[key] = contract_id
    return contract_id


def _shift_planned_before_first_event(conn, element_id: int) -> bool:
    """Самую РАННЮЮ запись "Запланирован" элемента поставить на минуту раньше
    самого раннего события другого статуса, если она оказалась позже него.
    Возвращает True, если сдвиг понадобился.

    Зачем — см. шапку модуля: запись "Запланирован" датирована моментом
    импорта чертежа, а не реальным планированием, поэтому проставленная
    задним числом дата поставки/монтажа оказывается РАНЬШЕ неё, и элемент
    по правилу "текущий статус = последняя по changed_at запись" откатывается
    в "Запланирован" — правка отменяет сама себя.

    Сдвигается ровно одна запись — самая ранняя из "Запланирован". Более
    поздние записи того же статуса — это РУЧНОЙ откат статуса
    (apply_status_change), настоящее событие, его двигать нельзя.
    """
    history = conn.execute(
        "SELECT id, status, changed_at FROM status_history WHERE element_id = ? "
        "ORDER BY changed_at, id",
        (element_id,),
    ).fetchall()
    planned = [r for r in history if r["status"] == PLANNED_STATUS]
    others = [r for r in history if r["status"] != PLANNED_STATUS]
    if not planned or not others:
        return False
    first_planned, earliest_other = planned[0], others[0]
    if earliest_other["changed_at"] >= first_planned["changed_at"]:
        return False
    moment = parse_moment(earliest_other["changed_at"])
    if moment is None:
        return False
    new_changed_at = (moment - timedelta(minutes=1)).strftime(MOMENT_FORMAT)
    conn.execute(
        "UPDATE status_history SET changed_at = ? WHERE id = ?",
        (new_changed_at, first_planned["id"]),
    )
    return True


def import_history(conn, source_file: str, rows: list, mode: str,
                   user=None, request_id: str = None):
    if mode not in ("replace", "merge", "sync"):
        raise HistoryImportError(422, "mode должен быть 'replace', 'merge' или 'sync'")

    handles = sorted({r["dxf_handle"] for r in rows})
    element_ids = {}
    for handle in handles:
        found = conn.execute(
            "SELECT id FROM elements WHERE source_file = ? AND dxf_handle = ?",
            (source_file, handle),
        ).fetchone()
        if found:
            element_ids[handle] = found["id"]

    unmatched_handles = [h for h in handles if h not in element_ids]

    if mode == "replace":
        matched_ids = list(element_ids.values())
        if matched_ids:
            placeholders = ",".join("?" * len(matched_ids))
            conn.execute(
                f"DELETE FROM status_history WHERE element_id IN ({placeholders})", matched_ids
            )

    inserted = updated = skipped_duplicate = skipped_unmatched = 0
    unpaired_existing = 0
    touched_element_ids = set()

    # Реквизиты контракта из файла (см. _resolve_contract_id). Счётчик
    # контрактов ДО импорта — чтобы в сводке честно показать, сколько было
    # создано новых, а не сколько всего упомянуто.
    contract_cache: dict = {}
    contract_warnings: list = []
    # Объект чертежа — на него заводятся договоры из реквизитов файла
    # (2026-08-12, см. _resolve_contract_id). None бывает только у
    # дообъектного наследия; тогда поведение прежнее, но с оговоркой в сводке.
    object_id = _object_for_source_file(conn, source_file)
    object_warnings: list = []
    if object_id is None:
        object_warnings.append(
            f"Чертёж «{source_file}» не привязан к объекту — договоры из файла заведены "
            f"без объекта (дообъектное наследие)"
        )
    contracts_before = conn.execute("SELECT COUNT(*) AS n FROM contracts").fetchone()["n"]
    # Покрытие ВСЕХ контрактов до импорта (2026-08-14, см.
    # app/contract_guard.py). Именно всех, а не затронутых: какие контракты
    # файл заденет, известно только после разбора строк — он их и создаёт
    # по ходу, — а контрактов в базе десятки, снимок дёшев.
    покрытие_до = {
        r["id"]: contract_guard.coverage_state(conn, r["id"])
        for r in conn.execute("SELECT id FROM contracts")
    }
    counterparty_by_lower = {
        r["short_name"].lower(): r["id"]
        for r in conn.execute("SELECT id, short_name FROM counterparties").fetchall()
        if r["short_name"]
    }
    rows_with_contract = 0

    def row_contract_id(row):
        nonlocal rows_with_contract
        contract_id = _resolve_contract_id(
            conn, row, contract_cache, contract_warnings, counterparty_by_lower,
            object_id, object_warnings,
        )
        if contract_id is not None:
            rows_with_contract += 1
        return contract_id

    def insert_row(element_id, row):
        nonlocal inserted
        conn.execute(
            "INSERT INTO status_history (element_id, status, changed_at, changed_by, comment, contract_id) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (element_id, row["status"], row["changed_at"], row["changed_by"], row["comment"],
             row_contract_id(row)),
        )
        inserted += 1
        touched_element_ids.add(element_id)

    def update_row(record, element_id, row):
        """Существующая запись истории приводится к строке файла.

        Пустая ячейка НЕ стирает уже записанное (кто изменил, комментарий,
        реквизиты контракта): у старых выгрузок этих колонок вообще нет, а
        наша пишет пустую строку вместо NULL — на обратном круге любое
        "нет значения" неотличимо от "значение убрали", и стирать в такой
        неоднозначности хуже, чем сохранить. Дата же приходит всегда (строки
        без даты отсеиваются при разборе) и обновляется безусловно — она и
        есть предмет коррекции.
        """
        nonlocal updated
        contract_id = row_contract_id(row)
        new_values = {
            "changed_at": row["changed_at"],
            "changed_by": row["changed_by"] if row["changed_by"] is not None else record["changed_by"],
            "comment": row["comment"] if row["comment"] is not None else record["comment"],
            "contract_id": contract_id if contract_id is not None else record["contract_id"],
        }
        if all(record[field] == value for field, value in new_values.items()):
            return  # запись уже в точности такая — ни UPDATE, ни пересчёт не нужны
        conn.execute(
            "UPDATE status_history SET changed_at = ?, changed_by = ?, comment = ?, contract_id = ? "
            "WHERE id = ?",
            (new_values["changed_at"], new_values["changed_by"], new_values["comment"],
             new_values["contract_id"], record["id"]),
        )
        updated += 1
        touched_element_ids.add(element_id)

    if mode == "sync":
        # Строки файла и существующие записи — по парам (элемент, статус),
        # внутри пары в хронологическом порядке (см. шапку модуля: своего
        # идентификатора у записи истории в выгрузке нет).
        existing_by_key: dict = defaultdict(list)
        matched_ids = sorted(element_ids.values())
        # Чанки — у SQLite есть предел числа параметров в запросе
        # (SQLITE_MAX_VARIABLE_NUMBER, у старых сборок 999), а элементов тут
        # тысячи.
        for start in range(0, len(matched_ids), 500):
            chunk = matched_ids[start:start + 500]
            placeholders = ",".join("?" * len(chunk))
            for record in conn.execute(
                "SELECT id, element_id, status, changed_at, changed_by, comment, contract_id "
                f"FROM status_history WHERE element_id IN ({placeholders}) "
                "ORDER BY changed_at, id",
                chunk,
            ).fetchall():
                existing_by_key[(record["element_id"], record["status"])].append(record)

        file_by_key: dict = defaultdict(list)
        for row in rows:
            element_id = element_ids.get(row["dxf_handle"])
            if element_id is None:
                skipped_unmatched += 1
                continue
            file_by_key[(element_id, row["status"])].append(row)

        paired_counts: dict = {}
        for key, file_rows in file_by_key.items():
            element_id = key[0]
            # Полностью совпадающие строки файла — это одна и та же запись,
            # продублированная в файле; иначе вторая такая строка не нашла бы
            # себе пары и добавилась бы новой записью, то есть режим коррекции
            # сам создал бы дубль.
            unique_rows, seen_moments = [], set()
            for row in sorted(file_rows, key=lambda r: r["changed_at"]):
                if row["changed_at"] in seen_moments:
                    skipped_duplicate += 1
                    continue
                seen_moments.add(row["changed_at"])
                unique_rows.append(row)
            paired_counts[key] = len(unique_rows)

            existing = existing_by_key.get(key, [])
            for index, row in enumerate(unique_rows):
                if index < len(existing):
                    update_row(existing[index], element_id, row)
                else:
                    insert_row(element_id, row)

        # Записи БД, которым в файле не нашлось строки. Не удаляем (это
        # "replace", а не коррекция), но и не молчим — иначе пользователь не
        # узнает, что в системе осталось событие, которого в файле нет.
        for key, existing in existing_by_key.items():
            unpaired_existing += max(0, len(existing) - paired_counts.get(key, 0))
    else:
        for row in rows:
            element_id = element_ids.get(row["dxf_handle"])
            if element_id is None:
                skipped_unmatched += 1
                continue

            date_part = row["changed_at"][:10]
            dup = conn.execute(
                "SELECT id FROM status_history WHERE element_id = ? AND status = ? "
                "AND substr(changed_at, 1, 10) = ?",
                (element_id, row["status"], date_part),
            ).fetchone()
            if dup:
                skipped_duplicate += 1
                continue

            insert_row(element_id, row)

    # Порядок жизненного цикла — см. _shift_planned_before_first_event: без
    # этого проставленная задним числом дата поставки/монтажа оказывается
    # раньше записи "Запланирован" от импорта чертежа, и элемент
    # откатывается в "Запланирован". Строго ДО пересчёта ниже — он и берёт
    # последнюю по changed_at запись.
    planned_shifted = 0
    for element_id in touched_element_ids:
        if _shift_planned_before_first_event(conn, element_id):
            planned_shifted += 1

    # recompute_status_and_actual_date (app/contracts.py) — тот же
    # пересчёт, что после обычной смены статуса/отката: current_status
    # ПЛЮС actual_delivery_date (момент последнего перехода в "Доставлено";
    # пусто у "Запланирован" и без такого перехода), не только
    # current_status в одиночку — иначе актуальная дата поставки молча
    # разошлась бы с восстановленной историей (см. Docs/backlog.md,
    # 2026-07-28, восстановление статусов).
    for element_id in touched_element_ids:
        effective_status, _ = recompute_status_and_actual_date(conn, element_id)
        # Событие НА КАЖДОЕ затронутое изделие, а не только сводка по
        # импорту (2026-08-03): иначе восстановленная история не видна ни в
        # отчёте «Моя работа», ни в фильтре «Изменения», ни в истории
        # изменений самого изделия — а именно это и меняет импорт. Общий
        # request_id связывает их со сводным событием операции.
        снимок = conn.execute(
            "SELECT element_type, subtype, mark FROM elements WHERE id = ?", (element_id,)
        ).fetchone()
        activity.log(
            "history_import", user=user, entity_type="element", entity_id=element_id,
            element_type=снимок["element_type"], subtype=снимок["subtype"], mark=снимок["mark"],
            new_value=effective_status, request_id=request_id,
            details={"режим": mode, "файл": source_file},
        )
        # Контракт принимаем ИЗ импортированных записей: здесь файл и есть
        # то, что восстанавливают, поэтому направление обратное обычному
        # (см. adopt_contract_from_history, app/contracts.py). Без этого
        # привязка осталась бы только в status_history, а схема и карточка
        # элемента показывали бы прежний контракт.
        adopt_contract_from_history(conn, element_id, effective_status)

    # Восстановленная привязка подчиняется тому же правилу, что и живая:
    # изделие держится за ПОЗИЦИЮ спецификации, и если её нет или на всех
    # не хватает — файл не применяется вовсе (2026-08-14, решение
    # пользователя; до этого импорт был единственным путём, где привязка
    # заводилась вообще без оглядки на контрактацию). Проверка после
    # adopt_contract_from_history: до неё контракт стоит только в записях
    # истории, а считается он по изделиям.
    беды = []
    for r in conn.execute("SELECT id FROM contracts"):
        беды.extend(contract_guard.regressions(
            покрытие_до.get(r["id"], {}), contract_guard.coverage_state(conn, r["id"])))
    if беды:
        raise HistoryImportError(
            409,
            "Файл восстанавливает привязки, которых не покрывает контрактация: "
            + "; ".join(беды[:10])
            + (f" (и ещё {len(беды) - 10})" if len(беды) > 10 else "")
            + ". Файл не загружен. Сначала загрузите контрактацию по этим позициям "
              "(или поправьте количества в справочнике контрактов).",
        )

    # Штамп перед фиксацией — иначе чужие вкладки не увидят импорт до
    # перезагрузки страницы (см. app.db.touch_elements).
    touch_elements(conn, touched_element_ids)
    conn.commit()

    contracts_after = conn.execute("SELECT COUNT(*) AS n FROM contracts").fetchone()["n"]

    return {
        "matched_elements": len(element_ids),
        "unmatched_elements": len(unmatched_handles),
        "unmatched_handles": unmatched_handles[:20],
        "inserted": inserted,
        "updated": updated,
        "planned_shifted": planned_shifted,
        "unpaired_existing": unpaired_existing,
        "skipped_duplicate": skipped_duplicate,
        "skipped_unmatched": skipped_unmatched,
        "rows_with_contract": rows_with_contract,
        "contracts_created": contracts_after - contracts_before,
        "contract_date_warnings": contract_warnings[:20],
        # Отдельно от предупреждений по датам (2026-08-12): «договор чужого
        # объекта» и «чертёж без объекта» — не опечатка в файле, а
        # расхождение с базой, и разбирается оно по-другому.
        "contract_object_warnings": object_warnings[:20],
        "object_id": object_id,
    }
